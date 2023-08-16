from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date
import datetime
import random

wb = load_workbook("data_wyn_gui.xlsx")
# ws = wb.active
ws = wb['Sheet1']
ws1 = wb['Sheet2']
ws2 = wb['Sheet3']
ws3 = wb['Sheet4']

data_number3 = ws3
data_number2 = ws2
data_number1 = ws1
data_number = ws

#row =
#column = 22
'''check_in_row = 3
check_in_col = 22'''
#print('\nMelvin Data Search \n')

def current_date():

    today = date.today()
    #print("today's date", today)
    return today

#def last_date_check_out (check_out_row, check_out_col):

def last_date_inactivity():
    #print('last activity date')
    check_out_account1 = []
    no_of_days3 = []

    for account_no in range(2, data_number2.max_row + 1):
        check_out_last = data_number2.cell(account_no, 23).value
        check_out_account = data_number2.cell(account_no, 1).value
        if check_out_account == None:
            continue
        #print(account_no)
        #print(check_out_last)
        check_out_account1.append(check_out_account)
        d = check_out_last
        #print(d)
        list_d = d.split('/')
        #print(list_d)
        b = 0
        y = []
        while b < len(list_d):
            # if b == 0:
            c = list_d[b]
            for x in c:
                if x.isnumeric() == True:
                    y.append(x)
            b += 1
        b = 0
        # print('split y ', y)
        output = [int(x) for x in y]
        # print('split  ', output)
        a4 = 0
        while a4 < len(output):
            if a4 <= 1:
                a1 = int(str(output[0]) + str(output[1]))
            elif a4 >= 1 and a4 <= 3:
                a2 = int(str(output[2]) + str(output[3]))
            elif a4 >= 3 and a4 <= 6:
                a3 = int(str(output[4]) + str(output[5]) + str(output[6]) + str(output[7]))
            a4 += 1
        today1 = date(a3, a2, a1)
        new_date = str(a3) + ',' + str(a2) + ',' + str(a1)
        # print('new date ', new_date)
        new_date1 = str(a1) + ',' + str(a2) + ',' + str(a3)
        # today1 = date(a3, a2, a1)
        check_out = today1.strftime("%Y %m %d")
        #print(check_out)

        check_in_last =today1
        check_in_current = current_date()
        no_of_days = check_in_current - check_in_last
        no_of_days1 = str(no_of_days.days)
        # print('no of days ', no_of_days1)
        no_of_days2 = int(''.join(no_of_days1))
        no_of_days3.append(no_of_days2)
        # print('number ',nums)

    check_out_account1 = [int(x)for x in check_out_account1]
    #print('check out ', check_out_account1)
    temp = list(dict.fromkeys(check_out_account1))
    #print('no_of_days3 ', no_of_days3)
    #print('check_out_account1 ', check_out_account1)
    #print('temp ', temp)
    temp1= []
    temp2= []
    b = 0
    c= 0

    while b < len(temp):
        temp1 = []
        y = 1
        for x in range(0, len(check_out_account1)):
            if b <len(temp):
                date_recent_post = data_number2.cell(y+1, 26).value
                #print('date_recent posssssssssssssssssssssssssst ', date_recent_post)
                #print(y)
                y += 1
                if date_recent_post == 'yes':
                    if temp[c] == check_out_account1[x]:
                        temp1.append(no_of_days3[x])
                        #print(' temp1 ', temp1)
                    else:
                       pass

                elif date_recent_post == 'no':
                    if temp[c] == check_out_account1[x]:
                        temp1.append(365)
        b+=1
        c +=1
        #print('temporary one ', temp1)
        if b <= len(temp):
            #temp2.append(temp1)
            temp2.append(min(temp1))
        #print('print temp2 ', temp2)
    #temp2.append(min(temp1))
    #temp2 = int(''.join(temp2))
    #print('temp2 ',temp2)
    z = 0

    for numdays in temp2:
        previous_date = datetime.datetime.today() - datetime.timedelta(days=numdays)
        str(previous_date.strftime("%Y %m %d"))
        #print('previous date ', previous_date)
        previous_date = str(previous_date)
        date_inactive = []
        date_inactivity = []
        for x in range(1):
            for y in previous_date:
                date_inactive.append(y)
            for x in range(0, len(date_inactive)):
                if x >= 0 and x <= 3:
                    a = str(date_inactive[0]) + str(date_inactive[1]) + str(date_inactive[2]) + str(date_inactive[3])
                elif x == 4:
                    b = '/'
                elif x >= 5 and x <= 6:
                    c = str(date_inactive[5]) + str(date_inactive[6])
                elif x == 7:
                    d = '/'
                elif x >= 8 and x <= 9:
                    e = str(date_inactive[8]) + str(date_inactive[9])
            date_inactivity = str('"') + str(e) + str(b) + str(c) + str(d) + str(a) + str('"')
            #print(date_inactivity)
            if z < len(temp):
                for account_no in range(2, data_number.max_row + 1):
                    check_out_account = data_number.cell(account_no,1).value
                    #print('check out account ',check_out_account)
                    if temp[z] == check_out_account:
                        #print(temp[z])
                        if temp2[z] != 365:
                            # different sheet no 1
                            data_number.cell(account_no, 14).value = date_inactivity
                            data_number.cell(account_no, 11).value = date_inactivity
                            date_inactivity1 = data_number.cell(account_no, 14).value
                            #print(date_inactivity1)
                        else:
                            date_inactivity1 = data_number.cell(account_no, 11).value
                            #print('value ',temp[z] , data_number.cell(account_no, 11).value)
                            #print(date_inactivity1)

                    else:
                        pass
            #print(len(temp))
            z += 1

    return date_inactivity1

last_date_inactivity()

def last_date(check_in_row, check_in_col):

    d = data_number.cell(check_in_row, check_in_col).value
    #print(d)
    list_d = d.split('/')

    b = 0
    y = []

    while b < len(list_d):
        #if b == 0:
        c = list_d[b]
        for x in c:
            if x.isnumeric() == True:
                y.append(x)
        b +=1
    b = 0
    #print('split y ', y)
    output = [int(x) for x in y]
    #print('split  ', output)

    a4 = 0
    while a4 < len(output):
        if a4 <= 1:
            a1 = int(str(output[0]) + str(output[1]))
        elif a4 >= 1 and a4 <=3:
            a2 = int(str(output[2]) + str(output[3]))
        elif a4 >= 3 and a4 <=6:
            a3 = int(str(output[4]) + str(output[5]) + str(output[6]) + str(output[7]))
        a4 += 1

    today1 = date(a3, a2, a1)
    new_date = str(a3) + ',' + str(a2) + ',' + str(a1)
    #print('new date ', new_date)
    new_date1 = str(a1) + ',' + str(a2) + ',' + str(a3)
    #today1 = date(a3, a2, a1)
    check_out = today1.strftime("%Y %m %d")
    return today1


#formula in post points
def last_date3(check_in_row, check_in_col):
    for account_no in range(2, data_number2.max_row + 1):
        recent_dates = data_number2.cell(account_no, 22).value
        #print('recent dates ',account_no , recent_dates)
    #print('recent dates ',check_in_row, check_in_col)
    d = data_number2.cell(check_in_row, check_in_col).value
    #1``print('recent datessssssssssssssssssssss ', d)
    list_d = d.split('/')
    #print(list_d)
    b = 0
    y = []

    while b < len(list_d):
        #if b == 0:
        c = list_d[b]
        for x in c:
            if x.isnumeric() == True:
                y.append(x)
        b +=1
    b = 0
    #print('split y ', y)
    output = [int(x) for x in y]
    #print('split  ', output)

    a4 = 0
    while a4 < len(output):
        if a4 <= 1:
            a1 = int(str(output[0]) + str(output[1]))
        elif a4 >= 1 and a4 <=3:
            a2 = int(str(output[2]) + str(output[3]))
        elif a4 >= 3 and a4 <=6:
            a3 = int(str(output[4]) + str(output[5]) + str(output[6]) + str(output[7]))
        a4 += 1

    today2 = date(a3, a2, a1)
    new_date = str(a3) + ',' + str(a2) + ',' + str(a1)
    #print('new date ', new_date)
    new_date1 = str(a1) + ',' + str(a2) + ',' + str(a3)
    #today1 = date(a3, a2, a1)
    check_out = today2.strftime("%Y %m %d")
    return today2

def last_date4(book_date):
    for account_no in range(2, data_number2.max_row + 1):
        recent_dates = data_number2.cell(account_no, 22).value
        #print('recent dates ',account_no , recent_dates)
    d = book_date
    #1``print('recent datessssssssssssssssssssss ', d)
    list_d = d.split('/')
    #print(list_d)
    b = 0
    y = []

    while b < len(list_d):
        #if b == 0:
        c = list_d[b]
        for x in c:
            if x.isnumeric() == True:
                y.append(x)
        b +=1
    b = 0
    #print('split y ', y)
    output = [int(x) for x in y]
    #print('split  ', output)

    a4 = 0
    while a4 < len(output):
        if a4 <= 1:
            a1 = int(str(output[0]) + str(output[1]))
        elif a4 >= 1 and a4 <=3:
            a2 = int(str(output[2]) + str(output[3]))
        elif a4 >= 3 and a4 <=6:
            a3 = int(str(output[4]) + str(output[5]) + str(output[6]) + str(output[7]))
        a4 += 1

    today2 = date(a3, a2, a1)
    new_date = str(a3) + ',' + str(a2) + ',' + str(a1)
    #print('new date ', new_date)
    new_date1 = str(a1) + ',' + str(a2) + ',' + str(a3)
    #today1 = date(a3, a2, a1)
    check_out = today2.strftime("%Y %m %d")
    return today2

#last_date1 = last_date(row,column)

def days_inactive(last_date1):
    check_in_last = last_date1
    check_in_current = current_date()
    no_of_days = check_in_current - check_in_last
    no_of_days1 = str(no_of_days.days)
    #print('no of days ', no_of_days1)
    no_of_days2 = int(''.join(no_of_days1))
    #print('number ',nums)
    #print('number ', no_of_days2)
    return no_of_days2

#date_that_inactive = days_inactive()
#print("date that inactive ", date_that_inactive)
#print('date today ', current_date())
#print('Last day of activity ', last_date1)


def date_post_inactivity(numdays):
    #previous_date = datetime.datetime.today() + datetime.timedelta(days=5)
    #print('future dates: ', previous_date)
    if numdays >= 150:
        numdays = numdays - 150
        previous_date = datetime.datetime.today() - datetime.timedelta(days = numdays)
        data_number.cell(account_no, 9).value = 'closed'

        forfeit_pts = data_number.cell(account_no, 10).value
        data_number.cell(account_no, 10).value = data_number.cell(account_no, 8).value + forfeit_pts
        data_number.cell(account_no, 8).value = 0
    elif numdays >= 40:
        numdays = numdays - 40
        previous_date = datetime.datetime.today() - datetime.timedelta(days = numdays)
        if data_number.cell(account_no, 9).value == 'inactive':
            pass
        else:
            data_number.cell(account_no, 9).value = 'inactive'
            forfeit_pts = data_number.cell(account_no, 10).value
            data_number.cell(account_no, 10).value = data_number.cell(account_no, 8).value + forfeit_pts
            data_number.cell(account_no, 8).value = 0
            #print(account_list.cell(account_no, 9).value)
    else:
        numdays = 39 - numdays
        previous_date = datetime.datetime.today() + datetime.timedelta(days = numdays)
        #data_number.cell(account_no, 9).value = 'active'
        #print(account_list.cell(account_no, 9).value)

    str(previous_date.strftime("%Y %m %d"))
    #print('previous date ', previous_date)
    previous_date = str(previous_date)
    date_inactive = []
    date_inactivity = []
    for x in range(1):
        for y in previous_date:
            date_inactive.append(y)
        for x in range(0,len(date_inactive)):
            if x >= 0 and x <= 3:
                a = str(date_inactive[0]) + str(date_inactive[1]) + str(date_inactive[2]) + str(date_inactive[3])
            elif x == 4:
                b = '/'
            elif x >= 5 and x <= 6:
                c = str(date_inactive[5]) + str(date_inactive[6])
            elif x == 7:
                d = '/'
            elif x >= 8 and x <= 9:
                e = str(date_inactive[8]) + str(date_inactive[9])

        date_inactivity = str('"') + str(e) + str(b) + str(c) + str(d) + str(a) + str('"')
    return date_inactivity

#date_post_inactivity1 =  date_post_inactivity(date_that_inactive)
#print('date_post_inactivity1 ', date_post_inactivity1)

#formula in post points
def date_post_inactivity3(fivedays):
    #previous_date = datetime.datetime.today() + datetime.timedelta(days=5)
    #print('future dates: ', previous_date)
    #if numdays >= 100:
    #print(fivedays)
    if fivedays > 5:
        fivedays = fivedays - 5
        previous_date = datetime.datetime.today() - datetime.timedelta(days = fivedays)
        #print('previous date ', previous_date)
        str(previous_date.strftime("%Y %m %d"))
        #print('previous date ', previous_date)
        previous_date = str(previous_date)
        date_inactive = []
        date_inactivity = []
        for x in range(1):
            for y in previous_date:
                date_inactive.append(y)
            for x in range(0,len(date_inactive)):
                if x >= 0 and x <= 3:
                    a = str(date_inactive[0]) + str(date_inactive[1]) + str(date_inactive[2]) + str(date_inactive[3])
                elif x == 4:
                    b = '/'
                elif x >= 5 and x <= 6:
                    c = str(date_inactive[5]) + str(date_inactive[6])
                elif x == 7:
                    d = '/'
                elif x >= 8 and x <= 9:
                    e = str(date_inactive[8]) + str(date_inactive[9])

            date_inactivity = str('"') + str(e) + str(b) + str(c) + str(d) + str(a) + str('"')

    else:
        for account_no in range(2, data_number2.max_row + 1):
            clear_dates = data_number2.cell(account_no, 25).value
            #print('clear dates ',clear_dates)
            if clear_dates == 'cancel':
                pass
            else:
                date_inactivity = 'no'

    return date_inactivity


def date_post_inactivity4(recent_days):

    previous_date = datetime.datetime.today() - datetime.timedelta(days = recent_days)
    #print('previous date ', previous_date)
    str(previous_date.strftime("%Y %m %d"))
    #print('previous date ', previous_date)
    previous_date = str(previous_date)
    date_inactive = []
    date_inactivity = []
    for x in range(1):
        for y in previous_date:
            date_inactive.append(y)
        for x in range(0,len(date_inactive)):
            if x >= 0 and x <= 3:
                a = str(date_inactive[0]) + str(date_inactive[1]) + str(date_inactive[2]) + str(date_inactive[3])
            elif x == 4:
                b = '/'
            elif x >= 5 and x <= 6:
                c = str(date_inactive[5]) + str(date_inactive[6])
            elif x == 7:
                d = '/'
            elif x >= 8 and x <= 9:
                e = str(date_inactive[8]) + str(date_inactive[9])

        date_inactivity = str('"') + str(e) + str(b) + str(c) + str(d) + str(a) + str('"')

    return date_inactivity

#Clear to Post or Ready Post
for account_no in range(2, data_number2.max_row + 1):
    mbr_no = data_number2.cell(account_no, 1).value
    resv_no = data_number2.cell(account_no, 2).value
    check_out = data_number2.cell(account_no, 23).value
    clear_post = data_number2.cell(account_no, 25).value
    post_stay = data_number2.cell(account_no, 26).value

    '''clear_post_date = last_date3(account_no, 23)
    no_days = days_inactive(clear_post_date)
    posted_dates = date_post_inactivity3(no_days)
    data_number2.cell(account_no, 27).value = posted_dates'''
    #print('no of days ', no_days)
    #print('posted days ', clear_post_date , posted_dates, data_number2.cell(account_no, 27).value)

    if clear_post == 'cancel':
        data_number2.cell(account_no, 26).value = 'no'
        #data_number2.cell(account_no, 27).value = 'cancelled'

    else:
        clear_post_date = last_date3(account_no, 23)
        no_days = days_inactive(clear_post_date)
        posted_dates = date_post_inactivity3(no_days)
        data_number2.cell(account_no, 27).value = posted_dates

        if data_number2.cell(account_no, 27).value != 'no':
            data_number2.cell(account_no, 25).value = 'yes'
            data_number2.cell(account_no, 27).value = 'ready'
            if post_stay == 'yes':
                data_number2.cell(account_no, 27).value = posted_dates

            #print('clear to post ',data_number2.cell(account_no, 25).value)
        else:
            data_number2.cell(account_no, 25).value = 'no'
            data_number2.cell(account_no, 27).value = 'not ready'
            #print('clear to post ', data_number2.cell(account_no, 25).value)


# post last stay date or inactivity posting
for account_no in range(2, data_number.max_row + 1):
    active_last_value = data_number.cell(account_no, 11).value
    active_last_initial = data_number.cell(account_no, 30).value
    #print('account no ', account_no)
    if active_last_initial != None:

        active_last_initial1 =  last_date(account_no, 30)
        #print('active_last_initial ', active_last_initial1)
        active_last_initial2 = days_inactive(active_last_initial1)
        #print('active_last_initial ', active_last_initial2)

        active_recent_initial1 = last_date(account_no, 11)
        #print('\nactive_recent_initial1 ', active_recent_initial1)
        active_recent_initial2 = days_inactive(active_recent_initial1)
        #print('active_recent_initial1 ', active_recent_initial2)

        if active_last_initial2 < active_recent_initial2:
            lates_date_activity = date_post_inactivity4(active_last_initial2)
            data_number.cell(account_no, 11).value = lates_date_activity
            #print('lates_date_activity ', lates_date_activity)
            #account_list.cell(account_no, 11).value = active_last_initial2

        else:
            lates_date_activity = date_post_inactivity4(active_recent_initial2)
            data_number.cell(account_no, 11).value = lates_date_activity
            #print('lates_date_activity ', lates_date_activity)
            #account_list.cell(account_no, 11).value = active_recent_initial2

    else:
        pass

# posting date expire
for account_no in range(2, data_number.max_row + 1):
    status_current = data_number.cell(account_no, 9).value
    active_last = data_number.cell(account_no, 11).value
    active_close = data_number.cell(account_no, 12).value
    #print(status_current)
    #print('account ',account_no)
    #print('account number ', active_last)
    #print(active_close)

    if status_current == 'Inactive':
        pass
    else:
        last_date1 = last_date(account_no , 11)
        #print('account number ', last_date1)
        date_that_inactive = days_inactive(last_date1)
        #print('account number ', date_that_inactive)
        date_post_inactivity1 = date_post_inactivity(date_that_inactive)
        #print('account number ', date_post_inactivity1)
        data_number.cell(account_no, 12).value = date_post_inactivity1
        #print('active ', date_post_inactivity1)

# deactivate account dates
sheet_one = []
for account_no in range(2, data_number.max_row + 1):
    acctno = data_number.cell(account_no, 1).value
    status = data_number.cell(account_no, 9).value
    sheet_one.append(status)

value = 0
for account_no in range(2, data_number1.max_row + 1):
    acctno = data_number1.cell(account_no, 1).value
    deactivate = data_number1.cell(account_no, 6).value
    status1 = data_number1.cell(account_no, 7).value
    #print('status ', deactivate)
    if sheet_one[value] == 'Inactive':
        if status1 == 'Active':
            data_number1.cell(account_no, 7).value = 'Inactive'
            today = date.today()
            today1 = today.strftime('"%d/%m/%Y"')
            data_number1.cell(account_no, 6).value = today1
            #print('status ',status1 ,sheet_one[value])
    else:
        data_number1.cell(account_no, 7).value = 'Active'
    value +=1


'''dates1 = [x for x in date_inactivity1 if x.isnumeric()]
a = str(dates1[4]) + str(dates1[5]) + str(dates1[6]) + str(dates1[7])
b = str(dates1[2]) + str(dates1[3])
c = str(dates1[0]) + str(dates1[1])
checkinn = date(int(a), int(b), int(c))
deactivate_date = checkinn + datetime.timedelta(days=40)
deactivate_date1 = deactivate_date.strftime('"%d/%m/%Y"')
data_number1.cell(account_no, 6).value = deactivate_date1'''

wb.save('data_wyn_gui.xlsx')






