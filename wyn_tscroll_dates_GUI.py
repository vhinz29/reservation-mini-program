from tkinter import *
import tkinter
import random
import datetime
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date
from time import sleep
from tkinter import messagebox

# select site dates

wb = load_workbook("data_wyn_gui.xlsx")
# ws = wb.active
ws = wb['Sheet1']
ws1 = wb['Sheet2']
ws2 = wb['Sheet3']
ws3 = wb['Sheet4']

data_number3 = ws3
data_number2 = ws2
data_number1 = ws1
data_number = ws

def book_dates(in_year ,in_month ,in_days ,out_year ,out_month ,out_days):
    print('wyn tscoll dates')
    valuedates = 0
    x = datetime.datetime.now()
    today = date.today()
    year = x.strftime('%Y')
    month = x.strftime('%m')
    days = x.strftime('%d')
    #print('year:',year ,'month:',month ,'days:',days)
    #print('booking check in ','year:',in_year ,'month:',in_month ,'days:',in_days)
    #print('booking check in ', 'year:', out_year, 'month:', out_month, 'days:', out_days)
    try:
        a = int(in_year)
        b = int(in_month)
        c = int(in_days)
        checkinn = date(a, b, c)
        valid_entry_in = checkinn - today
        valid_entry = valid_entry_in.days
        #print('valid entry number ', valid_entry_in.days)
        if int(valid_entry) >= 0:
            # print('valid entry check in date')
            pass
        else:
            # print('invalid entry check in date')
            #messagebox.showinfo(title='Check-in ', message='Invalid Check-in Dates')
            valuedates = 1

        x = int(out_year)
        y = int(out_month)
        z = int(out_days)
        checkout = date(x, y, z)

        valid_entry_out = checkout - checkinn
        #print('valid entry number ', valid_entry_out.days)
        valid_entry1 = valid_entry_out.days
        if int(valid_entry1) <= 366:
            if int(valid_entry1) >= 0:
                #print('valid entry check out date')
                pass
            else:
                #print('invalid entry check out date')
                #messagebox.showinfo(title='Check-out ', message='Invalid Check-out Dates')
                valuedates = 1
        else:
            valuedates = 1
            #messagebox.showinfo(title='Check-out ', message='Invalid Check-out Dates')

    except ValueError:
        #print('ValueError: day is out of range for month')
        valuedates = 1

    return valuedates

#book_dates(2023 ,5 ,25 ,2024 ,5 ,25)