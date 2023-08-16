from tkinter import *
import tkinter
from tkinter import ttk
import tkinter as tk
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date
from time import sleep
import random
import datetime
import wyn_inactivity_GUI
#import text_clicked
import wyn_tscroll_dates_GUI
import os
from tkinter import messagebox
import operator
import subprocess
import math
from tkcalendar import Calendar

wb = load_workbook("data_wyn_gui.xlsx")
#window.configure(background='red')
# ws = wb.active
ws = wb['Sheet1']
ws1 = wb['Sheet2']
ws2 = wb['Sheet3']
ws3 = wb['Sheet4']
ws4 = wb['Sheet5']

data_number4 = ws4
data_number3 = ws3
data_number2 = ws2
data_number1 = ws1
data_number = ws


gender = ["Male", "Female"]
phone_home = ["Phone", "Home"]
switching = []
get_name = []
new_entry_form = []

add_forget = []
add_find = []
enroll_new = []
enroll_new1 = []

# reservation
add_count = []
resv_forget = []
resv_left = []
resv_right = []
resv_entry_site = []
resv_mbrno = []
# booking reservation
book_rooms = []
book2_rooms = []
book_sites1 = []
book_resv = []
book_resv_forget = []
book_bed_rates = []
book_get_site = []
# stay find post stay
stay_post = []
but_post = []
# stay find list box
my_list = []
# site scrool
site_scrool1_forget = []
site_scrool1_loops = []
site_scrool_entry = []
site_prop = []
site_loops = [1]
site_condition = ['False']
# temporary reservation
temp_resv = []
temp_loop = []
temp_book = []
# confirm and back dates
back_confirm_forget = []
back_dates_return = []
back_dates_value = [1]
back_bookrooms = [0, 0, 0, 0, 0, 0, 0, 0]
back_return_value = [0]
back_resv_value = [0]
# view add and back dates
view_resvno = [0]
view_yes_no = ['yes']
def_booked_get_value = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
view_post1_payment = [0, 0, 0, 0]
view_return = []
view_return1 = []
# universal variable
universal_siteno = [0]
universal_bookcombo = ['Rates']
universal_bedscombo = ['Bed Class']
#chooose reservation
choose_resv = [0]
choose_acctno = [0]
choose_siteno = [0]
choose_count = [0]
choose_entry = [0]
choose_entry[0] = ''
choose_loopsite = [0]
choose_dates = [0,0,0,0,0,0,0,0,0,0]

#plus minus button
plus_date,plus_room1,plus_adult1,plus_child1  = [1],[1],[1],[0]
#change date
change_date,change_date1=[0],[0]
change_current_date = [0]
change_current_date[0] = datetime.datetime.now()
change_loop = [0]
change_forget= []
change_no_dates =[1]
change_lastdate = [0]
change_plus_minus = [0]
change_in_out_value =[0,0,0,0]
change_shift_resv = [0,0,0]

# add member
addnew_count = [0]
addnew_count[0] = 0

temp_addnewmbr = ['' for x in range(0,19)]
# modify reservation
modify_resv_number = [0]
modify_dates = [0,0,0,0,0,0,0,0,0,0]


def exit_program(event=None):
    # new_text = str.swapcase(lab["text"])
    print('hello')
    answer = messagebox.askquestion(title='Exit',message='Exit Program')
    if answer == 'yes':
        print('program exit', answer)
        window.destroy()
    else:
        print('program continue', answer)
        pass
    #messagebox.showinfo(title='Exit', message='Exit Window')
    # exit()

def exit_program_yellow(event=None):
    exit_prog.config(fg='yellow')
    #print('yellow color')

def exit_program_white(event=None):
    exit_prog.config(fg='white')
    #print('white color')


def member_search1(membernum):
    resv_nos = 0
    color_background = ' '
    text_background = ' '

    # print('member ',membernum)
    for data_search in range(2, data_number2.max_row + 1):
        member_no = str(data_number2.cell(data_search, 1).value)
        resv_no =(data_number2.cell(data_search, 2).value)
        posted = str(data_number2.cell(data_search, 26).value)
        if member_no == membernum:
            if posted == 'yes':
                resv_nos += 1
            # print('member number ', resv_no)
    #print('resv number ',resv_nos )
    if resv_nos == 0 or resv_nos == 1:
        color_background = 'lightblue'
        text_background = 'lightblue'
    elif resv_nos < 6:
        color_background = 'yellow'
        text_background = 'yellow'
    elif resv_nos < 11:
        color_background = '#969696'
        text_background = '#969696'
    elif resv_nos < 20:
        color_background = '#CFD8DC'
        text_background = '#CFD8DC'
    else:
        color_background = 'pink'
        text_background = 'pink'

    # print('color background', color_background)

    label9 = Label(window, background=color_background, padx=534, pady=30, bd=1, relief=SOLID)
    label9.place(x=200, y=85)

    label9 = Label(window, text='Name', font=('arial', 10, 'bold'), fg='black', bg=text_background)
    label9.place(x=220, y=86)

    label11 = Label(window, text='Member #', font=('arial', 10, 'bold'), fg='black', bg=text_background)
    label11.place(x=395, y=86)

    labelA1 = Label(window, text='Points Balance', font=('arial', 10, 'bold'), fg='black', bg=text_background)
    labelA1.place(x=505, y=86)

    labelA3 = Label(window, text='Level', font=('arial', 10, 'bold'), fg='black', bg=text_background)
    labelA3.place(x=645, y=86)

    labelA5 = Label(window, text='Enroll Date', font=('arial', 10, 'bold'), fg='black', bg=text_background)
    labelA5.place(x=730, y=86)

    labelA7 = Label(window, text='Status', font=('arial', 10, 'bold'), fg='black', bg=text_background)
    labelA7.place(x=850, y=86)

    labelA9 = Label(window, text='Comments', font=('arial', 10, 'bold'), fg='black', bg=text_background)
    labelA9.place(x=950, y=86)

    labelA11 = Label(window, text='Pin #', font=('arial', 10, 'bold'), fg='black', bg=text_background)
    labelA11.place(x=1090, y=86)

    labelA13 = Label(window, text='Barclay', font=('arial', 10, 'bold'), fg='black', bg=text_background)
    labelA13.place(x=1180, y=86)


def recent_contacts_names(new_name, names):
    # names =['Melvin','Mariel','Chervin','Angel','Misha']
    # new_name = 'Chervin'

    # print('list of names ',names)
    # print('new names ',new_name)
    value = 0
    for temp in range(0, len(names)):
        if new_name == names[temp]:
            value += 1
            # print('value ',value)
        else:
            pass

        if temp + 1 == len(names):
            if value == 0:

                names.insert(0, new_name)
                names.pop()
                # print('new list of names ',names)
            else:
                names.remove(new_name)
                names.insert(0, new_name)
                # print('new list of names ', names)
    # print('namessssssssssssssssssssssssssss ',names)
    return names


def cap_name(capname):
    temp = []
    temp1 = 0
    loop = 0
    for x in range(0, len(capname)):

        if x == 0:
            temp.append(capname[x].upper())

        elif capname[x] == ' ':
            # print('melvin')
            if loop == 0:
                temp.append(capname[x])
                loop += 1
            else:
                pass
            temp1 = x + 1

        elif x == temp1:
            loop = 0
            temp.append(capname[x].upper())
        else:
            temp.append(capname[x].lower())

    temp = "".join(temp)

    return temp


# conversion remove coma in a value
def conversion_numbers2(points):
    temp_point = [str(x) for x in str(points)]
    temp1_points = ''
    temp2_points = 0
    for x in temp_point:
        if x == ',':
            pass
        else:
            temp1_points += str(x)
    temp2_points = float(temp1_points)
    if temp2_points == 0:
        temp2_points = int(temp2_points)
    #print('temporary points ',temp2_points)
    return temp2_points

# conversion put coma and two decimal places in a value
def conversion_numbers1(points):
    decimal = 0
    decimal1 = []
    decimal2 = 0
    whole = []
    temp_point = [str(x) for x in str(points)]
    temp1_points = ''
    for x in temp_point:
        if x == ',':
            pass
        else:
            temp1_points += str(x)
    points = float(temp1_points)
    points = format(points, '.2f')
    #print('print coma ', points)

    points2 = [str(x) for x in str(points)]
    for x in range(0, len(points2)):
        if points2[x] == '.':
            decimal = x
            # print(decimal)

    for x in range(0, len(points2)):
        if decimal - 1 < int(x):
            decimal1.append(points2[x])
            # print('decimal places ', decimal1)

    for x in range(0, len(points2)):
        if decimal > int(x):
            whole.append(points2[x])
            # print('whole number ', whole)
        if decimal == 0:
            whole.append(points2[x])
            # print('whole number ', whole)

    whole = "".join(whole)
    decimal2 = "".join(decimal1)
    # print('decimal places ', decimal2)

    if str(decimal2) == str(points):
        decimal2 = ''

    c = [str(x) for x in str(whole)]
    c.reverse()
    e = c
    y = 3
    z = 3
    for x in range(0, len(c)):
        if x == y:
            e.insert(z, ',')
            y += 3
            z += 4

    e.reverse()
    points_coma = "".join(e) + decimal2
    return points_coma

# conversion put two decimal places in a value
def conversion_numbers(points):
    value = ''
    try:
        temp = [str(x) for x in points]
        for x in temp:
            if x == ',':
                pass
            else:
                value += x
        #print('points ', round(float(value)))
        points1 = round(float(value))
    except:
        points1 = round(points)

    c = [str(x) for x in str(points1)]
    c.reverse()
    e = c
    y = 3
    z = 3
    for x in range(0, len(c)):
        if x == y:
            e.insert(z, ',')
            y += 3
            z += 4

    e.reverse()
    points_coma = "".join(e)
    return points_coma

def resv_stay(username):
    resv_nos = 0
    text_background = ' '
    level_mbr = ' '
    for data_search in range(2, data_number2.max_row + 1):
        member_no = str(data_number2.cell(data_search, 1).value)
        resv_no = str(data_number2.cell(data_search, 2).value)
        posted = str(data_number2.cell(data_search, 26).value)
        if member_no == username:
            if posted == 'yes':
                resv_nos += 1

    if resv_nos == 0 or resv_nos == 1:
        text_background = 'lightblue'
        level_mbr = 'Blue'
    elif resv_nos < 6:
        text_background = 'yellow'
        level_mbr = 'Gold'
    elif resv_nos < 11:
        text_background = '#969696'
        level_mbr = 'Platinum'
    elif resv_nos < 20:
        level_mbr = 'Diamond'
        text_background = '#CFD8DC'
    else:
        level_mbr = 'pink'
        text_background = 'pink'
    return level_mbr, text_background

def member_search():
    print('member search')
    global resv_count
    resv_nos = 0
    text_background = ' '
    level_mbr = ' '
    loop = 0
    enroll = []
    reactivate1 = []
    inactive1 = []
    date_expire1 = []
    reactivate_acct1 = []
    deactivate_date1 = []
    pts_bal = []
    resv_count = 1
    # temp variable add member
    for x in range(0, len(temp_addnewmbr)):
        temp_addnewmbr[x] = ''
    #print('reservation count ',resv_count)
    for x in resv_forget:
        x.place_forget()
    resv_forget.clear()

    add_find.clear()
    add_find.append('member')
    #print('add find search ',add_find)
    get_name.clear()
    for x in enroll_new1:
        x.place_forget()
    enroll_new1.clear()
    for x in enroll_new:
        x.place_forget()
    enroll_new.clear()
    for x in switching:
        x.place_forget()
    switching.clear()
    username = label53.get()
    get_name.append(username)

    #print('username ', username, get_name)
    member_search1(username)
    valuex = 12
    valuey = 495
    recent_names = []
    recent_loop = 0
    # print('account name', username)
    #print('maximum row',   data_number.cell(data_number.max_row, 1).value , addnew_count[0])
    #new mbr display on input
    if addnew_count[0] == 1:
        username = data_number.cell(data_number.max_row, 1).value
        addnew_count[0] = 0
    else:
        pass

    for data_search in range(2, data_number.max_row + 1):
        member_no = str(data_number.cell(data_search, 1).value)
        first_name = data_number.cell(data_search, 4).value
        recent_contacts = data_number.cell(data_search, 20).value
        if username == str(member_no):
            new_first_name = data_number.cell(data_search, 4).value
            # print('first name ', new_first_name)
            # labelstay = Label(window1, background='lightgray', padx=300, pady=225, border=1, relief=SOLID)
            # labelstay.place(x=0, y=0)

            recent_loop = 1
        if data_search < 7:
            recent_names.append(recent_contacts)

    if recent_loop == 1:
        recent_contacts_names1 = recent_contacts_names(new_first_name, recent_names)
        # print('recent_contacts_names1 ',recent_contacts_names1)
        for data_search in range(2, data_number.max_row + 1):
            if data_search < 7:
                data_number.cell(data_search, 20).value = recent_contacts_names1[data_search - 2]
                recent_contacts_names2 = data_number.cell(data_search, 20).value
                recent_contacts = Label(window, text=recent_contacts_names2 + '                    ',
                                        font=('arial', 11, 'bold'),
                                        fg='darkgray')
                recent_contacts.place(x=valuex, y=valuey)
                valuey += 22

                # print('new value ',data_number.cell(data_search, 20).value)
    member_no_list = []
    for data_search in range(2, data_number.max_row + 1):
        member_no = str(data_number.cell(data_search, 1).value)
        member_no_list.append(member_no)

    #print('lenght of member number ', len(member_no_list))

    for data_search in range(2, data_number.max_row + 1):
        member_no = str(data_number.cell(data_search, 1).value)
        phone_no = data_number.cell(data_search, 2).value
        home_no = data_number.cell(data_search, 17).value
        last_name = data_number.cell(data_search, 3).value
        first_name = data_number.cell(data_search, 4).value
        middle_name = data_number.cell(data_search, 5).value
        enroll_date = data_number.cell(data_search, 13).value
        points_bal = data_number.cell(data_search, 8).value
        active_acct = data_number.cell(data_search, 9).value
        forfeit_pts = data_number.cell(data_search, 10).value
        gender_mf = data_number.cell(data_search, 16).value
        reactivate = data_number.cell(data_search, 11).value
        inactive = data_number.cell(data_search, 14).value
        date_expire = data_number.cell(data_search, 12).value
        ms_mr = data_number.cell(data_search, 18).value
        language = data_number.cell(data_search, 19).value
        phone_home = data_number.cell(data_search, 15).value
        recent_contacts = data_number.cell(data_search, 20).value
        birth_month = data_number.cell(data_search, 21).value
        birth_no = data_number.cell(data_search, 22).value
        email_add = data_number.cell(data_search, 29).value
        reactivate_acct = data_number.cell(data_search, 30).value

        no_home = data_number.cell(data_search, 23).value
        no_street = data_number.cell(data_search, 7).value
        no_brgy = data_number.cell(data_search, 24).value
        no_city = data_number.cell(data_search, 25).value
        no_prov = data_number.cell(data_search, 26).value
        no_country = data_number.cell(data_search, 27).value
        no_zipcode = data_number.cell(data_search, 28).value

        #different sheet no1
        deactivate_date = data_number1.cell(data_search, 6).value

        no_value = '                    '
        #print('member number: ',data_search, member_no)
        if str(username) == str(member_no):
            loop += 1
            # print('loop ',loop)
            # entry.delete(0, END)
            # label20.delete(0, None)
            # label22.insert(0, None)
            # abel24.insert(0, None)

            for x in enroll_date:
                if x != '"':
                    enroll.append(x)
            enroll = "".join(enroll)

            if deactivate_date != None:
                for x in deactivate_date:
                    if x != '"':
                        deactivate_date1.append(x)
                deactivate_date1 = "".join(deactivate_date1)

            else:
                pass

            if reactivate_acct != None:
                for x in reactivate_acct:
                    if x != '"':
                        reactivate_acct1.append(x)
                reactivate_acct1 = "".join(reactivate_acct1)

            else:
                pass

            if reactivate != None:
                for x in reactivate:
                    if x != '"':
                        reactivate1.append(x)
                reactivate1 = "".join(reactivate1)
            else:
                reactivate1 = None

            if inactive != None:
                for x in inactive:
                    if x != '"':
                        inactive1.append(x)
                inactive1 = "".join(inactive1)

            else:
                label46 = Label(window, text=no_value, font=('arial', 11), fg='black', bg='lightgray')
                label46.place(x=1120, y=330)

            if date_expire != None:
                for x in date_expire:
                    if x != '"':
                        date_expire1.append(x)
                date_expire1 = "".join(date_expire1)
            else:
                date_expire = None

            # add coma to points balance
            points_coma = conversion_numbers(points_bal)
            # print('points balance ',points_coma)

            # add coma to forfeit points
            forfeit_coma = conversion_numbers(forfeit_pts)
            # print('points balance ', forfeit_coma)

            # print(enroll)
            # print('mmmmmmmmmeeeeeeeeeeeeeeeeeeeeeeeelllllllllllllllllllll ',member_no)
            first_name = data_number.cell(data_search, 4).value
            last_name = data_number.cell(data_search, 3).value
            phone_no = data_number.cell(data_search, 2).value

            resv_stay1 = resv_stay(username)
            #print('reservation stay ',resv_stay1[0] ,resv_stay1[1])
            level_mbr = resv_stay1[0]
            text_background = resv_stay1[1]

            label10 = Label(window, text=cap_name(first_name) + ' ' + cap_name(last_name), font=('arial', 11),
                            fg='black', bg=text_background)
            label10.place(x=220, y=105)

            label12 = Label(window, text=member_no, font=('arial', 11), fg='black', bg=text_background)
            label12.place(x=395, y=105)

            labelA2 = Label(window, text=no_value, font=('arial', 11), fg='black', bg=text_background)
            labelA2.place(x=505, y=105)
            labelA2 = Label(window, text=points_coma, font=('arial', 11), fg='black', bg=text_background)
            labelA2.place(x=505, y=105)

            labelA4 = Label(window, text=cap_name(level_mbr), font=('arial', 11), fg='black', bg=text_background)
            labelA4.place(x=645, y=105)

            labelA6 = Label(window, text=enroll, font=('arial', 11), fg='black', bg=text_background)
            labelA6.place(x=730, y=105)

            labelA8 = Label(window, text=cap_name(active_acct), font=('arial', 11), fg='black', bg=text_background)
            labelA8.place(x=850, y=105)

            labelA10 = Label(window, text='Most Recent', font=('arial', 11), fg='black', bg=text_background)
            labelA10.place(x=950, y=105)

            labelA12 = Label(window, text='None', font=('arial', 11), fg='black', bg=text_background)
            labelA12.place(x=1090, y=105)

            labelA14 = Label(window, text='Yes', font=('arial', 11), fg='black', bg=text_background)
            labelA14.place(x=1180, y=105)

            label42 = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
            label42.place(x=1120, y=480)
            label42 = Label(window, text=reactivate_acct1, font=('arial', 11), fg='black', bg='lightgray')
            label42.place(x=1120, y=480)


            label46a = Label(window, text=no_value, font=('arial', 11), fg='black', bg='lightgray')
            label46a.place(x=872, y=520)
            label46a = Label(window, text=deactivate_date1, font=('arial', 11), fg='black', bg='lightgray')
            label46a.place(x=872, y=520)

            label46 = Label(window, text=no_value, font=('arial', 11), fg='black', bg='lightgray')
            label46.place(x=1120, y=520)
            label46 = Label(window, text=inactive1, font=('arial', 11), fg='black', bg='lightgray')
            label46.place(x=1120, y=520)

            label50a = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
            label50a.place(x=1120, y=560)
            label50a = Label(window, text=points_coma, font=('arial', 11), bg='lightgray')
            label50a.place(x=1120, y=560)

            label50b = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
            label50b.place(x=900, y=600)

            label50b = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
            label50b.place(x=900, y=600)
            label50b = Label(window, text=forfeit_coma, font=('arial', 11), bg='lightgray')
            label50b.place(x=900, y=600)

            label50c = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
            label50c.place(x=1120, y=600)
            label50c = Label(window, text=date_expire1, font=('arial', 11), bg='lightgray')
            label50c.place(x=1120, y=600)

            label16.delete(0, END)
            label16.insert(0, ms_mr)

            label17a.delete(0, END)
            label17a.insert(0, gender_mf)

            label20.delete(0, END)
            label20.insert(0, first_name)

            label22.delete(0, END)
            label22.insert(0, str(middle_name))

            label24.delete(0, END)
            label24.insert(0, str(last_name))

            # print('language ',language)
            label26.delete(0, END)
            label26.insert(0, str(language))

            label28.delete(0, END)
            label28.insert(0, str(birth_month))

            label29.delete(0, END)
            label29.insert(0, str(birth_no))

            label31a.delete(0, END)
            label31a.insert(0, str(phone_home))

            label35.delete(0, END)
            label35.insert(0, str(phone_no))

            label37.delete(0, END)
            label37.insert(0, str(home_no))

            label40.delete(0, END)
            label40.insert(0, str(active_acct))

            email_entry.delete(0, END)
            email_entry.insert(0, str(email_add))

            entry_homenumb.delete(0, END)
            entry_homenumb.insert(0, str(no_home))
            entry_homest.delete(0, END)
            entry_homest.insert(0, str(no_street))
            entry_homebarangay.delete(0, END)
            entry_homebarangay.insert(0, str(no_brgy))
            entry_homecity.delete(0, END)
            entry_homecity.insert(0, str(no_city))
            entry_homeprov.delete(0, END)
            entry_homeprov.insert(0, str(no_prov))
            entry_homecount.delete(0, END)
            entry_homecount.insert(0, str(no_country))
            entry_homezip.delete(0, END)
            entry_homezip.insert(0, str(no_zipcode))

            button_state = 'normal'
            # wb.save('data_wyn_gui.xlsx')
            # return username

    if loop == 0:
        label16.delete(0, END)
        label17a.delete(0, END)
        label20.delete(0, END)
        label22.delete(0, END)
        label24.delete(0, END)
        # print('language ',language)
        label26.delete(0, END)
        label28.delete(0, END)
        label29.delete(0, END)
        label31a.delete(0, END)
        label35.delete(0, END)
        label37.delete(0, END)
        label40.delete(0, END)
        labelc50h.delete(0, END)

        email_entry.delete(0, END)
        entry_homenumb.delete(0, END)
        entry_homest.delete(0, END)
        entry_homebarangay.delete(0, END)
        entry_homecity.delete(0, END)
        entry_homeprov.delete(0, END)
        entry_homecount.delete(0, END)
        entry_homezip.delete(0, END)


        '''email_entry = Entry(window, font=('arial', 11), width=30)
        email_entry.place(x=920, y=419)
        label43 = Label(window, text='Deactivate Date', font=('arial', 11), bg='lightgray')'''
        # member status & expiration

        label42 = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
        label42.place(x=1120, y=484)

        label46 = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
        label46.place(x=1120, y=520)

        label50a = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
        label50a.place(x=1120, y=560)

        label50b = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
        label50b.place(x=900, y=600)

        label50c = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
        label50c.place(x=1120, y=600)

        label43 = Label(window, text=no_value, font=('arial', 11), bg='lightgray')
        label43.place(x=872, y=520)

        button_state = 'disable'
        # tkinter.messagebox.showwarning(title="Error", message="Invalid Account Number")
    main_button = Button(window,
                         text="Update",
                         # command=click,
                         font=("arial", 11),
                         fg="black",
                         bg="lightgray",
                         activeforeground="black",
                         activebackground="lightgray",
                         padx=30,
                         pady=3,
                         width= 10,
                         state=button_state,
                         command=update)

    main_button.place(x=1110, y=177)

    buttond50i = Button(window,
                        text="Stay Find",
                        # command=click,
                        font=("arial", 11),
                        fg="black",
                        bg="lightgray",
                        activeforeground="black",
                        activebackground="lightgray",
                        padx=53,
                        pady=3,
                        width=6,
                        state=button_state,
                        command=stay_history)

    buttond50i.place(x=10, y=340)

    buttond50g = Button(window,
                        text="Enroll Member",
                        # command=click,
                        font=("arial", 11),
                        fg="black",
                        bg="lightgray",
                        activeforeground="black",
                        activebackground="lightgray",
                        padx=53,
                        pady=3,
                        width=6,
                        # state=button_state,
                        command=add_member)

    buttond50g.place(x=10, y=250)
    #wb.save('data_wyn_gui.xlsx')

# start post stay and modify stay
def scroll_stays(event):
    global get_resv_no, mod_loops
    mod_loops = 0

    #switching.clear()
    print('scrolls stays')
    get_value = []
    mylist_resv = []
    stay_post.clear()

    for x in my_list:
        if x.curselection() == ():
            my_list_value = str(0)
        else:
            my_list_value =x.curselection()
            my_list_value2 = list(x.get(x.curselection()))
    #print('my list value ',my_list_value2)

    #print(my_list_value)
    for x in my_list_value:
        my_list_value1 = x
        #print('my list value ', my_list_value1)

    if int(my_list_value1) == 0:
        #print('my list value ', my_list_value1)
        pass
    else:
        number_mbr = "".join(get_name)
        resv_stay1 = resv_stay(number_mbr)
        #print('reservation stay ', resv_stay1)

        level_mbr = resv_stay1[0]
        text_background = resv_stay1[1]

        labelstay11 = Label(window, background='#F0F0F0', padx=185, pady=205, border=1, relief=SOLID)
        labelstay11.place(x=935, y=216)
        switching.append(labelstay11)

        labelstay1 = Label(window, background='lightgray', padx=165, pady=205, border=1, relief=SOLID)
        labelstay1.place(x=935, y=216)
        switching.append(labelstay1)

        labelstay33 = Label(window, background='#F0F0F0', padx=50, pady=15)
        labelstay33.place(x=1267, y=216)
        switching.append(labelstay33)

        labelstay3 = Label(window, background=text_background, padx=165, pady=15, border=1, relief=SOLID)
        labelstay3.place(x=935, y=216)
        switching.append(labelstay3)

        labelstay333 = Label(window, background='#F0F0F0', padx=50, pady=15)
        labelstay333.place(x=1267, y=595)
        switching.append(labelstay333)

        label = Label(window, text="Reservation Details",
                      font=('arial', 17),
                      fg='blue',
                      bg=text_background,
                      padx=5, pady=0)
        label.place(x=950, y=223)
        switching.append(label)
        #print('my list value2 ',my_list_value2)
        for x in range(0,len(my_list_value2)):
            if x == 3 or x == 4 or x == 5:
                get_value.append(my_list_value2[x])
            else:
                pass
        my_resv_no = ''.join(get_value)
        for x in my_resv_no:
            if x.isnumeric() == True:
                mylist_resv.append(x)
            else:
                pass
        mylist_resv1 = "".join(mylist_resv)
        #print('reservation number is ', mylist_resv1)
        get_resv_no = mylist_resv1
        for data_search1 in range(2, data_number2.max_row + 1):
            no_member = data_number2.cell(data_search1, 1).value
            r_no = data_number2.cell(data_search1, 2).value
            last_name = data_number2.cell(data_search1, 4).value
            first_name = data_number2.cell(data_search1, 5).value
            name_hotel = data_number2.cell(data_search1, 7).value

            rate_code = data_number2.cell(data_search1, 10).value
            room_rate = data_number2.cell(data_search1, 11).value
            taxes = data_number2.cell(data_search1, 12).value
            points = data_number2.cell(data_search1, 13).value
            room_rate_total = data_number2.cell(data_search1, 14).value
            total_taxes = data_number2.cell(data_search1, 15).value
            total_points = data_number2.cell(data_search1, 16).value
            total_amount = data_number2.cell(data_search1, 17).value
            no_nyts = data_number2.cell(data_search1, 18).value

            adult = data_number2.cell(data_search1, 19).value
            child = data_number2.cell(data_search1, 20).value
            rooms_one_two = data_number2.cell(data_search1, 21).value
            checkin = data_number2.cell(data_search1, 22).value
            checkout = data_number2.cell(data_search1, 23).value

            rate_name = data_number2.cell(data_search1, 24).value
            clear_post = data_number2.cell(data_search1, 25).value
            ready_post = data_number2.cell(data_search1, 26).value
            postdate = data_number2.cell(data_search1, 27).value

            #print('reservation number ',r_no)
            if  str(r_no) == str(mylist_resv1):
                stay_post.append(''.join(mylist_resv))
                #print('reservation number post stay scroll ', ''.join(mylist_resv) , stay_post)
                #print('member number post stay scroll', number_mbr)
                if clear_post == 'cancel':
                    confirm = 'Cancelled'
                    clear_bg = 'red'
                else:
                    confirm = 'Confirmed'
                    clear_bg = '#50EA49'

                for data_search in range(2, data_number.max_row + 1):
                    member_numbers = str(data_number.cell(data_search, 1).value)
                    pts_balance = data_number.cell(data_search, 8).value
                    # print('member_numbers ', member_numbers ,no_member)
                    if str(member_numbers) == str(no_member):
                        points_data1_bal = pts_balance
                        # print('points balance ', points_data1_bal)

                d = data_number2.cell(data_search1, 22).value
                # print('recent datessssssssssssssssssssss ', d)
                list_d = d.split('/')
                # print(list_d)
                b = 0
                y = []

                while b < len(list_d):
                    # if b == 0:
                    c = list_d[b]
                    for x in c:
                        if x.isnumeric() == True:
                            y.append(x)
                    b += 1
                b = 0
                # print('split y ', y)
                output = [int(x) for x in y]
                # print('split  ', output)

                a4 = 0
                while a4 < len(output):
                    if a4 <= 1:
                        a1 = int(str(output[0]) + str(output[1]))
                    elif a4 >= 1 and a4 <= 3:
                        a2 = int(str(output[2]) + str(output[3]))
                    elif a4 >= 3 and a4 <= 6:
                        a3 = int(str(output[4]) + str(output[5]) + str(output[6]) + str(output[7]))
                    a4 += 1

                today2 = date(a3, a2, a1)
                new_date = str(a3) + ',' + str(a2) + ',' + str(a1)
                # print('new date ', new_date)
                new_date1 = str(a1) + ',' + str(a2) + ',' + str(a3)
                # today1 = date(a3, a2, a1)
                check_inn_format = today2.strftime("%a, %b %d - ")
                # print('check in format: ', check_inn_format)

                d = data_number2.cell(data_search1, 23).value
                # print('recent datessssssssssssssssssssss ', d)
                list_d = d.split('/')
                # print(list_d)
                b = 0
                y = []

                while b < len(list_d):
                    # if b == 0:
                    c = list_d[b]
                    for x in c:
                        if x.isnumeric() == True:
                            y.append(x)
                    b += 1
                b = 0
                # print('split y ', y)
                output = [int(x) for x in y]
                # print('split  ', output)

                a4 = 0
                while a4 < len(output):
                    if a4 <= 1:
                        a1 = int(str(output[0]) + str(output[1]))
                    elif a4 >= 1 and a4 <= 3:
                        a2 = int(str(output[2]) + str(output[3]))
                    elif a4 >= 3 and a4 <= 6:
                        a3 = int(str(output[4]) + str(output[5]) + str(output[6]) + str(output[7]))
                    a4 += 1

                today2 = date(a3, a2, a1)
                new_date = str(a3) + ',' + str(a2) + ',' + str(a1)
                # print('new date ', new_date)
                new_date1 = str(a1) + ',' + str(a2) + ',' + str(a3)
                # today1 = date(a3, a2, a1)
                check_outt_format = today2.strftime("%a, %b %d %Y")

                if len(name_hotel) < 35:
                    hotelname = name_hotel
                else:
                    hotelname = ''
                    word = name_hotel
                    words = []
                    word1 = ''
                    loops = 0

                    for x in range(0, len(word)):
                        word1 += word[x]
                        if word[x] == ' ':
                            words.append(word1)
                            word1 = ''
                    words.append(word1)

                    total = len(word)
                    loops = 0
                    total1 = 0
                    while loops == 0:
                        if total > 34:
                            total1 += len(words[-1])
                            total = len(word) - total1
                            words.remove(words[-1])
                        else:
                            loops = 1

                    temp = words[-1]
                    total1 = ''
                    for x in temp:
                        if x == ' ':
                            pass
                        else:
                            total1 += x

                    words[-1] = str(total1) + '...'
                    hotelname = ' '.join(words)
                # print('check out format: ',check_outt_format)
                # print('rrrrrrrrrrraaaaaaaaaattttttttttttttttteeeeeeeeeee ccccoooooodddddddeeeeeeeee ',rate_code)
                if rate_code == 'Go Fast':
                    total_amount1 = '$' + str(conversion_numbers1(room_rate)) + ' + ' + str(
                        conversion_numbers(points)) + ' PTS'

                    if len(str(conversion_numbers(total_points))) > 6:
                        total_value_format = 'USD ' + str(conversion_numbers1(room_rate_total)) + ' + '
                        total_value_format1 = str(conversion_numbers(total_points)) + ' PTS'

                    else:
                        total_value_format = 'USD ' + str(conversion_numbers1(room_rate_total)) + ' + ' \
                                             + str(conversion_numbers(total_points)) + ' PTS'

                    total_value_percent_format = 'USD ' + str(conversion_numbers1(total_taxes))

                    if len(str(conversion_numbers(total_points))) > 6:
                        total_value_payment_format = 'USD ' + str(conversion_numbers1(total_amount)) + ' + '
                        total_value_payment_format1 = str(conversion_numbers(total_points)) + ' PTS'

                    else:
                        total_value_payment_format = 'USD ' + str(conversion_numbers1(total_amount)) + ' + ' \
                                                     + str(conversion_numbers(total_points)) + ' PTS'

                    #print('lenght of value ', len(str(conversion_numbers(total_points))))

                elif rate_code == 'Go Free':
                    # print('gggggggooooooooooooo    fffffffffffrrrrrreeeeeeeeeeeeeeeeeeeeeeee')
                    total_amount1 = str(conversion_numbers(points)) + ' PTS'
                    total_value_format = str(conversion_numbers(total_points)) + ' PTS'
                    total_value_percent_format = 'USD ' + str(conversion_numbers1(total_taxes))
                    total_value_payment_format = str(conversion_numbers(total_points)) + ' PTS'

                else:
                    total_amount1 = '$' + str(conversion_numbers1(room_rate))
                    total_value_format = 'USD ' + str(conversion_numbers1(room_rate_total))
                    total_value_percent_format = 'USD ' + str(conversion_numbers1(total_taxes))
                    total_value_payment_format = 'USD ' + str(conversion_numbers1(total_amount))

                #print('reservation number: ',r_no)
                resv = Label(window, text=confirm, font=('arial', 11), fg='white', bg=clear_bg)
                resv.place(x=950, y=275)
                switching.append(resv)

                resv1 = Label(window, text=r_no, font=('arial', 11), fg='black', bg='lightgray')
                resv1.place(x=1031, y=275)
                switching.append(resv1)

                hotels = Label(window, text=hotelname, font=('arial', 11, 'bold'), fg='black', bg='lightgray')
                hotels.place(x=950, y=304)
                switching.append(hotels)

                hotel = Label(window, text=str(check_inn_format) + str(check_outt_format),
                              font=('arial', 10, 'bold'), fg='black', bg='lightgray')
                hotel.place(x=950, y=331)
                switching.append(hotel)

                adults = Label(window, text=str(adult) + ' adult ' + str(child) + ' child ',
                               font=('arial', 10),
                               fg='black', bg='lightgray')
                adults.place(x=950, y=356)
                switching.append(adults)

                fullname = Label(window, text=str(cap_name(first_name)) + ' ' + str(
                    cap_name(last_name)) + ' / ' + conversion_numbers(points_data1_bal) + ' PTS', font=('arial', 10),
                                 fg='blue', bg='lightgray')
                fullname.place(x=950, y=381)
                switching.append(fullname)

                rooms1 = Label(window, text=cap_name(rooms_one_two), font=('arial', 10), fg='black', bg='lightgray')
                rooms1.place(x=950, y=406)
                switching.append(rooms1)

                rate_rooms = Label(window, text=str(cap_name(rate_name)) + '  ' + str(rate_code), font=('arial', 10),
                                   fg='blue', bg='lightgray')
                rate_rooms.place(x=950, y=431)
                switching.append(rate_rooms)

                room_payment = Label(window, text=str(total_amount1) + '  avg / night ',
                                     font=('arial', 10), fg='black',
                                     bg='lightgray')
                room_payment.place(x=950, y=456)
                switching.append(room_payment)

                space_nyts1 = ''
                no_nyts1 = len(str(no_nyts))
                for x in range(0,no_nyts1):
                    space_nyts1 += space_nyts1

                if len(str(no_nyts)) == 1:
                    nyts_payment = Label(window,
                                         text='1 room ' + str(no_nyts) + ' night             '+ str(
                                             total_value_format), font=('arial', 10), fg='black',
                                         bg='lightgray')
                    nyts_payment.place(x=950, y=491)
                    switching.append(nyts_payment)

                elif len(str(no_nyts)) == 2:
                    nyts_payment = Label(window,
                                         text='1 room ' + str(no_nyts) + ' night           '+ str(
                                             total_value_format), font=('arial', 10), fg='black',
                                         bg='lightgray')
                    nyts_payment.place(x=950, y=491)
                    switching.append(nyts_payment)

                elif len(str(no_nyts)) == 3:
                    if rate_code == 'Go Fast':
                        nyts_payment = Label(window,
                                             text='1 room ' + str(no_nyts) + ' night           ' + str(
                        total_value_format), font = ('arial', 10), fg = 'black',bg = 'lightgray')
                        nyts_payment.place(x=950, y=491)
                        switching.append(nyts_payment)

                        nyts_payment1 = Label(window,text=total_value_format1, font = ('arial', 10),
                                             fg = 'black', bg = 'lightgray')
                        nyts_payment1.place(x=1090, y=514)
                        switching.append(nyts_payment1)

                        tax_fees = Label(window, text='taxes and fees              ' + str(
                            total_value_percent_format),font=('arial', 10), fg='black', bg='lightgray')
                        tax_fees.place(x=950, y=547)
                        switching.append(tax_fees)

                        tax_fees1 = Label(window, text='total                             ' + str(
                            total_value_payment_format),
                                          font=('arial', 10), fg='black', bg='lightgray')
                        tax_fees1.place(x=950, y=577)
                        switching.append(tax_fees1)

                        tax_fees2 = Label(window, text=total_value_payment_format1,
                                          font=('arial', 10), fg='black', bg='lightgray')
                        tax_fees2.place(x=1090, y=602)
                        switching.append(tax_fees1)

                    else:
                        nyts_payment = Label(window,
                                             text='1 room ' + str(no_nyts) + ' night         '+ str(
                                                 total_value_format), font=('arial', 10), fg='black',
                                             bg='lightgray')
                        nyts_payment.place(x=950, y=491)
                        switching.append(nyts_payment)

                if len(str(no_nyts)) != 3:
                    tax_fees = Label(window, text='taxes and fees            ' + str(
                            total_value_percent_format),font=('arial', 10), fg='black', bg='lightgray')
                    tax_fees.place(x=950, y=526)
                    switching.append(tax_fees)

                    tax_fees1 = Label(window, text='total                           ' + str(
                        total_value_payment_format),
                                      font=('arial', 10), fg='black', bg='lightgray')
                    tax_fees1.place(x=950, y=561)
                    switching.append(tax_fees1)

                stay_posting1_label = Label(window, background='lightgray', padx=120, pady=10)
                stay_posting1_label.place(x=450, y=599)
                switching.append(stay_posting1_label)

                post_button = Button(window,
                                     text="Post Stay",
                                     # command=click,
                                     font=("arial", 9),
                                     fg="black",
                                     bg="lightgray",
                                     activeforeground="black",
                                     activebackground="lightgray",
                                     padx=40,
                                     pady=3,
                                     # state='disabled',
                                     command=post_stay)

                post_button.place(x=215, y=600)
                # but_post.append(post_button)
                switching.append(post_button)



                cxl_button = Button(window,
                                    text="Cancel Stay",
                                    # command=click,
                                    font=("arial", 9),
                                    fg="black",
                                    bg="lightgray",
                                    activeforeground="black",
                                    activebackground="lightgray",
                                    padx=30,
                                    pady=3,
                                    # state='disabled',
                                    command=cancel_stay)

                cxl_button.place(x=770, y=600)
                switching.append(cxl_button)

                modified_button = Button(window,
                                         text="Modify Stay",
                                         # command=click,
                                         font=("arial", 9),
                                         fg="black",
                                         bg="lightgray",
                                         activeforeground="black",
                                         activebackground="lightgray",
                                         padx=30,
                                         pady=3,
                                         # state='disabled',
                                         command=modify_stay)
                modified_button.place(x=500, y=600)
                switching.append(modified_button)

def stay_history():
    print('stay history')
    global mod_acct_no, get_resv_no, mod_loops, temp_mod_siteno, temp_mod_siteno1, temp_mod_siteno2
    mod_acct_no ,get_resv_no, mod_loops ,temp_mod_siteno, temp_mod_siteno1, temp_mod_siteno2 = IntVar(), IntVar(), 0, 0, 0, 0
    patch_design = Label(window, background="#F0F0F0", padx=532, pady=230)
    patch_design.place(x=200, y=170)
    resv_forget.append(patch_design)
    my_list.clear()

    text_background = ' '
    level_mbr = ' '
    for x in range(0, len(temp_addnewmbr)):
        temp_addnewmbr[x] = ''

    for x in resv_forget:
        x.place_forget()
    resv_forget.clear()

    for x in enroll_new1:
        x.place_forget()
    enroll_new1.clear()

    for x in enroll_new:
        x.place_forget()
    enroll_new.clear()
    # number_mbr = member_search()
    stay_post.clear()

    stay_find = labelc50h.get()
    stay_post.append(stay_find)
    #print('reservation number append ', stay_post)
    #print('member number ', get_name)
    number_mbr = "".join(get_name)
    resv_stay1 = resv_stay(number_mbr)
    mod_acct_no = number_mbr
    #print('reservation stay ', number_mbr ,mod_acct_no)
    level_mbr = resv_stay1[0]
    text_background = resv_stay1[1]

    # window.geometry("1070x470")
    # dashboarb textbackgroun lightblue
    for x in add_find:
        if x == 'add':
            number_mbr = ''
            text_background ='lightblue'

    labelstay = Label(window, background='lightgray', padx=400, pady=225, border=1, relief=SOLID)
    labelstay.place(x=200, y=176)
    switching.append(labelstay)

    labelstay2 = Label(window, background=text_background, padx=400, pady=15, border=1, relief=SOLID)
    labelstay2.place(x=200, y=176)
    switching.append(labelstay2)

    labelstay1 = Label(window, background='lightgray', padx=165, pady=205, border=1, relief=SOLID)
    labelstay1.place(x=935, y=216)
    switching.append(labelstay1)

    labelstay3 = Label(window, background=text_background, padx=165, pady=15, border=1, relief=SOLID)
    labelstay3.place(x=935, y=216)
    switching.append(labelstay3)

    label = Label(window, text="Reservation Details",
                  font=('arial', 17),
                  fg='blue',
                  bg=text_background,
                  padx=5, pady=0)
    label.place(x=950, y=223)
    switching.append(label)

    label15 = Label(window, text="Reservation History",
                    font=('arial', 17),
                    fg='blue',
                    bg=text_background,
                    padx=5, pady=0)
    label15.place(x=210, y=183)
    switching.append(label15)

    main_button = Button(window,
                         text="Update",
                         # command=click,
                         font=("arial", 11),
                         fg="black",
                         bg="lightgray",
                         activeforeground="black",
                         activebackground="lightgray",
                         padx=30,
                         pady=3,
                         width=10,
                         state='disabled',
                         command=update)

    main_button.place(x=1110, y=177)

    # 200 ,175
    # clearscreen reservation details & history
    for x in add_find:
        if x == 'member':
            iteration = 0

    # print('member search', member_search())
    # print('stay number find ',stay_find)
    # add_find.clear()
    for data_search in range(2, data_number.max_row + 1):
        member_numbers = str(data_number.cell(data_search, 1).value)
        pts_balance = data_number.cell(data_search, 8).value

    rows, rows1, rows2, rows3, rows4, rows5 = 215, 290, 555, 654, 754, 854
    columns, columns1, columns2, columns3, columns4, columns5 = 270, 270, 270, 270, 270, 270

    two_array = []
    temp = []
    # print('sorting ',sorting)
    for account_no in range(2, data_number2.max_row + 1):
        mbr_no = data_number2.cell(account_no, 1).value
        resv_no = data_number2.cell(account_no, 2).value
        sort_checkin = data_number2.cell(account_no, 22).value
        if str(mbr_no) == str(number_mbr):
            temp = []
            sorts_dates = wyn_inactivity_GUI.last_date4(sort_checkin)
            #print('sorting date ', sorts_dates)
            sorts_inactive = wyn_inactivity_GUI.days_inactive(sorts_dates)
            #print('reservation & sorting date ',type(resv_no) ,type(sorts_inactive))
            temp.append(int(resv_no))
            temp.append(int(sorts_inactive))
            two_array.append(temp)

    #print('two dimensionallllllllll ', two_array)
    sort_two_array = sorted(two_array,key=operator.itemgetter(0), reverse=True)
    ##print('sorted two dimentional array ', sort_two_array)
    sort_two_array1 = sorted(sort_two_array,key=operator.itemgetter(1))
    #print('sorted two dimentional array ', sort_two_array1)
    #scrool bar and listbox
    scroll_iteration = []
    spaces = []
    spaces2 = []
    total = 0
    loop = 0
    loop1 = 0
    my_frame = Frame(window)
    switching.append(my_frame)
    scrollbar1 = Scrollbar(my_frame, orient=VERTICAL)
    scrollbar1.pack(side=RIGHT, fill=Y)

    scrollbar = Scrollbar(my_frame, orient=HORIZONTAL)
    scrollbar.pack(side=BOTTOM, fill=X)

    mylist = Listbox(my_frame, xscrollcommand=scrollbar.set,
                     yscrollcommand=scrollbar1.set,
                     font=('Consolas', 10, ''), fg='black', bg='lightblue',
                     activestyle="none", width=97, height=16)
    sort_arrs = []
    #list_dict = list(dict(sort_two_array1))
    #print('reservationnnnnnnnnnnnnnnnnnnnnn ',list_dict)
    for x in sort_two_array1:
        sort_arrs.append(x[0])
    #print('two arrays reservation value ', sort_arrs)
    #print('len sorted no ',len(sorted_dates))
    while len(sort_arrs) != 0:
        for data_search1 in range(2, data_number2.max_row + 1):
            listbox_no_member = data_number2.cell(data_search1, 1).value
            listbox_rno = data_number2.cell(data_search1, 2).value
            listbox_name_hotel = data_number2.cell(data_search1, 7).value
            listbox_room_rate_total = data_number2.cell(data_search1, 14).value
            listbox_checkin = data_number2.cell(data_search1, 22).value
            listbox_checkout = data_number2.cell(data_search1, 23).value
            listbox_postdate = data_number2.cell(data_search1, 27).value
            listbox_siteno = data_number2.cell(data_search1, 28).value

            spaces = []
            spaces2 =''
            #namelist = list(listbox_name_hotel)
            if listbox_name_hotel == 'Microtel Pampanga':
                total = 52 - len(str(listbox_name_hotel ))
                total1 = 3 - len(str(listbox_rno))
            elif listbox_name_hotel == 'Days Hotel':
                total = 45 - len(str(listbox_name_hotel ))
                total1 = 3 - len(str(listbox_rno))
            else:
                total = 55 - len(str(listbox_name_hotel))
                total1 = 3 - len(str(listbox_rno))

            #print(listbox_name_hotel ,len(listbox_name_hotel) , total)
            for x in range(0, total):
                spaces.append(' ')
            spaces1 = ''.join(spaces)

            if number_mbr == str(listbox_no_member):
                postdate1 = []
                if listbox_postdate != None:
                    for x in listbox_postdate:
                        if x != '"':
                            postdate1.append(x)
                    postdate1 = "".join(postdate1)
                else:
                    pass
            if number_mbr == str(listbox_no_member):
                checkin1 = []
                if listbox_checkin != None:
                    for x in listbox_checkin:
                        if x != '"':
                            checkin1.append(x)
                    checkin1 = "".join(checkin1)
                else:
                    pass
            if number_mbr == str(listbox_no_member):
                checkout1 = []
                if listbox_checkout != None:
                    for x in listbox_checkout:
                        if x != '"':
                            checkout1.append(x)
                    checkout1 = "".join(checkout1)
                else:
                    pass
            rating1 = 0
            if number_mbr == str(listbox_no_member):
                for data_search3 in range(2, data_number3.max_row + 1):
                    siteno = data_number3.cell(data_search3, 1).value
                    rating = data_number3.cell(data_search3, 5).value
                    if str(listbox_siteno) == str(siteno):
                        rating1 = rating
                if len(sort_arrs) == 0:
                    pass
                else:
                    if str(sort_arrs[loop1]) == str(listbox_rno):
                        #print('good ', str(sorted_dates[loop1]) ,str(listbox_rno))
                        if loop == 0:

                            mylist.insert(END, f"{'' : <0} {'Resv.#': <3} {'' : <3}{'Site #': <10} {'Property Name': <20} {'' : <35} "
                                               f"{'Rating': <13}{'Check-in': <16} {'Check-out': <16} {'Post Date': <16} "
                                               f"{'Earn Points': <16}")
                            loop +=1


                        #print('reservation number ', listbox_rno)
                        mylist.insert(END, f"{'' : <2} {listbox_rno :<3} {total1 *' ' : <3} {listbox_siteno: <10} {str(listbox_name_hotel): <20} "
                                           f"{total *' ' : <5} {rating1 : <12} {str(checkin1) : <16} {str(checkout1) : <16} "
                                           f"{str(postdate1) : <16} {str(conversion_numbers(int(listbox_room_rate_total * 10))) : <16}")
                        sort_arrs.remove(sort_arrs[loop1])
                        #print('sorted dates ',sorted_dates)
    loop1 += 1

    mylist.pack(side=LEFT)
    scrollbar1.config(command=mylist.yview, width=25)

    mylist.pack(side=TOP)
    scrollbar.config(command=mylist.xview, width=25)
    # my_frame.pack(anchor='w', padx=215, pady=270)
    my_frame.place(x=215, y=255)
    # my_frame.pack(anchor='w')
    mylist.bind('<ButtonRelease>', scroll_stays)
    #mylist.bind('<Double-1>', scroll_stays)
    #print('my list bind ',mylist)
    my_list.append(mylist)
    #print(' my list ', my_list)

    for data_search1 in range(2, data_number2.max_row + 1):
        no_member = data_number2.cell(data_search1, 1).value
        r_no = data_number2.cell(data_search1, 2).value
        last_name = data_number2.cell(data_search1, 4).value
        first_name = data_number2.cell(data_search1, 5).value
        name_hotel = data_number2.cell(data_search1, 7).value

        rate_code = data_number2.cell(data_search1, 10).value
        room_rate = data_number2.cell(data_search1, 11).value
        taxes = data_number2.cell(data_search1, 12).value
        points = data_number2.cell(data_search1, 13).value
        room_rate_total = data_number2.cell(data_search1, 14).value
        total_taxes = data_number2.cell(data_search1, 15).value
        total_points = data_number2.cell(data_search1, 16).value
        total_amount = data_number2.cell(data_search1, 17).value
        no_nyts = data_number2.cell(data_search1, 18).value

        adult = data_number2.cell(data_search1, 19).value
        child = data_number2.cell(data_search1, 20).value
        rooms_one_two = data_number2.cell(data_search1, 21).value
        checkin = data_number2.cell(data_search1, 22).value
        checkout = data_number2.cell(data_search1, 23).value

        rate_name = data_number2.cell(data_search1, 24).value
        clear_post = data_number2.cell(data_search1, 25).value
        ready_post = data_number2.cell(data_search1, 26).value
        postdate = data_number2.cell(data_search1, 27).value

        #print('check in and check out ',points ,total_points)
        #print('ready post ',ready_post)
        if number_mbr == str(no_member):
            postdate1 = []
            if postdate != None:
                for x in postdate:
                    if x != '"':
                        postdate1.append(x)
                postdate1 = "".join(postdate1)
            else:
                pass
        if number_mbr == str(no_member):
            checkin1 = []
            if checkin != None:
                for x in checkin:
                    if x != '"':
                        checkin1.append(x)
                checkin1 = "".join(checkin1)
            else:
                pass
        if number_mbr == str(no_member):
            checkout1 = []
            if checkout != None:
                for x in checkout:
                    if x != '"':
                        checkout1.append(x)
                checkout1 = "".join(checkout1)
            else:
                pass


            if rate_code == 'Go Free':
                earning_pts = 0
            else:
                earning_pts = room_rate_total * 10

            iteration += 1
            #print('stay find and resv. no ', stay_find, r_no + 1)
            if str(stay_find) == str(r_no):
                if clear_post == 'cancel':
                    confirm = 'Cancelled'
                    clear_bg = 'red'
                else:
                    confirm = 'Confirmed'
                    clear_bg = '#50EA49'
                #print('reservation number ', no_member)
                #print('stay find and resv. no ', stay_find, r_no)
                for data_search in range(2, data_number.max_row + 1):
                    member_numbers = str(data_number.cell(data_search, 1).value)
                    pts_balance = data_number.cell(data_search, 8).value
                    #print('member_numbers ', member_numbers ,no_member)
                    if str(member_numbers) == str(no_member):
                        points_data1_bal = pts_balance
                        # print('points balance ', points_data1_bal)
                # check-in dates
                #print('data search one ', data_search1 ,)
                #check_inn = wyn_inactivity_GUI.last_date3(data_search1, 22)
                #check_inn_format = check_inn.strftime("%a, %b %d - ")
                #print('check innnnnnnnnnnn ',check_inn_format)

                d = data_number2.cell(data_search1, 22).value
                # print('recent datessssssssssssssssssssss ', d)
                list_d = d.split('/')
                # print(list_d)
                b = 0
                y = []

                while b < len(list_d):
                    # if b == 0:
                    c = list_d[b]
                    for x in c:
                        if x.isnumeric() == True:
                            y.append(x)
                    b += 1
                b = 0
                # print('split y ', y)
                output = [int(x) for x in y]
                # print('split  ', output)

                a4 = 0
                while a4 < len(output):
                    if a4 <= 1:
                        a1 = int(str(output[0]) + str(output[1]))
                    elif a4 >= 1 and a4 <= 3:
                        a2 = int(str(output[2]) + str(output[3]))
                    elif a4 >= 3 and a4 <= 6:
                        a3 = int(str(output[4]) + str(output[5]) + str(output[6]) + str(output[7]))
                    a4 += 1

                today2 = date(a3, a2, a1)
                new_date = str(a3) + ',' + str(a2) + ',' + str(a1)
                # print('new date ', new_date)
                new_date1 = str(a1) + ',' + str(a2) + ',' + str(a3)
                # today1 = date(a3, a2, a1)
                check_inn_format = today2.strftime("%a, %b %d - ")
                #print('check in format: ', check_inn_format)

                # check-out dates
                #check_outt = wyn_inactivity_GUI.last_date3(data_search1, 23)
                #check_outt_format = check_outt.strftime("%a, %b %d %Y")
                #print('check outttttttttttttttttt ', check_outt_format)

                d = data_number2.cell(data_search1, 23).value
                # print('recent datessssssssssssssssssssss ', d)
                list_d = d.split('/')
                # print(list_d)
                b = 0
                y = []

                while b < len(list_d):
                    # if b == 0:
                    c = list_d[b]
                    for x in c:
                        if x.isnumeric() == True:
                            y.append(x)
                    b += 1
                b = 0
                # print('split y ', y)
                output = [int(x) for x in y]
                # print('split  ', output)

                a4 = 0
                while a4 < len(output):
                    if a4 <= 1:
                        a1 = int(str(output[0]) + str(output[1]))
                    elif a4 >= 1 and a4 <= 3:
                        a2 = int(str(output[2]) + str(output[3]))
                    elif a4 >= 3 and a4 <= 6:
                        a3 = int(str(output[4]) + str(output[5]) + str(output[6]) + str(output[7]))
                    a4 += 1

                today2 = date(a3, a2, a1)
                new_date = str(a3) + ',' + str(a2) + ',' + str(a1)
                # print('new date ', new_date)
                new_date1 = str(a1) + ',' + str(a2) + ',' + str(a3)
                # today1 = date(a3, a2, a1)
                check_outt_format = today2.strftime("%a, %b %d %Y")
                #print('check out format: ',check_outt_format)
                #print('rrrrrrrrrrraaaaaaaaaattttttttttttttttteeeeeeeeeee ccccoooooodddddddeeeeeeeee ',rate_code)
                if rate_code == 'Go Fast':
                    total_amount1 = '$' + str(conversion_numbers1(room_rate)) + ' + ' + str(conversion_numbers(points)) +' PTS'
                    total_value_format = 'USD ' + str(conversion_numbers1(room_rate_total)) + ' + ' \
                                         + str(conversion_numbers(total_points)) +' PTS'
                    total_value_percent_format = 'USD ' + str(conversion_numbers1(total_taxes))
                    total_value_payment_format = 'USD ' + str(conversion_numbers1(total_amount)) + ' + '\
                                                 + str(conversion_numbers(total_points)) +' PTS'

                elif rate_code == 'Go Free':
                    #print('gggggggooooooooooooo    fffffffffffrrrrrreeeeeeeeeeeeeeeeeeeeeeee')
                    total_amount1 = str(conversion_numbers(points)) + ' PTS'
                    total_value_format = str(conversion_numbers(total_points))+ ' PTS'
                    total_value_percent_format = 'USD ' + str(conversion_numbers1(total_taxes))
                    total_value_payment_format = str(conversion_numbers(total_points)) + ' PTS'

                else:
                    total_amount1 = '$' + str(conversion_numbers1(room_rate))
                    total_value_format = 'USD ' + str(conversion_numbers1(room_rate_total))
                    total_value_percent_format = 'USD ' + str(conversion_numbers1(total_taxes))
                    total_value_payment_format = 'USD ' + str(conversion_numbers1(total_amount))

                # total_amount = format(total_amount, '.2f')
                #round_total = round(total_amount) * 10
                # print('round total ', round(total_amount))
                #print('reservation number: ', r_no)
                resv = Label(window, text=confirm, font=('arial', 11), fg='white', bg=clear_bg)
                resv.place(x=950, y=275)
                switching.append(resv)

                resv1 = Label(window, text=r_no, font=('arial', 11), fg='black', bg='lightgray')
                resv1.place(x=1031, y=275)
                switching.append(resv1)

                hotels = Label(window, text=name_hotel, font=('arial', 11, 'bold'), fg='black', bg='lightgray')
                hotels.place(x=950, y=304)
                switching.append(hotels)

                hotel = Label(window, text=str(check_inn_format) + str(check_outt_format),
                              font=('arial', 10, 'bold'), fg='black', bg='lightgray')
                hotel.place(x=950, y=331)
                switching.append(hotel)

                adults = Label(window, text=str(adult) + ' adult ' + str(child) + ' child ',
                               font=('arial', 10),
                               fg='black', bg='lightgray')
                adults.place(x=950, y=356)
                switching.append(adults)

                fullname = Label(window, text=str(cap_name(first_name)) + ' ' + str(
                    cap_name(last_name)) + ' / ' + conversion_numbers(points_data1_bal) + ' PTS', font=('arial', 10),
                                 fg='blue', bg='lightgray')
                fullname.place(x=950, y=381)
                switching.append(fullname)

                rooms1 = Label(window, text=cap_name(rooms_one_two), font=('arial', 10), fg='black', bg='lightgray')
                rooms1.place(x=950, y=406)
                switching.append(rooms1)

                rate_rooms = Label(window, text=str(cap_name(rate_name)) + '  ' + str(rate_code), font=('arial', 10),
                                   fg='blue', bg='lightgray')
                rate_rooms.place(x=950, y=431)
                switching.append(rate_rooms)

                room_payment = Label(window, text=str(total_amount1) + '  avg / night ',
                                     font=('arial', 10), fg='black',
                                     bg='lightgray')
                room_payment.place(x=950, y=456)
                switching.append(room_payment)
                # canvas.create_line(100, 200, 200, 35, fill="green", width=5)

                nyts_payment = Label(window,
                                     text='1 room ' + str(no_nyts) + ' night             ' + str(
                                         total_value_format), font=('arial', 10), fg='black',
                                     bg='lightgray')
                nyts_payment.place(x=950, y=491)
                switching.append(nyts_payment)

                tax_fees = Label(window, text='taxes and fees             ' + str(
                    total_value_percent_format),
                                 font=('arial', 10), fg='black', bg='lightgray')
                tax_fees.place(x=950, y=526)
                switching.append(tax_fees)

                tax_fees1 = Label(window, text='total                            ' + str(
                    total_value_payment_format),
                                  font=('arial', 10), fg='black', bg='lightgray')
                tax_fees1.place(x=950, y=561)
                switching.append(tax_fees1)


                post_button = Button(window,
                                     text="Post Stay",font=("arial", 9), fg="black", bg="lightgray",
                                     activeforeground="black", activebackground="lightgray", padx=80, pady=1, width=1,
                                     #padx=40,
                                     #pady=3,
                                     # state='disabled',
                                     command=post_stay)

                post_button.place(x=215, y=600)
                #but_post.append(post_button)
                switching.append(post_button)

                cxl_button = Button(window,
                                     text="Cancel Stay", font=("arial", 9), fg="black", bg="lightgray",
                                     activeforeground="black", activebackground="lightgray", padx=80, pady=1, width=1,
                                     #padx=30,
                                     #pady=3,
                                     # state='disabled',
                                     command=cancel_stay)
                cxl_button.place(x=770, y=600)
                switching.append(cxl_button)

                modified_button = Button(window,
                                         text="Modify Stay", font=("arial", 9), fg="black", bg="lightgray",
                                         activeforeground="black", activebackground="lightgray", padx=80, pady=1, width=1,
                                         #padx=30,
                                         #pady=3,
                                         state='disabled',
                                         command=modify_stay)
                modified_button.place(x=500, y=600)
                switching.append(modified_button)

            else:
                pass
    #return number_mbr
    # window1 = mainloop()

def modify_stay():
    print('modify stay')
    global get_resv_no, mod_siteno, temp_mod_siteno ,temp_mod_siteno1, temp_mod_siteno2, mod_checkin, mod_checkout, mod_nonight, \
        mod_loops, mod_rooms, mod_adults, mod_childs
    mod_siteno, mod_checkin, mod_checkout, mod_nonight, mod_rooms, mod_adults, mod_childs = 0, 0, 0, 0, 1, 0, 0

    #print('reservation number get_resv_no', get_resv_no, mod_loops)
    number_mbr = "".join(get_name)
    modify_resv_number[0] = number_mbr
    #print('reservation stay ', number_mbr)
    # account = choose_acctno[0]
    reservation = get_resv_no
    #print('reservation number ', reservation)
    for data_search in range(2, data_number2.max_row + 1):
        resvno = data_number2.cell(data_search, 2).value
        checkin = data_number2.cell(data_search, 22).value
        if str(resvno) == str(reservation):
            a = [str(y) for y in checkin]
            c1 = str(a[1]) + str(a[2])
            c2 = str(a[4]) + str(a[5])
            c3 = str(a[7]) + str(a[8]) + str(a[9]) + str(a[10])
            c4 = date(int(c3), int(c2), int(c1))
            today = date.today()
            valid_entry_in = c4 - today
            valid_entry = valid_entry_in.days
            #print('current dates ', c4 ,today ,valid_entry)
    if int(valid_entry) < 0:
        messagebox.showinfo(title='Reservation ' + str(reservation), message='Stay cannot Modify')

    else:
        patch_design = Label(window, background="#F0F0F0", padx=532, pady=230)
        patch_design.place(x=200, y=170)
        resv_forget.append(patch_design)

        levelcolor = 'lightblue'
        nonight1 = 0
        adult1 = 0
        child1 = 0
        site = 0
        for data_search in range(2, data_number2.max_row + 1):
            resvno = data_number2.cell(data_search, 2).value
            room_rates = data_number2.cell(data_search, 10).value
            nonight = data_number2.cell(data_search, 18).value
            adult = data_number2.cell(data_search, 19).value
            child = data_number2.cell(data_search, 20).value
            room_type = data_number2.cell(data_search, 21).value
            checkin = data_number2.cell(data_search, 22).value
            checkout = data_number2.cell(data_search, 23).value
            siteno = data_number2.cell(data_search, 28).value
            if str(resvno) == str(reservation):
                nonight1 = nonight
                adult1 = adult
                child1 = child
                site = siteno
                value_book = room_rates
                value_beds = room_type
                a = [str(y) for y in checkin]
                b = [str(x) for x in checkout]

                c1 = str(a[1]) + str(a[2])
                c2 = str(a[4]) + str(a[5])
                c3 = str(a[7]) + str(a[8]) + str(a[9]) + str(a[10])
                c4 = date(int(c3), int(c2), int(c1))
                d1 = str(b[1]) + str(b[2])
                d2 = str(b[4]) + str(b[5])
                d3 = str(b[7]) + str(b[8]) + str(b[9]) + str(b[10])
                d4 = date(int(d3), int(d2), int(d1))
                mod_checkin = c4
                mod_checkout = d4
                mod_nonight = nonight

                if mod_loops == 1:
                    try:
                        c4 = date(int(modify_dates[2]), int(modify_dates[0]), int(modify_dates[1]))
                        adult1, child1 = modify_dates[6], modify_dates[7]
                    except:
                        checkinmonth = datetime.datetime.strptime(modify_dates[0], '%b').month
                        #print('check in month ',checkinmonth)
                        c4 = date(int(modify_dates[2]), int(checkinmonth), int(modify_dates[1]))
                        modify_dates[8] = modify_dates[8].get()
                        modify_dates[9] = modify_dates[9].get()
                        site = choose_siteno[0]
                        adult1, child1 = modify_dates[6], modify_dates[7]

                    try:
                        d4 = date(int(modify_dates[5]), int(modify_dates[3]), int(modify_dates[4]))
                    except:
                        checkoutmonth = datetime.datetime.strptime(modify_dates[3], '%b').month
                        #print('check out month ', checkoutmonth)
                        d4 = date(int(modify_dates[5]), int(checkoutmonth), int(modify_dates[4]))

                    mod_checkin = c4
                    mod_checkout = d4
                    mod_entry_in = mod_checkout - mod_checkin
                    mod_nonight = mod_entry_in.days
                    for x in change_forget:
                        x.place_forget()
                    change_forget.clear()
                else:
                    modify_dates[0] = c4.strftime("%b")
                    modify_dates[1] = c4.strftime("%d")
                    modify_dates[2] = c4.strftime("%Y")
                    modify_dates[3] = d4.strftime("%b")
                    modify_dates[4] = d4.strftime("%d")
                    modify_dates[5] = d4.strftime("%Y")
                    modify_dates[6] = adult1
                    modify_dates[7] = child1
                    modify_dates[8] = value_book
                    modify_dates[9] = value_beds

        temp_mod_nonight = 0
        temp_mod_nonight = d4 - c4
        temp_mod_nonight = temp_mod_nonight.days

        if mod_nonight == 0:
            mod_nonight = 1
        mod_loops = 1
        #print('modify dates ',modify_dates)
        # today = date.today()
        mod_date = c4
        year = mod_date.strftime('%Y')
        month = mod_date.strftime('%m')
        days = mod_date.strftime('%d')
        current_dates = mod_date.strftime("%b %d, %Y")
        current_dates1 = mod_date.strftime("%B %d %Y")
        current_dates2 = mod_date.strftime("%d %m %Y")
        # get check in choose dates
        modify_dates[0] = mod_date.strftime("%b")
        modify_dates[1] = mod_date.strftime("%d")
        modify_dates[2] = mod_date.strftime("%Y")

        change_in_out_value[0] = current_dates1
        change_in_out_value[2] = current_dates2
        for y in month:
            month_checkin = y
        date_tom = 0
        last_date1 = 0
        #if change_loop[0] == 0:
        change_current_date[0] = mod_date
        #print('check out date ', mod_date, temp_mod_nonight)
        plus_adult1[0] = int(adult1)
        plus_child1[0] = int(child1)
        mod_adults, mod_childs = plus_adult1[0], plus_child1[0]

        # x = datetime.datetime.now()
        #plus_date[0] = int(nonight1)
        # print('days number ', plus_date[0] ,nonight1)
        date_tom = mod_date + datetime.timedelta(days=temp_mod_nonight)
        #date_tom = d4
        #print('date tomorrow ', date_tom)

        last_date1 = date_tom.strftime("%b %d, %Y")
        out_year = date_tom.strftime('%Y')
        out_month = date_tom.strftime('%m')
        out_days = date_tom.strftime('%d')
        for y in out_month:
            month_checkout = y
        last_date2 = date_tom.strftime("%B %d %Y")
        last_date3 = date_tom.strftime("%d %m %Y")
        change_in_out_value[1] = last_date2
        change_in_out_value[3] = last_date3
        # get check out choose dates
        modify_dates[3] = date_tom.strftime("%b")
        modify_dates[4] = date_tom.strftime("%d")
        modify_dates[5] = date_tom.strftime("%Y")
        '''modify_dates[6] = plus_adult1[0]
        modify_dates[7] = plus_child1[0]'''
        # print('check-in and check-out dates ', change_in_out_value[0], change_in_out_value[1])

        my_frame = Frame(window)
        canvas = Canvas(my_frame)

        canvas = Canvas(my_frame, width=373, height=440, background="lightgray")
        canvas.pack(fill="both", expand=True)

        vsb = Scrollbar(my_frame, orient="vertical", command=canvas.yview, width=25)
        hsb = Scrollbar(my_frame, orient="horizontal", command=canvas.xview, width=25)

        window.grid_rowconfigure(0, weight=1)
        window.grid_columnconfigure(0, weight=1)
        canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")

        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        resv_head = Label(canvas, text="   Reservation Details ", font=('arial', 15),
                          bg=levelcolor, foreground='blue', padx=0, pady=10)

        site_no = Label(canvas, text="   Site Number", font=('arial', 10),
                        bg='lightgray', foreground='blue', padx=0, pady=10, )

        site_entry_modify = Entry(canvas, font=('arial', 11), width=10)
        temp_mod_siteno = site_entry_modify
        #site_entry.insert(0, choose_entry[0])
        #choose_entry[0] = site_entry

        site_entry_button = Button(canvas, text='Select', font=('arial', 10), fg='black', bg="#F0F0F0",
                                   activeforeground="#00695C", activebackground="#F0F0F0", padx=15, pady=1, width=5,
                                   # wraplength= 1,
                                   # state='disable',
                                   command=mod_confirm_resv2)
        canvas.create_window(230, 73, anchor="nw", window=site_entry_button)
        # check in and check out
        checkin_design = Label(canvas, bg='#EEEEEE', padx=142, pady=20, highlightbackground="lightgray",
                               highlightthickness=1, border=0)
        checkin_dates = Label(canvas, text='check-in', font=('arial', 9), fg='grey', bg='#EEEEEE')
        checkin_dates1 = Label(canvas, text=str(current_dates), font=('arial', 10), fg='#4D4D4D', bg='#EEEEEE')

        checkout_dates = Label(canvas, text='check-out', font=('arial', 9), fg='grey', bg='#EEEEEE')
        checkin_design1 = Label(canvas, fg='#EEEEEE', bg='#EEEEEE', padx=0, pady=20, highlightbackground="lightgray",
                                highlightthickness=1, border=0)
        checkout_dates1 = Label(canvas, text=last_date1, font=('arial', 10), fg='#4D4D4D', bg='#EEEEEE')
        photo_calendar = Label(canvas, image=photo1, bg='#EEEEEE', height=24, width=24)
        photo_calendar1 = Label(canvas, image=photo1, bg='#EEEEEE', height=24, width=24)
        # nyt
        checkin_design_nyts = Label(canvas, bg='#EEEEEE', padx=142, pady=15, highlightbackground="lightgray",
                                    highlightthickness=1, border=0)
        checkin_nyts = Label(canvas, text='Nights', font=('arial', 9), fg='grey', bg='#EEEEEE')
        checkin_nytsno = Label(canvas, text=mod_nonight, font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')

        plus_nyts = Button(canvas, text='+', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                           activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                           command=plus_button)

        minus_nyts = Button(canvas, text='-', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                            activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                            command=minus_button)
        # room
        checkin_design_room = Label(canvas, bg='#EEEEEE', padx=142, pady=15, highlightbackground="lightgray",
                                    highlightthickness=1, border=0)
        rooms = Label(canvas, text='Rooms', font=('arial', 9), fg='grey', bg='#EEEEEE')
        rooms_no1 = Label(canvas, text=plus_room1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
        plus_room = Button(canvas, text='+', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                           activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                           command=room_button)

        minus_room = Button(canvas, text='-', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                            activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                            command=room_button1)

        # adult
        checkin_design_adult = Label(canvas, bg='#EEEEEE', padx=142, pady=15, highlightbackground="lightgray",
                                     highlightthickness=1, border=0)
        adults = Label(canvas, text='Adults', font=('arial', 9), fg='grey', bg='#EEEEEE')
        adults_no = Label(canvas, text=plus_adult1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
        plus_adults = Button(canvas, text='+', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                             activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                             command=room_adults)
        minus_adults = Button(canvas, text='-', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                              activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                              command=room_adults1)

        # child
        checkin_design_child = Label(canvas, bg='#EEEEEE', padx=142, pady=15, highlightbackground="lightgray",
                                     highlightthickness=1, border=0)
        child = Label(canvas, text='Child', font=('arial', 9), fg='grey', bg='#EEEEEE')
        child_no = Label(canvas, text=plus_child1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
        plus_child = Button(canvas, text='+', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                            activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                            command=room_child)
        minus_child = Button(canvas, text='-', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                             activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                             command=room_child1)

        # spacing
        k31 = Label(canvas, text='               ', font=('arial', 10), bg="lightgray", foreground='black', padx=0, pady=10)
        k32 = Label(canvas, text='               ', font=('arial', 10), bg="lightgray", foreground='black', padx=0, pady=10)

        canvas.create_window(3, 3, anchor="nw", window=resv_head)
        canvas.create_window(10, 70, anchor="nw", window=site_no)
        canvas.create_window(110, 77, anchor="nw", window=site_entry_modify)
        canvas.create_window(20, 120, anchor="nw", window=checkin_design)
        canvas.create_window(30, 127, anchor="nw", window=checkin_dates)
        canvas.create_window(30, 150, anchor="nw", window=checkin_dates1)
        canvas.create_window(168, 127, anchor="nw", window=checkout_dates)
        canvas.create_window(160, 120, anchor="nw", window=checkin_design1)
        canvas.create_window(168, 150, anchor="nw", window=checkout_dates1)
        canvas.create_window(122, 123, anchor="nw", window=photo_calendar)
        canvas.create_window(264, 123, anchor="nw", window=photo_calendar1)
        # plus minus nyts
        canvas.create_window(20, 185, anchor="nw", window=checkin_design_nyts)
        canvas.create_window(30, 190, anchor="nw", window=checkin_nyts)
        canvas.create_window(30, 208, anchor="nw", window=checkin_nytsno)
        canvas.create_window(210, 195, anchor="nw", window=minus_nyts)
        canvas.create_window(250, 195, anchor="nw", window=plus_nyts)
        # room
        canvas.create_window(20, 240, anchor="nw", window=checkin_design_room)
        canvas.create_window(30, 245, anchor="nw", window=rooms)
        canvas.create_window(30, 263, anchor="nw", window=rooms_no1)
        canvas.create_window(210, 250, anchor="nw", window=minus_room)
        canvas.create_window(250, 250, anchor="nw", window=plus_room)
        # adults
        canvas.create_window(20, 295, anchor="nw", window=checkin_design_adult)
        canvas.create_window(30, 300, anchor="nw", window=adults)
        canvas.create_window(30, 318, anchor="nw", window=adults_no)
        canvas.create_window(210, 305, anchor="nw", window=minus_adults)
        canvas.create_window(250, 305, anchor="nw", window=plus_adults)
        # child
        canvas.create_window(20, 355, anchor="nw", window=checkin_design_child)
        canvas.create_window(30, 360, anchor="nw", window=child)
        canvas.create_window(30, 377, anchor="nw", window=child_no)
        canvas.create_window(210, 365, anchor="nw", window=minus_child)
        canvas.create_window(250, 365, anchor="nw", window=plus_child)

        photo_calendar.bind("<Button-1>", modify_fetch_date)
        photo_calendar1.bind("<Button-1>", modify_fetch_date)

        my_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
        my_frame.place(x=870, y=176)
        resv_forget.append(my_frame)

        my_frame1 = Frame(window)
        canvas = Canvas(my_frame1)

        canvas = Canvas(my_frame1, width=625, height=380, background="lightgray")
        canvas.pack(fill="both", expand=True)

        vsb = Scrollbar(my_frame1, orient="vertical", command=canvas.yview, width=25)

        hsb = Scrollbar(my_frame1, orient="horizontal", command=canvas.xview, width=25)

        window.grid_rowconfigure(0, weight=1)
        window.grid_columnconfigure(0, weight=1)
        canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")

        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # value_resv_stay = resv_stay(label53.get())
        value_resv_stay = 26
        # print('reservation valueeeeeeeeeeeeeeeeeee ', value_resv_stay)
        # levelcolor = value_resv_stay[1]
        levelcolor = 'lightblue'
        property = Label(canvas, text='  Property Name  ', font=('arial', 15), bg=levelcolor, fg='blue', padx=0,
                         pady=10, )
        canvas.create_window(10, 5, anchor="nw", window=property)
        #print('modify confirm resv22222222222222222222 ', temp_mod_siteno2, temp_mod_siteno1)
        if temp_mod_siteno1 == 1:
            site = 0
            mod_siteno = temp_mod_siteno2
            site = mod_siteno
            temp_mod_siteno1 = 0
        else:
            mod_siteno = site

        for data_search in range(2, data_number3.max_row + 1):
            siteno = data_number3.cell(data_search, 1).value
            hotelname = data_number3.cell(data_search, 2).value
            address = data_number3.cell(data_search, 3).value
            phoneno = data_number3.cell(data_search, 4).value
            rating = data_number3.cell(data_search, 5).value

            rrodking = data_number3.cell(data_search, 8).value
            swriking = data_number3.cell(data_search, 10).value
            arrpking = data_number3.cell(data_search, 12).value
            fastking = data_number3.cell(data_search, 14).value

            gofastpoints = data_number3.cell(data_search, 16).value
            freeking = data_number3.cell(data_search, 19).value

            rrodqueens = data_number3.cell(data_search, 9).value
            swriqueens = data_number3.cell(data_search, 11).value
            arrpqueens = data_number3.cell(data_search, 13).value
            fastqueens = data_number3.cell(data_search, 15).value

            # print('data search ',data_search)

            if str(siteno) == str(site):
                z = 1
                x = 60
                temp_siteno = siteno
                count_rooms = mod_nonight

                prop = Label(canvas, text='  ' + '(' + str(siteno) + ')  ' + hotelname, font=('arial', 12),
                             fg='black',
                             bg='lightgray')
                canvas.create_window(10, x, anchor="nw", window=prop)

                prop1 = Label(canvas, text='   ' + phoneno + ',  ' + rating, font=('arial', 8), fg='blue',
                              bg='lightgray')
                canvas.create_window(10, x + 22, anchor="nw", window=prop1)

                prop2 = Label(canvas, text='   ' + address, font=('arial', 9), fg='black', bg='lightgray', )
                canvas.create_window(10, x + 42, anchor="nw", window=prop2)

                prop3 = Label(canvas, text='  ' + 'Rates', font=('arial', 11), fg='blue', bg='lightgray')
                canvas.create_window(10, x + 72, anchor="nw", window=prop3)

                prop4 = Label(canvas, text='One King', font=('arial', 11), fg='blue', bg='lightgray')
                canvas.create_window(150, x + 72, anchor="nw", window=prop4)

                prop5 = Label(canvas, text='Taxes', font=('arial', 11), fg='blue', bg="lightgray")
                canvas.create_window(300, x + 72, anchor="nw", window=prop5)

                prop6 = Label(canvas, text='Total', font=('arial', 11), fg='blue', bg="lightgray")
                canvas.create_window(420, x + 72, anchor="nw", window=prop6)

                prop7 = Label(canvas, text='Two Queens', font=('arial', 11), fg='blue', bg="lightgray")
                canvas.create_window(650, x + 72, anchor="nw", window=prop7)

                prop8 = Label(canvas, text='Taxes', font=('arial', 11), fg='blue', bg="lightgray")
                canvas.create_window(800, x + 72, anchor="nw", window=prop8)

                prop9 = Label(canvas, text='Total', font=('arial', 11), fg='blue', bg="lightgray")
                canvas.create_window(920, x + 72, anchor="nw", window=prop9)

                prop10 = Label(canvas, text='   rrod', font=('arial', 9), fg='blue', bg='lightgray')
                canvas.create_window(10, x + 100, anchor="nw", window=prop10)

                prop11 = Label(canvas, text='   swr1', font=('arial', 9), fg='blue', bg='lightgray')
                canvas.create_window(10, x + 120, anchor="nw", window=prop11)

                prop12 = Label(canvas, text='   aarp', font=('arial', 9), fg='blue', bg="lightgray")
                canvas.create_window(10, x + 140, anchor="nw", window=prop12)

                prop13 = Label(canvas, text='   go fast', font=('arial', 9), fg='blue', bg="lightgray")
                canvas.create_window(10, x + 160, anchor="nw", window=prop13)

                prop14 = Label(canvas, text='   go free', font=('arial', 9), fg='blue', bg="lightgray")
                canvas.create_window(10, x + 180, anchor="nw", window=prop14)

                # king rates
                prop15 = Label(canvas, text='$' + str(conversion_numbers1(rrodking)), font=('arial', 9), fg='black',
                               bg='lightgray')
                canvas.create_window(150, x + 100, anchor="nw", window=prop15)

                prop16 = Label(canvas, text='$' + str(conversion_numbers1(swriking)), font=('arial', 9), fg='black',
                               bg='lightgray')
                canvas.create_window(150, x + 120, anchor="nw", window=prop16)

                prop17 = Label(canvas, text='$' + str(conversion_numbers1(arrpking)), font=('arial', 9), fg='black',
                               bg="lightgray")
                canvas.create_window(150, x + 140, anchor="nw", window=prop17)

                prop18 = Label(canvas, text='$' + str(conversion_numbers1(fastking)), font=('arial', 9), fg='black',
                               bg="lightgray")
                canvas.create_window(150, x + 160, anchor="nw", window=prop18)

                prop19 = Label(canvas, text=str(conversion_numbers(freeking)) + ' pts', font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(150, x + 180, anchor="nw", window=prop19)

                # king taxes rate
                prop20 = Label(canvas, text='$' + str(conversion_numbers1(rrodking * .05)), font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(300, x + 100, anchor="nw", window=prop20)

                prop21 = Label(canvas, text='$' + str(conversion_numbers1(swriking * .05)), font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(300, x + 120, anchor="nw", window=prop21)

                prop22 = Label(canvas, text='$' + str(conversion_numbers1(arrpking * .05)), font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(300, x + 140, anchor="nw", window=prop22)

                prop23 = Label(canvas, text='$' + str(conversion_numbers1(fastking * .05)), font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(300, x + 160, anchor="nw", window=prop23)

                prop24 = Label(canvas, text='$' + str(conversion_numbers1(0)), font=('arial', 9), fg='black',
                               bg="lightgray")
                canvas.create_window(300, x + 180, anchor="nw", window=prop24)

                # king total rate
                prop25 = Label(canvas, text='$' + str(
                    conversion_numbers1((rrodking * count_rooms) + ((rrodking * count_rooms) * .05))),
                               font=('arial', 9), fg='black', bg='lightgray')
                canvas.create_window(420, x + 100, anchor="nw", window=prop25)

                prop26 = Label(canvas, text='$' + str(
                    conversion_numbers1((swriking * count_rooms) + ((swriking * count_rooms) * .05))),
                               font=('arial', 9), fg='black', bg='lightgray')
                canvas.create_window(420, x + 120, anchor="nw", window=prop26)

                prop27 = Label(canvas, text='$' + str(
                    conversion_numbers1((arrpking * count_rooms) + ((arrpking * count_rooms) * .05))),
                               font=('arial', 9), fg='black', bg="lightgray")
                canvas.create_window(420, x + 140, anchor="nw", window=prop27)

                prop28 = Label(canvas, text='$' + str(
                    conversion_numbers1((fastking * count_rooms) + ((fastking * count_rooms) * .05))) + ' + ' + str(
                    conversion_numbers((gofastpoints * count_rooms))) + ' pts', font=('arial', 9), fg='black',
                               bg="lightgray")
                canvas.create_window(420, x + 160, anchor="nw", window=prop28)

                prop29 = Label(canvas, text=str(conversion_numbers((freeking * count_rooms))) + ' pts',
                               font=('arial', 9), fg='black',
                               bg="lightgray")
                canvas.create_window(420, x + 180, anchor="nw", window=prop29)

                # two queens beds
                prop29 = Label(canvas, text='$' + str(conversion_numbers1(rrodqueens)),
                               font=('arial', 9), fg='black', bg='lightgray')
                canvas.create_window(650, x + 100, anchor="nw", window=prop29)

                prop30 = Label(canvas, text='$' + str(conversion_numbers1(swriqueens)), font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(650, x + 120, anchor="nw", window=prop30)

                prop31 = Label(canvas, text='$' + str(conversion_numbers1(arrpqueens)), font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(650, x + 140, anchor="nw", window=prop31)

                prop32 = Label(canvas, text='$' + str(conversion_numbers1(fastqueens)), font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(650, x + 160, anchor="nw", window=prop32)

                prop33 = Label(canvas, text=str(conversion_numbers(freeking)) + ' pts', font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(650, x + 180, anchor="nw", window=prop33)

                # two queens beds taxes
                prop34 = Label(canvas, text='$' + str(conversion_numbers1(rrodqueens * .05)), font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(800, x + 100, anchor="nw", window=prop34)

                prop35 = Label(canvas, text='$' + str(conversion_numbers1(swriqueens * .05)), font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(800, x + 120, anchor="nw", window=prop35)

                prop36 = Label(canvas, text='$' + str(conversion_numbers1(arrpqueens * .05)), font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(800, x + 140, anchor="nw", window=prop36)

                prop37 = Label(canvas, text='$' + str(conversion_numbers1(fastqueens * .05)), font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(800, x + 160, anchor="nw", window=prop37)

                prop38 = Label(canvas, text='$' + str(conversion_numbers1(0)), font=('arial', 9), fg='black',
                               bg="lightgray")
                canvas.create_window(800, x + 180, anchor="nw", window=prop38)

                # two queens beds total taxes
                prop39 = Label(canvas, text='$' + str(
                    conversion_numbers1((rrodqueens * count_rooms) + ((rrodqueens * count_rooms) * .05))),
                               font=('arial', 9), fg='black', bg='lightgray')
                canvas.create_window(920, x + 100, anchor="nw", window=prop39)

                prop40 = Label(canvas, text='$' + str(
                    conversion_numbers1((swriqueens * count_rooms) + ((swriqueens * count_rooms) * .05))),
                               font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(920, x + 120, anchor="nw", window=prop40)

                prop41 = Label(canvas, text='$' + str(
                    conversion_numbers1((arrpqueens * count_rooms) + ((arrpqueens * count_rooms) * .05))),
                               font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(920, x + 140, anchor="nw", window=prop41)

                prop42 = Label(canvas,
                               text='$' + str(conversion_numbers1(
                                   (fastqueens * count_rooms) + ((fastqueens * count_rooms) * .05))) + ' + ' + str(
                                   conversion_numbers((gofastpoints * count_rooms))) + ' pts', font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(920, x + 160, anchor="nw", window=prop42)

                prop43 = Label(canvas, text=str(conversion_numbers((freeking * count_rooms))) + ' pts',
                               font=('arial', 9), fg='black',
                               bg="lightgray")
                canvas.create_window(920, x + 180, anchor="nw", window=prop43)

                ratebookcombo = Label(canvas, text="  Rate Code", font=('arial', 10), bg='lightgray',
                                      foreground='blue')
                bookcombo = ttk.Combobox(canvas, font=('arial', 10),
                                         values=['Rates', 'RROD', 'AARP', 'SWR1', 'Go Fast', 'Go Free'],
                                         width=10, foreground='#263238')
                bookcombo.insert(0, modify_dates[8])
                modify_dates[8] = bookcombo

                ratebedscombo = Label(canvas, text="  Bed Class", font=('arial', 10), bg='lightgray',
                                      foreground='blue')
                bedscombo = ttk.Combobox(canvas, font=('arial', 10),
                                         values=['Bed Class', 'One Bed', 'Two Beds'],
                                         width=10, foreground='#263238')
                bedscombo.insert(0, modify_dates[9])
                modify_dates[9] = bedscombo

                canvas_nyts = Label(canvas, text="Night", font=('arial', 10), bg='lightgray', foreground='blue')
                canvas_nyts1 = Label(canvas, text=count_rooms, font=('arial', 9), background='white', padx=12,
                                     pady=1, border=1, relief='sunken')

                canvas.create_window(10, x + 240, anchor="nw", window=ratebookcombo)
                canvas.create_window(100, x + 240, anchor="nw", window=bookcombo)
                canvas.create_window(230, x + 240, anchor="nw", window=ratebedscombo)
                canvas.create_window(320, x + 240, anchor="nw", window=bedscombo)

                canvas.create_window(450, x + 240, anchor="nw", window=canvas_nyts)
                canvas.create_window(500, x + 240, anchor="nw", window=canvas_nyts1)

                spacing = Label(canvas, text=' ', font=('arial', 9), fg='black',
                                bg="lightgray")
                # canvas.create_window(1150, x + 330, anchor="nw", window=spacing)

        my_frame1.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
        my_frame1.place(x=200, y=176)
        resv_forget.append(my_frame1)

        mod_change_resv = Button(window, text='Change Property', font=('arial', 9), fg='black', bg='lightgray',
                             activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                             # wraplength= 1,
                             # state='disable',
                             command=mod_change_resv1)
        mod_change_resv.place(x=650, y=585)
        resv_forget.append(mod_change_resv)

        mod_confirm_resv = Button(window, text='Confirm Reservation', font=('arial', 9), fg='black', bg='lightgray',
                              activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                              # wraplength= 1,
                              # state='disable',
                              command=mod_confirm_resv1)
        mod_confirm_resv.place(x=650, y=615)
        resv_forget.append(mod_confirm_resv)

def mod_confirm_resv2():
    global temp_mod_siteno, temp_mod_siteno1, temp_mod_siteno2
    temp_mod_siteno2 = 0
    print('modify confirm resv2')
    loops = 0
    temp_mod_siteno.get()
    for data_search in range(2, data_number3.max_row + 1):
        siteno = data_number3.cell(data_search, 1).value
        if str(siteno) == str(temp_mod_siteno.get()):
            loops = 1
    if loops == 1:
        temp_mod_siteno2 = temp_mod_siteno.get()
        temp_mod_siteno1 = 1
        #print('modify confirm resv2 ', temp_mod_siteno2, temp_mod_siteno1)
        modify_stay()
    else:
        #print('Invalid Entry')
        pass


def mod_confirm_resv1():
    global get_resv_no, mod_siteno, mod_checkin, mod_checkout, mod_nonight, mod_loops,changedate, changedate1,\
        mod_rooms, mod_adults, mod_childs
    changedate, changedate1 = 0, 0
    print('modify confirm resv1 ')
    #print ('site number and reservation ',mod_siteno ,get_resv_no)
    #print('number of nights, room, adults, childs ',mod_nonight ,mod_rooms, mod_adults, mod_childs)
    modify_dates[6], modify_dates[7] = mod_adults, mod_childs
    '''for x in range (0 ,len(modify_dates)):
        if x == 8 or x == 9:
            print('modify date ', x, modify_dates[x].get())
        else:
            print('modify date ', x ,modify_dates[x])'''
    choose_modify()

def mod_change_resv1():
    print('modify change reservation 1')
    guest_next_modify()

def modify_fetch_date(event=None):
    print('modify fetch date')
    global mod_checkin, mod_checkout, mod_nonight,changedate, changedate1

    print('check in and check out dates ',mod_checkin , mod_checkout)
    diff_dates = 1
    # tkc calendar
    #x = change_current_date[0]
    x = mod_checkin
    print('xxxxxxxxxxxxxxxxxx ',x)
    today = date.today()
    years = eval(x.strftime('%Y'))
    month = x.strftime('%m')
    day = x.strftime('%d')
    current_dates = x.strftime("%B %d, %Y")
    date_tom = x + datetime.timedelta(days=0)

    years_1 = date_tom.strftime('%Y')
    months_1 = date_tom.strftime('%m')
    day_1 = date_tom.strftime('%d')

    years_10 = round(float(years_1))
    months10 = round(float(months_1))
    days10 = round(float(day_1))
    print('number of dayssssssssssssssssss ',days10)
    months = round(float(month))
    days = round(float(day))
    # tkc mindate and maxdate
    y = datetime.datetime.now()
    years9 = eval(y.strftime('%Y'))
    month = y.strftime('%m')
    day = y.strftime('%d')

    current_dates = x.strftime("%B %d, %Y")
    date_tom = y + datetime.timedelta(days=1)
    months9 = round(float(month))
    days9 = round(float(day))

    # print('booking check in ', 'year:', years, 'month:',months, 'days:', days)
    tkc = Calendar(window, selectmode="day", year=years, month=months, day=days, mindate=date(years9, months9, days9),
                   maxdate=date(years9 + 1, months9, days9))
    tkc.place(x=750, y=100)
    changedate= tkc
    change_forget.append(tkc)
    view_return.append(tkc)
    resv_forget.append(tkc)

    get_the_dates = mod_checkout
    date_tom1 = get_the_dates
    print('get the dates and date tomorrow ',mod_checkout ,mod_checkin)
    if mod_checkout == mod_checkin:
        date_tom1 = get_the_dates + datetime.timedelta(days=1)
    #date_tom1 = get_the_dates + datetime.timedelta(days=mod_nonight)


    out_year = eval(date_tom1.strftime('%Y'))
    out_month = date_tom1.strftime('%m')
    out_days = date_tom1.strftime('%d')
    out_month1 = round(float(out_month))
    out_days1 = round(float(out_days))

    tkc1 = Calendar(window, selectmode="day", year=out_year, month=out_month1, day=out_days1,
                    mindate=date(years_10, months10, days10), maxdate=date(years9 + 1, months9, days9))
    # tkc1 = Calendar(window, selectmode="day", year=out_year, month=out_month1, day=out_days1)
    tkc1.place(x=1000, y=100)
    changedate1 = tkc1
    change_forget.append(tkc1)
    view_return.append(tkc1)
    resv_forget.append(tkc1)

    but = Button(window, text="Select Date", command=modify_fetch_date1, bg="#EEEEEE", fg='#4D4D4D',
                 activeforeground='#4D4D4D', activebackground='#EEEEEE', padx=15, pady=2, width=5)
    but.place(x=1178, y=285)
    change_forget.append(but)
    view_return.append(but)
    resv_forget.append(but)


# choose fetch date1 / post stay and modify stay end of define funtions
def modify_fetch_date1(event=None):
    print('fetch date1')
    global changedate, changedate1, mod_loops
    #modify_fetch_date()
    mod_loops = 1
    '''site = choose_entry[0].get()
    choose_loopsite[0] = 0
    loops = 0
    for data_search in range(2, data_number2.max_row + 1):
        siteno = data_number2.cell(data_search, 28).value
        if str(siteno) == str(site):
            choose_loopsite[0] = 1
            loops = 1
    if loops == 0:
        choose_loopsite[0] = 0
        choose_entry[0] = ''
    else:
        choose_entry[0] = site

    change_plus_minus[0] = 0'''
    print('check out dates ', modify_dates)
    get_the_dates = changedate.selection_get()
    print('check in dates ', get_the_dates)
    #change_current_date[0] = get_the_dates
    get_the_dates1 = changedate1.selection_get()
    print('check out dates ', get_the_dates1)

    #checkinmonth = datetime.datetime.strptime(modify_dates[0], '%b').month
    #checkoutmonth = datetime.datetime.strptime(modify_dates[3], '%b').month
    modify_dates[0] = get_the_dates.strftime("%m")
    modify_dates[1] = get_the_dates.strftime("%d")
    modify_dates[2] = get_the_dates.strftime("%Y")
    modify_dates[3] = get_the_dates1.strftime("%m")
    modify_dates[4] = get_the_dates1.strftime("%d")
    modify_dates[5] = get_the_dates1.strftime("%Y")
    '''modify_dates[6] = modify_dates[6]
    modify_dates[7] = modify_dates[7]
    modify_dates[8] = modify_dates[8]
    modify_dates[9] = modify_dates[9]'''

    for x in range(0 ,len(modify_dates)):
        if x == 8:
            modify_dates[8] = modify_dates[x].get()
            print('modify dates ', modify_dates[8])
        elif x == 9:
            modify_dates[9] = modify_dates[x].get()
            print('modify dates ', modify_dates[9])
        else:
            print('modify dates ',modify_dates[x])

    modify_stay()
    #change_loop[0] = 1


def cancel_stay():
    stay_post1 = ''.join(stay_post)
    get_name1 = ''.join(get_name)
    number_mbr = "".join(get_name)
    print('Cancel Stay')
    #print('stay post1, get name1, number mbr ',stay_post1 ,get_name1 , number_mbr)
    for account_no in range(2, data_number2.max_row + 1):
        mbr_no = data_number2.cell(account_no, 1).value
        resv_no = data_number2.cell(account_no, 2).value
        rates = data_number2.cell(account_no, 10).value
        earn_pts = data_number2.cell(account_no, 14).value
        use_pts = data_number2.cell(account_no, 16).value
        check_in = data_number2.cell(account_no, 22).value
        check_out = data_number2.cell(account_no, 23).value
        clear_post = data_number2.cell(account_no, 25).value
        post_stay_points = data_number2.cell(account_no, 26).value
        post_date = data_number2.cell(account_no, 27).value
        point_post = data_number2.cell(account_no, 29).value

        #print('clear post ',clear_post)
        if str(resv_no) == str(stay_post1):
            if clear_post == 'yes':
                if post_stay_points == 'yes':
                    #print('Stay Already Posted')
                    '''postname = 'resv.# ' + str(resv_no) + '   stay posted   '
                    stay_posting1 = Label(window, text=postname, font=('arial', 12),
                                          background='white', padx=20, pady=6, border=2, relief='sunken')
                    stay_posting1.place(x=450, y=599, )
                    switching.append(stay_posting1)'''
                    messagebox.showinfo(title='Reservation '+ str(resv_no), message='Stay is Already Post')

                else:
                    #print('Ready for Posting')
                    messagebox.showinfo(title='Reservation ' + str(resv_no), message='Stay is Ready to Post')

            elif clear_post == 'no':
                #print('Stay is cancelled')
                dates1 = [x for x in check_in if x.isnumeric()]
                a = str(dates1[4]) + str(dates1[5]) + str(dates1[6]) + str(dates1[7])
                b = str(dates1[2]) + str(dates1[3])
                c = str(dates1[0]) + str(dates1[1])
                checkinn = date(int(a), int(b), int(c))

                cxldate = datetime.datetime.now()
                cxldate = cxldate.strftime('"%d/%m/%Y"')

                cxldate1 = date.today()
                total_no_days =  checkinn - cxldate1
                #print('today date ', total_no_days.days)
                if total_no_days.days < 0:
                    #print('cannot cxl')
                    messagebox.showinfo(title='Reservation ' + str(resv_no), message='Stay Cannot Cancel')

                else:
                    resv = Label(window, text="Cancelled ", font=('arial', 11), fg='white', bg='red')
                    resv.place(x=950, y=275)
                    switching.append(resv)

                    data_number2.cell(account_no, 25).value = 'cancel'
                    data_number2.cell(account_no, 27).value = cxldate
                    messagebox.showinfo(title='Reservation ' + str(resv_no), message='Stay is now Cancel')

            elif clear_post == 'cancel':
                #print('stay already cancel ')
                messagebox.showinfo(title='Reservation ' + str(resv_no), message='Stay is already Cancel')

            else:
                pass
    wb.save('data_wyn_gui.xlsx')

def post_stay():

    print('Post Stay')
    #print('reservation number post stay ', stay_post)
    #print('member number post stay', get_name)
    stay_post1 = ''.join(stay_post)
    get_name1 = ''.join(get_name)
    number_mbr = "".join(get_name)
    resv_stay1 = resv_stay(number_mbr)
    # print('reservation stay ', resv_stay1)
    level_mbr = resv_stay1[0]
    text_background = resv_stay1[1]
    add_points = 'no'
    get_earn_pts = 0
    total_use_pts = 0
    rows4 = 754
    columns4 = 270
    postdate1 = []
    postname =''
    min_no = []
    already_post_dates = 'no'
    for account_no in range(2, data_number2.max_row + 1):
        mbr_no = data_number2.cell(account_no, 1).value
        resv_no = data_number2.cell(account_no, 2).value
        rates = data_number2.cell(account_no, 10).value
        earn_pts = data_number2.cell(account_no, 14).value
        use_pts = data_number2.cell(account_no, 16).value
        check_out = data_number2.cell(account_no, 23).value
        clear_post = data_number2.cell(account_no, 25).value
        post_stay_points = data_number2.cell(account_no, 26).value
        post_date = data_number2.cell(account_no, 27).value
        point_post = data_number2.cell(account_no, 29).value

        #print('points post ',point_post)
        if str(get_name1) == str(mbr_no):
            if str(stay_post1) == str(resv_no):
                if clear_post == 'yes':
                    data_number2.cell(account_no, 26).value = 'yes'
                    clear_post_date = wyn_inactivity_GUI.last_date3(account_no, 23)
                    no_days = wyn_inactivity_GUI.days_inactive(clear_post_date)
                    posted_dates = wyn_inactivity_GUI.date_post_inactivity3(no_days)
                    for x in posted_dates:
                        if x != '"':
                            postdate1.append(x)
                    postdate1 = "".join(postdate1)
                    data_number2.cell(account_no, 27).value = postdate1
                    data_number2.cell(account_no, 27).value = '"' + str(postdate1) + '"'
                    #postname = 'resv.# ' + str(resv_no) + '   stay posted   '
                    if point_post == 'no':
                        messagebox.showinfo(title='Reservation ' + str(resv_no), message='Stay is now Post')

                    else:
                        messagebox.showinfo(title='Reservation ' + str(resv_no), message='Stay is already Post')

                elif clear_post == 'cancel':
                    messagebox.showinfo(title='Reservation ' + str(resv_no), message='Stay is already Cancel')

                else:
                    messagebox.showinfo(title='Reservation ' + str(resv_no), message='Stay is not ready to Post')

                if data_number2.cell(account_no, 26).value == 'yes':
                    if point_post == 'no':
                        add_points = 'yes'
                        if rates == 'Go Free':
                            total_use_pts = use_pts
                        elif rates == 'Go Fast':
                            total_use_pts = use_pts

                        get_earn_pts = earn_pts
                        point_post = data_number2.cell(account_no, 29).value = 'yes'
                        already_post_dates = 'yes'
                    else:
                        pass
                        #print('pts is already post')
                else:
                    pass
                    #print('not ready to post')

            columns4 += 30
            '''stay_posting1_label = Label(window, background='lightgray', padx=120, pady=10)
            stay_posting1_label.place(x=450, y=599, )
            switching.append(stay_posting1_label)

            stay_posting1 = Label(window,text=postname, font=('arial', 12),
                                  background='white', padx=20, pady=6, border=2, relief='sunken')
            stay_posting1.place(x=450, y=599,)
            switching.append(stay_posting1)'''

    print('\n')

    for account_no1 in range(2, data_number2.max_row + 1):
        mbr_no = data_number2.cell(account_no1, 1).value
        point_bal = data_number2.cell(account_no1, 8).value
        posted_dates= data_number2.cell(account_no1, 27).value
        if str(get_name1) == str(mbr_no):
            if posted_dates == 'ready':
                pass
            elif posted_dates == 'not ready':
                pass
            elif posted_dates == 'cancelled':
                pass
            elif posted_dates == None or '':
                pass
            else:
                #print('posted datesssssss', data_number2.cell(account_no1, 27).value)
                list_d = posted_dates.split('/')
                # print(list_d)
                b = 0
                y = []

                while b < len(list_d):
                    # if b == 0:
                    c = list_d[b]
                    for x in c:
                        if x.isnumeric() == True:
                            y.append(x)
                    b += 1
                b = 0
                # print('split y ', y)
                output = [int(x) for x in y]
                # print('split  ', output)

                a4 = 0
                while a4 < len(output):
                    if a4 <= 1:
                        a1 = int(str(output[0]) + str(output[1]))
                    elif a4 >= 1 and a4 <= 3:
                        a2 = int(str(output[2]) + str(output[3]))
                    elif a4 >= 3 and a4 <= 6:
                        a3 = int(str(output[4]) + str(output[5]) + str(output[6]) + str(output[7]))
                    a4 += 1

                today2 = date(a3, a2, a1)
                #print('convert post dates ',today2)
                date_post_number = wyn_inactivity_GUI.days_inactive(today2) + 5
                #print('date post number ', date_post_number)
                min_no.append(date_post_number)
                #print('minimium number ',min(min_no))
                recent_date = wyn_inactivity_GUI.date_post_inactivity4(min(min_no))
                date_expires = wyn_inactivity_GUI.date_post_inactivity4((min(min_no))-39)
                #recent_date = wyn_inactivity_GUI.date_post_inactivity4(date_post_number)
                #print('recent date ',recent_date , date_expires)
                recent_date1 = []
                for account_no in range(2, data_number.max_row + 1):
                    mbr_no1 = data_number.cell(account_no, 1).value
                    resv_no = data_number.cell(account_no, 2).value
                    date_expired = data_number.cell(account_no, 12).value
                    recent_stay = data_number.cell(account_no, 14).value

                    if str(get_name1) == str(mbr_no1):
                        if already_post_dates == 'yes':
                            data_number.cell(account_no, 14).value = recent_date
                            data_number.cell(account_no, 12).value = date_expires
                            #print('date expire ', date_expires)
                            #print('recent stay ',recent_date)

    resv_stay1 = resv_stay(number_mbr)
    # print('reservation stay ', resv_stay1)
    level_mbr = resv_stay1[0]
    text_background = resv_stay1[1]

    for account_no1 in range(2, data_number.max_row + 1):
        mbr_no = data_number.cell(account_no1, 1).value
        first_name = data_number.cell(account_no1, 4).value
        last_name = data_number.cell(account_no1, 3).value
        point_bal = data_number.cell(account_no1, 8).value
        active_acct = data_number.cell(account_no1, 9).value
        enroll_date = data_number.cell(account_no1, 13).value

        enroll = []
        for x in enroll_date:
            if x != '"':
                enroll.append(x)
        enroll = "".join(enroll)

        if str(get_name1) == str(mbr_no):
            if add_points == 'yes':

                labelstay2 = Label(window, background=text_background, padx=400, pady=15, border=1, relief=SOLID)
                labelstay2.place(x=200, y=176)
                switching.append(labelstay2)

                labelstay3 = Label(window, background=text_background, padx=265, pady=15, border=1, relief=SOLID)
                labelstay3.place(x=935, y=216)
                switching.append(labelstay3)

                label = Label(window, text="Reservation Details", font=('arial', 17), fg='blue', bg=text_background,
                              padx=5, pady=0)
                label.place(x=950, y=223)
                switching.append(label)

                label15 = Label(window, text="Reservation History", font=('arial', 17), fg='blue', bg=text_background,
                                padx=5, pady=0)
                label15.place(x=210, y=183)
                switching.append(label15)

                total_points = int(point_bal) + int(get_earn_pts * 10) - int(total_use_pts)
                data_number.cell(account_no1, 8).value = total_points
                points_bal = data_number.cell(account_no1, 8).value
                no_value = '                 '
                points_coma = conversion_numbers(points_bal)

                label9 = Label(window, background=text_background, padx=534, pady=30, bd=1, relief=SOLID)
                label9.place(x=200, y=85)

                label9 = Label(window, text='Name', font=('arial', 10, 'bold'), fg='black', bg=text_background)
                label9.place(x=220, y=86)

                label11 = Label(window, text='Member #', font=('arial', 10, 'bold'), fg='black', bg=text_background)
                label11.place(x=395, y=86)

                labelA1 = Label(window, text='Points Balance', font=('arial', 10, 'bold'), fg='black',
                                bg=text_background)
                labelA1.place(x=505, y=86)

                labelA3 = Label(window, text='Level', font=('arial', 10, 'bold'), fg='black', bg=text_background)
                labelA3.place(x=645, y=86)

                labelA5 = Label(window, text='Enroll Date', font=('arial', 10, 'bold'), fg='black', bg=text_background)
                labelA5.place(x=730, y=86)

                labelA7 = Label(window, text='Status', font=('arial', 10, 'bold'), fg='black', bg=text_background)
                labelA7.place(x=850, y=86)

                labelA9 = Label(window, text='Comments', font=('arial', 10, 'bold'), fg='black', bg=text_background)
                labelA9.place(x=950, y=86)

                labelA11 = Label(window, text='Pin #', font=('arial', 10, 'bold'), fg='black', bg=text_background)
                labelA11.place(x=1090, y=86)

                labelA13 = Label(window, text='Barclay', font=('arial', 10, 'bold'), fg='black', bg=text_background)
                labelA13.place(x=1180, y=86)

                label10 = Label(window, text=cap_name(first_name) + ' ' + cap_name(last_name), font=('arial', 11),
                                fg='black', bg=text_background)
                label10.place(x=220, y=105)

                label12 = Label(window, text=mbr_no, font=('arial', 11), fg='black', bg=text_background)
                label12.place(x=395, y=105)

                labelA4 = Label(window, text=no_value, font=('arial', 11), fg='black', bg=text_background)
                labelA4.place(x=645, y=105)
                labelA4 = Label(window, text=cap_name(level_mbr), font=('arial', 11), fg='black', bg=text_background)
                labelA4.place(x=645, y=105)

                labelA6 = Label(window, text= enroll, font=('arial', 11), fg='black', bg=text_background)
                labelA6.place(x=730, y=105)

                labelA8 = Label(window, text=cap_name(active_acct), font=('arial', 11), fg='black', bg=text_background)
                labelA8.place(x=850, y=105)

                labelA10 = Label(window, text='Most Recent', font=('arial', 11), fg='black', bg=text_background)
                labelA10.place(x=950, y=105)

                labelA12 = Label(window, text='None', font=('arial', 11), fg='black', bg=text_background)
                labelA12.place(x=1090, y=105)

                labelA14 = Label(window, text='Yes', font=('arial', 11), fg='black', bg=text_background)
                labelA14.place(x=1180, y=105)

                labelA2 = Label(window, text=no_value, font=('arial', 11), fg='black', bg=text_background)
                labelA2.place(x=505, y=105)
                labelA2 = Label(window, text=points_coma, font=('arial', 11), fg='black', bg=text_background)
                labelA2.place(x=505, y=105)
            else:
                pass

    wb.save('data_wyn_gui.xlsx')

def update():
    print('update')
    reactivate_acct1 = []
    deactivate_acct1 = []
    deactivate_sheet1 = []
    for x in add_find:
        if x == 'member':
            title = label16.get()
            # print('title ', title)

            gender = label17a.get()
            # print('gender', gender)

            firstname = label20.get()
            # print('First Name', firstname)
            middlename = label22.get()
            # print('Middle Name', middlename)
            lastname = label24.get()
            # print('Last Name', lastname)

            language = label26.get()
            # print('language ', language)
            months = label28.get()
            # print('months ', months)
            days = label29.get()
            # print('days ', days)
            primary_phone = label31a.get()
            # print('primary_phone ', primary_phone)
            phoneno = label35.get()
            # print('phoneno ', phoneno)
            homeno = label37.get()
            # print('homeno ', homeno)
            status = label40.get()
            #print('status ', status)
            emailaddress = email_entry.get()

            no_home = entry_homenumb.get()
            no_street = entry_homest.get()
            no_brgy = entry_homebarangay.get()
            no_city = entry_homecity.get()
            no_prov = entry_homeprov.get()
            no_country = entry_homecount.get()
            no_zipcode = entry_homezip.get()
            mbr_no = "".join(get_name)

            resv_stay1 = resv_stay(mbr_no)
            # print('reservation stay ',resv_stay(mbr_no))
            text_background = resv_stay1[1]
            # print('account number: ', mbr_no)
            for data_search1 in range(2, data_number.max_row + 1):
                account_no = data_number.cell(data_search1, 1).value
                pts_bal = data_number.cell(data_search1, 8).value
                current_status = data_number.cell(data_search1, 9).value
                forfeited_pts = data_number.cell(data_search1, 10).value
                reactivate_acct = data_number.cell(data_search1, 11).value
                deactivate_acct = data_number.cell(data_search1, 12).value

                reactivate_date = data_number.cell(data_search1, 30).value
                #sheet 2
                deactivate_sheet = data_number1.cell(data_search1, 6).value
                #print('statusssssssssssssssssssss ', deactivate_sheet)
                # first_name  = data_number.cell(data_search1, 4).value
                # last_name = data_number.cell(data_search1, 3).value
                if str(account_no) == str(mbr_no):
                    data_number.cell(data_search1, 2).value = phoneno
                    data_number.cell(data_search1, 3).value = lastname
                    data_number.cell(data_search1, 4).value = firstname
                    data_number.cell(data_search1, 5).value = middlename
                    data_number.cell(data_search1, 9).value = status
                    data_number.cell(data_search1, 15).value = primary_phone
                    data_number.cell(data_search1, 16).value = gender
                    data_number.cell(data_search1, 17).value = homeno
                    data_number.cell(data_search1, 18).value = title
                    data_number.cell(data_search1, 19).value = language
                    data_number.cell(data_search1, 21).value = months
                    data_number.cell(data_search1, 22).value = days
                    data_number.cell(data_search1, 29).value = emailaddress

                    data_number.cell(data_search1, 7).value = no_street
                    data_number.cell(data_search1, 23).value = no_home
                    data_number.cell(data_search1, 24).value = no_brgy
                    data_number.cell(data_search1, 25).value = no_city
                    data_number.cell(data_search1, 26).value = no_prov
                    data_number.cell(data_search1, 27).value = no_country
                    data_number.cell(data_search1, 28).value = no_zipcode

                    #data_number1.cell(data_search1, 6).value = status
                    if current_status == status:
                        #print('No Updates in Status')
                        pass

                    else:
                        if status == 'Active':
                            #print('account is active')
                            todays_date = datetime.datetime.now()
                            #print(todays_date)
                            todays_date1 = todays_date.strftime('"%d/%m/%Y"')
                            #print('Current Dates ', todays_date1)
                            #print('current status and new status ',current_status ,status)
                            data_number.cell(data_search1, 8).value = forfeited_pts + pts_bal
                            data_number.cell(data_search1, 10).value = 0

                            data_number.cell(data_search1, 11).value = todays_date1

                            reactivate_acct = data_number.cell(data_search1, 11).value

                            data_number.cell(data_search1, 30).value = todays_date1
                            reactivate_date = data_number.cell(data_search1, 30).value

                            date_post_inactivity1 = wyn_inactivity_GUI.date_post_inactivity(0)
                            # print('account number ', date_post_inactivity1)
                            data_number.cell(data_search1, 12).value = date_post_inactivity1
                            #print('Inactivity Date ', date_post_inactivity1)
                            deactivate_acct = date_post_inactivity1
                            #print('pts. bal ', data_number.cell(data_search1, 8).value )
                            #print('forfeited pts. ', data_number.cell(data_search1, 10).value)

                        elif status == 'Inactive':
                            #print('account is inactive')
                            todays_date = datetime.datetime.now()
                            #print(todays_date)
                            todays_date1 = todays_date.strftime('"%d/%m/%Y"')
                            #print('Current Dates ', todays_date1)
                            data_number.cell(data_search1, 10).value = pts_bal + forfeited_pts
                            data_number.cell(data_search1, 8).value = 0
                            data_number.cell(data_search1, 12).value = todays_date1
                            deactivate_acct = data_number.cell(data_search1, 12).value

                            data_number1.cell(data_search1, 6).value = todays_date1
                            deactivate_sheet = todays_date1
                            #print('pts. bal ', data_number.cell(data_search1, 8).value)
                            #print('forfeited pts. ', data_number.cell(data_search1, 10).value)

                        elif status == 'Closed':
                            pass
                            #print('account is closed')
                        else:
                            pass

                    if reactivate_acct != None:
                        for x in reactivate_acct:
                            if x != '"':
                                reactivate_acct1.append(x)
                        if reactivate_date == None:
                            reactivate_acct1 = None
                        else:
                            reactivate_acct1 = "".join(reactivate_acct1)
                    else:
                        #reactivate_acct1 = None
                        pass


                    if deactivate_acct != None:
                        for x in deactivate_acct:
                            if x != '"':
                                deactivate_acct1.append(x)
                        deactivate_acct1 = "".join(deactivate_acct1)
                    else:
                        #deactivate_acct1 = None
                        pass


                    if deactivate_sheet != None:
                        for x in deactivate_sheet:
                            if x != '"':
                                deactivate_sheet1.append(x)
                        deactivate_sheet1 = "".join(deactivate_sheet1)
                    else:
                        #deactivate_acct1 = None
                        pass


                    # print('new mbr name ', data_number.cell(data_search1, 4).value)
                    label10 = Label(window, text='                                     ', font=('arial', 11), fg='black',
                                    bg=text_background)
                    label10.place(x=220, y=105)
                    label10 = Label(window, text=cap_name(firstname) + ' ' + cap_name(lastname), font=('arial', 11), fg='black',
                                    bg=text_background)
                    label10.place(x=220, y=105)

                    labelA2 = Label(window, text='                  ', font=('arial', 11), fg='black', bg=text_background)
                    labelA2.place(x=505, y=105)
                    labelA2 = Label(window, text=conversion_numbers(data_number.cell(data_search1, 8).value), font=('arial', 11), fg='black', bg=text_background)
                    labelA2.place(x=505, y=105)

                    labelA8 = Label(window, text='             ', font=('arial', 11), fg='black', bg=text_background)
                    labelA8.place(x=850, y=105)
                    labelA8 = Label(window, text=cap_name(status), font=('arial', 11), fg='black', bg=text_background)
                    labelA8.place(x=850, y=105)

                    #actvie and inactive status
                    label42 = Label(window, text='                            ', font=('arial', 11), fg='black', bg='lightgray')
                    label42.place(x=1120, y=480)
                    label42 = Label(window, text=reactivate_acct1, font=('arial', 11), fg='black', bg='lightgray')
                    label42.place(x=1120, y=480)

                    label46a = Label(window, text='                            ', font=('arial', 11), fg='black',bg='lightgray')
                    label46a.place(x=872, y=520)
                    label46a = Label(window, text=deactivate_sheet1, font=('arial', 11), fg='black', bg='lightgray')
                    label46a.place(x=872, y=520)

                    label50a = Label(window, text='                  ', font=('arial', 11), fg='black',bg='lightgray')
                    label50a.place(x=1120, y=560)
                    label50a = Label(window, text=conversion_numbers(data_number.cell(data_search1, 8).value), font=('arial', 11), bg='lightgray')
                    label50a.place(x=1120, y=560)

                    label50b = Label(window, text='                       ', font=('arial', 11), fg='black',bg='lightgray')
                    label50b.place(x=900, y=600)
                    label50b = Label(window, text=conversion_numbers(data_number.cell(data_search1, 10).value), font=('arial', 11), bg='lightgray')
                    label50b.place(x=900, y=600)

                    label50c = Label(window, text='                  ', font=('arial', 11), fg='black',bg='lightgray')
                    label50c.place(x=1120, y=600)
                    label50c = Label(window, text=deactivate_acct1, font=('arial', 11), bg='lightgray')
                    label50c.place(x=1120, y=600)


                    # print('primary phone ',primary_phone.lower())
                    if primary_phone == 'Phone':
                        select = phoneno

                    elif primary_phone == 'Home':
                        select = homeno


            for data_search1 in range(2, data_number2.max_row + 1):
                account_no = data_number2.cell(data_search1, 1).value
                if str(account_no) == str(mbr_no):
                    data_number2.cell(data_search1, 3).value = select
                    data_number2.cell(data_search1, 5).value = firstname
                    data_number2.cell(data_search1, 4).value = lastname
                    data_number2.cell(data_search1, 6).value = middlename
                    # print('new name replace ', str(firstname) + ' ' + str(lastname))


    wb.save('data_wyn_gui.xlsx')
    # member_search()
    # new_update()

def find_member(mbr_no):

    member_search = label53.get()
    search = 0
    for data_search1 in range(2, data_number.max_row + 1):
        account_no = data_number2.cell(data_search1, 1).value
        if str(account_no) == str(member_search):
            search = 1
            # print('member search number ', member_search)

    if search == 0:
        member_search = None
        # print('member search number ', member_search)
    return member_search

def add_member():
    print('add member')
    y = 0
    '''for x in temp_addnewmbr:
        print(y ,x)
        y +=1'''

    for x in enroll_new1:
        x.place_forget()
    enroll_new1.clear()

    for x in enroll_new:
        x.place_forget()
    enroll_new.clear()

    for x in change_forget:
        x.place_forget()
    change_forget.clear()

    add_find.clear()
    add_find.append('add')
    # member no.
    label53.delete(0, END)
    # reservation number
    labelc50h.delete(0, END)

    for x in switching:
        x.place_forget()
    switching.clear()

    # add_forget.clear()
    label13 = Label(window, background='#F0F0F0', padx=535, pady=228)
    label13.place(x=200, y=170)
    enroll_new.append(label13)

    label13 = Label(window, background='lightgray', padx=300, pady=202, border=1, relief=RIDGE)
    label13.place(x=200, y=220)
    enroll_new.append(label13)

    label14 = Label(window, background='lightgray', padx=266, pady=225, border=1, relief=RIDGE)
    label14.place(x=735, y=175)
    enroll_new.append(label14)

    # username = label53.get()
    add_me = Label(window, text='Member Enrollment', font=('blue', 17), fg='blue', bg='#F0F0F0')
    add_me.place(x=200, y=188)
    enroll_new.append(add_me)

    label16a = Label(window, text="Title", font=('arial', 11),
                     bg='lightgray',
                     padx=0,
                     pady=10, )
    label16a.place(x=210, y=240)
    enroll_new.append(label16a)

    label16 = ttk.Combobox(window, font=('arial', 10), values=["Select", "Mr.", "Ms.", "Mrs."], width=10)
    label16.place(x=250, y=250)
    label16.insert(0, temp_addnewmbr[0])
    enroll_new1.append(label16)

    label17 = tkinter.Label(window, text="Gender", font=('arial', 11), bg='lightgray')
    label17.place(x=500, y=250)
    enroll_new.append(label17)

    label17a = ttk.Combobox(window, font=('arial', 10), values=["Select", "Male", "Female"], width=10)
    label17a.place(x=560, y=250)
    label17a.insert(0, temp_addnewmbr[1])
    enroll_new1.append(label17a)

    label19 = Label(window, text="First Name", font=('arial', 11),
                    bg='lightgray',
                    padx=0,
                    pady=10)
    label19.place(x=210, y=280)
    enroll_new.append(label19)

    label20 = Entry(window, font=('arial', 11))
    label20.place(x=210, y=315)
    label20.insert(0, temp_addnewmbr[2])
    enroll_new1.append(label20)

    label21 = Label(window, text="Middle Name", font=('arial', 11),
                    bg='lightgray',
                    padx=0,
                    pady=10)
    label21.place(x=450, y=280)
    enroll_new.append(label21)

    label22 = Entry(window, font=('arial', 11))
    label22.place(x=450, y=315)
    label22.insert(0, temp_addnewmbr[3])
    enroll_new1.append(label22)

    label23 = Label(window, text="Last Name", font=('arial', 11),
                    bg='lightgray',
                    padx=0,
                    pady=10)

    label23.place(x=210, y=340)
    enroll_new.append(label23)

    label24 = Entry(window, font=('arial', 11), width=35)
    # label24.insert(END,'last name')
    label24.place(x=210, y=375)
    label24.insert(0, temp_addnewmbr[4])
    enroll_new1.append(label24)


    label27 = Label(window, text="Birthday", font=('arial', 11),
                    bg='lightgray',
                    padx=0,
                    pady=10)
    label27.place(x=210, y=400)
    enroll_new.append(label27)

    label28 = ttk.Combobox(window, font=('arial', 10), values=["                 Month   ", "January", "February",
                                                               "March", "April", "May", "June",
                                                               "July", "August", "September", "October",
                                                               "November", "December"])
    label28.place(x=210, y=435)
    label28.insert(0, temp_addnewmbr[5])
    enroll_new1.append(label28)

    label29 = ttk.Combobox(window, font=('arial', 10),
                           values=["   day ", 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15,
                                   16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31],
                           width=5)
    label29.place(x=450, y=435)
    label29.insert(0, temp_addnewmbr[6])
    enroll_new1.append(label29)

    label30 = Label(window, text="Phone Number",
                    font=('arial', 17),
                    fg='blue',
                    bg='lightgray',
                    padx=5, pady=0)
    label30.place(x=205, y=480)
    enroll_new.append(label30)

    label31 = Label(window, text="Select Primary Phone", font=('arial', 11),
                    bg='lightgray',
                    padx=0,
                    pady=10)
    label31.place(x=210, y=515)
    enroll_new.append(label31)

    label31a = ttk.Combobox(window, font=('arial', 11), values=["Select", "Phone", "Home"], width=10)
    label31a.place(x=450, y=523)
    label31a.insert(0, temp_addnewmbr[7])
    enroll_new1.append(label31a)

    label34 = Label(window, text="Phone Number", font=('arial', 11),
                    bg='lightgray',
                    padx=0,
                    pady=10)
    label34.place(x=210, y=550)
    enroll_new.append(label34)

    label35 = Entry(window, font=('arial', 11), width=20)
    label35.place(x=210, y=585)
    label35.insert(0, temp_addnewmbr[8])
    enroll_new1.append(label35)

    label36 = Label(window, text="Home Number", font=('arial', 11),
                    bg='lightgray',
                    padx=0,
                    pady=10)
    label36.place(x=450, y=550)
    enroll_new.append(label36)

    label37 = Entry(window, font=('arial', 11), width=20)
    label37.place(x=450, y=585)
    label37.insert(0, temp_addnewmbr[9])
    enroll_new1.append(label37)

    homeadd = Label(window, text="Home Address",
                    font=('arial', 17),
                    fg='blue',
                    bg='lightgray',
                    padx=5, pady=0)
    homeadd.place(x=747, y=190)
    enroll_new.append(homeadd)

    homenumb = Label(window, text="Number", font=('arial', 11),
                     bg='lightgray',
                     padx=0,
                     pady=10, )
    homenumb.place(x=750, y=230)
    enroll_new.append(homenumb)

    entry_homenumb = Entry(window, font=('arial', 11), width=10)
    entry_homenumb.place(x=750, y=265)
    entry_homenumb.insert(0, temp_addnewmbr[10])
    enroll_new1.append(entry_homenumb)


    homest = Label(window, text="Street", font=('arial', 11),
                     bg='lightgray',
                     padx=0,
                     pady=10, )
    homest.place(x=870, y=230)
    enroll_new.append(homest)

    entry_homest = Entry(window, font=('arial', 11), width=20)
    entry_homest.place(x=870, y=264)
    entry_homest.insert(0, temp_addnewmbr[11])
    enroll_new1.append(entry_homest)

    homebarangay = Label(window, text="Barangay", font=('arial', 11),
                   bg='lightgray',
                   padx=0,
                   pady=10, )
    homebarangay.place(x=1065, y=230)
    enroll_new.append(homebarangay)

    entry_homebarangay = Entry(window, font=('arial', 11), width=20)
    entry_homebarangay.place(x=1065, y=263)
    entry_homebarangay.insert(0, temp_addnewmbr[12])
    enroll_new1.append(entry_homebarangay)


    homecity = Label(window, text="City", font=('arial', 11),
                   bg='lightgray',
                   padx=0,
                   pady=10, )
    homecity.place(x=750, y=290)
    enroll_new.append(homecity)

    entry_homecity = Entry(window, font=('arial', 11), width=25)
    entry_homecity.place(x=750, y=325)
    entry_homecity.insert(0, temp_addnewmbr[13])
    enroll_new1.append(entry_homecity)

    homeprov = Label(window, text="Province", font=('arial', 11),
                     bg='lightgray',
                     padx=0,
                     pady=10, )
    homeprov.place(x=1030, y=290)
    enroll_new.append(homeprov)

    entry_homeprov = Entry(window, font=('arial', 11), width=25)
    entry_homeprov.place(x=1030, y=325)
    entry_homeprov.insert(0, temp_addnewmbr[14])
    enroll_new1.append(entry_homeprov)

    homecount = Label(window, text="Country", font=('arial', 11),
                     bg='lightgray',
                     padx=0,
                     pady=10, )
    homecount.place(x=750, y=355)
    enroll_new.append(homecount)

    entry_homecount = Entry(window, font=('arial', 11), width=25)
    entry_homecount.place(x=750, y=390)
    entry_homecount.insert(0, temp_addnewmbr[15])
    enroll_new1.append(entry_homecount)

    homezip = Label(window, text="Country Code", font=('arial', 11),
                      bg='lightgray',
                      padx=0,
                      pady=10, )
    homezip.place(x=1030, y=355)
    enroll_new.append(homezip)

    entry_homezip = Entry(window, font=('arial', 11), width=15)
    entry_homezip.place(x=1030, y=390)
    entry_homezip.insert(0, temp_addnewmbr[16])
    enroll_new1.append(entry_homezip)


    emailadd = Label(window, text="Email Address",
                    font=('arial', 17),
                    fg='blue',
                    bg='lightgray',
                    padx=5, pady=0)
    emailadd.place(x=747, y=435)
    enroll_new.append(emailadd)

    emaillabeladd = Label(window, text="Email Address", font=('arial', 11),
                    bg='lightgray',
                    padx=0,
                    pady=10, )
    emaillabeladd.place(x=750, y=470)
    enroll_new.append(emaillabeladd)

    entry_emailadd = Entry(window, font=('arial', 11), width=30)
    entry_emailadd.place(x=870, y=478)
    entry_emailadd.insert(0, temp_addnewmbr[17])
    enroll_new1.append(entry_emailadd)


    homelanguage = Label(window, text="Language", font=('arial', 17),
                    fg='blue',
                    bg='lightgray',
                    padx=0,
                    pady=10)
    homelanguage.place(x=750, y=515)
    enroll_new.append(homelanguage)

    labelanguage = Label(window, text="Language", font=('arial', 11),
                    bg='lightgray',
                    padx=0,
                    pady=10)
    labelanguage.place(x=750, y=560)
    enroll_new.append(labelanguage)

    combolanguage = ttk.Combobox(window, font=('arial', 10), values=["Select", "English", "Filipino"])
    combolanguage.place(x=835, y=567)
    combolanguage.insert(0, temp_addnewmbr[18])
    enroll_new1.append(combolanguage)


    #add_member_input()
    addbutton = Button(window,
                    text="Add",
                    # command=click,
                    font=("arial", 11),
                    fg="black",
                    bg="lightgray",
                    activeforeground="black",
                    activebackground="lightgray",
                    padx=53,
                    pady=3,
                    width=6,
                    # state='disable',
                    command=add_member_input)
    addbutton.place(x=565, y=185)
    enroll_new.append(addbutton)

    username = 1
    member_search1(username)

def add_member_input():
    print('Add member input')
    # 0 to 11
    loops = 0
    search = 0
    search1  = 0
    count_none = 0
    count_none1 = 0
    value_enroll = []
    for x in range(0, len(enroll_new1)):
        value ,temp_addnewmbr[x] = enroll_new1[x].get(),enroll_new1[x].get()
        value_enroll.append(value)
        if value == '':
            pass
        else:
            loops += 1
            #print('value ',str(x),'=> ' ,enroll_new1[x].get())

    #print(type(value))
    list_member = []
    list_member2 = []
    pick_listno = 0
    loop_infor = 1
    for x in range(1,41):
        list_member2.append(x)
    for data_search in range(2, data_number.max_row + 1):
        member_no = data_number.cell(data_search, 1).value
        if member_no != None:
            list_member.append(member_no)
    list_member2 =  [x for x in list_member2 if x not in list_member]
    pick_listno = random.choice(list_member2)

    x = datetime.datetime.now()
    date_now = x.strftime('"%d/%m/%Y"')
    #print('print today ',date_now)
    #print('lenght of value enroll ', len(value_enroll))
    if loops == 19:
        for data_search in range(2, data_number.max_row + 1):
            member_no = data_number.cell(data_search, 1).value
            search += 1
            #print('member number ', data_search, member_no, search)
            if member_no == None:
                count_none +=1
                if count_none == 1:
                    data_number.cell(data_search, 1).value = int(pick_listno)
                    data_number.cell(data_search + 1, 18).value = value_enroll[0]
                    data_number.cell(data_search + 1, 16).value = value_enroll[1]
                    data_number.cell(data_search + 1, 4).value = value_enroll[2]
                    data_number.cell(data_search + 1, 5).value = value_enroll[3]
                    data_number.cell(data_search + 1, 3).value = value_enroll[4]
                    data_number.cell(data_search + 1, 21).value = value_enroll[5]
                    data_number.cell(data_search + 1, 22).value = value_enroll[6]
                    data_number.cell(data_search + 1, 15).value = value_enroll[7]
                    data_number.cell(data_search + 1, 2).value = value_enroll[8]
                    data_number.cell(data_search + 1, 17).value = value_enroll[9]
                    data_number.cell(data_search + 1, 23).value = value_enroll[10]
                    data_number.cell(data_search + 1, 7).value = value_enroll[11]
                    data_number.cell(data_search + 1, 24).value = value_enroll[12]
                    data_number.cell(data_search + 1, 25).value = value_enroll[13]
                    data_number.cell(data_search + 1, 26).value = value_enroll[14]
                    data_number.cell(data_search + 1, 27).value = value_enroll[15]
                    data_number.cell(data_search + 1, 28).value = value_enroll[16]
                    data_number.cell(data_search + 1, 29).value = value_enroll[17]
                    data_number.cell(data_search + 1, 19).value = value_enroll[18]
                    data_number.cell(data_search + 1, 8).value = 200000
                    data_number.cell(data_search + 1, 9).value = 'Active'
                    data_number.cell(data_search + 1, 10).value = 0
                    data_number.cell(data_search + 1, 13).value = date_now
                    data_number.cell(data_search + 1, 11).value = date_now

                    count_none1 = 1
                    #print('None')

        if member_no != None:
            if count_none1 == 0:
                data_number.cell(data_search + 1, 1).value = int(pick_listno)
                for x in range(0, len(value_enroll)):
                    data_number.cell(data_search + 1, 18).value = value_enroll[0]
                    data_number.cell(data_search + 1, 16).value = value_enroll[1]
                    data_number.cell(data_search + 1, 4).value = value_enroll[2]
                    data_number.cell(data_search + 1, 5).value = value_enroll[3]
                    data_number.cell(data_search + 1, 3).value = value_enroll[4]
                    data_number.cell(data_search + 1, 21).value = value_enroll[5]
                    data_number.cell(data_search + 1, 22).value = value_enroll[6]
                    data_number.cell(data_search + 1, 15).value = value_enroll[7]
                    data_number.cell(data_search + 1, 2).value = value_enroll[8]
                    data_number.cell(data_search + 1, 17).value = value_enroll[9]
                    data_number.cell(data_search + 1, 23).value = value_enroll[10]
                    data_number.cell(data_search + 1, 7).value = value_enroll[11]
                    data_number.cell(data_search + 1, 24).value = value_enroll[12]
                    data_number.cell(data_search + 1, 25).value = value_enroll[13]
                    data_number.cell(data_search + 1, 26).value = value_enroll[14]
                    data_number.cell(data_search + 1, 27).value = value_enroll[15]
                    data_number.cell(data_search + 1, 28).value = value_enroll[16]
                    data_number.cell(data_search + 1, 29).value = value_enroll[17]
                    data_number.cell(data_search + 1, 19).value = value_enroll[18]
                    data_number.cell(data_search + 1, 8).value = 200000
                    data_number.cell(data_search + 1, 9).value = 'Active'
                    data_number.cell(data_search + 1, 10).value = 0
                    data_number.cell(data_search + 1, 13).value = date_now
                    data_number.cell(data_search + 1, 11).value = date_now

    else:
        messagebox.showinfo(title='Member Information', message='Incomplete Information')
        loop_infor = 0
        add_member()

    #print('lllloooooopppsss ',loops)

    '''value 0 mr.
    value 1 male
    value 2 vhinz
    value 3 g
    value 4 diaz
    value 5 november
    value 6 29
    value 7 phonehome
    value 8 111
    value 9 222
    value 10 number
    value 11 street
    value 12 barangay
    value 13 city
    value 14 province
    value 15 country
    value 16 zip
    value 17 email
    value 18 english
    '''
    addbutton1 = Button(window,
                    text="Add",
                    # command=click,
                    font=("arial", 11),
                    fg="black",
                    bg="lightgray",
                    activeforeground="black",
                    activebackground="lightgray",
                    padx=53,
                    pady=3,
                    width=6,
                    #state='disable',
                    command=add_member_input)
    addbutton1.place(x=565, y=185)
    enroll_new.append(addbutton1)
    wb.save('data_wyn_gui.xlsx')

    if loop_infor == 1:
        addnew_count[0] = 1
        member_search()

# start of reservation // start of reservation // start of reservation
def resv_information():
    pass

def temp_modified(get_resvno):
    print('temp modifed')
    modified_nyts1 = 0
    modified_site1 = 0
    temp_book.clear()
    loop = 1
    modified_site1 = 0
    for data_search in range(2, data_number4.max_row + 1):
        accntno = data_number4.cell(data_search, 1).value
        modified_resv = data_number4.cell(data_search, 2).value
        hotel_name = data_number4.cell(data_search, 7).value
        modified_nyts = data_number4.cell(data_search, 18).value
        modified_site = data_number4.cell(data_search, 28).value
        checkin = data_number4.cell(data_search, 22).value
        checkout = data_number4.cell(data_search, 23).value
        ratecode = data_number4.cell(data_search, 10).value
        roomtype = data_number4.cell(data_search, 21).value
        adults = data_number4.cell(data_search, 19).value
        child = data_number4.cell(data_search, 20).value
        # print('modified ', modified_site)
        if str(modified_resv) == str(get_resvno):
            modified_site1 = modified_site
            # print('modified ',modified_site ,modified_resv ,str(get_resvno))
            getdate1 = list(checkin)
            getdate2 = []
            getdate = ['', '', '']

            outgetdate1 = list(checkout)
            outgetdate2 = []
            outgetdate = ['', '', '', '']

            z = 0
            z1 = 0
            for x in getdate1:
                if x == ',' or x == '-':
                    pass
                else:
                    getdate2.append(x)

            for x in getdate2:
                if x == ' ':
                    z += 1
                else:
                    getdate[z] += x

            for x in outgetdate1:
                if x == ',' or x == '-':
                    pass
                else:
                    outgetdate2.append(x)

            for x in range(0, len(outgetdate2)):
                if x == 0:
                    pass
                elif outgetdate2[x] == ' ':
                    z1 += 1
                else:
                    outgetdate[z1] += outgetdate2[x]

            inyear = outgetdate[3]
            longmonth = datetime.datetime.strptime(getdate[1], '%b').month
            indays = getdate[2]

            total_ins = datetime.datetime(int(inyear), int(longmonth), int(indays))
            total_ins1 = total_ins.strftime("%Y %B %d ")

            outyear = outgetdate[3]
            outlongmonth = datetime.datetime.strptime(outgetdate[1], '%b').month

            outdays = outgetdate[2]

            total_outs1 = datetime.datetime(int(outyear), int(outlongmonth), int(outdays))
            total_outs = total_outs1.strftime("%Y %B %d ")

            total_in = total_ins.strftime("%d")
            y = 0
            remove_int = [str(y) + str(x) for x in range(1, 10)]
            for x in range(0, len(remove_int)):
                if remove_int[x] == total_in:
                    total_in = x + 1

            total_out = total_outs1.strftime("%d")
            remove_int = [str(y) + str(x) for x in range(1, 10)]
            for x in range(0, len(remove_int)):
                if remove_int[x] == total_out:
                    total_out = x + 1

            back_bookrooms[0] = total_ins.strftime("%B")
            back_bookrooms[1] = total_in
            back_bookrooms[2] = total_outs1.strftime("%Y")
            back_bookrooms[3] = total_outs1.strftime("%B")
            back_bookrooms[4] = total_out
            back_bookrooms[5] = total_outs1.strftime("%Y")
            back_bookrooms[6] = adults
            back_bookrooms[7] = child
            back_resv_value[0] = get_resvno

    for data_search in range(2, data_number3.max_row + 1):
        resvno = data_number3.cell(data_search, 1).value
        if str(resvno) == str(modified_site1):
            # print('before site prop', site_prop)
            site_prop[0] = data_search - 1
            # print('after site entry ', site_scrool_entry)
            # print('after site prop ', site_prop, data_search - 1)

            site_scrools2()
            site_scrolls1()

def modified1():
    print('modified one')
    my_frame = Frame(window)
    canvas = Canvas(my_frame)

    canvas = Canvas(my_frame, width=373, height=440, background="lightgray")
    canvas.pack(fill="both", expand=True)

    vsb = Scrollbar(my_frame, orient="vertical", command=canvas.yview, width=25)
    hsb = Scrollbar(my_frame, orient="horizontal", command=canvas.xview, width=25)

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)
    canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
    canvas.grid(row=0, column=0, sticky="nsew")

    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    resv_head = Label(canvas, text="   Reservation Details ", font=('arial', 15),
                      bg='lightblue', foreground='blue', padx=0, pady=10, )

    site_no = Label(canvas, text="   Site Number", font=('arial', 10),
                    bg='lightgray', foreground='blue', padx=0, pady=10, )

    site_entry = Entry(canvas, font=('arial', 11), width=10)

    checkin = Label(canvas, text="   Check-In Date ", font=('arial', 10),
                    bg='lightgray', foreground='blue', padx=0, pady=10, )

    checkinmonths = ttk.Combobox(canvas, font=('arial', 10),
                                 values=['Month', 'January', 'February', 'March', 'April', 'May', 'June', 'July',
                                         'August', 'September', 'October', 'November', 'December'], width=12,
                                 foreground='#263238')
    checkinmonths.insert(0, 'Month')
    temp_book.append(checkinmonths)

    checkin_days = ttk.Combobox(canvas, font=('arial', 10),
                                values=['Days', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13',
                                        '14', '15',
                                        '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28',
                                        '29', '30', '31'],
                                width=10, foreground='#263238')

    checkin_days.insert(0, 'Days')
    temp_book.append(checkin_days)

    checkin_year = ttk.Combobox(canvas, font=('arial', 10), values=['Year', '2023', '2024', '2025', '2026', '2027'],
                                width=10, foreground='#263238')
    checkin_year.insert(0, 'Year')
    temp_book.append(checkin_year)

    checkout = Label(canvas, text="   Check-Out Date ", font=('arial', 10), bg='lightgray', foreground='blue', padx=0,
                     pady=10, )

    checkoutmonths = ttk.Combobox(canvas, font=('arial', 10),
                                  values=['Month', 'January', 'February', 'March', 'April', 'May', 'June', 'July',
                                          'August', 'September', 'October', 'November', 'December'], width=12,
                                  foreground='#263238')
    checkoutmonths.insert(0, 'Month')
    temp_book.append(checkoutmonths)

    checkout_days = ttk.Combobox(canvas, font=('arial', 10),
                                 values=['Days', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13',
                                         '14', '15',
                                         '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28',
                                         '29', '30', '31'],
                                 width=10, foreground='#263238')
    checkout_days.insert(0, 'Days')
    temp_book.append(checkout_days)

    checkout_year = ttk.Combobox(canvas, font=('arial', 10), values=['Year', '2023', '2024', '2025', '2026', '2027'],
                                 width=10, foreground='#263238')
    checkout_year.insert(0, 'Year')
    temp_book.append(checkout_year)

    adults = Label(canvas, text="   Adults ", font=('arial', 10), bg='lightgray', foreground='blue', padx=0, pady=10)

    adultscombo = ttk.Combobox(canvas, font=('arial', 10), values=['1', '2', '3', '4', '5', '6', '7', '8', '9'],
                               width=8, foreground='#263238')
    adultscombo.insert(0, '1')
    temp_book.append(adultscombo)

    children = Label(canvas, text="   Child", font=('arial', 10), bg='lightgray', foreground='blue', padx=0, pady=10)

    childrencombo = ttk.Combobox(canvas, font=('arial', 10), values=['0', '1', '2', '3', '4', '5', '6', '7', '8', '9'],
                                 width=8, foreground='#263238')
    childrencombo.insert(0, '0')
    temp_book.append(childrencombo)

    # spacing
    k31 = Label(canvas, text='               ', font=('arial', 10), bg="lightgray", foreground='black', padx=0, pady=10)
    k32 = Label(canvas, text='               ', font=('arial', 10), bg="lightgray", foreground='black', padx=0, pady=10)

    canvas.create_window(10, 50, anchor="nw", window=resv_head)
    # canvas.create_window(10, 105, anchor="nw", window=site_no)
    # canvas.create_window(110, 112, anchor="nw", window=site_entry)
    canvas.create_window(10, 140, anchor="nw", window=checkin)
    canvas.create_window(20, 185, anchor="nw", window=checkinmonths)
    canvas.create_window(150, 185, anchor="nw", window=checkin_days)
    canvas.create_window(270, 185, anchor="nw", window=checkin_year)
    canvas.create_window(10, 215, anchor="nw", window=checkout)
    canvas.create_window(20, 255, anchor="nw", window=checkoutmonths)
    canvas.create_window(150, 255, anchor="nw", window=checkout_days)
    canvas.create_window(270, 255, anchor="nw", window=checkout_year)
    canvas.create_window(10, 290, anchor="nw", window=adults)
    canvas.create_window(90, 297, anchor="nw", window=adultscombo)
    canvas.create_window(10, 330, anchor="nw", window=children)
    canvas.create_window(90, 337, anchor="nw", window=childrencombo)
    # canvas.create_window(10, 370, anchor="nw", window=rooms)
    # canvas.create_window(90, 377, anchor="nw", window=rooms_no)

    modified_date = Button(canvas, text="Modify", font=("arial", 10), fg="black", bg="#F0F0F0",
                           activeforeground="green", activebackground="#F0F0F0", border=2, relief=RAISED, padx=6,
                           pady=2,
                           width=7,
                           # wraplength= 1,
                           # state='disable',
                           command=temp_modified_dates1)
    canvas.create_window(160, 470, anchor="nw", window=modified_date)

    '''back_dates = Button(window, text='Backssssssssss', font=('arial', 10), fg='black', bg="#F0F0F0",
                            activeforeground="green", activebackground="#F0F0F0", padx=5, pady=2, width=5,
                            # wraplength= 1,
                            # state='disable',
                            command=back_dates_forget)
    back_dates.place(x=500, y=600)'''

    my_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
    my_frame.place(x=870, y=176)
    resv_forget.append(my_frame)


def back_dates_forget():
    pass
    # post()


def temp_modified_dates1():
    correct_date = 'True'
    print('temp modified dates ')
    # print('temp modified dates ', temp_resv)
    # print('count of temp book ', len(temp_book))
    for x in temp_book:
        pass
        # print(x.get())
    try:
        for x in range(0, len(temp_book)):
            ratebookcombo = temp_book[0].get()
            ratebedscombo = temp_book[1].get()
            checkinmonth1 = temp_book[2].get()
            checkindays = temp_book[3].get()
            checkinyear = temp_book[4].get()
            checkoutmonth1 = temp_book[5].get()
            checkoutdays = temp_book[6].get()
            checkoutyear = temp_book[7].get()
            adults = temp_book[8].get()
            childs = temp_book[9].get()

        # print('book and bed ',ratebookcombo ,ratebedscombo)
        # print('check in and check out month ',checkoutmonth1 ,checkoutmonth1)

        checkinmonth = datetime.datetime.strptime(checkinmonth1, '%B').month
        checkoutmonth = datetime.datetime.strptime(checkoutmonth1, '%B').month
        dateins = date(int(checkinyear), int(checkinmonth), int(checkindays))

        dateouts = date(int(checkoutyear), int(checkoutmonth), int(checkoutdays))
        bookrooms = dateouts - dateins
        bookrooms = bookrooms.days

        check_in11 = datetime.datetime(int(checkinyear), int(checkinmonth), int(checkindays))
        check_in12 = check_in11.strftime("%a, %b %d  -")

        # check out dates
        check_out11 = datetime.datetime(int(checkoutyear), int(checkoutmonth), int(checkoutdays))
        check_out12 = check_out11.strftime(" %a, %b %d %Y")

        if bookrooms == 0:
            bookrooms = 1
        else:
            pass

        # check in - toda date
        today = date.today()
        today_checkin = dateins - today
        # print('check in dates ', today_checkin.days)
        if ratebookcombo == 'Rates':
            correct_date = 'False'

        if ratebedscombo == 'Bed Class':
            correct_date = 'False'

        if today_checkin.days >= 0:
            # print('correct check in dates ', today_checkin.days)
            pass
        else:
            # print('incorrect check in dates ', today_checkin.days)
            correct_date = 'False'

        # check out date - check in date

        today_checkout = dateouts - dateins
        # print('check out dates ', today_checkout.days)
        if today_checkout.days <= 366:
            if today_checkout.days >= 0:
                # print('correct check out dates ', today_checkout.days)
                pass
            else:
                # print('incorrect check out dates ', today_checkout.days)
                correct_date = 'False'
        else:
            # print('incorrect check out dates ', today_checkout.days)
            correct_date = 'False'

        if correct_date == 'False':
            messagebox.showinfo(title='Modified Reservation', message='Invalid Entry')

        elif correct_date == 'True':
            for x in temp_resv:
                temp = x
            for data_search in range(2, data_number4.max_row + 1):
                resvno = data_number4.cell(data_search, 2).value
                update_site = data_number4.cell(data_search, 28).value
                if str(resvno) == str(temp):
                    new_site = update_site

            # print('reservation & site number ', temp , new_site)
            merge_value = str(ratebookcombo) + ' ' + str(ratebedscombo)
            value_free_fast, total_free_fast, room_rate, room_tax, total_rooom_rate, total_taxes, total_amount \
                = 0, 0, 0, 0, 0, 0, 0
            rate_name = ''
            for data_search1 in range(2, data_number3.max_row + 1):
                numsite = data_number3.cell(data_search1, 1).value
                rrodone = data_number3.cell(data_search1, 8).value
                rrodtwo = data_number3.cell(data_search1, 9).value
                swr1one = data_number3.cell(data_search1, 10).value
                swr1two = data_number3.cell(data_search1, 11).value
                arrpone = data_number3.cell(data_search1, 12).value
                arrptwo = data_number3.cell(data_search1, 13).value
                fastone = data_number3.cell(data_search1, 14).value
                fasttwo = data_number3.cell(data_search1, 15).value
                room_gofast = data_number3.cell(data_search1, 16).value
                room_gofree = data_number3.cell(data_search1, 19).value

                if numsite == new_site:
                    if ratebookcombo == 'Go Free':
                        value_free_fast = room_gofree
                        total_free_fast = value_free_fast * bookrooms
                        rate_name = 'Wyndham rewards go free rate'
                        # print('room go fast and go freeeeeeeeeeee ',value_free_fast ,total_free_fast )


                    elif ratebookcombo == 'Go Fast':
                        value_free_fast = room_gofast
                        total_free_fast = value_free_fast * bookrooms
                        rate_name = 'Wyndham rewards go fast rate'
                        # print('room go fasttttttttt and go free ', value_free_fast, total_free_fast)

                    if merge_value == 'RROD One Bed':
                        room_rate = rrodone
                        room_tax = room_rate * .05
                        total_rooom_rate = room_rate * bookrooms
                        total_taxes = room_tax * bookrooms
                        total_amount = total_rooom_rate + total_taxes
                        rate_name = 'Wyndham rewards flexible rate'
                        # print('site and title name value ',room_rate, room_tax, total_rooom_rate, total_taxes, total_amount)

                    elif merge_value == 'RROD Two Beds':
                        room_rate = rrodtwo
                        room_tax = room_rate * .05
                        total_rooom_rate = room_rate * bookrooms
                        total_taxes = room_tax * bookrooms
                        total_amount = total_rooom_rate + total_taxes
                        rate_name = 'Wyndham rewards flexible rate'
                        # print('site and title name value ', room_rate, room_tax, total_rooom_rate, total_taxes,total_amount)

                    elif merge_value == 'SWR1 One Bed':
                        room_rate = swr1one
                        room_tax = room_rate * .05
                        total_rooom_rate = room_rate * bookrooms
                        total_taxes = room_tax * bookrooms
                        total_amount = total_rooom_rate + total_taxes
                        rate_name = 'Wyndham rewards member rate'
                        # print('site and title name value ', room_rate, room_tax, total_rooom_rate, total_taxes,total_amount)

                    elif merge_value == 'SWR1 Two Beds':
                        room_rate = swr1two
                        room_tax = room_rate * .05
                        total_rooom_rate = room_rate * bookrooms
                        total_taxes = room_tax * bookrooms
                        total_amount = total_rooom_rate + total_taxes
                        rate_name = 'Wyndham rewards member rate'
                        # print('site and title name value ', room_rate, room_tax, total_rooom_rate, total_taxes,total_amount)

                    elif merge_value == 'AARP One Bed':
                        room_rate = arrpone
                        room_tax = room_rate * .05
                        total_rooom_rate = room_rate * bookrooms
                        total_taxes = room_tax * bookrooms
                        total_amount = total_rooom_rate + total_taxes
                        rate_name = 'Wyndham rewards aarp rate'
                        # print('site and title name value ', room_rate, room_tax, total_rooom_rate, total_taxes,total_amount)

                    elif merge_value == 'AARP Two Beds':
                        room_rate = arrptwo
                        room_tax = room_rate * .05
                        total_rooom_rate = room_rate * bookrooms
                        total_taxes = room_tax * bookrooms
                        total_amount = total_rooom_rate + total_taxes
                        rate_name = 'Wyndham rewards aarp rate'
                        # print('site and title name value ', room_rate, room_tax, total_rooom_rate, total_taxes,total_amount)

                    elif merge_value == 'Go Fast One Bed':
                        room_rate = fastone
                        room_tax = room_rate * .05
                        total_rooom_rate = room_rate * bookrooms
                        total_taxes = room_tax * bookrooms
                        total_amount = total_rooom_rate + total_taxes
                        # print('site and title name value ', room_rate, room_tax, total_rooom_rate, total_taxes,total_amount)

                    elif merge_value == 'Go Fast Two Beds':
                        room_rate = fasttwo
                        room_tax = room_rate * .05
                        total_rooom_rate = room_rate * bookrooms
                        total_taxes = room_tax * bookrooms
                        total_amount = total_rooom_rate + total_taxes
                        # print('site and title name value ', room_rate, room_tax, total_rooom_rate, total_taxes,total_amount)

            for data_search in range(2, data_number4.max_row + 1):
                modified_resv = data_number4.cell(data_search, 2).value
                modified_nyts = data_number4.cell(data_search, 18).value
                modified_site = data_number4.cell(data_search, 28).value
                if str(modified_resv) == str(temp):
                    # print('reservation number ',modified_resv)
                    data_number4.cell(data_search, 10).value = ratebookcombo
                    data_number4.cell(data_search, 11).value = room_rate
                    data_number4.cell(data_search, 12).value = room_tax
                    data_number4.cell(data_search, 13).value = value_free_fast
                    data_number4.cell(data_search, 14).value = total_rooom_rate
                    data_number4.cell(data_search, 15).value = total_taxes
                    data_number4.cell(data_search, 16).value = total_free_fast
                    data_number4.cell(data_search, 17).value = total_amount
                    data_number4.cell(data_search, 18).value = bookrooms
                    data_number4.cell(data_search, 19).value = adults
                    data_number4.cell(data_search, 20).value = childs
                    data_number4.cell(data_search, 21).value = ratebedscombo
                    data_number4.cell(data_search, 22).value = check_in12
                    data_number4.cell(data_search, 23).value = check_out12
                    data_number4.cell(data_search, 24).value = rate_name

            # wb.save('data_wyn_gui.xlsx')
            # post()
        return bookrooms, correct_date

    except ValueError:
        correct_date = 'False'
        if correct_date == 'False':
            messagebox.showinfo(title='Modified Reservation', message='Invalid Entry')
        # print('ValueError: invalid literal for int() with base 10: Year')


def post():
    # post1()
    print('Post')
    my_frame = Frame(window)
    canvas = Canvas(my_frame)

    canvas = Canvas(my_frame, width=385, height=440, background="lightgray")
    canvas.pack(fill="both", expand=True)

    vsb = Scrollbar(my_frame, orient="vertical", command=canvas.yview, width=25)
    hsb = Scrollbar(my_frame, orient="horizontal", command=canvas.xview, width=25)

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)
    canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
    canvas.grid(row=0, column=0, sticky="nsew")

    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    a = 0
    b = 0
    c = 1
    space_color = 0
    resv_head = Label(canvas, text="  Reservation Details  ", font=('arial', 15),
                      bg='lightblue', foreground='blue', padx=0, pady=10)
    canvas.create_window(10, 50, anchor="nw", window=resv_head)
    loops = 1
    data_value = 0
    for data_search in range(2, data_number4.max_row + 1):
        acctno = data_number4.cell(data_search, 1).value
        if acctno != None:
            loops += 1

    if data_number4.max_row == 1:
        data_value = data_number4.max_row

    elif data_number4.max_row == loops:
        data_value = data_number4.max_row

    elif data_number4.max_row != loops:
        data_value = loops

    temp_loop.clear()

    for data_search in range(1, data_value):
        acctno = data_number4.cell(data_value - b, 1).value
        resvno = data_number4.cell(data_value - b, 2).value
        siteno = data_number4.cell(data_value - b, 28).value
        hotelname1 = data_number4.cell(data_value - b, 7).value
        check_ins1 = data_number4.cell(data_value - b, 22).value
        check_out1 = data_number4.cell(data_value - b, 23).value
        bookadults = data_number4.cell(data_value - b, 19).value
        bookchild = data_number4.cell(data_value - b, 20).value
        bookfirst = data_number4.cell(data_value - b, 5).value
        booklast = data_number4.cell(data_value - b, 4).value
        bedsize = data_number4.cell(data_value - b, 21).value
        ratenames = data_number4.cell(data_value - b, 24).value
        bookrates = data_number4.cell(data_value - b, 10).value
        room_price = data_number4.cell(data_value - b, 11).value
        points = data_number4.cell(data_value - b, 13).value
        total_amount = data_number4.cell(data_value - b, 14).value
        taxes = data_number4.cell(data_value - b, 15).value
        total_points = data_number4.cell(data_value - b, 16).value
        payment = data_number4.cell(data_value - b, 17).value
        no_nyts = data_number4.cell(data_value - b, 18).value
        confirmation1 = data_number4.cell(data_value - b, 30).value

        if confirmation1 == 'confirmed':
            confirm_status = " Confirmed "
            cofirm_color = '#50EA49'
            space_color = 115
        else:
            confirm_status = " Cancelled "
            cofirm_color = 'red'
            space_color = 115

        # word of lenght greater than 34 hotel name
        if len(hotelname1) < 35:
            hotelname = hotelname1
        else:
            hotelname = ''
            word = hotelname1
            words = []
            word1 = ''
            loops = 0

            for x in range(0, len(word)):
                word1 += word[x]
                if word[x] == ' ':
                    words.append(word1)
                    word1 = ''
            words.append(word1)

            total = len(word)
            loops = 0
            total1 = 0
            while loops == 0:
                if total > 34:
                    total1 += len(words[-1])
                    total = len(word) - total1
                    words.remove(words[-1])
                else:
                    loops = 1

            temp = words[-1]
            total1 = ''
            for x in temp:
                if x == ' ':
                    pass
                else:
                    total1 += x

            words[-1] = str(total1) + '...'
            hotelname = ' '.join(words)
        if str(choose_resv[0]) == str(resvno):

            line_frame = Label(canvas, text=data_search, foreground='lightgray', bg='lightgray', padx=185, pady=187,
                               border=2, relief="solid")
            canvas.create_window(15, 105 + a, anchor="nw", window=line_frame)


        else:
            line_frame = Label(canvas, text=data_search, foreground='lightgray', bg='lightgray', padx=185, pady=187,
                               border=1, relief="solid")
            canvas.create_window(15, 105 + a, anchor="nw", window=line_frame)

        confirm = Label(canvas, text=confirm_status, font=('arial', 11), fg='white', bg=cofirm_color)
        resv1 = Label(canvas, text=resvno, font=('arial', 11),
                      fg='black',
                      bg='lightgray')
        hotels = Label(canvas, text='(' + str(siteno) + ')' + '  ' + str(hotelname),
                       font=('arial', 11, 'bold'), fg='black', bg='lightgray')
        dates = Label(canvas, text=str(check_ins1) + str(check_out1), font=('arial', 10, 'bold'),
                      fg='black', bg='lightgray')
        adults = Label(canvas, text=str(bookadults) + ' adults ' + str(bookchild) + ' child',
                       font=('arial', 10), fg='black', bg='lightgray')

        fullname = Label(canvas, text=str(cap_name(bookfirst)) + ' ' + str(cap_name(booklast)),
                         font=('arial', 10), fg='blue', bg='lightgray')

        rooms1 = Label(canvas, text=bedsize, font=('arial', 10), fg='black', bg='lightgray')
        rate_rooms = Label(canvas, text=str(ratenames) + '  ' + str(bookrates), font=('arial', 10),
                           fg='blue', bg='lightgray')

        if bookrates == 'Go Free':
            room_payment = Label(canvas, text=str(points) + ' pts ' + 'avg / night ', font=('arial', 10),
                                 fg='black', bg='lightgray')
        elif bookrates == 'Go Fast':
            room_payment = Label(canvas, text=str(points) + ' pts + $' + str(room_price) + ' avg / night ',
                                 font=('arial', 10),
                                 fg='black', bg='lightgray')
        else:
            room_payment = Label(canvas, text='$' + str(room_price) + ' avg / night ', font=('arial', 10),
                                 fg='black', bg='lightgray')

        if bookrates == 'Go Free':
            nyts_payment = Label(canvas, text='1 room ' + str(no_nyts) + ' night',
                                 font=('arial', 10), fg='black', bg='lightgray')
            nyts_payment1 = Label(canvas, text=str(total_points) + ' pts', font=('arial', 10),
                                  fg='black',
                                  bg='lightgray')
            tax_fees = Label(canvas, text=str('taxes and fees '), font=('arial', 10), fg='black',
                             bg='lightgray')
            tax_fees1 = Label(canvas, text='USD ' + str(taxes), font=('arial', 10), fg='black',
                              bg='lightgray')
            tax_total = Label(canvas, text="tax and total payment", font=('arial', 10), fg='black',
                              bg='lightgray')
            tax_total1 = Label(canvas, text=str(total_points) + ' pts', font=('arial', 10), fg='black',
                               bg='lightgray')


        elif bookrates == 'Go Fast':
            nyts_payment = Label(canvas, text='1 room ' + str(no_nyts) + ' night',
                                 font=('arial', 10), fg='black', bg='lightgray')
            nyts_payment1 = Label(canvas, text='USD ' + str(total_amount) + ' + ' + str(total_points) + ' pts',
                                  font=('arial', 10),
                                  fg='black',
                                  bg='lightgray')
            tax_fees = Label(canvas, text=str('taxes and fees '), font=('arial', 10), fg='black',
                             bg='lightgray')
            tax_fees1 = Label(canvas, text='USD ' + str(taxes), font=('arial', 10), fg='black',
                              bg='lightgray')
            tax_total = Label(canvas, text="tax and total payment", font=('arial', 10), fg='black',
                              bg='lightgray')
            tax_total1 = Label(canvas, text='USD ' + str(payment) + ' + ' + str(total_points) + ' pts',
                               font=('arial', 10),
                               fg='black',
                               bg='lightgray')


        else:
            nyts_payment = Label(canvas, text='1 room ' + str(no_nyts) + ' night',
                                 font=('arial', 10), fg='black', bg='lightgray')
            nyts_payment1 = Label(canvas, text='USD ' + str(total_amount), font=('arial', 10), fg='black',
                                  bg='lightgray')
            tax_fees = Label(canvas, text=str('taxes and fees '), font=('arial', 10), fg='black',
                             bg='lightgray')
            tax_fees1 = Label(canvas, text='USD ' + str(taxes), font=('arial', 10), fg='black', bg='lightgray')
            tax_total = Label(canvas, text="tax and total payment", font=('arial', 10), fg='black',
                              bg='lightgray')
            tax_total1 = Label(canvas, text='USD ' + str(payment), font=('arial', 10), fg='black', bg='lightgray')

        space = Label(canvas, text='', font=('arial', 10), fg='black', bg='lightgray')
        canvas.create_window(23, 115 + a, anchor="nw", window=confirm)
        canvas.create_window(space_color, 115 + a, anchor="nw", window=resv1)
        canvas.create_window(23, 143 + a, anchor="nw", window=hotels)
        canvas.create_window(23, 168 + a, anchor="nw", window=dates)
        canvas.create_window(23, 193 + a, anchor="nw", window=adults)
        canvas.create_window(23, 218 + a, anchor="nw", window=fullname)
        canvas.create_window(23, 243 + a, anchor="nw", window=rooms1)
        canvas.create_window(23, 268 + a, anchor="nw", window=rate_rooms)
        canvas.create_window(23, 293 + a, anchor="nw", window=room_payment)
        canvas.create_window(23, 348 + a, anchor="nw", window=nyts_payment)
        canvas.create_window(190, 348 + a, anchor="nw", window=nyts_payment1)
        canvas.create_window(23, 378 + a, anchor="nw", window=tax_fees)
        canvas.create_window(190, 378 + a, anchor="nw", window=tax_fees1)
        canvas.create_window(23, 408 + a, anchor="nw", window=tax_total)
        canvas.create_window(190, 408 + a, anchor="nw", window=tax_total1)

        def select(event=None, number=data_search - 2, account_no=acctno, rownos=data_number4.max_row - c):
            choose_entry[0] = ''
            choose_loopsite[0] = 0
            #print('reservation choose number ', number ,account_no ,rownos ,data_number4.max_row)
            #print('reservation choose number ', data_number4.cell(data_number4.max_row, 2).value)
            resvno = data_number4.cell(data_number4.max_row, 2).value
            resvno1 = int(resvno) - int(number)
            choose_resv[0] = int(resvno1) - 1
            #print('reservation choose number ', int(resvno), resvno1 ,choose_resv[0])
            choose_acctno[0] = account_no
            choose_siteno[0] = data_number4.cell(rownos + 1, 28).value
            # print('reservation choose number ',choose_siteno[0])
            choose_count[0] = 1
            post()
            guest_search3()
            # temp_modified(resvno1)

        def select1(number=data_search - 1):
            resvno = data_number4.cell(data_number4.max_row, 2).value
            # print('data number max row:> ', data_number4.max_row)
            # print('clicked resrvation number:> ',resvno - number)
            resvno1 = resvno - number
            # temp_modified_dates()
            # print('reservation number ',resvno1)
            cxl_resv(resvno1)

        line_frame.bind("<Button-1>", select)

        a += 430
        b += 1
        c += 1
        my_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
        my_frame.place(x=863, y=176)
        view_return.append(my_frame)
        resv_forget.append(my_frame)

        next_button_next = Label(window, text='                                  ', bg='#F0F0F0',
                                 activebackground="lightgray", padx=85, pady=4, width=2, )
        next_button_next.place(x=210, y=585)
        resv_forget.append(next_button_next)

        change_info = Button(window, text='Back', font=('arial', 9), fg='black', bg='lightgray',
                             activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                             # wraplength= 1,
                             # state='disable',
                             command=return_view)
        change_info.place(x=210, y=615)
        view_return.append(change_info)
        resv_forget.append(change_info)


def view_reservation1():
    pass


def add_reservation():
    # remove button cxl resv
    cancel_resv = Label(window, font=('arial', 10), bg="#F0F0F0", activebackground="#F0F0F0", padx=17, pady=5, width=6)
    cancel_resv.place(x=530, y=600)
    resv_forget.append(cancel_resv)
    site_scrools2()
    site_scrolls()


def wrap_up():
    print('Wrap up')
    if data_number4.max_row == 1:
        pass
    else:
        answer = messagebox.askquestion(title='New Reservation', message='Wrap-up Stay')
        if answer == 'yes':
            print('Wrap-up', answer)
            for data_search in range(2, data_number4.max_row + 1):
                resvno = data_number4.cell(data_search, 2).value
                data_number4.delete_rows(2)
                wb.save('data_wyn_gui.xlsx')
        else:
            print('Not Wrap-up', answer)
            pass
    reservation()


def view_reservation():
    value = 0
    print('view reservation')
    change_loop[0] = 0

    if data_number4.cell(2, 2).value == None:
        value = 1

    if data_number4.max_row > 1 and value == 0:
        post()

    for x in view_return1:
        x.place_forget()
    view_return1.clear()

    view_resv = Button(window, text='View Reservation', font=('arial', 9), fg='black', bg='lightgray',
                       activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                       # wraplength= 1,
                       # state='disable',
                       command=return_view)
    view_resv.place(x=650, y=615)
    resv_forget.append(view_resv)

def return_view():
    print('return view')
    for x in view_return:
        x.place_forget()
    view_return.clear()
    print('shift reservation ', change_shift_resv[0])
    # wrap_patch = Label(window, text='           ', font=('arial', 11), bg='black', padx=20, pady=5)
    next_button_back = Label(window, text='                                  ', bg='#F0F0F0',
                             activebackground="lightgray", padx=90, pady=4, width=2, )
    next_button_back.place(x=635, y=585)
    resv_forget.append(next_button_back)

    view_resv = Button(window, text='View Reservation', font=('arial', 9), fg='black', bg='lightgray',
                       activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                       # wraplength= 1,
                       # state='disable',
                       command=view_reservation)
    view_resv.place(x=650, y=615)
    resv_forget.append(view_resv)

    '''guest_search2_dates = Button(window, text='View', font=('arial', 10), fg='black', bg='lightgray',
                                 activeforeground="#00695C", activebackground="lightgray", padx=15, pady=2, width=5,
                                 # wraplength= 1,
                                 # state='disable',
                                 command=view_reservation)
    guest_search2_dates.place(x=740, y=600)'''

    if change_shift_resv[0] == 1:
        guest_search2()
    elif change_shift_resv[1] == 1:
        post1()
    elif change_shift_resv[2] == 1:
        next_button = Button(window, text='Next', font=('arial', 9), fg='black', bg='lightgray',
                             activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                             # wraplength= 1,
                             # state='disable',
                             command=confirmed_dates2)
        next_button.place(x=210, y=585)
        resv_forget.append(next_button)
        # site_scrool1_forget.append(next_button)


def cxl_resv():
    account = choose_acctno[0]
    reservation = choose_resv[0]
    print('cancellation of reservation ')
    for data_search in range(2, data_number4.max_row + 1):
        resvno = data_number4.cell(data_search, 2).value
        if str(resvno) == str(reservation):
            data_number4.cell(data_search, 30).value = 'cancel'
            data_number2.cell(int(resvno) + 1, 25).value = 'cancel'
            wb.save('data_wyn_gui.xlsx')
            post()


def post1():
    print('post 1')
    change_shift_resv[0] = 0
    change_shift_resv[1] = 1
    change_shift_resv[2] == 0
    value = 1
    if view_resvno[0] == 0:
        pass
    else:
        resvno = view_resvno[0].get()
        for data_search in range(2, data_number.max_row + 1):
            accntno = data_number.cell(data_search, 1).value
            if str(accntno) == str(resvno):
                value = 0
                guest_search1()

    if value == 1:
        points = view_post1_payment[0]
        room_price = view_post1_payment[1]
        total_points = view_post1_payment[2]
        total_roomprice = view_post1_payment[3]
        if str(points) != '0' and str(room_price) != '0.00':
            compute_room_price = str(points) + ' pts + ' + str(room_price)
            compute_total_room_price = str(total_points) + ' pts + ' + str(total_roomprice)

        elif str(points) == '0' and str(room_price) != '0.00':
            compute_room_price = room_price
            compute_total_room_price = total_roomprice

        elif str(points) != '0' and str(room_price) == '0.00':
            compute_room_price = str(points) + ' pts'
            compute_total_room_price = total_points + ' pts'

        my_frame1 = Frame(window)
        canvas = Canvas(my_frame1)

        canvas = Canvas(my_frame1, width=625, height=380, background="lightgray")
        canvas.pack(fill="both", expand=True)

        vsb = Scrollbar(my_frame1, orient="vertical", command=canvas.yview, width=25)

        hsb = Scrollbar(my_frame1, orient="horizontal", command=canvas.xview, width=25)

        window.grid_rowconfigure(0, weight=1)
        window.grid_columnconfigure(0, weight=1)
        canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")

        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        property = Label(canvas, text='  Add Reservation Details  ', font=('arial', 15), bg='lightgray', fg='black',
                         padx=0,
                         pady=10, )
        canvas.create_window(0, 3, anchor="nw", window=property)
        y = 60

        for x in range(7):
            design_bg = 'm' + str(x)
            design_bg = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=10)
            canvas.create_window(0, y, anchor="nw", window=design_bg)
            if x == 2:
                design_bg = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=42)
                canvas.create_window(0, y, anchor="nw", window=design_bg)
            if x == 6:
                design_bg = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=14)
                canvas.create_window(0, y, anchor="nw", window=design_bg)
            y += 60
        design_bg8 = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=10)
        canvas.create_window(0, 485, anchor="nw", window=design_bg8)

        guest_required = Label(canvas, text='Guest', font=('arial', 11), fg='black')
        guest_require = Label(canvas, text='Required', font=('arial', 11), fg='#FFB300')
        guest_search = Button(canvas, text='Enter Member', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                              activeforeground='#00695C', activebackground='#EEEEEE', padx=5, pady=3, width=10,
                              command=guest_search1)
        guest_entry_search = Entry(canvas, font=('arial', 11), width=12)
        view_resvno[0] = guest_entry_search

        guest_agent = Label(canvas, text='Travel Agent or Booker', font=('arial', 11), fg='black')
        guest_agent1 = Button(canvas, text='Find', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                              activeforeground='#00695C', activebackground='#EEEEEE', padx=5, pady=3, width=10,
                              command=guest_agent11)
        guest_agent2 = Button(canvas, text='Add New', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                              activeforeground='#00695C', activebackground='#EEEEEE', padx=5, pady=3, width=10,
                              command=guest_agent12)

        guest_details = Label(canvas, text='Details', font=('arial', 11), fg='black')
        guest_details1 = Label(canvas,
                               text='Request to Hotel, Notification to Guest / Booker, Transportation, Marketing Info, Source',
                               font=('arial', 8), fg='black')
        guest_details2 = Label(canvas, font=('arial', 1), bg='darkgray', padx=303, pady=1)
        guest_details3 = Label(canvas, text='Market Segment', font=('arial', 11), fg='black')
        guest_details4 = Label(canvas, text='Loyalty', font=('arial', 10), fg='black')

        guest_coupon = Label(canvas, text='Coupons', font=('arial', 11), fg='black')
        guest_payment = Label(canvas, text='Payment', font=('arial', 11), fg='black')
        guest_payment2 = Label(canvas, text='ADD GUEST TO ADD PAYMENT', font=('arial', 11), fg='#FFB300')
        guest_payment1 = Button(canvas, text='Add', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                                activeforeground='#00695C', activebackground='#EEEEEE', padx=5, pady=3, width=10,
                                state="disabled",
                                command=guest_payment11)
        guest_dispay_payment = Label(canvas, text=str(compute_room_price) + ' avg. / night', font=('arial', 11),
                                     fg='black')
        guest_dispay_payment1 = Label(canvas, text=str(compute_total_room_price) + ' total with taxes & fees',
                                      font=('arial', 8), fg='black')
        guest_policies = Label(canvas, text='Policies', font=('arial', 11), fg='black')
        guest_policies1 = Label(canvas, text='Cancel 3pm day of arrival to avoid one night charge', font=('arial', 8),
                                fg='black')

        guest_space = Label(canvas, font=('arial', 1), bg='lightgray', padx=303, pady=1)

        canvas.create_window(10, 70, anchor="nw", window=guest_required)
        canvas.create_window(56, 70, anchor="nw", window=guest_require)
        canvas.create_window(400, 67, anchor="nw", window=guest_search)
        canvas.create_window(510, 73, anchor="nw", window=guest_entry_search)
        canvas.create_window(10, 130, anchor="nw", window=guest_agent)
        canvas.create_window(400, 130, anchor="nw", window=guest_agent1)
        canvas.create_window(520, 130, anchor="nw", window=guest_agent2)
        canvas.create_window(10, 190, anchor="nw", window=guest_details)
        canvas.create_window(10, 210, anchor="nw", window=guest_details1)
        canvas.create_window(7, 230, anchor="nw", window=guest_details2)
        canvas.create_window(10, 242, anchor="nw", window=guest_details3)
        canvas.create_window(10, 262, anchor="nw", window=guest_details4)
        canvas.create_window(10, 310, anchor="nw", window=guest_coupon)
        canvas.create_window(10, 370, anchor="nw", window=guest_payment)
        canvas.create_window(520, 370, anchor="nw", window=guest_payment1)
        canvas.create_window(76, 370, anchor="nw", window=guest_payment2)
        canvas.create_window(10, 425, anchor="nw", window=guest_dispay_payment)
        canvas.create_window(10, 450, anchor="nw", window=guest_dispay_payment1)
        canvas.create_window(10, 494, anchor="nw", window=guest_policies)
        canvas.create_window(350, 498, anchor="nw", window=guest_policies1)
        canvas.create_window(10, 540, anchor="nw", window=guest_space)

        my_frame1.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
        my_frame1.place(x=200, y=176)
        resv_forget.append(my_frame1)

        next_button_next = Label(window, text='                                  ', bg='#F0F0F0',
                                 activebackground="lightgray", padx=85, pady=4, width=2, )
        next_button_next.place(x=210, y=585)
        resv_forget.append(next_button_next)

        change_info = Button(window, text='Change Information', font=('arial', 9), fg='black', bg='lightgray',
                             activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                             # wraplength= 1,
                             # state='disable',
                             command=site_scrolls3)
        change_info.place(x=210, y=615)
        resv_forget.append(change_info)
        site_scrool1_forget.append(change_info)

def guest_agent11():
    pass


def guest_agent12():
    pass


def guest_payment11():
    pass


def add_guest_remove():
    print('remove reservation number ')
    view_resvno[0] = 0
    post1()
    booked()


def add_guest_button12():
    pass


def notify_friends(event=None):
    print('notify friends')


def add_another(event=None):
    print('add_another')


def guest_search1():
    print('guest search 1')
    loop = 1
    resvno = view_resvno[0].get()
    # print('Reservation view number ', resvno)
    for data_search in range(2, data_number.max_row + 1):
        accntno = data_number.cell(data_search, 1).value
        if str(accntno) == str(resvno):
            loop = 0
            view_yes_no = 'no'
            guest_search2()
    if loop == 1:
        view_yes_no = 'yes'


def guest_search2():
    print('guest search 2')
    resvno = view_resvno[0].get()
    for data_search in range(2, data_number.max_row + 1):
        accntno = data_number.cell(data_search, 1).value
        phoneno = data_number.cell(data_search, 2).value
        lastname = data_number.cell(data_search, 3).value
        firstname = data_number.cell(data_search, 4).value
        home_address = data_number.cell(data_search, 7).value
        if str(accntno) == str(resvno):
            phone = phoneno
            lastname1 = lastname
            firstname1 = firstname
            home_address1 = home_address

    points = view_post1_payment[0]
    room_price = view_post1_payment[1]
    total_points = view_post1_payment[2]
    total_roomprice = view_post1_payment[3]

    if str(points) != '0' and str(room_price) != '0.00':
        compute_room_price = str(points) + ' pts + ' + str(room_price)
        compute_total_room_price = str(total_points) + ' pts + ' + str(total_roomprice)

    elif str(points) == '0' and str(room_price) != '0.00':
        compute_room_price = room_price
        compute_total_room_price = total_roomprice

    elif str(points) != '0' and str(room_price) == '0.00':
        compute_room_price = str(points) + ' pts'
        compute_total_room_price = total_points + ' pts'

    my_frame1 = Frame(window)
    canvas = Canvas(my_frame1)

    canvas = Canvas(my_frame1, width=625, height=380, background="lightgray")
    canvas.pack(fill="both", expand=True)

    vsb = Scrollbar(my_frame1, orient="vertical", command=canvas.yview, width=25)

    hsb = Scrollbar(my_frame1, orient="horizontal", command=canvas.xview, width=25)

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)
    canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
    canvas.grid(row=0, column=0, sticky="nsew")

    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    property = Label(canvas, text='  Add Reservation Details  ', font=('arial', 15), bg='lightgray', fg='black', padx=0,
                     pady=10, )
    canvas.create_window(0, 3, anchor="nw", window=property)

    design_bg = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=80)
    canvas.create_window(0, 60, anchor="nw", window=design_bg)

    design_bg1 = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=10)
    canvas.create_window(0, 260, anchor="nw", window=design_bg1)

    design_bg2 = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=42)
    canvas.create_window(0, 320, anchor="nw", window=design_bg2)

    design_bg3 = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=10)
    canvas.create_window(0, 443, anchor="nw", window=design_bg3)

    design_bg4 = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=10)
    canvas.create_window(0, 503, anchor="nw", window=design_bg4)

    design_bg5 = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=16)
    canvas.create_window(0, 563, anchor="nw", window=design_bg5)

    design_bg6 = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=10)
    canvas.create_window(0, 633, anchor="nw", window=design_bg6)

    guest_required = Label(canvas, text='Guest', font=('arial', 11), fg='black')
    guest_required1 = Label(canvas, text='i', font=('arial', 12), fg='#00695C')
    guest_required2 = Label(canvas, text=str(cap_name(firstname1)) + ' ' + str(cap_name(lastname1)), font=('arial', 11),
                            fg='black')
    guest_required3 = Label(canvas, text=home_address1, font=('arial', 8), fg='black')
    guest_required4 = Label(canvas, text=phoneno, font=('arial', 8), fg='black')
    guest_required5 = Label(canvas, text='+', font=('arial', 12), fg='#00695C')
    guest_required6 = Label(canvas, text='+', font=('arial', 12), fg='#00695C')
    guest_required7 = Label(canvas, text='Notify Friends', font=('arial', 10), fg='#00695C')
    guest_required7.bind("<Button-1>", notify_friends)
    guest_required8 = Label(canvas, text='Add Another Guest', font=('arial', 10), fg='#00695C')
    guest_required8.bind("<Button-1>", add_another)
    add_guest_button = Button(canvas, text='Remove', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                              activeforeground='#00695C', activebackground='#EEEEEE', padx=5, pady=3, width=10,
                              command=add_guest_remove)
    add_guest_button1 = Button(canvas, text='Edit Guest Info', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                               activeforeground='#00695C', activebackground='#EEEEEE', padx=5, pady=3, width=10,
                               command=add_guest_button12)

    guest_agent = Label(canvas, text='Travel Agent or Booker', font=('arial', 11), fg='black')
    guest_agent1 = Button(canvas, text='Find', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                          activeforeground='#00695C', activebackground='#EEEEEE', padx=5, pady=3, width=10,
                          command=guest_agent11)
    guest_agent2 = Button(canvas, text='Add New', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                          activeforeground='#00695C', activebackground='#EEEEEE', padx=5, pady=3, width=10,
                          command=guest_agent12)

    guest_details = Label(canvas, text='Details', font=('arial', 11), fg='black')
    guest_details1 = Label(canvas,
                           text='Request to Hotel, Notification to Guest / Booker, Transportation, Marketing Info, Source',
                           font=('arial', 8), fg='black')
    guest_details2 = Label(canvas, font=('arial', 1), bg='darkgray', padx=303, pady=1)
    guest_details3 = Label(canvas, text='Market Segment', font=('arial', 11), fg='black')
    guest_details4 = Label(canvas, text='Loyalty', font=('arial', 10), fg='black')

    guest_coupon = Label(canvas, text='Coupons', font=('arial', 11), fg='black')
    guest_payment = Label(canvas, text='Payment', font=('arial', 11), fg='black')
    guest_payment2 = Label(canvas, text='ADD GUEST TO ADD PAYMENT', font=('arial', 11), fg='#FFB300')
    guest_payment1 = Button(canvas, text='Add', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                            activeforeground='#00695C', activebackground='#EEEEEE', padx=5, pady=3, width=10,
                            state="disabled",
                            command=guest_payment11)
    guest_dispay_payment = Label(canvas, text=str(compute_room_price) + ' avg. / night', font=('arial', 11), fg='black')
    guest_dispay_payment1 = Label(canvas, text=str(compute_total_room_price) + ' total with taxes & fees',
                                  font=('arial', 8), fg='black')
    guest_policies = Label(canvas, text='Policies', font=('arial', 11), fg='black')
    guest_policies1 = Label(canvas, text='Cancel 3pm day of arrival to avoid one night charge', font=('arial', 8),
                            fg='black')

    guest_space = Label(canvas, font=('arial', 1), bg='lightgray', padx=303, pady=1)
    canvas.create_window(10, 70, anchor="nw", window=guest_required)
    canvas.create_window(11, 95, anchor="nw", window=guest_required1)
    canvas.create_window(30, 95, anchor="nw", window=guest_required2)
    canvas.create_window(30, 115, anchor="nw", window=guest_required3)
    canvas.create_window(30, 132, anchor="nw", window=guest_required4)
    canvas.create_window(30, 178, anchor="nw", window=guest_required5)
    canvas.create_window(30, 203, anchor="nw", window=guest_required6)
    canvas.create_window(45, 180, anchor="nw", window=guest_required7)
    canvas.create_window(45, 205, anchor="nw", window=guest_required8)
    canvas.create_window(410, 95, anchor="nw", window=add_guest_button)
    canvas.create_window(510, 95, anchor="nw", window=add_guest_button1)

    canvas.create_window(10, 269, anchor="nw", window=guest_agent)
    canvas.create_window(400, 269, anchor="nw", window=guest_agent1)
    canvas.create_window(520, 269, anchor="nw", window=guest_agent2)

    canvas.create_window(10, 328, anchor="nw", window=guest_details)
    canvas.create_window(10, 348, anchor="nw", window=guest_details1)
    canvas.create_window(7, 373, anchor="nw", window=guest_details2)
    canvas.create_window(10, 383, anchor="nw", window=guest_details3)
    canvas.create_window(10, 403, anchor="nw", window=guest_details4)
    canvas.create_window(10, 457, anchor="nw", window=guest_coupon)
    canvas.create_window(10, 517, anchor="nw", window=guest_payment)
    canvas.create_window(520, 513, anchor="nw", window=guest_payment1)
    canvas.create_window(76, 517, anchor="nw", window=guest_payment2)
    canvas.create_window(10, 572, anchor="nw", window=guest_dispay_payment)
    canvas.create_window(10, 597, anchor="nw", window=guest_dispay_payment1)
    canvas.create_window(10, 650, anchor="nw", window=guest_policies)
    canvas.create_window(350, 650, anchor="nw", window=guest_policies1)
    canvas.create_window(10, 685, anchor="nw", window=guest_space)

    my_frame1.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
    my_frame1.place(x=200, y=176)
    resv_forget.append(my_frame1)

    booked_resv = Button(window, text='Booked Reservation', font=('arial', 9), fg='black', bg='lightgray',
                         activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                         # wraplength= 1,
                         # state='disable',
                         command=temp_booked)
    booked_resv.place(x=210, y=585)
    resv_forget.append(booked_resv)
    site_scrool1_forget.append(booked_resv)

    change_info = Button(window, text='Change Information', font=('arial', 9), fg='black', bg='lightgray',
                         activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                         # wraplength= 1,
                         # state='disable',
                         command=site_scrolls3)
    change_info.place(x=210, y=615)
    resv_forget.append(change_info)
    site_scrool1_forget.append(change_info)

    if change_shift_resv[0] == 1:
        booked_resv1 = Button(window, text='Booked Reservation', font=('arial', 9), fg='black', bg='lightgray',
                              activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                              # wraplength= 1,
                              # state='disable',
                              command=temp_booked)
        booked_resv1.place(x=210, y=585)
        resv_forget.append(booked_resv1)

    else:
        def_booked_value()


def guest_search3():
    account = choose_acctno[0]
    reservation = choose_resv[0]
    print('guest search three')
    for data_search in range(2, data_number.max_row + 1):
        accntno = data_number.cell(data_search, 1).value
        phoneno = data_number.cell(data_search, 2).value
        lastname = data_number.cell(data_search, 3).value
        firstname = data_number.cell(data_search, 4).value
        home_address = data_number.cell(data_search, 7).value
        if str(accntno) == str(account):
            phone = phoneno
            lastname1 = lastname
            firstname1 = firstname
            home_address2 = home_address

    for data_search in range(2, data_number4.max_row + 1):
        no_site = data_number4.cell(data_search, 1).value
        resv = data_number4.cell(data_search, 2).value
        phoneno = data_number4.cell(data_search, 3).value
        home_address1 = data_number4.cell(data_search, 7).value
        hotel_address = data_number4.cell(data_search, 8).value
        lastname1 = data_number4.cell(data_search, 4).value
        firstname1 = data_number4.cell(data_search, 5).value
        mbr_rate = data_number4.cell(data_search, 10).value
        room_price = data_number4.cell(data_search, 11).value
        taxes = data_number4.cell(data_search, 12).value
        points = data_number4.cell(data_search, 13).value
        room_price1 = data_number4.cell(data_search, 14).value
        total_points = data_number4.cell(data_search, 16).value
        total_roomprice = data_number4.cell(data_search, 17).value
        norooms = data_number4.cell(data_search, 18).value
        adults = data_number4.cell(data_search, 19).value
        child = data_number4.cell(data_search, 20).value
        room_type = data_number4.cell(data_search, 21).value
        checkin = data_number4.cell(data_search, 22).value
        checkout = data_number4.cell(data_search, 23).value
        rate_name = data_number4.cell(data_search, 24).value
        siteno = data_number4.cell(data_search, 28).value
        if str(resv) == str(reservation):
            # print('total paymenttttttttttttttttttttttttttttttttttt ', points, room_price, total_points, total_roomprice)
            # compute_room_price = 0
            # compute_total_room_price = 0
            if str(points) != '0' and str(room_price) != '0.00':
                compute_room_price = str(points) + ' pts + ' + str(room_price)
                compute_total_room_price = str(total_points) + ' pts + ' + str(total_roomprice)

            elif str(points) == '0' and str(room_price) != '0.00':
                compute_room_price = room_price
                compute_total_room_price = total_roomprice

            elif str(points) != '0' and str(room_price) == '0.00':
                compute_room_price = str(points) + ' pts'
                compute_total_room_price = total_points + ' pts'

            a2 = [str(x) for x in checkin]
            b1 = [str(x) for x in checkout]

            a2.remove(a2[-1])
            a2[-1] = ','
            a2.append(' '), a2.append(b1[-4]), a2.append(b1[-3]), a2.append(b1[-2]), a2.append(b1[-1])
            a3 = ''.join(a2)

            my_frame1 = Frame(window)
            canvas = Canvas(my_frame1)

            canvas = Canvas(my_frame1, width=625, height=380, background="lightgray")
            canvas.pack(fill="both", expand=True)

            vsb = Scrollbar(my_frame1, orient="vertical", command=canvas.yview, width=25)

            hsb = Scrollbar(my_frame1, orient="horizontal", command=canvas.xview, width=25)

            window.grid_rowconfigure(0, weight=1)
            window.grid_columnconfigure(0, weight=1)
            canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
            canvas.grid(row=0, column=0, sticky="nsew")

            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")

            property = Label(canvas, text='  Add Reservation Details  ', font=('arial', 15), bg='lightgray', fg='black',
                             padx=0,
                             pady=10, )
            canvas.create_window(0, 3, anchor="nw", window=property)

            design_bg = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=250)
            canvas.create_window(0, 60, anchor="nw", window=design_bg)

            design_bg1 = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=10)
            canvas.create_window(0, 600, anchor="nw", window=design_bg1)

            design_bg2 = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=42)
            canvas.create_window(0, 320, anchor="nw", window=design_bg2)

            design_bg3 = Label(canvas, font=('arial', 15), bg='#EEEEEE', fg='white', padx=310, pady=10)
            canvas.create_window(0, 350, anchor="nw", window=design_bg3)

            guest_resv = Label(canvas, text=resv, font=('consolas', 10, 'bold'), fg='black')
            guest_resv1 = Label(canvas, text='Booking Channel Voice', font=('arial', 9), fg='black')
            guest_resv2 = Label(canvas, text='Booked On:  ' + str(a3), font=('arial', 9), fg='black')
            guest_resv3 = Label(canvas, text='Booked by:  ' + 'TP Agents', font=('arial', 9), fg='black')
            guest_resv4 = Label(canvas, text='i', font=('arial', 12), fg='#00695C')
            guest_resv5 = Label(canvas, text=str(home_address1) + ' (' + str(siteno) + ')', font=('Consolas', 10),
                                fg='#00695C')

            guest_resv5a = Label(canvas, text=str(home_address1), font=('Consolas', 10, 'bold'), fg='black')
            canvas.create_window(30, 165, anchor="nw", window=guest_resv5)
            canvas.create_window(30, 165, anchor="nw", window=guest_resv5a)

            guest_resv6 = Label(canvas, text=hotel_address, font=('arial', 8), fg='black')
            guest_resv7 = Label(canvas, text=phoneno, font=('arial', 8), fg='black')
            guest_resv8 = Label(canvas, font=('arial', 1), bg='darkgray', padx=303, pady=1)
            guest_resv9 = Label(canvas, text='Room', font=('arial', 9), fg='black')
            guest_resv10 = Label(canvas, text='i', font=('arial', 12), fg='#00695C')
            guest_resv11 = Label(canvas, text=str(checkin) + ' ' + str(checkout), font=('Consolas', 10, 'bold'),
                                 fg='black')
            if norooms > 1:
                nyt = 'nights'
            else:
                nyt = 'night'

            if int(adults) > 1:
                adult = 'adults'
            else:
                adult = 'adult'

            if int(child) > 1:
                childs = 'childrens'
            else:
                childs = 'children'

            guest_resv12 = Label(canvas, text=str(norooms) + ' ' + str(nyt) + ', 1 room, ' + str(adults) + ' ' +
                                              str(adult) + ', ' + str(child) + ' ' + str(childs), font=('arial', 8),
                                 fg='black')
            guest_resv13 = Label(canvas, text=str(room_type) + ', Non-Smoking', font=('arial', 8), fg='black')
            guest_resv14 = Label(canvas, text=str(mbr_rate) + ', ' + str(rate_name), font=('arial', 8), fg='black')
            guest_resv14a = Label(canvas, text='Wyndham Rewards - ' + str(no_site), font=('arial', 8), fg='#00695C')

            guest_resv15 = Label(canvas, text=str(room_price) + ' avg./night', font=('arial', 9), fg='black')
            guest_resv16 = Label(canvas, text=str(total_roomprice) + ' total with taxes & fees', font=('arial', 8),
                                 fg='black')
            guest_resv17 = Label(canvas, text=str(taxes) + ' Tax', font=('arial', 8), fg='black')

            guest_resv18 = Label(canvas, text=str(points) + ', ' + str(total_points) + ', points & total points',
                                 font=('arial', 8), fg='black')
            canvas.create_window(400, 345, anchor="nw", window=guest_resv18)

            guest_resv19 = Label(canvas, text='Daily Price Breakdown', font=('arial', 9), fg='#00695C')
            canvas.create_window(400, 365, anchor="nw", window=guest_resv18)

            temp_price = ''
            temp1 = conversion_numbers(room_price1)
            temp = [str(x) for x in temp1]
            for x in temp:
                if x.isnumeric() == True:
                    temp_price += x

            if str(temp_price) == '0':
                points_round = 'pt.'
            else:
                points_round = 'pts.'
            guest_resv20 = Label(canvas,
                                 text='Points Earned ' + str(conversion_numbers(int(temp_price) * 10)) + ' ' + str(
                                     points_round),
                                 font=('arial', 9), fg='#C0CA33')
            guest_resv21 = Label(canvas, font=('arial', 1), bg='darkgray', padx=303, pady=1)
            canvas.create_window(10, 70, anchor="nw", window=guest_resv)
            canvas.create_window(10, 90, anchor="nw", window=guest_resv1)
            canvas.create_window(10, 110, anchor="nw", window=guest_resv2)
            canvas.create_window(10, 130, anchor="nw", window=guest_resv3)
            canvas.create_window(10, 162, anchor="nw", window=guest_resv4)
            # canvas.create_window(30, 165, anchor="nw", window=guest_resv5)
            canvas.create_window(30, 185, anchor="nw", window=guest_resv6)
            canvas.create_window(30, 205, anchor="nw", window=guest_resv7)
            canvas.create_window(10, 235, anchor="nw", window=guest_resv8)
            canvas.create_window(10, 250, anchor="nw", window=guest_resv9)
            canvas.create_window(10, 282, anchor="nw", window=guest_resv10)
            canvas.create_window(30, 285, anchor="nw", window=guest_resv11)
            canvas.create_window(30, 305, anchor="nw", window=guest_resv12)
            canvas.create_window(30, 325, anchor="nw", window=guest_resv13)
            canvas.create_window(30, 345, anchor="nw", window=guest_resv14)
            canvas.create_window(30, 365, anchor="nw", window=guest_resv14a)
            canvas.create_window(400, 285, anchor="nw", window=guest_resv15)
            canvas.create_window(400, 305, anchor="nw", window=guest_resv16)
            canvas.create_window(400, 325, anchor="nw", window=guest_resv17)
            canvas.create_window(400, 345, anchor="nw", window=guest_resv18)
            canvas.create_window(400, 365, anchor="nw", window=guest_resv19)
            canvas.create_window(400, 390, anchor="nw", window=guest_resv20)
            canvas.create_window(10, 420, anchor="nw", window=guest_resv21)

            guest_required1 = Label(canvas, text='i', font=('arial', 12), fg='#00695C')
            guest_required2 = Label(canvas, text=str(cap_name(firstname1)) + ' ' + str(cap_name(lastname1)),
                                    font=('consolas', 10, 'bold'),
                                    fg='black')
            guest_required3 = Label(canvas, text=home_address2, font=('arial', 8), fg='black')
            guest_required4 = Label(canvas, text=phoneno, font=('arial', 8), fg='black')
            guest_required5 = Label(canvas, text='Company Name: None', font=('arial', 8), fg='black')
            guest_required6 = Label(canvas, text='Guest Info', font=('arial', 8), fg='#00695C')
            guest_required7 = Label(canvas, text='Wyndham Rewards' + ' (' + str(no_site) + ')', font=('consolas', 9),
                                    fg='#00695C')
            guest_required7a = Label(canvas, text='Wyndham Rewards', font=('consolas', 9, 'bold'), fg='black')
            guest_policies = Label(canvas, text='Policies', font=('arial', 9), fg='black')
            guest_policies1 = Label(canvas, text='Cancel 3pm day of arrival to avoid one night charge',
                                    font=('arial', 8),
                                    fg='black')
            guest_space = Label(canvas, font=('arial', 1), bg='lightgray', padx=303, pady=1)

            canvas.create_window(10, 433, anchor="nw", window=guest_required1)
            canvas.create_window(30, 435, anchor="nw", window=guest_required2)
            canvas.create_window(30, 455, anchor="nw", window=guest_required3)
            canvas.create_window(30, 475, anchor="nw", window=guest_required4)
            canvas.create_window(30, 495, anchor="nw", window=guest_required5)
            canvas.create_window(30, 515, anchor="nw", window=guest_required6)
            canvas.create_window(30, 540, anchor="nw", window=guest_required7)
            canvas.create_window(30, 540, anchor="nw", window=guest_required7a)
            canvas.create_window(10, 610, anchor="nw", window=guest_policies)
            canvas.create_window(350, 610, anchor="nw", window=guest_policies1)
            canvas.create_window(10, 650, anchor="nw", window=guest_space)

            my_frame1.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
            my_frame1.place(x=200, y=176)
            view_return.append(my_frame1)
            resv_forget.append(my_frame1)

            next_button_next = Label(window, text='                                  ', bg='#F0F0F0',
                                     activebackground="lightgray", padx=85, pady=4, width=2, )
            next_button_next.place(x=210, y=585)
            resv_forget.append(next_button_next)


            cxl_reservation = Button(window, text='Cancel Reservation', font=('arial', 9), fg='black', bg='lightgray',
                                     activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                                     # wraplength= 1,
                                     # state='disable',
                                     command=cxl_resv)
            cxl_reservation.place(x=650, y=585)
            resv_forget.append(cxl_reservation)

            view_resv = Button(window, text='Modify Reservation', font=('arial', 9), fg='black', bg='lightgray',
                               activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                               # wraplength= 1,
                               # state='disable',
                               command=guest_search_modify)
            view_resv.place(x=650, y=615)
            resv_forget.append(view_resv)


# choose guest_search_modify site button
def choose_site_entry_button():
    print('choose entry site number ')
    site = ''
    site = choose_entry[0].get()
    print('choose site number ', choose_entry[0], site)
    choose_loopsite[0] = 0
    loops = 0
    temp_site = 0
    for data_search in range(2, data_number2.max_row + 1):
        siteno = data_number2.cell(data_search, 28).value
        if str(siteno) == str(site):
            temp_site = siteno
            choose_loopsite[0] = 1
            loops = 1
    if loops == 0:
        choose_loopsite[0] = 0
        choose_entry[0] = ''
        guest_search_modify()
    else:
        choose_entry[0] = temp_site
        print('choose site number ',choose_entry[0] ,site)
        guest_search_modify()


# choose fetch date
def fetch_date(event=None):
    diff_dates = 1
    print('check in fetch date')
    # tkc calendar
    x = change_current_date[0]
    print('check in datessssssssssssssssssssssssss ',x)
    today = date.today()
    years = eval(x.strftime('%Y'))
    month = x.strftime('%m')
    day = x.strftime('%d')
    current_dates = x.strftime("%B %d, %Y")
    date_tom = x + datetime.timedelta(days=0)
    years_1 = date_tom.strftime('%Y')
    months_1 = date_tom.strftime('%m')
    day_1 = date_tom.strftime('%d')
    years_10 = round(float(years_1))
    months10 = round(float(months_1))
    days10 = round(float(day_1))
    print('number of dayssssssssssssssssss ', days10)

    #print('date tommorow ', date_tom, years, years_10 ,months10, days10)

    months = round(float(month))
    days = round(float(day))
    # tkc mindate and maxdate
    y = datetime.datetime.now()
    years9 = eval(y.strftime('%Y'))
    month = y.strftime('%m')
    day = y.strftime('%d')
    current_dates = x.strftime("%B %d, %Y")
    date_tom = y + datetime.timedelta(days=1)
    months9 = round(float(month))
    days9 = round(float(day))



    # print('booking check in ', 'year:', years, 'month:',months, 'days:', days)
    tkc = Calendar(window, selectmode="day", year=years, month=months, day=days, mindate=date(years9, months9, days9),
                   maxdate=date(years9 + 1, months9, days9))
    tkc.place(x=750, y=100)
    change_date[0] = tkc
    change_forget.append(tkc)
    view_return.append(tkc)
    resv_forget.append(tkc)

    get_the_dates = change_date[0].selection_get()
    print('out days number ', get_the_dates)
    # date_tom1 = get_the_dates + datetime.timedelta(days=change_no_dates[0])
    date_tom1 = get_the_dates + datetime.timedelta(days=plus_date[0])
    #print('get the dates and date tomorrow ',date_tom1 ,change_no_dates[0])
    if change_plus_minus[0] == 1:
        date_tom1 = get_the_dates + datetime.timedelta(days=plus_date[0])
    out_year = eval(date_tom1.strftime('%Y'))
    out_month = date_tom1.strftime('%m')
    out_days = date_tom1.strftime('%d')
    out_month1 = round(float(out_month))
    out_days1 = round(float(out_days))
    print('out days number ',out_days1)

    tkc1 = Calendar(window, selectmode="day", year=out_year, month=out_month1, day=out_days1,
                    mindate=date(years_10, months10, days10), maxdate=date(years9 + 1, months9, days9))
    # tkc1 = Calendar(window, selectmode="day", year=out_year, month=out_month1, day=out_days1)
    tkc1.place(x=1000, y=100)
    change_date1[0] = tkc1
    change_forget.append(tkc1)
    view_return.append(tkc1)
    resv_forget.append(tkc1)

    but = Button(window, text="Select Date", command=fetch_date1, bg="#EEEEEE", fg='#4D4D4D',
                 activeforeground='#4D4D4D', activebackground='#EEEEEE', padx=15, pady=2, width=5)
    but.place(x=1178, y=285)
    change_forget.append(but)
    view_return.append(but)
    resv_forget.append(but)


# choose fetch date1
def fetch_date1(event=None):
    print('fetch date1')

    site = choose_entry[0].get()
    choose_loopsite[0] = 0
    loops = 0
    for data_search in range(2, data_number2.max_row + 1):
        siteno = data_number2.cell(data_search, 28).value
        if str(siteno) == str(site):
            choose_loopsite[0] = 1
            loops = 1
    if loops == 0:
        choose_loopsite[0] = 0
        choose_entry[0] = ''
    else:
        choose_entry[0] = site

    change_plus_minus[0] = 0

    get_the_dates = change_date[0].selection_get()
    # print('check in dates ', get_the_dates)
    change_current_date[0] = get_the_dates
    get_the_dates1 = change_date1[0].selection_get()
    # print('check out dates ', get_the_dates1)
    change_loop[0] = 1
    guest_search_modify()


# plus  room button
def room_button():
    global mod_loops, mod_rooms
    plus_room1[0] = plus_room1[0] + 1
    if mod_loops == 1:
        mod_rooms = plus_room1[0]
    # print('plus room button ',plus_room1[0])
    rooms_no1 = Label(window, text='      ', font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    rooms_no1.place(x=900, y=439)
    resv_forget.append(rooms_no1)
    view_return.append(rooms_no1)
    rooms_no1 = Label(window, text=plus_room1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    rooms_no1.place(x=900, y=439)
    resv_forget.append(rooms_no1)
    view_return.append(rooms_no1)


# minus room button
def room_button1():
    global mod_loops, mod_rooms
    plus_room1[0] = plus_room1[0] - 1
    if plus_room1[0] == 0:
        plus_room1[0] = 1
    if mod_loops == 1:
        mod_rooms = plus_room1[0]
    # print('plus room button ', plus_room1[0])
    rooms_no1 = Label(window, text='      ', font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    rooms_no1.place(x=900, y=439)
    resv_forget.append(rooms_no1)
    view_return.append(rooms_no1)
    rooms_no1 = Label(window, text=plus_room1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    rooms_no1.place(x=900, y=439)
    resv_forget.append(rooms_no1)
    view_return.append(rooms_no1)

# plus adult button
def room_adults():
    global mod_loops, mod_adults
    plus_adult1[0] = plus_adult1[0] + 1
    # print('plus room button ', plus_adult1[0])
    choose_dates[6] = plus_adult1[0]
    if mod_loops == 1:
        mod_adults = plus_adult1[0]
        modify_dates[6] = mod_adults

    adults_no = Label(window, text='      ', font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    adults_no.place(x=900, y=494)
    resv_forget.append(adults_no)
    view_return.append(adults_no)
    adults_no = Label(window, text=plus_adult1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    adults_no.place(x=900, y=494)
    resv_forget.append(adults_no)
    view_return.append(adults_no)


# minus adult button
def room_adults1():
    global mod_loops, mod_adults
    plus_adult1[0] = plus_adult1[0] - 1
    if plus_adult1[0] == 0:
        plus_adult1[0] = 1
    choose_dates[6] = plus_adult1[0]
    if mod_loops == 1:
        mod_adults = plus_adult1[0]
        modify_dates[6] = mod_adults

    # print('plus room button ', plus_adult1[0])
    adults_no = Label(window, text='      ', font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    adults_no.place(x=900, y=494)
    resv_forget.append(adults_no)
    view_return.append(adults_no)
    adults_no = Label(window, text=plus_adult1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    adults_no.place(x=900, y=494)
    resv_forget.append(adults_no)
    view_return.append(adults_no)


# plus child button
def room_child():
    global mod_loops, mod_childs
    plus_child1[0] = plus_child1[0] + 1
    # print('plus room button ', plus_child1[0])
    choose_dates[7] = plus_child1[0]
    if mod_loops == 1:
        mod_childs = plus_child1[0]
        modify_dates[7] = mod_childs
    child_no = Label(window, text='      ', font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    child_no.place(x=900, y=553)
    resv_forget.append(child_no)
    view_return.append(child_no)
    child_no = Label(window, text=plus_child1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    child_no.place(x=900, y=553)
    resv_forget.append(child_no)
    view_return.append(child_no)


# minus child button
def room_child1():
    global mod_loops, mod_childs
    # print('plus room button ', plus_child1[0])
    if plus_child1[0] == 0:
        plus_child1[0] = 0
    else:
        plus_child1[0] = plus_child1[0] - 1
    choose_dates[7] = plus_child1[0]
    if mod_loops == 1:
        mod_childs = plus_child1[0]
        modify_dates[7] = mod_childs
    child_no = Label(window, text='      ', font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    child_no.place(x=900, y=553)
    resv_forget.append(child_no)
    view_return.append(child_no)
    child_no = Label(window, text=plus_child1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    child_no.place(x=900, y=553)
    resv_forget.append(child_no)
    view_return.append(child_no)


def plus_button():
    global mod_nonight, mod_loops, mod_checkout
    temp_add = 1
    change_plus_minus[0] = 1
    plus_date[0] = plus_date[0] + 1

    if mod_loops == 1:# find stay and modify funtion condition
        print('modify loops ',mod_loops)
        mod_nonight += 1
        plus_date[0] = mod_nonight
        '''modify_dates[0] = get_the_dates.strftime("%m")
        modify_dates[1] = get_the_dates.strftime("%d")
        modify_dates[2] = get_the_dates.strftime("%Y")
        modify_dates[3] = get_the_dates1.strftime("%m")
        modify_dates[4] = get_the_dates1.strftime("%d")
        modify_dates[5] = get_the_dates1.strftime("%Y")'''
        checkinmonth = datetime.datetime.strptime(modify_dates[0], '%b').month
        change_current_date[0] = date(int(modify_dates[2]), int(checkinmonth), int(modify_dates[1]))

        temp_checkinmonth = datetime.datetime.strptime(modify_dates[0], '%b').month
        temp_choose_datein = date(int(modify_dates[2]), int(temp_checkinmonth), int(modify_dates[1]))
        temp_checkoutmonth = datetime.datetime.strptime(modify_dates[3], '%b').month
        temp_choose_dateout = date(int(modify_dates[5]), int(temp_checkoutmonth), int(modify_dates[4]))

        if temp_choose_datein == temp_choose_dateout:
            mod_nonight = 1
            plus_date[0] = mod_nonight

    else:
        temp_checkinmonth = datetime.datetime.strptime(choose_dates[0], '%b').month
        temp_choose_datein = date(int(choose_dates[2]), int(temp_checkinmonth), int(choose_dates[1]))
        temp_checkoutmonth = datetime.datetime.strptime(choose_dates[3], '%b').month
        temp_choose_dateout = date(int(choose_dates[5]), int(temp_checkoutmonth), int(choose_dates[4]))

        if temp_choose_datein == temp_choose_dateout:
            plus_date[0] = 1

    checkin_dates1 = Label(window, text='         ', font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    checkin_dates1.place(x=900, y=384)
    resv_forget.append(checkin_dates1)
    view_return.append(checkin_dates1)

    checkin_dates1 = Label(window, text=plus_date[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    checkin_dates1.place(x=900, y=384)
    resv_forget.append(checkin_dates1)
    view_return.append(checkin_dates1)
    #print('change current dateeeeeeeeeeeee ',change_current_date[0])
    date_tom = change_current_date[0] + datetime.timedelta(plus_date[0])

    last_date1 = date_tom.strftime("%b %d, %Y")
    last_date2 = date_tom.strftime("%b %m %Y")
    last_date3 = date_tom.strftime("%d %m %Y")
    # update choose dates
    choose_dates[3] = date_tom.strftime("%b")
    choose_dates[4] = date_tom.strftime("%d")
    choose_dates[5] = date_tom.strftime("%Y")

    if mod_loops == 1:# find stay and modify funtion condition
        modify_dates[3] = date_tom.strftime("%b")
        modify_dates[4] = date_tom.strftime("%d")
        modify_dates[5] = date_tom.strftime("%Y")
        mod_checkout = date(int(modify_dates[5]), int(date_tom.strftime("%m")), int(modify_dates[4]))
        print('change date 1 ', modify_dates)

    # print('check in and check out date ', change_current_date[0].strftime("%Y-%m-%d") ,date_tom.strftime("%Y-%m-%d"))
    checkout_dates = Label(window, text=last_date1, font=('arial', 10), fg='#4D4D4D', bg='#EEEEEE')
    checkout_dates.place(x=1038, y=326)
    resv_forget.append(checkout_dates)
    view_return.append(checkout_dates)

    change_in_out_value[1] = last_date2
    change_in_out_value[3] = last_date3
    # print('check-in and check-out dates ', change_in_out_value[0], change_in_out_value[1])

    minus_nyts = Button(window, text='-', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                        activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                        # state='disable',
                        command=minus_button)
    minus_nyts.place(x=1080, y=371)
    resv_forget.append(minus_nyts)
    view_return.append(minus_nyts)


def compute_plus_minus():
    date_tom1 = change_current_date[0] + datetime.timedelta(days=plus_date[0])
    change_lastdate[0] = date_tom1
    # print('change current date ',change_lastdate[0])


def minus_button():
    global mod_nonight, mod_loops, mod_checkout
    change_plus_minus[0] = 1

    if mod_loops == 1:
        temp_checkinmonth = datetime.datetime.strptime(modify_dates[0], '%b').month
        temp_choose_datein = date(int(modify_dates[2]), int(temp_checkinmonth), int(modify_dates[1]))
        temp_checkoutmonth = datetime.datetime.strptime(modify_dates[3], '%b').month
        temp_choose_dateout = date(int(modify_dates[5]), int(temp_checkoutmonth), int(modify_dates[4]))
    else:
        temp_checkinmonth = datetime.datetime.strptime(choose_dates[0], '%b').month
        temp_choose_datein = date(int(choose_dates[2]), int(temp_checkinmonth), int(choose_dates[1]))
        temp_checkoutmonth = datetime.datetime.strptime(choose_dates[3], '%b').month
        temp_choose_dateout = date(int(choose_dates[5]), int(temp_checkoutmonth), int(choose_dates[4]))

    if temp_choose_datein == temp_choose_dateout:
        pass

    else:
        print('minus button')
        compute_plus_minus()
        plus_date[0] = plus_date[0] - 1
        if plus_date[0] == 0:
            plus_date[0] = 1

        if mod_loops == 1:# find stay and modify funtion condition
            print('modify loops ',mod_loops)
            mod_nonight -= 1
            if mod_nonight == 0:
                mod_nonight = 1
            plus_date[0] = mod_nonight
            '''modify_dates[0] = get_the_dates.strftime("%m")
            modify_dates[1] = get_the_dates.strftime("%d")
            modify_dates[2] = get_the_dates.strftime("%Y")
            modify_dates[3] = get_the_dates1.strftime("%m")
            modify_dates[4] = get_the_dates1.strftime("%d")
            modify_dates[5] = get_the_dates1.strftime("%Y")'''
            checkinmonth = datetime.datetime.strptime(modify_dates[3], '%b').month
            change_lastdate[0] = date(int(modify_dates[5]), int(checkinmonth), int(modify_dates[4]))

        checkin_dates1 = Label(window, text='         ', font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
        checkin_dates1.place(x=900, y=384)
        resv_forget.append(checkin_dates1)
        view_return.append(checkin_dates1)

        checkin_dates1 = Label(window, text=plus_date[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
        checkin_dates1.place(x=900, y=384)
        resv_forget.append(checkin_dates1)
        view_return.append(checkin_dates1)

        print('minus date ', change_lastdate[0])
        date_tom = change_lastdate[0] - datetime.timedelta(days=1)
        print('minus date tomorrow ', change_current_date[0] ,change_lastdate[0])
        print('change current date ', change_current_date[0])

        if date_tom.strftime("%Y-%m-%d") == change_current_date[0].strftime("%Y-%m-%d"):
            # print('minus date tomorrowwwwwwwwwwwwww ', date_tom.strftime("%Y-%m-%d"))
            plus_date[0] = 1
            temp = date_tom.strftime("%Y-%m-%d")
            last_date1 = change_lastdate[0].strftime("%b %d, %Y")
            last_date2 = change_lastdate[0].strftime("%Y-%m-%d")
            last_date3 = change_lastdate[0].strftime("%d %m %Y")
            # update choose dates
            choose_dates[3] = change_lastdate[0].strftime("%b")
            choose_dates[4] = change_lastdate[0].strftime("%d")
            choose_dates[5] = change_lastdate[0].strftime("%Y")
            date_tom1 = date_tom + datetime.timedelta(days=plus_date[0])

            checkout_dates = Label(window, text=last_date1, font=('arial', 10), fg='#4D4D4D', bg='#EEEEEE')
            resv_forget.append(checkout_dates)
            checkout_dates.place(x=1038, y=326)
            view_return.append(checkout_dates)
            # print('check in and check out date ', change_current_date[0].strftime("%Y-%m-%d"), last_date1)

        else:
            temp = date_tom
            # print('check in and check out date ', change_current_date[0].strftime("%Y-%m-%d") ,temp.strftime("%Y-%m-%d"))
            # print('number of nyts ', plus_date[0])
            # change_date1[0]= temp
            # print('change dates ',change_date1[0])
            last_date1 = temp.strftime("%b %d, %Y")
            last_date2 = temp.strftime("%Y-%m-%d")
            last_date3 = temp.strftime("%d %m %Y")
            # update choose dates
            choose_dates[3] = temp.strftime("%b")
            choose_dates[4] = temp.strftime("%d")
            choose_dates[5] = temp.strftime("%Y")

            if mod_loops == 1:# find stay and modify funtion condition
                modify_dates[3] = date_tom.strftime("%b")
                modify_dates[4] = date_tom.strftime("%d")
                modify_dates[5] = date_tom.strftime("%Y")
                mod_checkout = date(int(modify_dates[5]), int(date_tom.strftime("%m")), int(modify_dates[4]))
                print('change date 1 ', modify_dates)

            # change_lastdate[0] = last_date1
            checkout_dates = Label(window, text=last_date1, font=('arial', 10), fg='#4D4D4D', bg='#EEEEEE')
            checkout_dates.place(x=1038, y=326)
            resv_forget.append(checkout_dates)
            view_return.append(checkout_dates)

        change_in_out_value[1] = last_date2
        change_in_out_value[3] = last_date3
        '''for x in range(0, len(choose_dates)):
            if x == 8 or x == 9:
                print('choose dates ', choose_dates[x].get())
            else:
                print('choose dates ', choose_dates[x])
        print('check-in and check-out dates ', change_in_out_value[0], change_in_out_value[1])'''


# choose modify
def guest_search_modify(event=None):
    print('guest search modify')
    # choose_entry[0] = 0
    account = choose_acctno[0]
    reservation = choose_resv[0]
    #print('reservaation number ', reservation ,account)
    levelcolor = 'lightblue'
    nonight1 = 0
    adult1 = 0
    child1 = 0
    site = 0
    for data_search in range(2, data_number2.max_row + 1):
        resvno = data_number2.cell(data_search, 2).value
        room_rates = data_number2.cell(data_search, 10).value
        nonight = data_number2.cell(data_search, 18).value
        adult = data_number2.cell(data_search, 19).value
        child = data_number2.cell(data_search, 20).value
        room_type = data_number2.cell(data_search, 21).value
        checkin = data_number2.cell(data_search, 22).value
        checkout = data_number2.cell(data_search, 23).value
        siteno = data_number2.cell(data_search, 28).value
        if str(resvno) == str(reservation):
            nonight1 = nonight
            adult1 = adult
            child1 = child
            site = siteno
            value_book = room_rates
            value_beds = room_type
            a = [str(y) for y in checkin]
            b = [str(x) for x in checkout]

            c1 = str(a[1]) + str(a[2])
            c2 = str(a[4]) + str(a[5])
            c3 = str(a[7]) + str(a[8]) + str(a[9]) + str(a[10])
            c4 = date(int(c3), int(c2), int(c1))
            d1 = str(b[1]) + str(b[2])
            d2 = str(b[4]) + str(b[5])
            d3 = str(b[7]) + str(b[8]) + str(b[9]) + str(b[10])
            d4 = date(int(d3), int(d2), int(d1))

            choose_dates[0] = c4.strftime("%b")
            choose_dates[1] = c4.strftime("%d")
            choose_dates[2] = c4.strftime("%Y")
            choose_dates[3] = d4.strftime("%b")
            choose_dates[4] = d4.strftime("%d")
            choose_dates[5] = d4.strftime("%Y")
            choose_dates[6] = adult1
            choose_dates[7] = child1
            # print('value beds and rates ',value_book ,value_beds)
    site = choose_siteno[0]
    if choose_loopsite[0] == 1:
        site = choose_entry[0]
        #choose_entry[0] = site
        choose_siteno[0] = site
    #print('site number ',site)

    for x in change_forget:
        x.place_forget()
    change_forget.clear()
    #print('change loop ',change_loop[0])
    if change_loop[0] == 0:
        mod_date = c4
    else:
        mod_date = change_current_date[0]
        value_book = choose_dates[-2].get()
        value_beds = choose_dates[-1].get()
    #print('modify date ',mod_date)
    # today = date.today()
    year = mod_date.strftime('%Y')
    month = mod_date.strftime('%m')
    days = mod_date.strftime('%d')
    current_dates = mod_date.strftime("%b %d, %Y")
    current_dates1 = mod_date.strftime("%B %d %Y")
    current_dates2 = mod_date.strftime("%d %m %Y")
    # get check in choose dates
    choose_dates[0] = mod_date.strftime("%b")
    choose_dates[1] = mod_date.strftime("%d")
    choose_dates[2] = mod_date.strftime("%Y")

    change_in_out_value[0] = current_dates1
    change_in_out_value[2] = current_dates2
    for y in month:
        month_checkin = y
    date_tom = 0
    last_date1 = 0
    if change_loop[0] == 0:
        change_current_date[0] = mod_date
        plus_adult1[0] = int(adult1)
        plus_child1[0] = int(child1)
        # x = datetime.datetime.now()
        plus_date[0] = int(nonight1)
        temp_mod_nonight = 0
        temp_mod_nonight = d4 - c4
        temp_mod_nonight = temp_mod_nonight.days
        date_tom = mod_date + datetime.timedelta(days=temp_mod_nonight)
        #print('date tomorrow if condition ',date_tom ,plus_date[0])
        # plus_date[0] = 1
    else:
        x = change_current_date[0]
        diff_dates = change_date1[0].selection_get() - change_current_date[0]
        #print('current datessssssssssssssssss ',change_date1[0].selection_get() ,change_current_date[0])
        change_no_dates[0] = diff_dates.days
        date_tom = change_current_date[0] + datetime.timedelta(days=diff_dates.days)
        #print('date tomorrow if condition ', date_tom, plus_date[0])

        if diff_dates.days == 0:
            plus_date[0] = 1
        else:
            plus_date[0] = diff_dates.days
        #print('date tomorrow else condition ', date_tom, plus_date[0])

    last_date1 = date_tom.strftime("%b %d, %Y")
    out_year = date_tom.strftime('%Y')
    out_month = date_tom.strftime('%m')
    out_days = date_tom.strftime('%d')
    for y in out_month:
        month_checkout = y
    last_date2 = date_tom.strftime("%B %d %Y")
    last_date3 = date_tom.strftime("%d %m %Y")
    change_in_out_value[1] = last_date2
    change_in_out_value[3] = last_date3
    # get check out choose dates
    choose_dates[3] = date_tom.strftime("%b")
    choose_dates[4] = date_tom.strftime("%d")
    choose_dates[5] = date_tom.strftime("%Y")
    choose_dates[6] = plus_adult1[0]
    choose_dates[7] = plus_child1[0]
    #print('check-in and check-out dates ', change_in_out_value[0], change_in_out_value[1])
    #print('check-in and check-out dates ', c4, d4)

    '''if plus_date[0] < 0:
        current_dates = c4.strftime("%b %d, %Y")
        last_date1 = d4.strftime("%b %d, %Y")'''

    my_frame = Frame(window)
    canvas = Canvas(my_frame)

    canvas = Canvas(my_frame, width=373, height=440, background="lightgray")
    canvas.pack(fill="both", expand=True)

    vsb = Scrollbar(my_frame, orient="vertical", command=canvas.yview, width=25)
    hsb = Scrollbar(my_frame, orient="horizontal", command=canvas.xview, width=25)

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)
    canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
    canvas.grid(row=0, column=0, sticky="nsew")

    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    resv_head = Label(canvas, text="   Reservation Details ", font=('arial', 15),
                      bg=levelcolor, foreground='blue', padx=0, pady=10)

    site_no = Label(canvas, text="   Site Number", font=('arial', 10),
                    bg='lightgray', foreground='blue', padx=0, pady=10, )

    site_entry = Entry(canvas, font=('arial', 11), width=10)
    #site_entry.insert(0, choose_entry[0])
    choose_entry[0] = site_entry

    site_entry_button = Button(canvas, text='Select', font=('arial', 10), fg='black', bg="#F0F0F0",
                               activeforeground="#00695C", activebackground="#F0F0F0", padx=15, pady=1, width=5,
                               # wraplength= 1,
                               # state='disable',
                               command=choose_site_entry_button)
    canvas.create_window(230, 73, anchor="nw", window=site_entry_button)
    # check in and check out
    checkin_design = Label(canvas, bg='#EEEEEE', padx=142, pady=20, highlightbackground="lightgray",
                           highlightthickness=1, border=0)
    checkin_dates = Label(canvas, text='check-in', font=('arial', 9), fg='grey', bg='#EEEEEE')
    checkin_dates1 = Label(canvas, text=str(current_dates), font=('arial', 10), fg='#4D4D4D', bg='#EEEEEE')

    checkout_dates = Label(canvas, text='check-out', font=('arial', 9), fg='grey', bg='#EEEEEE')
    checkin_design1 = Label(canvas, fg='#EEEEEE', bg='#EEEEEE', padx=0, pady=20, highlightbackground="lightgray",
                            highlightthickness=1, border=0)
    checkout_dates1 = Label(canvas, text=last_date1, font=('arial', 10), fg='#4D4D4D', bg='#EEEEEE')
    photo_calendar = Label(canvas, image=photo1, bg='#EEEEEE', height=24, width=24)
    photo_calendar1 = Label(canvas, image=photo1, bg='#EEEEEE', height=24, width=24)
    # nyt
    checkin_design_nyts = Label(canvas, bg='#EEEEEE', padx=142, pady=15, highlightbackground="lightgray",
                                highlightthickness=1, border=0)
    checkin_nyts = Label(canvas, text='Nights', font=('arial', 9), fg='grey', bg='#EEEEEE')
    checkin_nytsno = Label(canvas, text=plus_date[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')

    plus_nyts = Button(canvas, text='+', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                       activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                       command=plus_button)

    minus_nyts = Button(canvas, text='-', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                        activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                        command=minus_button)
    # room
    checkin_design_room = Label(canvas, bg='#EEEEEE', padx=142, pady=15, highlightbackground="lightgray",
                                highlightthickness=1, border=0)
    rooms = Label(canvas, text='Rooms', font=('arial', 9), fg='grey', bg='#EEEEEE')
    rooms_no1 = Label(canvas, text=plus_room1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    plus_room = Button(canvas, text='+', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                       activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                       command=room_button)

    minus_room = Button(canvas, text='-', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                        activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                        command=room_button1)

    # adult
    checkin_design_adult = Label(canvas, bg='#EEEEEE', padx=142, pady=15, highlightbackground="lightgray",
                                 highlightthickness=1, border=0)
    adults = Label(canvas, text='Adults', font=('arial', 9), fg='grey', bg='#EEEEEE')
    adults_no = Label(canvas, text=plus_adult1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    plus_adults = Button(canvas, text='+', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                         activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                         command=room_adults)
    minus_adults = Button(canvas, text='-', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                          activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                          command=room_adults1)

    # child
    checkin_design_child = Label(canvas, bg='#EEEEEE', padx=142, pady=15, highlightbackground="lightgray",
                                 highlightthickness=1, border=0)
    child = Label(canvas, text='Child', font=('arial', 9), fg='grey', bg='#EEEEEE')
    child_no = Label(canvas, text=plus_child1[0], font=('arial', 9), fg='#4D4D4D', bg='#EEEEEE')
    plus_child = Button(canvas, text='+', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                        activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                        command=room_child)
    minus_child = Button(canvas, text='-', font=('arial', 10), fg='#00695C', bg='#EEEEEE',
                         activeforeground='#00695C', activebackground='#EEEEEE', padx=12, pady=1, width=1,
                         command=room_child1)

    # spacing
    k31 = Label(canvas, text='               ', font=('arial', 10), bg="lightgray", foreground='black', padx=0, pady=10)
    k32 = Label(canvas, text='               ', font=('arial', 10), bg="lightgray", foreground='black', padx=0, pady=10)

    canvas.create_window(3, 3, anchor="nw", window=resv_head)
    canvas.create_window(10, 70, anchor="nw", window=site_no)
    canvas.create_window(110, 77, anchor="nw", window=site_entry)
    canvas.create_window(20, 120, anchor="nw", window=checkin_design)
    canvas.create_window(30, 127, anchor="nw", window=checkin_dates)
    canvas.create_window(30, 150, anchor="nw", window=checkin_dates1)
    canvas.create_window(168, 127, anchor="nw", window=checkout_dates)
    canvas.create_window(160, 120, anchor="nw", window=checkin_design1)
    canvas.create_window(168, 150, anchor="nw", window=checkout_dates1)
    canvas.create_window(122, 123, anchor="nw", window=photo_calendar)
    canvas.create_window(264, 123, anchor="nw", window=photo_calendar1)
    # plus minus nyts
    canvas.create_window(20, 185, anchor="nw", window=checkin_design_nyts)
    canvas.create_window(30, 190, anchor="nw", window=checkin_nyts)
    canvas.create_window(30, 208, anchor="nw", window=checkin_nytsno)
    canvas.create_window(210, 195, anchor="nw", window=minus_nyts)
    canvas.create_window(250, 195, anchor="nw", window=plus_nyts)
    # room
    canvas.create_window(20, 240, anchor="nw", window=checkin_design_room)
    canvas.create_window(30, 245, anchor="nw", window=rooms)
    canvas.create_window(30, 263, anchor="nw", window=rooms_no1)
    canvas.create_window(210, 250, anchor="nw", window=minus_room)
    canvas.create_window(250, 250, anchor="nw", window=plus_room)
    # adults
    canvas.create_window(20, 295, anchor="nw", window=checkin_design_adult)
    canvas.create_window(30, 300, anchor="nw", window=adults)
    canvas.create_window(30, 318, anchor="nw", window=adults_no)
    canvas.create_window(210, 305, anchor="nw", window=minus_adults)
    canvas.create_window(250, 305, anchor="nw", window=plus_adults)
    # child
    canvas.create_window(20, 355, anchor="nw", window=checkin_design_child)
    canvas.create_window(30, 360, anchor="nw", window=child)
    canvas.create_window(30, 377, anchor="nw", window=child_no)
    canvas.create_window(210, 365, anchor="nw", window=minus_child)
    canvas.create_window(250, 365, anchor="nw", window=plus_child)

    photo_calendar.bind("<Button-1>", fetch_date)
    photo_calendar1.bind("<Button-1>", fetch_date)

    my_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
    my_frame.place(x=870, y=176)
    view_return.append(my_frame)
    resv_forget.append(my_frame)

    my_frame1 = Frame(window)
    canvas = Canvas(my_frame1)

    canvas = Canvas(my_frame1, width=625, height=380, background="lightgray")
    canvas.pack(fill="both", expand=True)

    vsb = Scrollbar(my_frame1, orient="vertical", command=canvas.yview, width=25)

    hsb = Scrollbar(my_frame1, orient="horizontal", command=canvas.xview, width=25)

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)
    canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
    canvas.grid(row=0, column=0, sticky="nsew")

    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    # value_resv_stay = resv_stay(label53.get())
    value_resv_stay = 26
    # print('reservation valueeeeeeeeeeeeeeeeeee ', value_resv_stay)
    # levelcolor = value_resv_stay[1]
    levelcolor = 'lightblue'
    property = Label(canvas, text='  Property Name  ', font=('arial', 15), bg=levelcolor, fg='blue', padx=0,
                     pady=10, )
    canvas.create_window(10, 5, anchor="nw", window=property)
    for data_search in range(2, data_number3.max_row + 1):
        siteno = data_number3.cell(data_search, 1).value
        hotelname = data_number3.cell(data_search, 2).value
        address = data_number3.cell(data_search, 3).value
        phoneno = data_number3.cell(data_search, 4).value
        rating = data_number3.cell(data_search, 5).value

        rrodking = data_number3.cell(data_search, 8).value
        swriking = data_number3.cell(data_search, 10).value
        arrpking = data_number3.cell(data_search, 12).value
        fastking = data_number3.cell(data_search, 14).value

        gofastpoints = data_number3.cell(data_search, 16).value
        freeking = data_number3.cell(data_search, 19).value

        rrodqueens = data_number3.cell(data_search, 9).value
        swriqueens = data_number3.cell(data_search, 11).value
        arrpqueens = data_number3.cell(data_search, 13).value
        fastqueens = data_number3.cell(data_search, 15).value

        # print('data search ',data_search)

        if str(siteno) == str(site):
            z = 1
            x = 60
            temp_siteno = siteno
            count_rooms = nonight1

            prop = Label(canvas, text='  ' + '(' + str(siteno) + ')  ' + hotelname, font=('arial', 12),
                         fg='black',
                         bg='lightgray')
            canvas.create_window(10, x, anchor="nw", window=prop)

            prop1 = Label(canvas, text='   ' + phoneno + ',  ' + rating, font=('arial', 8), fg='blue',
                          bg='lightgray')
            canvas.create_window(10, x + 22, anchor="nw", window=prop1)

            prop2 = Label(canvas, text='   ' + address, font=('arial', 9), fg='black', bg='lightgray', )
            canvas.create_window(10, x + 42, anchor="nw", window=prop2)

            prop3 = Label(canvas, text='  ' + 'Rates', font=('arial', 11), fg='blue', bg='lightgray')
            canvas.create_window(10, x + 72, anchor="nw", window=prop3)

            prop4 = Label(canvas, text='One King', font=('arial', 11), fg='blue', bg='lightgray')
            canvas.create_window(150, x + 72, anchor="nw", window=prop4)

            prop5 = Label(canvas, text='Taxes', font=('arial', 11), fg='blue', bg="lightgray")
            canvas.create_window(300, x + 72, anchor="nw", window=prop5)

            prop6 = Label(canvas, text='Total', font=('arial', 11), fg='blue', bg="lightgray")
            canvas.create_window(420, x + 72, anchor="nw", window=prop6)

            prop7 = Label(canvas, text='Two Queens', font=('arial', 11), fg='blue', bg="lightgray")
            canvas.create_window(650, x + 72, anchor="nw", window=prop7)

            prop8 = Label(canvas, text='Taxes', font=('arial', 11), fg='blue', bg="lightgray")
            canvas.create_window(800, x + 72, anchor="nw", window=prop8)

            prop9 = Label(canvas, text='Total', font=('arial', 11), fg='blue', bg="lightgray")
            canvas.create_window(920, x + 72, anchor="nw", window=prop9)

            prop10 = Label(canvas, text='   rrod', font=('arial', 9), fg='blue', bg='lightgray')
            canvas.create_window(10, x + 100, anchor="nw", window=prop10)

            prop11 = Label(canvas, text='   swr1', font=('arial', 9), fg='blue', bg='lightgray')
            canvas.create_window(10, x + 120, anchor="nw", window=prop11)

            prop12 = Label(canvas, text='   aarp', font=('arial', 9), fg='blue', bg="lightgray")
            canvas.create_window(10, x + 140, anchor="nw", window=prop12)

            prop13 = Label(canvas, text='   go fast', font=('arial', 9), fg='blue', bg="lightgray")
            canvas.create_window(10, x + 160, anchor="nw", window=prop13)

            prop14 = Label(canvas, text='   go free', font=('arial', 9), fg='blue', bg="lightgray")
            canvas.create_window(10, x + 180, anchor="nw", window=prop14)

            # king rates
            prop15 = Label(canvas, text='$' + str(conversion_numbers1(rrodking)), font=('arial', 9), fg='black',
                           bg='lightgray')
            canvas.create_window(150, x + 100, anchor="nw", window=prop15)

            prop16 = Label(canvas, text='$' + str(conversion_numbers1(swriking)), font=('arial', 9), fg='black',
                           bg='lightgray')
            canvas.create_window(150, x + 120, anchor="nw", window=prop16)

            prop17 = Label(canvas, text='$' + str(conversion_numbers1(arrpking)), font=('arial', 9), fg='black',
                           bg="lightgray")
            canvas.create_window(150, x + 140, anchor="nw", window=prop17)

            prop18 = Label(canvas, text='$' + str(conversion_numbers1(fastking)), font=('arial', 9), fg='black',
                           bg="lightgray")
            canvas.create_window(150, x + 160, anchor="nw", window=prop18)

            prop19 = Label(canvas, text=str(conversion_numbers(freeking)) + ' pts', font=('arial', 9),
                           fg='black',
                           bg="lightgray")
            canvas.create_window(150, x + 180, anchor="nw", window=prop19)

            # king taxes rate
            prop20 = Label(canvas, text='$' + str(conversion_numbers1(rrodking * .05)), font=('arial', 9),
                           fg='black', bg='lightgray')
            canvas.create_window(300, x + 100, anchor="nw", window=prop20)

            prop21 = Label(canvas, text='$' + str(conversion_numbers1(swriking * .05)), font=('arial', 9),
                           fg='black', bg='lightgray')
            canvas.create_window(300, x + 120, anchor="nw", window=prop21)

            prop22 = Label(canvas, text='$' + str(conversion_numbers1(arrpking * .05)), font=('arial', 9),
                           fg='black', bg="lightgray")
            canvas.create_window(300, x + 140, anchor="nw", window=prop22)

            prop23 = Label(canvas, text='$' + str(conversion_numbers1(fastking * .05)), font=('arial', 9),
                           fg='black', bg="lightgray")
            canvas.create_window(300, x + 160, anchor="nw", window=prop23)

            prop24 = Label(canvas, text='$' + str(conversion_numbers1(0)), font=('arial', 9), fg='black',
                           bg="lightgray")
            canvas.create_window(300, x + 180, anchor="nw", window=prop24)

            # king total rate
            prop25 = Label(canvas, text='$' + str(
                conversion_numbers1((rrodking * count_rooms) + ((rrodking * count_rooms) * .05))),
                           font=('arial', 9), fg='black', bg='lightgray')
            canvas.create_window(420, x + 100, anchor="nw", window=prop25)

            prop26 = Label(canvas, text='$' + str(
                conversion_numbers1((swriking * count_rooms) + ((swriking * count_rooms) * .05))),
                           font=('arial', 9), fg='black', bg='lightgray')
            canvas.create_window(420, x + 120, anchor="nw", window=prop26)

            prop27 = Label(canvas, text='$' + str(
                conversion_numbers1((arrpking * count_rooms) + ((arrpking * count_rooms) * .05))),
                           font=('arial', 9), fg='black', bg="lightgray")
            canvas.create_window(420, x + 140, anchor="nw", window=prop27)

            prop28 = Label(canvas, text='$' + str(
                conversion_numbers1((fastking * count_rooms) + ((fastking * count_rooms) * .05))) + ' + ' + str(
                conversion_numbers((gofastpoints * count_rooms))) + ' pts', font=('arial', 9), fg='black',
                           bg="lightgray")
            canvas.create_window(420, x + 160, anchor="nw", window=prop28)

            prop29 = Label(canvas, text=str(conversion_numbers((freeking * count_rooms))) + ' pts',
                           font=('arial', 9), fg='black',
                           bg="lightgray")
            canvas.create_window(420, x + 180, anchor="nw", window=prop29)

            # two queens beds
            prop29 = Label(canvas, text='$' + str(conversion_numbers1(rrodqueens)),
                           font=('arial', 9), fg='black', bg='lightgray')
            canvas.create_window(650, x + 100, anchor="nw", window=prop29)

            prop30 = Label(canvas, text='$' + str(conversion_numbers1(swriqueens)), font=('arial', 9),
                           fg='black', bg='lightgray')
            canvas.create_window(650, x + 120, anchor="nw", window=prop30)

            prop31 = Label(canvas, text='$' + str(conversion_numbers1(arrpqueens)), font=('arial', 9),
                           fg='black', bg="lightgray")
            canvas.create_window(650, x + 140, anchor="nw", window=prop31)

            prop32 = Label(canvas, text='$' + str(conversion_numbers1(fastqueens)), font=('arial', 9),
                           fg='black',
                           bg="lightgray")
            canvas.create_window(650, x + 160, anchor="nw", window=prop32)

            prop33 = Label(canvas, text=str(conversion_numbers(freeking)) + ' pts', font=('arial', 9),
                           fg='black',
                           bg="lightgray")
            canvas.create_window(650, x + 180, anchor="nw", window=prop33)

            # two queens beds taxes
            prop34 = Label(canvas, text='$' + str(conversion_numbers1(rrodqueens * .05)), font=('arial', 9),
                           fg='black', bg='lightgray')
            canvas.create_window(800, x + 100, anchor="nw", window=prop34)

            prop35 = Label(canvas, text='$' + str(conversion_numbers1(swriqueens * .05)), font=('arial', 9),
                           fg='black', bg='lightgray')
            canvas.create_window(800, x + 120, anchor="nw", window=prop35)

            prop36 = Label(canvas, text='$' + str(conversion_numbers1(arrpqueens * .05)), font=('arial', 9),
                           fg='black', bg="lightgray")
            canvas.create_window(800, x + 140, anchor="nw", window=prop36)

            prop37 = Label(canvas, text='$' + str(conversion_numbers1(fastqueens * .05)), font=('arial', 9),
                           fg='black', bg="lightgray")
            canvas.create_window(800, x + 160, anchor="nw", window=prop37)

            prop38 = Label(canvas, text='$' + str(conversion_numbers1(0)), font=('arial', 9), fg='black',
                           bg="lightgray")
            canvas.create_window(800, x + 180, anchor="nw", window=prop38)

            # two queens beds total taxes
            prop39 = Label(canvas, text='$' + str(
                conversion_numbers1((rrodqueens * count_rooms) + ((rrodqueens * count_rooms) * .05))),
                           font=('arial', 9), fg='black', bg='lightgray')
            canvas.create_window(920, x + 100, anchor="nw", window=prop39)

            prop40 = Label(canvas, text='$' + str(
                conversion_numbers1((swriqueens * count_rooms) + ((swriqueens * count_rooms) * .05))),
                           font=('arial', 9),
                           fg='black', bg='lightgray')
            canvas.create_window(920, x + 120, anchor="nw", window=prop40)

            prop41 = Label(canvas, text='$' + str(
                conversion_numbers1((arrpqueens * count_rooms) + ((arrpqueens * count_rooms) * .05))),
                           font=('arial', 9),
                           fg='black', bg="lightgray")
            canvas.create_window(920, x + 140, anchor="nw", window=prop41)

            prop42 = Label(canvas,
                           text='$' + str(conversion_numbers1(
                               (fastqueens * count_rooms) + ((fastqueens * count_rooms) * .05))) + ' + ' + str(
                               conversion_numbers((gofastpoints * count_rooms))) + ' pts', font=('arial', 9),
                           fg='black',
                           bg="lightgray")
            canvas.create_window(920, x + 160, anchor="nw", window=prop42)

            prop43 = Label(canvas, text=str(conversion_numbers((freeking * count_rooms))) + ' pts',
                           font=('arial', 9), fg='black',
                           bg="lightgray")
            canvas.create_window(920, x + 180, anchor="nw", window=prop43)

            ratebookcombo = Label(canvas, text="  Rate Code", font=('arial', 10), bg='lightgray',
                                  foreground='blue')
            bookcombo = ttk.Combobox(canvas, font=('arial', 10),
                                     values=['Rates', 'RROD', 'AARP', 'SWR1', 'Go Fast', 'Go Free'],
                                     width=10, foreground='#263238')
            bookcombo.insert(0, value_book)
            choose_dates[8] = bookcombo

            ratebedscombo = Label(canvas, text="  Bed Class", font=('arial', 10), bg='lightgray',
                                  foreground='blue')
            bedscombo = ttk.Combobox(canvas, font=('arial', 10),
                                     values=['Bed Class', 'One Bed', 'Two Beds'],
                                     width=10, foreground='#263238')
            bedscombo.insert(0, value_beds)
            choose_dates[9] = bedscombo

            canvas_nyts = Label(canvas, text="Night", font=('arial', 10), bg='lightgray', foreground='blue')
            canvas_nyts1 = Label(canvas, text=count_rooms, font=('arial', 9), background='white', padx=12,
                                 pady=1, border=1, relief='sunken')

            canvas.create_window(10, x + 240, anchor="nw", window=ratebookcombo)
            canvas.create_window(100, x + 240, anchor="nw", window=bookcombo)
            canvas.create_window(230, x + 240, anchor="nw", window=ratebedscombo)
            canvas.create_window(320, x + 240, anchor="nw", window=bedscombo)

            canvas.create_window(450, x + 240, anchor="nw", window=canvas_nyts)
            canvas.create_window(500, x + 240, anchor="nw", window=canvas_nyts1)

            spacing = Label(canvas, text=' ', font=('arial', 9), fg='black',
                            bg="lightgray")
            # canvas.create_window(1150, x + 330, anchor="nw", window=spacing)

    my_frame1.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
    my_frame1.place(x=200, y=176)
    view_return.append(my_frame1)
    resv_forget.append(my_frame1)

    # remove button cxl resv

    '''guest_search3_dates = Button(window, text='confirm', font=('arial', 10), fg='black', bg='lightgray',
                                 activeforeground="#00695C", activebackground="lightgray", padx=15, pady=2, width=5,
                                 # wraplength= 1,
                                 #state='disable',
                                 command=choose_modify)
    guest_search3_dates.place(x=740, y=600)

    guest_back_mod = Button(window, text='Back Mod', font=('arial', 10), fg='black', bg="lightgray",
                            activeforeground="#00695C", activebackground="lightgray", padx=20, pady=2, width=5,
                            # wraplength= 1,
                            # state='disable',
                            command=guest_next_modify)
    #guest_back_mod.place(x=210, y=600)
    #guest_back_mod.place(x=530, y=600)
    guest_back_mod.place(x=635, y=600)'''

    change_resv = Button(window, text='Change Property', font=('arial', 9), fg='black', bg='lightgray',
                         activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                         # wraplength= 1,
                         # state='disable',
                         command=guest_next_modify)
    change_resv.place(x=650, y=585)
    resv_forget.append(change_resv)

    confirm_resv = Button(window, text='Confirm Reservation', font=('arial', 9), fg='black', bg='lightgray',
                          activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                          # wraplength= 1,
                          # state='disable',
                          command=choose_modify)
    confirm_resv.place(x=650, y=615)
    resv_forget.append(confirm_resv)


# choose dates confirm
def choose_modify():
    print('\nchoose modify reservation confirmation')
    global mod_acct_no, get_resv_no, mod_siteno, mod_nonight, mod_loops, mod_rooms, mod_adults, mod_childs

    # choose_resv[0] = resvno
    # choose_acctno[0] = account_no
    # choose_siteno[0] = site no
    validate_date = 0
    if mod_loops == 1:
        choose_dates[0] = modify_dates[0]
        choose_dates[1] = modify_dates[1]
        choose_dates[2] = modify_dates[2]
        choose_dates[3] = modify_dates[3]
        choose_dates[4] = modify_dates[4]
        choose_dates[5] = modify_dates[5]
        choose_dates[6] = modify_dates[6]
        choose_dates[7] = modify_dates[7]
        choose_dates[8] = modify_dates[8]
        choose_dates[9] = modify_dates[9]

        '''for x in range (0, len(choose_dates)):
            if x == 8 or x == 9:
                print('modify datessssss ', choose_dates[x].get())
            else:
                print('modify datesssssssss ',choose_dates[x])'''

        #valid_entry1 = mod_nonight
        choose_siteno[0] = mod_siteno
        choose_acctno[0] = mod_acct_no
        choose_resv[0] = get_resv_no
        #print('account no, site no, reservation  no , valid entry1 ', mod_acct_no ,mod_siteno , choose_resv[0], mod_nonight)

    inmonth = datetime.datetime.strptime(choose_dates[0], '%b').month
    outmonth = datetime.datetime.strptime(choose_dates[3], '%b').month
    validate_date = wyn_tscroll_dates_GUI.book_dates(choose_dates[2], inmonth, choose_dates[1],
                                                     choose_dates[5], outmonth, choose_dates[4])
    a = int(choose_dates[2])
    b = int(inmonth)
    c = int(choose_dates[1])
    checkinn = date(a, b, c)
    x = int(choose_dates[5])
    y = int(outmonth)
    z = int(choose_dates[4])
    checkout = date(x, y, z)
    valid_entry_out = checkout - checkinn
    valid_entry1 = valid_entry_out.days

    if mod_loops == 1:
        valid_entry1 = mod_nonight


    check_ins = datetime.datetime(int(a), int(b), int(c))
    check_ins1 = check_ins.strftime("%a, %b %d -")

    check_out = datetime.datetime(int(x), int(y), int(z))
    check_out1 = check_out.strftime(" %a, %b %d %Y")

    # confirmation dates
    confirmed_dates = check_ins.strftime("%B %d, %Y")

    if choose_dates[-2].get() == 'Rates':
        validate_date = 1
    else:
        bookrates = choose_dates[-2].get()

    if choose_dates[-1].get() == 'Bed Class':
        validate_date = 1

    if valid_entry1 == 0:
        valid_entry1 = 1
    elif valid_entry1 < 0:
        validate_date == 1


    if validate_date == 1:
        messagebox.showinfo(title='Validation Entry ', message='Invalid Entry')
        #print('book rooms ',choose_dates)
        #post()
        #guest_search_modify()

    else:
        #print('book rooms ', choose_dates)
        nonyts = valid_entry1
        if int(nonyts) == 1:
            night_message = 'night'
        else:
            night_message = 'nights'

        name_hotel1 = ''
        for data_search in range(2, data_number3.max_row + 1):
            no_site = data_number3.cell(data_search, 1).value
            name_hotel = data_number3.cell(data_search, 2).value
            if str(no_site) == str(choose_siteno[0]):
                name_hotel1 = name_hotel

        answer = messagebox.askquestion(title='Mofify Reservation',
                                        message='I have you Confirmed in ' + str(name_hotel1) +
                                                ' on ' + str(confirmed_dates) +
                                                ' for ' + str(nonyts) + ' ' + str(night_message))

        if answer == 'yes':
            book_rooms = choose_dates
            for x in range(0, len(book_rooms)):
                # print('reservation temporary booked ',num1 ,book_rooms[x].get())
                checkinmonth1 = book_rooms[0]
                checkindays = book_rooms[1]
                checkinyear = book_rooms[2]
                checkoutmonth1 = book_rooms[3]
                checkoutdays = book_rooms[4]
                checkoutyear = book_rooms[5]
                bookadults = book_rooms[6]
                bookchild = book_rooms[7]
                bookrates = book_rooms[8].get()
                bookbeds = book_rooms[9].get()

            checkinmonth = datetime.datetime.strptime(checkinmonth1, '%b').month
            checkoutmonth = datetime.datetime.strptime(checkoutmonth1, '%b').month
            date1 = date(int(checkinyear), int(checkinmonth), int(checkindays))
            date2 = date(int(checkoutyear), int(checkoutmonth), int(checkoutdays))
            date1_in = str(date1.strftime('"%d/%m/%Y"'))
            date2_out = str(date2.strftime('"%d/%m/%Y"'))

            # print('Reservataion Number ', choose_resv[0])
            # print('Number of row data number4 ', data_number4.max_row)

            # rate names
            ratenames = bookrates
            if bookrates == 'RROD':
                ratenames = 'Wyndham rewards flexible rate'
            elif bookrates == 'SWR1':
                ratenames = 'Wyndham rewards member rate'
            elif bookrates == 'AARP':
                ratenames = 'Wyndham rewards aarp rate'
            elif bookrates == 'Go Fast':
                ratenames = 'Wyndham rewards go fast rate'
            elif bookrates == 'Go Free':
                ratenames = 'Wyndham rewards go free rate'
            else:
                messagebox.showinfo(title='Bookrates ', message='Invalid Book Rates')
                pass

            accountno = choose_acctno[0]
            for data_search1 in range(2, data_number.max_row + 1):
                acctno = data_number.cell(data_search1, 1).value
                phone_data = data_number.cell(data_search1, 2).value
                lastname = data_number.cell(data_search1, 3).value
                firstname = data_number.cell(data_search1, 4).value
                middlename = data_number.cell(data_search1, 5).value
                home_data = data_number.cell(data_search1, 17).value
                if str(acctno) == str(accountno):
                    savecell = phone_data
                    booklast = lastname
                    bookfirst = firstname
                    bookmiddle = middlename

            if book_rooms[9].get() == 'One Bed':
                bookbeds = 'One King Bed'

            elif book_rooms[9].get() == 'One King Bed':
                bookbeds = 'One King Bed'
            else:
                bookbeds = 'Two Queen Beds'

            data_count2, data_count4 = 0, 0
            for data_search in range(2, data_number2.max_row + 1):
                resvno_data2 = data_number2.cell(data_search, 2).value
                if str(resvno_data2) == str(choose_resv[0]):
                    data_count2 = data_search

            for data_search1 in range(2, data_number4.max_row + 1):
                resvno_data4 = data_number4.cell(data_search1, 2).value
                if str(resvno_data4) == str(choose_resv[0]):
                    data_count4 = data_search1
            #print('data counttttttttttttttttttttttt ',data_count2 ,data_count4)
            # print('universal site no ', choose_siteno[0])
            siteno = choose_siteno[0]
            bookrooms = valid_entry1
            hotelname1 = ''
            newaddress = ''
            savecell1 = ''

            for data_search in range(2, data_number3.max_row + 1):
                no_site = data_number3.cell(data_search, 1).value
                hotelname = data_number3.cell(data_search, 2).value
                address = data_number3.cell(data_search, 3).value
                phoneno = data_number3.cell(data_search, 4).value
                hotelstar = data_number3.cell(data_search, 5).value

                rrodking = data_number3.cell(data_search, 8).value
                swriking = data_number3.cell(data_search, 10).value
                arrpking = data_number3.cell(data_search, 12).value
                fastking = data_number3.cell(data_search, 14).value
                gofastpoints = data_number3.cell(data_search, 16).value
                freeking = data_number3.cell(data_search, 19).value
                rrodqueens = data_number3.cell(data_search, 9).value
                swriqueens = data_number3.cell(data_search, 11).value
                arrpqueens = data_number3.cell(data_search, 13).value
                fastqueens = data_number3.cell(data_search, 15).value
                loops = 0
                if str(no_site) == str(siteno):
                    newaddress = address
                    savecell1 = phoneno
                    hotelname1 = hotelname
                    loops = 1
                    if bookrates == 'RROD':
                        if bookbeds == 'One King Bed':
                            bookrooms = int(bookrooms)
                            total_amount = '$ ' + str(conversion_numbers1(rrodking))
                            total_amount1 = 'USD ' + str(conversion_numbers1(rrodking * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((rrodking * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((rrodking * bookrooms) * .05 + (rrodking * bookrooms)))

                            newtotal_amount = rrodking
                            newtotal_amount1 = rrodking * bookrooms
                            newtaxes_new = (rrodking * bookrooms) * .05
                            newpayment = ((rrodking * bookrooms) * .05) + (rrodking * bookrooms)
                            newtaxes = rrodking * .05
                            newgofast = 0
                            newgofreetotal = 0

                        else:
                            bookrooms = int(bookrooms)
                            total_amount = '$ ' + str(conversion_numbers1(rrodqueens))
                            total_amount1 = 'USD ' + str(conversion_numbers1(rrodqueens * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((rrodqueens * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((rrodqueens * bookrooms) * .05 + (rrodqueens * bookrooms)))

                            newtotal_amount = rrodqueens
                            newtotal_amount1 = rrodqueens * bookrooms
                            newtaxes_new = (rrodqueens * bookrooms) * .05
                            newpayment = ((rrodqueens * bookrooms) * .05) + (rrodqueens * bookrooms)
                            newtaxes = rrodqueens * .05
                            newgofast = 0
                            newgofreetotal = 0

                    elif bookrates == 'SWR1':
                        bookrooms = int(bookrooms)
                        if bookbeds == 'One King Bed':
                            total_amount = '$ ' + str(conversion_numbers1(swriking))
                            total_amount1 = 'USD ' + str(conversion_numbers1(swriking * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((swriking * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((swriking * bookrooms) * .05 + (swriking * bookrooms)))

                            newtotal_amount = swriking
                            newtotal_amount1 = swriking * bookrooms
                            newtaxes_new = (swriking * bookrooms) * .05
                            newpayment = ((swriking * bookrooms) * .05) + (swriking * bookrooms)
                            newtaxes = swriking * .05
                            newgofast = 0
                            newgofreetotal = 0
                        else:
                            bookrooms = int(bookrooms)
                            total_amount = '$ ' + str(conversion_numbers1(swriqueens))
                            total_amount1 = 'USD ' + str(conversion_numbers1(swriqueens * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((swriqueens * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((swriqueens * bookrooms) * .05 + (swriqueens * bookrooms)))

                            newtotal_amount = swriqueens
                            newtotal_amount1 = swriqueens * bookrooms
                            newtaxes_new = (swriqueens * bookrooms) * .05
                            newpayment = ((swriqueens * bookrooms) * .05) + (swriqueens * bookrooms)
                            newtaxes = swriqueens * .05
                            newgofast = 0
                            newgofreetotal = 0

                    elif bookrates == 'AARP':

                        bookrooms = int(bookrooms)
                        if bookbeds == 'One King Bed':
                            # print('type of number ', arrpking, )
                            total_amount = '$ ' + str(conversion_numbers1(arrpking))
                            total_amount1 = 'USD ' + str(conversion_numbers1(arrpking * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((arrpking * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((arrpking * bookrooms) * .05 + (arrpking * bookrooms)))

                            newtotal_amount = arrpking
                            newtotal_amount1 = arrpking * bookrooms
                            newtaxes_new = (arrpking * bookrooms) * .05
                            newpayment = ((arrpking * bookrooms) * .05) + (arrpking * bookrooms)
                            newtaxes = arrpking * .05
                            newgofast = 0
                            newgofreetotal = 0
                        else:
                            bookrooms = int(bookrooms)
                            total_amount = '$ ' + str(conversion_numbers1(arrpqueens))
                            total_amount1 = 'USD ' + str(conversion_numbers1(arrpqueens * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((arrpqueens * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((arrpqueens * bookrooms) * .05 + (arrpqueens * bookrooms)))

                            newtotal_amount = arrpqueens
                            newtotal_amount1 = arrpqueens * bookrooms
                            newtaxes_new = (arrpqueens * bookrooms) * .05
                            newpayment = ((arrpqueens * bookrooms) * .05) + (arrpqueens * bookrooms)
                            newtaxes = arrpqueens * .05
                            newgofast = 0
                            newgofreetotal = 0

                    elif bookrates == 'Go Fast':

                        bookrooms = int(bookrooms)
                        if bookbeds == 'One King Bed':
                            total_amount = '$' + str(conversion_numbers1(fastking)) + ' + ' + str(
                                conversion_numbers(gofastpoints)) + ' pts'
                            total_amount1 = 'USD ' + str(conversion_numbers1(fastking * bookrooms)) + ' + ' + str(
                                conversion_numbers(int(gofastpoints) * bookrooms)) + ' pts'
                            taxes = 'USD ' + str(conversion_numbers1((fastking * bookrooms) * .05))
                            payment = 'USD ' + str(conversion_numbers1(
                                (fastking * bookrooms) * .05 + (fastking * bookrooms))) + ' + ' + str(
                                conversion_numbers(gofastpoints * bookrooms)) + ' pts'

                            newtotal_amount = fastking
                            newtotal_amount1 = fastking * bookrooms
                            newtaxes_new = (fastking * bookrooms) * .05
                            newpayment = ((fastking * bookrooms) * .05) + (fastking * bookrooms)
                            newtaxes = fastking * .05
                            newgofast = gofastpoints
                            newgofreetotal = int(gofastpoints) * int(bookrooms)

                        else:
                            bookrooms = int(bookrooms)

                            total_amount = '$' + str(conversion_numbers1(fastqueens)) + ' + ' + str(
                                conversion_numbers(gofastpoints)) + ' pts'
                            total_amount1 = 'USD ' + str(conversion_numbers1(fastqueens * bookrooms)) + ' + ' + str(
                                conversion_numbers(int(gofastpoints) * bookrooms)) + ' pts'
                            taxes = 'USD ' + str(conversion_numbers1((fastqueens * bookrooms) * .05))
                            payment = 'USD ' + str(conversion_numbers1(
                                (fastqueens * bookrooms) * .05 + (fastqueens * bookrooms))) + ' + ' + str(
                                conversion_numbers(int(gofastpoints) * bookrooms)) + ' pts'

                            newtotal_amount = fastqueens
                            newtotal_amount1 = fastqueens * bookrooms
                            newtaxes_new = (fastqueens * bookrooms) * .05
                            newpayment = ((fastqueens * bookrooms) * .05) + (fastqueens * bookrooms)
                            newtaxes = fastqueens * .05
                            newgofast = conversion_numbers(gofastpoints)
                            newgofreetotal = conversion_numbers(int(gofastpoints) * int(bookrooms))

                    elif bookrates == 'Go Free':

                        bookrooms = int(bookrooms)
                        total_amount = str(conversion_numbers(freeking)) + ' pts'
                        total_amount1 = str(conversion_numbers(int(freeking) * int(bookrooms))) + ' pts'
                        taxes = 'USD ' + str(conversion_numbers1(0))
                        payment = str(conversion_numbers(int(freeking) * int(bookrooms))) + ' pts'

                        newtotal_amount = 0
                        newtotal_amount1 = 0
                        newtaxes_new = 0
                        newpayment = 0
                        newtaxes = 0
                        newgofast = freeking
                        newgofreetotal = int(freeking) * int(bookrooms)

                    else:
                        pass

                if loops == 1:
                    if data_count4 != 0:
                        #print('modify stay ', data_count4, data_count2, int(choose_acctno[0]))
                        data_number4.cell(data_count4, 1).value, data_number2.cell(data_count2, 1).value = int(choose_acctno[0]), \
                            int(choose_acctno[0])
                        data_number4.cell(data_count4, 2).value, data_number2.cell(data_count2, 2).value \
                            = str(choose_resv[0]), str(choose_resv[0])
                        data_number4.cell(data_count4, 3).value, data_number2.cell(data_count2,
                                                                                   3).value = savecell, savecell
                        data_number4.cell(data_count4, 4).value, data_number2.cell(data_count2,
                                                                                   4).value = booklast, booklast
                        data_number4.cell(data_count4, 5).value, data_number2.cell(data_count2,
                                                                                   5).value = bookfirst, bookfirst
                        data_number4.cell(data_count4, 6).value, data_number2.cell(data_count2,
                                                                                   6).value = bookmiddle, bookmiddle
                        data_number4.cell(data_count4, 7).value, data_number2.cell(data_count2,
                                                                                   7).value = hotelname1, hotelname1
                        data_number4.cell(data_count4, 8).value, data_number2.cell(data_count2,
                                                                                   8).value = newaddress, newaddress
                        data_number4.cell(data_count4, 9).value, data_number2.cell(data_count2,
                                                                                   9).value = savecell1, savecell1
                        data_number4.cell(data_count4, 10).value, data_number2.cell(data_count2,
                                                                                    10).value = bookrates, bookrates
                        data_number4.cell(data_count4, 11).value, data_number2.cell(data_count2,
                                                                                    11).value = str(
                            conversion_numbers1(newtotal_amount)), float(conversion_numbers2(newtotal_amount))
                        data_number4.cell(data_count4, 12).value, data_number2.cell(data_count2,
                                                                                    12).value = str(
                            conversion_numbers1(newtaxes)), float(conversion_numbers2(newtaxes))
                        data_number4.cell(data_count4, 13).value, data_number2.cell(data_count2,
                                                                                    13).value = str(
                            conversion_numbers(newgofast)), float(conversion_numbers2(newgofast))
                        data_number4.cell(data_count4, 14).value, data_number2.cell(data_count2,
                                                                                    14).value = str(
                            conversion_numbers1(newtotal_amount1)), float(conversion_numbers2(newtotal_amount1))
                        data_number4.cell(data_count4, 15).value, data_number2.cell(data_count2,
                                                                                    15).value = str(
                            conversion_numbers1(newtaxes_new)), float(conversion_numbers2(newtaxes_new))
                        data_number4.cell(data_count4, 16).value, data_number2.cell(data_count2,
                                                                                    16).value = str(
                            conversion_numbers(newgofreetotal)), float(conversion_numbers2(newgofreetotal))
                        data_number4.cell(data_count4, 17).value, data_number2.cell(data_count2,
                                                                                    17).value = str(
                            conversion_numbers1(newpayment)), float(conversion_numbers2(newpayment))
                        data_number4.cell(data_count4, 18).value, data_number2.cell(data_count2, 18).value = nonyts, nonyts
                        data_number4.cell(data_count4, 19).value, data_number2.cell(data_count2,
                                                                                    19).value = bookadults, bookadults
                        data_number4.cell(data_count4, 20).value, data_number2.cell(data_count2,
                                                                                    20).value = bookchild, bookchild
                        data_number4.cell(data_count4, 21).value, data_number2.cell(data_count2,
                                                                                    21).value = bookbeds, bookbeds
                        data_number4.cell(data_count4, 22).value, data_number2.cell(data_count2, 22).value = \
                            check_ins1, date1_in
                        data_number4.cell(data_count4, 23).value, data_number2.cell(data_count2, 23).value = \
                            check_out1, date2_out
                        data_number4.cell(data_count4, 24).value, data_number2.cell(data_count2, 24).value = \
                            ratenames, ratenames
                        data_number4.cell(data_count4, 25).value, data_number2.cell(data_count2, 25).value = 'no', 'no'
                        data_number4.cell(data_count4, 26).value, data_number2.cell(data_count2, 26).value = 'no', 'no'
                        data_number4.cell(data_count4, 27).value, data_number2.cell(data_count2,
                                                                                    27).value = 'not ready', 'not ready'
                        data_number4.cell(data_count4, 28).value, data_number2.cell(data_count2, 28).value = siteno, siteno
                        data_number4.cell(data_count4, 29).value, data_number2.cell(data_count2, 29).value = 'no', 'no'
                        data_number4.cell(data_count4, 30).value = 'confirmed'
                        # choose_entry[0] = ''
                        # choose_loopsite[0] = 0
                        # choose_siteno[0]
                        wb.save('data_wyn_gui.xlsx')
                    else:
                        print('modify stay ', data_count4, data_count2, int(choose_acctno[0]))
                        data_number2.cell(data_count2, 1).value = int(choose_acctno[0])

                        data_number2.cell(data_count2, 2).value = str(choose_resv[0])

                        data_number2.cell(data_count2, 3).value = savecell

                        data_number2.cell(data_count2, 4).value =  booklast

                        data_number2.cell(data_count2, 5).value =  bookfirst

                        data_number2.cell(data_count2, 6).value = bookmiddle

                        data_number2.cell(data_count2, 7).value = hotelname1

                        data_number2.cell(data_count2, 8).value = newaddress

                        data_number2.cell(data_count2, 9).value = savecell1

                        data_number2.cell(data_count2, 10).value = bookrates

                        data_number2.cell(data_count2, 11).value = float(conversion_numbers2(newtotal_amount))

                        data_number2.cell(data_count2, 12).value = float(conversion_numbers2(newtaxes))

                        data_number2.cell(data_count2, 13).value = float(conversion_numbers2(newgofast))

                        data_number2.cell(data_count2, 14).value = float(conversion_numbers2(newtotal_amount1))

                        data_number2.cell(data_count2, 15).value = float(conversion_numbers2(newtaxes_new))

                        data_number2.cell(data_count2, 16).value = float(conversion_numbers2(newgofreetotal))

                        data_number2.cell(data_count2, 17).value = float(conversion_numbers2(newpayment))

                        data_number2.cell(data_count2, 18).value = nonyts

                        data_number2.cell(data_count2, 19).value = bookadults

                        data_number2.cell(data_count2, 20).value = bookchild

                        data_number2.cell(data_count2, 21).value = bookbeds

                        data_number2.cell(data_count2, 22).value = date1_in

                        data_number2.cell(data_count2, 23).value = date2_out

                        data_number2.cell(data_count2, 24).value = ratenames

                        data_number2.cell(data_count2, 25).value = 'no'

                        data_number2.cell(data_count2, 26).value = 'no'

                        data_number2.cell(data_count2, 27).value = 'not ready'

                        data_number2.cell(data_count2, 28).value = siteno

                        data_number2.cell(data_count2, 29).value = 'no'

                        # choose_entry[0] = ''
                        # choose_loopsite[0] = 0
                        # choose_siteno[0]
                        wb.save('data_wyn_gui.xlsx')
    if validate_date == 1:
        pass
    else:
        my_frame1 = Frame(window)
        canvas = Canvas(my_frame1)

        canvas = Canvas(my_frame1, width=625, height=380, background="lightgray")
        canvas.pack(fill="both", expand=True)

        vsb = Scrollbar(my_frame1, orient="vertical", command=canvas.yview, width=25)

        hsb = Scrollbar(my_frame1, orient="horizontal", command=canvas.xview, width=25)

        window.grid_rowconfigure(0, weight=1)
        window.grid_columnconfigure(0, weight=1)
        canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")

        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # value_resv_stay = resv_stay(label53.get())
        value_resv_stay = 26

        levelcolor = 'lightblue'
        property = Label(canvas, text='  Property Name  ', font=('arial', 15), bg=levelcolor, fg='blue',
                         padx=0,
                         pady=10, )
        canvas.create_window(10, 5, anchor="nw", window=property)
        for data_search in range(2, data_number3.max_row + 1):
            siteno = data_number3.cell(data_search, 1).value
            hotelname = data_number3.cell(data_search, 2).value
            address = data_number3.cell(data_search, 3).value
            phoneno = data_number3.cell(data_search, 4).value
            rating = data_number3.cell(data_search, 5).value

            rrodking = data_number3.cell(data_search, 8).value
            swriking = data_number3.cell(data_search, 10).value
            arrpking = data_number3.cell(data_search, 12).value
            fastking = data_number3.cell(data_search, 14).value

            gofastpoints = data_number3.cell(data_search, 16).value
            freeking = data_number3.cell(data_search, 19).value

            rrodqueens = data_number3.cell(data_search, 9).value
            swriqueens = data_number3.cell(data_search, 11).value
            arrpqueens = data_number3.cell(data_search, 13).value
            fastqueens = data_number3.cell(data_search, 15).value

            # print('data search ',data_search)

            if str(siteno) == str(choose_siteno[0]):
                z = 1
                x = 60
                temp_siteno = siteno
                count_rooms = nonyts

                prop = Label(canvas, text='  ' + '(' + str(siteno) + ')  ' + hotelname, font=('arial', 12),
                             fg='black',
                             bg='lightgray')
                canvas.create_window(10, x, anchor="nw", window=prop)

                prop1 = Label(canvas, text='   ' + phoneno + ',  ' + rating, font=('arial', 8), fg='blue',
                              bg='lightgray')
                canvas.create_window(10, x + 22, anchor="nw", window=prop1)

                prop2 = Label(canvas, text='   ' + address, font=('arial', 9), fg='black', bg='lightgray', )
                canvas.create_window(10, x + 42, anchor="nw", window=prop2)

                prop3 = Label(canvas, text='  ' + 'Rates', font=('arial', 11), fg='blue', bg='lightgray')
                canvas.create_window(10, x + 72, anchor="nw", window=prop3)

                prop4 = Label(canvas, text='One King', font=('arial', 11), fg='blue', bg='lightgray')
                canvas.create_window(150, x + 72, anchor="nw", window=prop4)

                prop5 = Label(canvas, text='Taxes', font=('arial', 11), fg='blue', bg="lightgray")
                canvas.create_window(300, x + 72, anchor="nw", window=prop5)

                prop6 = Label(canvas, text='Total', font=('arial', 11), fg='blue', bg="lightgray")
                canvas.create_window(420, x + 72, anchor="nw", window=prop6)

                prop7 = Label(canvas, text='Two Queens', font=('arial', 11), fg='blue', bg="lightgray")
                canvas.create_window(650, x + 72, anchor="nw", window=prop7)

                prop8 = Label(canvas, text='Taxes', font=('arial', 11), fg='blue', bg="lightgray")
                canvas.create_window(800, x + 72, anchor="nw", window=prop8)

                prop9 = Label(canvas, text='Total', font=('arial', 11), fg='blue', bg="lightgray")
                canvas.create_window(920, x + 72, anchor="nw", window=prop9)

                prop10 = Label(canvas, text='   rrod', font=('arial', 9), fg='blue', bg='lightgray')
                canvas.create_window(10, x + 100, anchor="nw", window=prop10)

                prop11 = Label(canvas, text='   swr1', font=('arial', 9), fg='blue', bg='lightgray')
                canvas.create_window(10, x + 120, anchor="nw", window=prop11)

                prop12 = Label(canvas, text='   aarp', font=('arial', 9), fg='blue', bg="lightgray")
                canvas.create_window(10, x + 140, anchor="nw", window=prop12)

                prop13 = Label(canvas, text='   go fast', font=('arial', 9), fg='blue', bg="lightgray")
                canvas.create_window(10, x + 160, anchor="nw", window=prop13)

                prop14 = Label(canvas, text='   go free', font=('arial', 9), fg='blue', bg="lightgray")
                canvas.create_window(10, x + 180, anchor="nw", window=prop14)

                # king rates
                prop15 = Label(canvas, text='$' + str(conversion_numbers1(rrodking)), font=('arial', 9),
                               fg='black',
                               bg='lightgray')
                canvas.create_window(150, x + 100, anchor="nw", window=prop15)

                prop16 = Label(canvas, text='$' + str(conversion_numbers1(swriking)), font=('arial', 9),
                               fg='black',
                               bg='lightgray')
                canvas.create_window(150, x + 120, anchor="nw", window=prop16)

                prop17 = Label(canvas, text='$' + str(conversion_numbers1(arrpking)), font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(150, x + 140, anchor="nw", window=prop17)

                prop18 = Label(canvas, text='$' + str(conversion_numbers1(fastking)), font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(150, x + 160, anchor="nw", window=prop18)

                prop19 = Label(canvas, text=str(conversion_numbers(freeking)) + ' pts', font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(150, x + 180, anchor="nw", window=prop19)

                # king taxes rate
                prop20 = Label(canvas, text='$' + str(conversion_numbers1(rrodking * .05)),
                               font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(300, x + 100, anchor="nw", window=prop20)

                prop21 = Label(canvas, text='$' + str(conversion_numbers1(swriking * .05)),
                               font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(300, x + 120, anchor="nw", window=prop21)

                prop22 = Label(canvas, text='$' + str(conversion_numbers1(arrpking * .05)),
                               font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(300, x + 140, anchor="nw", window=prop22)

                prop23 = Label(canvas, text='$' + str(conversion_numbers1(fastking * .05)),
                               font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(300, x + 160, anchor="nw", window=prop23)

                prop24 = Label(canvas, text='$' + str(conversion_numbers1(0)), font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(300, x + 180, anchor="nw", window=prop24)

                # king total rate
                prop25 = Label(canvas, text='$' + str(
                    conversion_numbers1((rrodking * count_rooms) + ((rrodking * count_rooms) * .05))),
                               font=('arial', 9), fg='black', bg='lightgray')
                canvas.create_window(420, x + 100, anchor="nw", window=prop25)

                prop26 = Label(canvas, text='$' + str(
                    conversion_numbers1((swriking * count_rooms) + ((swriking * count_rooms) * .05))),
                               font=('arial', 9), fg='black', bg='lightgray')
                canvas.create_window(420, x + 120, anchor="nw", window=prop26)

                prop27 = Label(canvas, text='$' + str(
                    conversion_numbers1((arrpking * count_rooms) + ((arrpking * count_rooms) * .05))),
                               font=('arial', 9), fg='black', bg="lightgray")
                canvas.create_window(420, x + 140, anchor="nw", window=prop27)

                prop28 = Label(canvas, text='$' + str(
                    conversion_numbers1(
                        (fastking * count_rooms) + ((fastking * count_rooms) * .05))) + ' + ' + str(
                    conversion_numbers((gofastpoints * count_rooms))) + ' pts', font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(420, x + 160, anchor="nw", window=prop28)

                prop29 = Label(canvas, text=str(conversion_numbers((freeking * count_rooms))) + ' pts',
                               font=('arial', 9), fg='black',
                               bg="lightgray")
                canvas.create_window(420, x + 180, anchor="nw", window=prop29)

                # two queens beds
                prop29 = Label(canvas, text='$' + str(conversion_numbers1(rrodqueens)),
                               font=('arial', 9), fg='black', bg='lightgray')
                canvas.create_window(650, x + 100, anchor="nw", window=prop29)

                prop30 = Label(canvas, text='$' + str(conversion_numbers1(swriqueens)), font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(650, x + 120, anchor="nw", window=prop30)

                prop31 = Label(canvas, text='$' + str(conversion_numbers1(arrpqueens)), font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(650, x + 140, anchor="nw", window=prop31)

                prop32 = Label(canvas, text='$' + str(conversion_numbers1(fastqueens)), font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(650, x + 160, anchor="nw", window=prop32)

                prop33 = Label(canvas, text=str(conversion_numbers(freeking)) + ' pts', font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(650, x + 180, anchor="nw", window=prop33)

                # two queens beds taxes
                prop34 = Label(canvas, text='$' + str(conversion_numbers1(rrodqueens * .05)),
                               font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(800, x + 100, anchor="nw", window=prop34)

                prop35 = Label(canvas, text='$' + str(conversion_numbers1(swriqueens * .05)),
                               font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(800, x + 120, anchor="nw", window=prop35)

                prop36 = Label(canvas, text='$' + str(conversion_numbers1(arrpqueens * .05)),
                               font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(800, x + 140, anchor="nw", window=prop36)

                prop37 = Label(canvas, text='$' + str(conversion_numbers1(fastqueens * .05)),
                               font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(800, x + 160, anchor="nw", window=prop37)

                prop38 = Label(canvas, text='$' + str(conversion_numbers1(0)), font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(800, x + 180, anchor="nw", window=prop38)

                # two queens beds total taxes
                prop39 = Label(canvas, text='$' + str(
                    conversion_numbers1((rrodqueens * count_rooms) + ((rrodqueens * count_rooms) * .05))),
                               font=('arial', 9), fg='black', bg='lightgray')
                canvas.create_window(920, x + 100, anchor="nw", window=prop39)

                prop40 = Label(canvas, text='$' + str(
                    conversion_numbers1((swriqueens * count_rooms) + ((swriqueens * count_rooms) * .05))),
                               font=('arial', 9),
                               fg='black', bg='lightgray')
                canvas.create_window(920, x + 120, anchor="nw", window=prop40)

                prop41 = Label(canvas, text='$' + str(
                    conversion_numbers1((arrpqueens * count_rooms) + ((arrpqueens * count_rooms) * .05))),
                               font=('arial', 9),
                               fg='black', bg="lightgray")
                canvas.create_window(920, x + 140, anchor="nw", window=prop41)

                prop42 = Label(canvas,
                               text='$' + str(conversion_numbers1(
                                   (fastqueens * count_rooms) + (
                                           (fastqueens * count_rooms) * .05))) + ' + ' + str(
                                   conversion_numbers((gofastpoints * count_rooms))) + ' pts',
                               font=('arial', 9),
                               fg='black',
                               bg="lightgray")
                canvas.create_window(920, x + 160, anchor="nw", window=prop42)

                prop43 = Label(canvas, text=str(conversion_numbers((freeking * count_rooms))) + ' pts',
                               font=('arial', 9), fg='black',
                               bg="lightgray")
                canvas.create_window(920, x + 180, anchor="nw", window=prop43)

                ratebookcombo = Label(canvas, text="  Rate Code", font=('arial', 10), bg='lightgray',
                                      foreground='blue')
                bookcombo = ttk.Combobox(canvas, font=('arial', 10),
                                         values=['Rates', 'RROD', 'AARP', 'SWR1', 'Go Fast', 'Go Free'],
                                         width=10, foreground='#263238')
                bookcombo.insert(0, choose_dates[8].get())
                # choose_dates[8] = bookcombo

                ratebedscombo = Label(canvas, text="  Bed Class", font=('arial', 10), bg='lightgray',
                                      foreground='blue')
                bedscombo = ttk.Combobox(canvas, font=('arial', 10),
                                         values=['Bed Class', 'One Bed', 'Two Beds'],
                                         width=10, foreground='#263238')
                bedscombo.insert(0, choose_dates[9].get())
                # choose_dates[9] = bedscombo

                canvas_nyts = Label(canvas, text="Night", font=('arial', 10), bg='lightgray',
                                    foreground='blue')
                canvas_nyts1 = Label(canvas, text=count_rooms, font=('arial', 9), background='white',
                                     padx=12,
                                     pady=1, border=1, relief='sunken')

                canvas.create_window(10, x + 240, anchor="nw", window=ratebookcombo)
                canvas.create_window(100, x + 240, anchor="nw", window=bookcombo)
                canvas.create_window(230, x + 240, anchor="nw", window=ratebedscombo)
                canvas.create_window(320, x + 240, anchor="nw", window=bedscombo)

                canvas.create_window(450, x + 240, anchor="nw", window=canvas_nyts)
                canvas.create_window(500, x + 240, anchor="nw", window=canvas_nyts1)

                spacing = Label(canvas, text=' ', font=('arial', 9), fg='black',
                                bg="lightgray")
                # canvas.create_window(1150, x + 330, anchor="nw", window=spacing)

        my_frame1.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
        my_frame1.place(x=200, y=176)
        view_return.append(my_frame1)
        resv_forget.append(my_frame1)
        if mod_loops == 1:
            mod_loops = 0
            modify_stay()
        else:
            post()

# choose dates
def guest_dates_modify1():
    print('guess dates modify')

    guest_back_mod = Button(window, text='Back Mod', font=('arial', 10), fg='black', bg="lightgray",
                            activeforeground="#00695C", activebackground="lightgray", padx=20, pady=2, width=5,
                            # wraplength= 1,
                            # state='disable',
                            command=guest_next_modify)
    # guest_back_mod.place(x=210, y=600)
    # guest_back_mod.place(x=530, y=600)
    guest_back_mod.place(x=635, y=600)
    resv_forget.append(guest_back_mod)

    '''guest_next_mod = Button(window, text='Next Mod', font=('arial', 10), fg='black', bg='lightgray',
                            activeforeground="#00695C", activebackground="lightgray", padx=15, pady=2, width=5,
                            # wraplength= 1,
                            state='disable',
                            command=guest_search_modify)
    guest_next_mod.place(x=310, y=600)'''


# choose site no
def guest_next_modify():
    global mod_loops
    print('guest next modify')
    value_resv_stay = 26
    levelcolor = 'lightblue'

    my_frame = Frame(window)
    canvas = Canvas(my_frame)

    canvas = Canvas(my_frame, width=625, height=380, background="lightgray")
    canvas.pack(fill="both", expand=True)

    vsb = Scrollbar(my_frame, orient="vertical", command=canvas.yview, width=25)

    hsb = Scrollbar(my_frame, orient="horizontal", command=canvas.xview, width=25)

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)
    canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
    canvas.grid(row=0, column=0, sticky="nsew")

    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    property = Label(canvas, text='  Property Name  ', font=('arial', 15), bg=levelcolor, fg='blue', padx=0, pady=10, )
    canvas.create_window(10, 5, anchor="nw", window=property)
    x = 60
    y = 1
    for data_search in range(2, data_number3.max_row + 1):
        siteno = data_number3.cell(data_search, 1).value
        hotelname = data_number3.cell(data_search, 2).value
        address = data_number3.cell(data_search, 3).value
        phoneno = data_number3.cell(data_search, 4).value
        rating = data_number3.cell(data_search, 5).value

        rrodking = data_number3.cell(data_search, 8).value
        swriking = data_number3.cell(data_search, 10).value
        arrpking = data_number3.cell(data_search, 12).value
        fastking = data_number3.cell(data_search, 14).value

        gofastpoints = data_number3.cell(data_search, 16).value
        freeking = data_number3.cell(data_search, 19).value

        rrodqueens = data_number3.cell(data_search, 9).value
        swriqueens = data_number3.cell(data_search, 11).value
        arrpqueens = data_number3.cell(data_search, 13).value
        fastqueens = data_number3.cell(data_search, 15).value

        prop = str(property) + str(y)
        prop1 = str(property) + str(y)
        prop2 = str(property) + str(y)

        prop = Label(canvas, text='  ' + '(' + str(siteno) + ')  ' + str(hotelname), font=('arial', 12), fg='black',
                     bg='lightgray')
        canvas.create_window(10, x, anchor="nw", window=prop)

        prop1 = Label(canvas, text='   ' + str(phoneno) + ',  ' + str(rating), font=('arial', 8), fg='blue',
                      bg='lightgray')
        canvas.create_window(10, x + 22, anchor="nw", window=prop1)

        prop2 = Label(canvas, text='   ' + str(address), font=('arial', 9), fg='black', bg='lightgray', )
        canvas.create_window(10, x + 42, anchor="nw", window=prop2)

        def select(number=data_search):
            choose_entry[0] = ''
            choose_loopsite[0] = 0
            choose_siteno[0] = number - 1 + 9000
            # print('choose site number ',choose_siteno[0])
            # site_condition[0] = 'True'
            # print(number - 1, 'clickeddddddddddddd')
            if mod_loops == 1:
                modify_stay()
            else:
                guest_search_modify()

        stay_buttons = Button(canvas, text="Select", font=("arial", 10), fg="black", bg="#F0F0F0",
                              activeforeground="Green", activebackground="#F0F0F0", padx=6, pady=2, width=7,
                              # wraplength= 1,
                              # state='disable',
                              command=select)
        canvas.create_window(520, x + 100, anchor="nw", window=stay_buttons)

        space = Label(canvas, text='', font=('arial', 9), fg='black', bg='lightgray', )
        canvas.create_window(10, x + 130, anchor="nw", window=space)
        x += 190
        y += 1

    my_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
    my_frame.place(x=200, y=176)
    resv_forget.append(my_frame)


def temp_next_modify():
    pass


def site_add_detais():
    print('site and details')
    back_return_value[0] = 1
    modified_nyts1 = 0
    modified_site1 = 0
    temp_book.clear()
    loop = 1
    for x in back_bookrooms:
        pass
        # print('temp modified ',x.get())
    get_resvno = data_number4.max_row - 1

    modified_site1 = 0
    for data_search in range(2, data_number4.max_row + 1):
        accntno = data_number4.cell(data_search, 1).value
        modified_resv = data_number4.cell(data_search, 2).value
        hotel_name = data_number4.cell(data_search, 7).value
        modified_nyts = data_number4.cell(data_search, 18).value
        modified_site = data_number4.cell(data_search, 28).value
        checkin = data_number4.cell(data_search, 22).value
        checkout = data_number4.cell(data_search, 23).value
        ratecode = data_number4.cell(data_search, 10).value
        roomtype = data_number4.cell(data_search, 21).value
        adults = data_number4.cell(data_search, 19).value
        child = data_number4.cell(data_search, 20).value
        # print('modified ', modified_site)
        if str(modified_resv) == str(get_resvno):
            modified_site1 = modified_site
            # print('modified ',modified_site ,modified_resv ,str(get_resvno))
            getdate1 = list(checkin)
            getdate2 = []
            getdate = ['', '', '']

            outgetdate1 = list(checkout)
            outgetdate2 = []
            outgetdate = ['', '', '', '']

            z = 0
            z1 = 0
            for x in getdate1:
                if x == ',' or x == '-':
                    pass
                else:
                    getdate2.append(x)

            for x in getdate2:
                if x == ' ':
                    z += 1
                else:
                    getdate[z] += x

            for x in outgetdate1:
                if x == ',' or x == '-':
                    pass
                else:
                    outgetdate2.append(x)

            for x in range(0, len(outgetdate2)):
                if x == 0:
                    pass
                elif outgetdate2[x] == ' ':
                    z1 += 1
                else:
                    outgetdate[z1] += outgetdate2[x]

            inyear = outgetdate[3]
            longmonth = datetime.datetime.strptime(getdate[1], '%b').month
            indays = getdate[2]

            total_ins = datetime.datetime(int(inyear), int(longmonth), int(indays))
            total_ins1 = total_ins.strftime("%Y %B %d ")

            outyear = outgetdate[3]
            outlongmonth = datetime.datetime.strptime(outgetdate[1], '%b').month
            outdays = outgetdate[2]

            total_outs1 = datetime.datetime(int(outyear), int(outlongmonth), int(outdays))
            total_outs = total_outs1.strftime("%Y %B %d ")

            total_in = total_ins.strftime("%d")
            y = 0
            remove_int = [str(y) + str(x) for x in range(1, 10)]
            for x in range(0, len(remove_int)):
                if remove_int[x] == total_in:
                    total_in = x + 1

            total_out = total_outs1.strftime("%d")
            remove_int = [str(y) + str(x) for x in range(1, 10)]
            for x in range(0, len(remove_int)):
                if remove_int[x] == total_out:
                    total_out = x + 1
            back_bookrooms[0] = total_ins.strftime("%B")
            back_bookrooms[1] = total_in
            back_bookrooms[2] = total_outs1.strftime("%Y")
            back_bookrooms[3] = total_outs1.strftime("%B")
            back_bookrooms[4] = total_out
            back_bookrooms[5] = total_outs1.strftime("%Y")
            back_bookrooms[6] = adults
            back_bookrooms[7] = child
            back_resv_value[0] = get_resvno

    for data_search in range(2, data_number3.max_row + 1):
        resvno = data_number3.cell(data_search, 1).value
        if str(resvno) == str(modified_site1):
            # print('before site prop', site_prop)
            site_prop[0] = data_search - 1
            # print('after site entry ', site_scrool_entry)
            # print('after site prop ', site_prop, data_search - 1)
            site_scrools2()
            site_scrolls1()


def reservation():
    global mod_loops
    patch_design = Label(window, background="#F0F0F0", padx=532, pady=230)
    patch_design.place(x=200, y=170)
    resv_forget.append(patch_design)
    back_bookrooms[0] = 0
    back_dates_value[0] = 1
    universal_siteno[0] = 0
    view_resvno[0] = 0
    site_prop.clear()
    site_scrool_entry.clear()
    # print('value of reservation no. ', len(site_prop), len(site_scrool_entry))
    universal_bookcombo[0] = 'Rates'
    universal_bedscombo[0] = 'Bed Class'
    back_return_value = [0]
    change_shift_resv = [0, 0]
    change_shift_resv = [0, 0, 0]

    mod_loops = 0
    #calendar
    for x in change_forget:
        x.place_forget()
    change_forget.clear()
    site_scrools2()
    site_scrolls()



def site_scrools2():
    print('Site Scrools 2')
    value_resv_stay = 26
    # value_resv_stay = resv_stay(label53.get())
    # print('reservation valueeeeeeeeeeeeeeeeeee ', value_resv_stay)
    levelcolor = 'lightblue'
    # levelcolor = value_resv_stay[1]

    # temp variable add member
    for x in range(0, len(temp_addnewmbr)):
        temp_addnewmbr[x] = ''
    # print('count rooms ',book_rooms)
    book_rooms.clear()
    book2_rooms.clear()
    site_scrool_entry.clear()

    my_frame = Frame(window)
    canvas = Canvas(my_frame)

    canvas = Canvas(my_frame, width=373, height=440, background="lightgray")
    canvas.pack(fill="both", expand=True)

    vsb = Scrollbar(my_frame, orient="vertical", command=canvas.yview, width=25)
    hsb = Scrollbar(my_frame, orient="horizontal", command=canvas.xview, width=25)

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)
    canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
    canvas.grid(row=0, column=0, sticky="nsew")

    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    resv_head = Label(canvas, text="   Reservation Details ", font=('arial', 15),
                      bg=levelcolor, foreground='blue', padx=0, pady=10, )

    site_no = Label(canvas, text="   Site Number", font=('arial', 10),
                    bg='lightgray', foreground='blue', padx=0, pady=10, )

    site_entry = Entry(canvas, font=('arial', 11), width=10)
    site_scrool_entry.append(site_entry)
    ##site_entry.insert(0, END)
    # book_rooms.append(site_entry)

    site_entry_button = Button(canvas, text='Select', font=('arial', 10), fg='black', bg="#F0F0F0",
                               activeforeground="#00695C", activebackground="#F0F0F0", padx=15, pady=1, width=5,
                               # wraplength= 1,
                               # state='disable',
                               command=site_scrolls1)
    canvas.create_window(230, 107, anchor="nw", window=site_entry_button)
    # today and tommorows date
    if back_bookrooms[0] == 0:
        today = date.today()
        dated_checkin_month = today.strftime("%B")
        dated_checkin_day = int(today.strftime("%d"))
        dated_checkin_year = today.strftime("%Y")
        date_tom = datetime.date.today() + datetime.timedelta(days=1)
        dated_checkout_month = date_tom.strftime("%B")
        dated_checkout_day = int(date_tom.strftime("%d"))
        dated_checkout_year = date_tom.strftime("%Y")
        adults1 = 1
        child1 = 0
        nonights = back_dates_value[0]

    else:
        try:
            dated_checkin_month = back_bookrooms[0].get()
            dated_checkin_day = back_bookrooms[1].get()
            dated_checkin_year = back_bookrooms[2].get()
            dated_checkout_month = back_bookrooms[3].get()
            dated_checkout_day = back_bookrooms[4].get()
            dated_checkout_year = back_bookrooms[5].get()
            adults1 = back_bookrooms[6].get()
            child1 = back_bookrooms[7].get()

            checkinmonth = datetime.datetime.strptime(dated_checkin_month, '%B').month
            checkoutmonth = datetime.datetime.strptime(dated_checkout_month, '%B').month
            a = int(dated_checkin_year)
            b = int(checkinmonth)
            c = int(dated_checkin_day)
            checkinn = date(a, b, c)

            x = int(dated_checkout_year)
            y = int(checkoutmonth)
            z = int(dated_checkout_day)
            checkout = date(x, y, z)

            valid_entry_in = checkout - checkinn
            back_dates_value[0] = valid_entry_in.days
            nonights = back_dates_value[0]

        except:
            dated_checkin_month = back_bookrooms[0]
            dated_checkin_day = back_bookrooms[1]
            dated_checkin_year = back_bookrooms[2]
            dated_checkout_month = back_bookrooms[3]
            dated_checkout_day = back_bookrooms[4]
            dated_checkout_year = back_bookrooms[5]
            adults1 = back_bookrooms[6]
            child1 = back_bookrooms[7]

            checkinmonth = datetime.datetime.strptime(dated_checkin_month, '%B').month
            checkoutmonth = datetime.datetime.strptime(dated_checkout_month, '%B').month
            a = int(dated_checkin_year)
            b = int(checkinmonth)
            c = int(dated_checkin_day)
            checkinn = date(a, b, c)

            x = int(dated_checkout_year)
            y = int(checkoutmonth)
            z = int(dated_checkout_day)
            checkout = date(x, y, z)
            nonights = back_dates_value[0]

    checkin = Label(canvas, text="   Check-In Date ", font=('arial', 10),
                    bg='lightgray', foreground='blue', padx=0, pady=10, )

    checkinmonths = ttk.Combobox(canvas, font=('arial', 10),
                                 values=['Month', 'January', 'February', 'March', 'April', 'May', 'June', 'July',
                                         'August', 'September', 'October', 'November', 'December'], width=12,
                                 foreground='#263238')
    checkinmonths.insert(0, dated_checkin_month)
    book_rooms.append(checkinmonths)
    book2_rooms.append(checkinmonths)
    back_bookrooms[0] = checkinmonths

    checkin_days = ttk.Combobox(canvas, font=('arial', 10),
                                values=['Days', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13',
                                        '14', '15',
                                        '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28',
                                        '29', '30', '31'],
                                width=10, foreground='#263238')

    checkin_days.insert(0, dated_checkin_day)
    book_rooms.append(checkin_days)
    book2_rooms.append(checkin_days)
    back_bookrooms[1] = checkin_days

    checkin_year = ttk.Combobox(canvas, font=('arial', 10), values=['Year', '2023', '2024', '2025', '2026', '2027'],
                                width=10, foreground='#263238')
    checkin_year.insert(0, dated_checkin_year)
    book_rooms.append(checkin_year)
    book2_rooms.append(checkin_year)
    back_bookrooms[2] = checkin_year

    checkout = Label(canvas, text="   Check-Out Date ", font=('arial', 10), bg='lightgray', foreground='blue', padx=0,
                     pady=10, )

    checkoutmonths = ttk.Combobox(canvas, font=('arial', 10),
                                  values=['Month', 'January', 'February', 'March', 'April', 'May', 'June', 'July',
                                          'August', 'September', 'October', 'November', 'December'], width=12,
                                  foreground='#263238')
    checkoutmonths.insert(0, dated_checkout_month)
    book_rooms.append(checkoutmonths)
    book2_rooms.append(checkoutmonths)
    back_bookrooms[3] = checkoutmonths

    checkout_days = ttk.Combobox(canvas, font=('arial', 10),
                                 values=['Days', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13',
                                         '14', '15',
                                         '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28',
                                         '29', '30', '31'],
                                 width=10, foreground='#263238')
    checkout_days.insert(0, dated_checkout_day)
    book_rooms.append(checkout_days)
    book2_rooms.append(checkout_days)
    back_bookrooms[4] = checkout_days

    # checkout_days.bind("<Button-1>", post_no_stay1)
    # checkout_days.bind("<Enter>", post_no_stay2)
    # checkout_days.bind("<Leave>", post_no_stay3)

    checkout_year = ttk.Combobox(canvas, font=('arial', 10), values=['Year', '2023', '2024', '2025', '2026', '2027'],
                                 width=10, foreground='#263238')
    checkout_year.insert(0, dated_checkout_year)
    book_rooms.append(checkout_year)
    book2_rooms.append(checkout_year)
    back_bookrooms[5] = checkout_year

    adults = Label(canvas, text="   Adults ", font=('arial', 10), bg='lightgray', foreground='blue', padx=0, pady=10)

    adultscombo = ttk.Combobox(canvas, font=('arial', 10), values=['1', '2', '3', '4', '5', '6', '7', '8', '9'],
                               width=8, foreground='#263238')
    adultscombo.insert(0, adults1)
    book_rooms.append(adultscombo)
    # book2_rooms.append(adultscombo)
    back_bookrooms[6] = adultscombo

    children = Label(canvas, text="   Child", font=('arial', 10), bg='lightgray', foreground='blue', padx=0, pady=10)

    childrencombo = ttk.Combobox(canvas, font=('arial', 10), values=['0', '1', '2', '3', '4', '5', '6', '7', '8', '9'],
                                 width=8, foreground='#263238')
    childrencombo.insert(0, child1)
    book_rooms.append(childrencombo)
    back_bookrooms[7] = childrencombo

    rooms = Label(canvas, text="   Rooms", font=('arial', 10), bg='lightgray', foreground='blue', padx=0, pady=10)

    rooms_no = ttk.Combobox(canvas, font=('arial', 10), values=['0', '1', '2', '3', '4', '5', '6', '7', '8', '9'],
                            width=8, foreground='#263238')
    rooms_no.insert(0, '0')

    canvas_nyts = Label(canvas, text="Night", font=('arial', 10), bg='lightgray', foreground='blue')
    no_nights = ttk.Combobox(canvas, font=('arial', 10), values=['1', '2', '3', '4', '5', '6', '7', '8', '9'],
                             width=8, foreground='#263238')
    no_nights.insert(0, nonights)
    # back_dates_value[0] = nonights

    # book_rooms.append(rooms_no)

    # spacing
    k31 = Label(canvas, text='               ', font=('arial', 10), bg="lightgray", foreground='black', padx=0, pady=10)
    k32 = Label(canvas, text='               ', font=('arial', 10), bg="lightgray", foreground='black', padx=0, pady=10)

    canvas.create_window(10, 50, anchor="nw", window=resv_head)
    canvas.create_window(10, 105, anchor="nw", window=site_no)
    canvas.create_window(110, 112, anchor="nw", window=site_entry)
    canvas.create_window(10, 140, anchor="nw", window=checkin)
    canvas.create_window(20, 185, anchor="nw", window=checkinmonths)
    canvas.create_window(150, 185, anchor="nw", window=checkin_days)
    canvas.create_window(270, 185, anchor="nw", window=checkin_year)
    canvas.create_window(10, 215, anchor="nw", window=checkout)
    canvas.create_window(20, 255, anchor="nw", window=checkoutmonths)
    canvas.create_window(150, 255, anchor="nw", window=checkout_days)
    canvas.create_window(270, 255, anchor="nw", window=checkout_year)
    canvas.create_window(10, 290, anchor="nw", window=adults)
    canvas.create_window(90, 297, anchor="nw", window=adultscombo)
    canvas.create_window(10, 330, anchor="nw", window=children)
    canvas.create_window(90, 337, anchor="nw", window=childrencombo)
    canvas.create_window(10, 370, anchor="nw", window=rooms)
    canvas.create_window(90, 377, anchor="nw", window=rooms_no)

    canvas.create_window(220, 299, anchor="nw", window=canvas_nyts)
    canvas.create_window(270, 299, anchor="nw", window=no_nights)

    back_dates = Button(canvas, text="Back", font=("arial", 10), fg="black", bg="#F0F0F0",
                        activeforeground="#00695C", activebackground="#F0F0F0", border=2, relief=RAISED, padx=6, pady=2,
                        width=7,
                        # wraplength= 1,
                        # state='disable',
                        command=dates_back)
    canvas.create_window(120, 470, anchor="nw", window=back_dates)

    # canvas.create_window(10, 510 , anchor="nw", window=space)

    confirm_dates = Button(canvas, text="Next", font=("arial", 10), fg="black", bg="#F0F0F0",
                           activeforeground="#00695C", activebackground="#F0F0F0", border=2, relief=RAISED, padx=6,
                           pady=2, width=7,
                           # wraplength= 1,
                           # state='disable',
                           command=dates_confirm)
    canvas.create_window(200, 470, anchor="nw", window=confirm_dates)

    # canvas.create_window(10, 510 + a, anchor="nw", window=space)

    my_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
    my_frame.place(x=870, y=176)
    resv_forget.append(canvas)
    resv_forget.append(my_frame)

    view_resv = Button(window, text='View Reservation', font=('arial', 9), fg='black', bg='lightgray',
                       activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                       # wraplength= 1,
                       # state='disable',
                       command=view_reservation)
    view_resv.place(x=650, y=615)
    resv_forget.append(view_resv)

    wrap_up1 = Button(window, text='Wrap-up', font=('arial', 9), fg='black', bg='lightgray',
                      activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                      # wraplength= 1,
                      # state='disable',
                      command=wrap_up)
    wrap_up1.place(x=430, y=615)
    resv_forget.append(wrap_up1)

    new_resv = Button(window, text='New Reservation', font=('arial', 9), fg='black', bg='lightgray',
                      activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                      # wraplength= 1,
                      # state='disable',
                      command=reservation)
    new_resv.place(x=430, y=585)
    resv_forget.append(new_resv)


def dates_confirm():
    print('dates confirm')
    count_list = []
    count = 0
    for x in range(0, len(back_bookrooms)):
        if x == 0:
            print('back rooms ', back_bookrooms[x].get())

    for x in book2_rooms:
        count_list.append(x.get())

    for x in count_list:
        if x == 'Month':
            count += 1
        elif x == 'Days':
            count += 1
        elif x == 'Year':
            count += 1
        elif x == 'Month':
            count += 1
        elif x == 'Days':
            count += 1
        elif x == 'Year':
            count += 1

    if count == 0:
        no_nyts = no_of_nyts_post()

        if no_nyts[1] == 'True':
            if no_nyts[0] >= 1:
                count_rooms = no_nyts[0]
            elif no_nyts[0] <= 0:
                count_rooms = no_nyts[0]
                count = 6
            elif no_nyts == '':
                pass
        else:
            count = 6

        value_resv_stay = 26
        # value_resv_stay = resv_stay(label53.get())
        # print('reservation valueeeeeeeeeeeeeeeeeee ', value_resv_stay)
        levelcolor = 'lightblue'
        # levelcolor = value_resv_stay[1]
        for x in book_rooms:
            count_list.append(x.get())

        checkinmonth = datetime.datetime.strptime(book_rooms[0].get(), '%B').month
        checkoutmonth = datetime.datetime.strptime(book_rooms[3].get(), '%B').month
        a = int(book_rooms[2].get())
        b = int(checkinmonth)
        c = int(book_rooms[1].get())
        checkinn = date(a, b, c)

        x = int(book_rooms[5].get())
        y = int(checkoutmonth)
        z = int(book_rooms[4].get())
        checkout = date(x, y, z)

        valid_entry_in = checkout - checkinn
        back_dates_value[0] = valid_entry_in.days

        my_frame = Frame(window)
        canvas = Canvas(my_frame)

        canvas = Canvas(my_frame, width=373, height=440, background="lightgray")
        canvas.pack(fill="both", expand=True)

        vsb = Scrollbar(my_frame, orient="vertical", command=canvas.yview, width=25)
        hsb = Scrollbar(my_frame, orient="horizontal", command=canvas.xview, width=25)

        window.grid_rowconfigure(0, weight=1)
        window.grid_columnconfigure(0, weight=1)
        canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")

        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        resv_head = Label(canvas, text="   Reservation Details ", font=('arial', 15),
                          bg=levelcolor, foreground='blue', padx=0, pady=10, )

        site_no = Label(canvas, text="   Site Number", font=('arial', 10),
                        bg='lightgray', foreground='blue', padx=0, pady=10, )

        site_entry = Entry(canvas, font=('arial', 11), width=10)
        site_scrool_entry.append(site_entry)
        ##site_entry.insert(0, END)
        # book_rooms.append(site_entry)
        # today and tommorows date

        checkin = Label(canvas, text="   Check-In Date ", font=('arial', 10),
                        bg='lightgray', foreground='blue', padx=0, pady=10, )

        checkinmonths = ttk.Combobox(canvas, font=('arial', 10),
                                     values=['Month', 'January', 'February', 'March', 'April', 'May', 'June', 'July',
                                             'August', 'September', 'October', 'November', 'December'], width=12,
                                     foreground='#263238')
        checkinmonths.insert(0, back_bookrooms[0].get())

        checkin_days = ttk.Combobox(canvas, font=('arial', 10),
                                    values=['Days', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13',
                                            '14', '15',
                                            '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27',
                                            '28', '29', '30', '31'],
                                    width=10, foreground='#263238')

        checkin_days.insert(0, back_bookrooms[1].get())

        checkin_year = ttk.Combobox(canvas, font=('arial', 10), values=['Year', '2023', '2024', '2025', '2026', '2027'],
                                    width=10, foreground='#263238')
        checkin_year.insert(0, back_bookrooms[2].get())

        checkout = Label(canvas, text="   Check-Out Date ", font=('arial', 10), bg='lightgray', foreground='blue',
                         padx=0, pady=10, )

        checkoutmonths = ttk.Combobox(canvas, font=('arial', 10),
                                      values=['Month', 'January', 'February', 'March', 'April', 'May', 'June', 'July',
                                              'August', 'September', 'October', 'November', 'December'], width=12,
                                      foreground='#263238')
        checkoutmonths.insert(0, back_bookrooms[3].get())

        checkout_days = ttk.Combobox(canvas, font=('arial', 10),
                                     values=['Days', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12',
                                             '13', '14', '15',
                                             '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27',
                                             '28', '29', '30', '31'],
                                     width=10, foreground='#263238')
        checkout_days.insert(0, back_bookrooms[4].get())

        checkout_year = ttk.Combobox(canvas, font=('arial', 10),
                                     values=['Year', '2023', '2024', '2025', '2026', '2027'],
                                     width=10, foreground='#263238')
        checkout_year.insert(0, back_bookrooms[5].get())

        adults = Label(canvas, text="   Adults ", font=('arial', 10), bg='lightgray', foreground='blue', padx=0,
                       pady=10)

        adultscombo = ttk.Combobox(canvas, font=('arial', 10), values=['1', '2', '3', '4', '5', '6', '7', '8', '9'],
                                   width=8, foreground='#263238')
        adultscombo.insert(0, back_bookrooms[6].get())

        children = Label(canvas, text="   Child", font=('arial', 10), bg='lightgray', foreground='blue', padx=0,
                         pady=10)
        childrencombo = ttk.Combobox(canvas, font=('arial', 10),
                                     values=['0', '1', '2', '3', '4', '5', '6', '7', '8', '9'],
                                     width=8, foreground='#263238')
        childrencombo.insert(0, back_bookrooms[7].get())

        rooms = Label(canvas, text="   Rooms", font=('arial', 10), bg='lightgray', foreground='blue', padx=0, pady=10)
        rooms_no = ttk.Combobox(canvas, font=('arial', 10), values=['0', '1', '2', '3', '4', '5', '6', '7', '8', '9'],
                                width=8, foreground='#263238')
        rooms_no.insert(0, '0')

        canvas_nyts = Label(canvas, text="Night", font=('arial', 10), bg='lightgray', foreground='blue')
        no_nights = ttk.Combobox(canvas, font=('arial', 10), values=['1', '2', '3', '4', '5', '6', '7', '8', '9'],
                                 width=8, foreground='#263238')
        no_nights.insert(0, back_dates_value[0])

        # spacing
        k31 = Label(canvas, text='               ', font=('arial', 10), bg="lightgray", foreground='black', padx=0,
                    pady=10)
        k32 = Label(canvas, text='               ', font=('arial', 10), bg="lightgray", foreground='black', padx=0,
                    pady=10)

        canvas.create_window(10, 50, anchor="nw", window=resv_head)
        canvas.create_window(10, 105, anchor="nw", window=site_no)
        canvas.create_window(110, 112, anchor="nw", window=site_entry)
        canvas.create_window(10, 140, anchor="nw", window=checkin)
        canvas.create_window(20, 185, anchor="nw", window=checkinmonths)
        canvas.create_window(150, 185, anchor="nw", window=checkin_days)
        canvas.create_window(270, 185, anchor="nw", window=checkin_year)
        canvas.create_window(10, 215, anchor="nw", window=checkout)
        canvas.create_window(20, 255, anchor="nw", window=checkoutmonths)
        canvas.create_window(150, 255, anchor="nw", window=checkout_days)
        canvas.create_window(270, 255, anchor="nw", window=checkout_year)
        canvas.create_window(10, 290, anchor="nw", window=adults)
        canvas.create_window(90, 297, anchor="nw", window=adultscombo)
        canvas.create_window(10, 330, anchor="nw", window=children)
        canvas.create_window(90, 337, anchor="nw", window=childrencombo)
        canvas.create_window(10, 370, anchor="nw", window=rooms)
        canvas.create_window(90, 377, anchor="nw", window=rooms_no)

        canvas.create_window(220, 299, anchor="nw", window=canvas_nyts)
        canvas.create_window(270, 299, anchor="nw", window=no_nights)

        back_dates = Button(canvas, text="Back", font=("arial", 10), fg="black", bg="#F0F0F0",
                            activeforeground="green", activebackground="#F0F0F0", border=2, relief=RAISED, padx=6,
                            pady=2,
                            width=7,
                            # wraplength= 1,
                            # state='disable',
                            command=dates_back)
        canvas.create_window(120, 470, anchor="nw", window=back_dates)

        # canvas.create_window(10, 510 , anchor="nw", window=space)

        confirm_dates = Button(canvas, text="Next", font=("arial", 10), fg="black", bg="#F0F0F0",
                               activeforeground="green", activebackground="#F0F0F0", border=2, relief=RAISED, padx=6,
                               pady=2, width=7,
                               # wraplength= 1,
                               # state='disable',
                               command=dates_confirm)
        canvas.create_window(200, 470, anchor="nw", window=confirm_dates)

        # canvas.create_window(10, 510 + a, anchor="nw", window=space)

        my_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
        my_frame.place(x=870, y=176)
        resv_forget.append(canvas)
        resv_forget.append(my_frame)

        back_dates_return.append(canvas)
        back_dates_return.append(my_frame)

        if site_condition[0] == "False":
            pass
        else:
            site_scrolls1()


def dates_back():
    print('melvin')
    '''for x in back_dates_return:
        x.place_forget()
    back_dates_return.clear()'''
    site_scrools2()


def site_scrolls3():
    print('site scrolls three')
    site_scrools2()
    site_scrolls1()


def site_scrolls():
    # change_shift_resv[0] = 0
    change_shift_resv[2] = 0
    print('site scrolls')
    value_resv_stay = 26
    # value_resv_stay = resv_stay(label53.get())
    # print('reservation valueeeeeeeeeeeeeeeeeee ', value_resv_stay)
    levelcolor = 'lightblue'
    # levelcolor = value_resv_stay[1]

    my_frame = Frame(window)
    canvas = Canvas(my_frame)

    canvas = Canvas(my_frame, width=625, height=380, background="lightgray")
    canvas.pack(fill="both", expand=True)

    vsb = Scrollbar(my_frame, orient="vertical", command=canvas.yview, width=25)

    hsb = Scrollbar(my_frame, orient="horizontal", command=canvas.xview, width=25)

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)
    canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
    canvas.grid(row=0, column=0, sticky="nsew")

    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    property = Label(canvas, text='  Property Name  ', font=('arial', 15), bg=levelcolor, fg='blue', padx=0, pady=10, )
    canvas.create_window(0, 5, anchor="nw", window=property)
    x = 60
    y = 1
    for data_search in range(2, data_number3.max_row + 1):
        siteno = data_number3.cell(data_search, 1).value
        hotelname = data_number3.cell(data_search, 2).value
        address = data_number3.cell(data_search, 3).value
        phoneno = data_number3.cell(data_search, 4).value
        rating = data_number3.cell(data_search, 5).value

        rrodking = data_number3.cell(data_search, 8).value
        swriking = data_number3.cell(data_search, 10).value
        arrpking = data_number3.cell(data_search, 12).value
        fastking = data_number3.cell(data_search, 14).value

        gofastpoints = data_number3.cell(data_search, 16).value
        freeking = data_number3.cell(data_search, 19).value

        rrodqueens = data_number3.cell(data_search, 9).value
        swriqueens = data_number3.cell(data_search, 11).value
        arrpqueens = data_number3.cell(data_search, 13).value
        fastqueens = data_number3.cell(data_search, 15).value

        prop = str(property) + str(y)
        prop1 = str(property) + str(y)
        prop2 = str(property) + str(y)

        prop = Label(canvas, text='  ' + '(' + str(siteno) + ')  ' + str(hotelname), font=('arial', 12), fg='black',
                     bg='lightgray')
        canvas.create_window(10, x, anchor="nw", window=prop)

        prop1 = Label(canvas, text='   ' + str(phoneno) + ',  ' + str(rating), font=('arial', 8), fg='blue',
                      bg='lightgray')
        canvas.create_window(10, x + 22, anchor="nw", window=prop1)

        prop2 = Label(canvas, text='   ' + str(address), font=('arial', 9), fg='black', bg='lightgray', )
        canvas.create_window(10, x + 42, anchor="nw", window=prop2)

        def select(number=data_search):
            site_prop.append(number - 1)
            site_loops.append(1)
            site_scrolls1()
            site_condition[0] = 'True'
            # print(number - 1, 'clicked')

        stay_buttons = Button(canvas, text="Select", font=("arial", 10), fg="black", bg="#F0F0F0",
                              activeforeground="Green", activebackground="#F0F0F0", padx=6, pady=2, width=7,
                              # wraplength= 1,
                              # state='disable',
                              command=select)
        resv_forget.append(stay_buttons)
        canvas.create_window(520, x + 100, anchor="nw", window=stay_buttons)

        space = Label(canvas, text='', font=('arial', 9), fg='black', bg='lightgray', )
        canvas.create_window(10, x + 130, anchor="nw", window=space)
        x += 190
        y += 1

    my_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
    my_frame.place(x=200, y=176)
    resv_forget.append(canvas)
    resv_forget.append(my_frame)

    next_button_next = Label(window, text='                                  ', bg='#F0F0F0',
                             activebackground="lightgray", padx=85, pady=4, width=2, )
    next_button_next.place(x=210, y=585)
    resv_forget.append(next_button_next)

    next_button_back = Label(window, text='                                  ', bg='#F0F0F0',
                             activebackground="lightgray", padx=85, pady=4, width=2, )
    next_button_back.place(x=210, y=615)
    resv_forget.append(next_button_back)

def site_scrolls1():
    print('Site Scrools 1')
    change_shift_resv[0] = 0
    change_shift_resv[1] = 0
    change_shift_resv[2] = 1
    book_bed_rates.clear()
    # site_scrool1_forget
    site_true = 0
    count_rooms = back_dates_value[-1]
    loops = 0
    site_no = 0
    site_select1 = 0
    y = 0
    value_book = ''
    value_beds = ''
    if universal_bookcombo[0] == 'Rates':
        value_book = 'Rates'
    else:
        value_book = universal_bookcombo[0].get()

    if universal_bedscombo[0] == 'Bed Class':
        value_beds = 'Bed Class'
    else:
        value_beds = universal_bedscombo[0].get()

    sited_loops = 0
    if len(site_loops) > 1:
        site_loops.remove(site_loops[0])
        sited_loops = 1

    if sited_loops == 0:
        for x in site_scrool_entry:  # entry site number
            site_select1 = x.get()

        for x in site_prop:  # entry all of hotels
            site_no = x
        if len(site_prop) > 1:
            site_prop.remove(site_prop[0])

        for data_search in range(2, data_number3.max_row + 1):
            siteno = data_number3.cell(data_search, 1).value
            if str(siteno) == str(site_select1):
                site_selects = data_search - 1
                site_prop.append(int(site_selects))

    for x in site_prop:
        site_no = x

    if len(site_prop) > 1:
        # _prop[1] = 0
        site_prop.remove(site_prop[0])

    # site_prop.clear()
    # site_scrool_entry.clear()
    for data_search in range(2, data_number3.max_row + 1):
        siteno = data_number3.cell(data_search, 1).value
        if str(data_search - 1) == str(site_no):
            site_select = str(siteno)
            loops = 1

    if loops == 1:
        pass
    else:
        try:
            site_select = 0
            count += 1
        except:
            pass

    for data_search in range(2, data_number3.max_row + 1):
        siteno = data_number3.cell(data_search, 1).value

        if str(site_select) == str(siteno):
            universal_siteno[0] = siteno
            book_get_site.append(siteno)
            book_sites1.append(siteno)
            site_true = 1

    if site_true == 1:

        my_frame1 = Frame(window)
        canvas = Canvas(my_frame1)

        canvas = Canvas(my_frame1, width=625, height=380, background="lightgray")
        canvas.pack(fill="both", expand=True)

        vsb = Scrollbar(my_frame1, orient="vertical", command=canvas.yview, width=25)

        hsb = Scrollbar(my_frame1, orient="horizontal", command=canvas.xview, width=25)

        window.grid_rowconfigure(0, weight=1)
        window.grid_columnconfigure(0, weight=1)
        canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")

        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # value_resv_stay = resv_stay(label53.get())
        value_resv_stay = 26
        # print('reservation valueeeeeeeeeeeeeeeeeee ', value_resv_stay)
        # levelcolor = value_resv_stay[1]
        levelcolor = 'lightblue'
        property = Label(canvas, text='  Property Name  ', font=('arial', 15), bg=levelcolor, fg='blue', padx=0,
                         pady=10, )
        canvas.create_window(10, 5, anchor="nw", window=property)
        x = 60
        y = 1
        z = 0
        data_search = 1
        # for data_search in range(2, data_number3.max_row + 1):
        while data_search <= data_number3.max_row:
            siteno = data_number3.cell(data_search, 1).value
            hotelname = data_number3.cell(data_search, 2).value
            address = data_number3.cell(data_search, 3).value
            phoneno = data_number3.cell(data_search, 4).value
            rating = data_number3.cell(data_search, 5).value

            rrodking = data_number3.cell(data_search, 8).value
            swriking = data_number3.cell(data_search, 10).value
            arrpking = data_number3.cell(data_search, 12).value
            fastking = data_number3.cell(data_search, 14).value

            gofastpoints = data_number3.cell(data_search, 16).value
            freeking = data_number3.cell(data_search, 19).value

            rrodqueens = data_number3.cell(data_search, 9).value
            swriqueens = data_number3.cell(data_search, 11).value
            arrpqueens = data_number3.cell(data_search, 13).value
            fastqueens = data_number3.cell(data_search, 15).value
            # print('data search ',data_search)

            if z == 0:
                if str(site_select) == str(siteno):
                    z = 1
                    # print('site search ', siteno)
                    temp_siteno = siteno
                    data_search = 1
                    prop = str(property) + str(y)
                    prop1 = str(property) + str(y)
                    prop2 = str(property) + str(y)

                    prop = Label(canvas, text='  ' + '(' + str(siteno) + ')  ' + hotelname, font=('arial', 12),
                                 fg='black',
                                 bg='lightgray')
                    canvas.create_window(10, x, anchor="nw", window=prop)

                    prop1 = Label(canvas, text='   ' + phoneno + ',  ' + rating, font=('arial', 8), fg='blue',
                                  bg='lightgray')
                    canvas.create_window(10, x + 22, anchor="nw", window=prop1)

                    prop2 = Label(canvas, text='   ' + address, font=('arial', 9), fg='black', bg='lightgray', )
                    canvas.create_window(10, x + 42, anchor="nw", window=prop2)

                    prop3 = Label(canvas, text='  ' + 'Rates', font=('arial', 11), fg='blue', bg='lightgray')
                    canvas.create_window(10, x + 72, anchor="nw", window=prop3)

                    prop4 = Label(canvas, text='One King', font=('arial', 11), fg='blue', bg='lightgray')
                    canvas.create_window(150, x + 72, anchor="nw", window=prop4)

                    prop5 = Label(canvas, text='Taxes', font=('arial', 11), fg='blue', bg="lightgray")
                    canvas.create_window(300, x + 72, anchor="nw", window=prop5)

                    prop6 = Label(canvas, text='Total', font=('arial', 11), fg='blue', bg="lightgray")
                    canvas.create_window(420, x + 72, anchor="nw", window=prop6)

                    prop7 = Label(canvas, text='Two Queens', font=('arial', 11), fg='blue', bg="lightgray")
                    canvas.create_window(650, x + 72, anchor="nw", window=prop7)

                    prop8 = Label(canvas, text='Taxes', font=('arial', 11), fg='blue', bg="lightgray")
                    canvas.create_window(800, x + 72, anchor="nw", window=prop8)

                    prop9 = Label(canvas, text='Total', font=('arial', 11), fg='blue', bg="lightgray")
                    canvas.create_window(920, x + 72, anchor="nw", window=prop9)

                    prop10 = Label(canvas, text='   rrod', font=('arial', 9), fg='blue', bg='lightgray')
                    canvas.create_window(10, x + 100, anchor="nw", window=prop10)

                    prop11 = Label(canvas, text='   swr1', font=('arial', 9), fg='blue', bg='lightgray')
                    canvas.create_window(10, x + 120, anchor="nw", window=prop11)

                    prop12 = Label(canvas, text='   aarp', font=('arial', 9), fg='blue', bg="lightgray")
                    canvas.create_window(10, x + 140, anchor="nw", window=prop12)

                    prop13 = Label(canvas, text='   go fast', font=('arial', 9), fg='blue', bg="lightgray")
                    canvas.create_window(10, x + 160, anchor="nw", window=prop13)

                    prop14 = Label(canvas, text='   go free', font=('arial', 9), fg='blue', bg="lightgray")
                    canvas.create_window(10, x + 180, anchor="nw", window=prop14)

                    # king rates
                    prop15 = Label(canvas, text='$' + str(conversion_numbers1(rrodking)), font=('arial', 9), fg='black',
                                   bg='lightgray')
                    canvas.create_window(150, x + 100, anchor="nw", window=prop15)

                    prop16 = Label(canvas, text='$' + str(conversion_numbers1(swriking)), font=('arial', 9), fg='black',
                                   bg='lightgray')
                    canvas.create_window(150, x + 120, anchor="nw", window=prop16)

                    prop17 = Label(canvas, text='$' + str(conversion_numbers1(arrpking)), font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(150, x + 140, anchor="nw", window=prop17)

                    prop18 = Label(canvas, text='$' + str(conversion_numbers1(fastking)), font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(150, x + 160, anchor="nw", window=prop18)

                    prop19 = Label(canvas, text=str(conversion_numbers(freeking)) + ' pts', font=('arial', 9),
                                   fg='black',
                                   bg="lightgray")
                    canvas.create_window(150, x + 180, anchor="nw", window=prop19)

                    # king taxes rate
                    prop20 = Label(canvas, text='$' + str(conversion_numbers1(rrodking * .05)), font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(300, x + 100, anchor="nw", window=prop20)

                    prop21 = Label(canvas, text='$' + str(conversion_numbers1(swriking * .05)), font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(300, x + 120, anchor="nw", window=prop21)

                    prop22 = Label(canvas, text='$' + str(conversion_numbers1(arrpking * .05)), font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(300, x + 140, anchor="nw", window=prop22)

                    prop23 = Label(canvas, text='$' + str(conversion_numbers1(fastking * .05)), font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(300, x + 160, anchor="nw", window=prop23)

                    prop24 = Label(canvas, text='$' + str(conversion_numbers1(0)), font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(300, x + 180, anchor="nw", window=prop24)

                    # king total rate
                    prop25 = Label(canvas, text='$' + str(
                        conversion_numbers1((rrodking * count_rooms) + ((rrodking * count_rooms) * .05))),
                                   font=('arial', 9), fg='black', bg='lightgray')
                    canvas.create_window(420, x + 100, anchor="nw", window=prop25)

                    prop26 = Label(canvas, text='$' + str(
                        conversion_numbers1((swriking * count_rooms) + ((swriking * count_rooms) * .05))),
                                   font=('arial', 9), fg='black', bg='lightgray')
                    canvas.create_window(420, x + 120, anchor="nw", window=prop26)

                    prop27 = Label(canvas, text='$' + str(
                        conversion_numbers1((arrpking * count_rooms) + ((arrpking * count_rooms) * .05))),
                                   font=('arial', 9), fg='black', bg="lightgray")
                    canvas.create_window(420, x + 140, anchor="nw", window=prop27)

                    prop28 = Label(canvas, text='$' + str(
                        conversion_numbers1((fastking * count_rooms) + ((fastking * count_rooms) * .05))) + ' + ' + str(
                        conversion_numbers((gofastpoints * count_rooms))) + ' pts', font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(420, x + 160, anchor="nw", window=prop28)

                    prop29 = Label(canvas, text=str(conversion_numbers((freeking * count_rooms))) + ' pts',
                                   font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(420, x + 180, anchor="nw", window=prop29)

                    # two queens beds
                    prop29 = Label(canvas, text='$' + str(conversion_numbers1(rrodqueens)),
                                   font=('arial', 9), fg='black', bg='lightgray')
                    canvas.create_window(650, x + 100, anchor="nw", window=prop29)

                    prop30 = Label(canvas, text='$' + str(conversion_numbers1(swriqueens)), font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(650, x + 120, anchor="nw", window=prop30)

                    prop31 = Label(canvas, text='$' + str(conversion_numbers1(arrpqueens)), font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(650, x + 140, anchor="nw", window=prop31)

                    prop32 = Label(canvas, text='$' + str(conversion_numbers1(fastqueens)), font=('arial', 9),
                                   fg='black',
                                   bg="lightgray")
                    canvas.create_window(650, x + 160, anchor="nw", window=prop32)

                    prop33 = Label(canvas, text=str(conversion_numbers(freeking)) + ' pts', font=('arial', 9),
                                   fg='black',
                                   bg="lightgray")
                    canvas.create_window(650, x + 180, anchor="nw", window=prop33)

                    # two queens beds taxes
                    prop34 = Label(canvas, text='$' + str(conversion_numbers1(rrodqueens * .05)), font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(800, x + 100, anchor="nw", window=prop34)

                    prop35 = Label(canvas, text='$' + str(conversion_numbers1(swriqueens * .05)), font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(800, x + 120, anchor="nw", window=prop35)

                    prop36 = Label(canvas, text='$' + str(conversion_numbers1(arrpqueens * .05)), font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(800, x + 140, anchor="nw", window=prop36)

                    prop37 = Label(canvas, text='$' + str(conversion_numbers1(fastqueens * .05)), font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(800, x + 160, anchor="nw", window=prop37)

                    prop38 = Label(canvas, text='$' + str(conversion_numbers1(0)), font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(800, x + 180, anchor="nw", window=prop38)

                    # two queens beds total taxes
                    prop39 = Label(canvas, text='$' + str(
                        conversion_numbers1((rrodqueens * count_rooms) + ((rrodqueens * count_rooms) * .05))),
                                   font=('arial', 9), fg='black', bg='lightgray')
                    canvas.create_window(920, x + 100, anchor="nw", window=prop39)

                    prop40 = Label(canvas, text='$' + str(
                        conversion_numbers1((swriqueens * count_rooms) + ((swriqueens * count_rooms) * .05))),
                                   font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(920, x + 120, anchor="nw", window=prop40)

                    prop41 = Label(canvas, text='$' + str(
                        conversion_numbers1((arrpqueens * count_rooms) + ((arrpqueens * count_rooms) * .05))),
                                   font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(920, x + 140, anchor="nw", window=prop41)

                    prop42 = Label(canvas,
                                   text='$' + str(conversion_numbers1(
                                       (fastqueens * count_rooms) + ((fastqueens * count_rooms) * .05))) + ' + ' + str(
                                       conversion_numbers((gofastpoints * count_rooms))) + ' pts', font=('arial', 9),
                                   fg='black',
                                   bg="lightgray")
                    canvas.create_window(920, x + 160, anchor="nw", window=prop42)

                    prop43 = Label(canvas, text=str(conversion_numbers((freeking * count_rooms))) + ' pts',
                                   font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(920, x + 180, anchor="nw", window=prop43)

                    ratebookcombo = Label(canvas, text="  Rate Code", font=('arial', 10), bg='lightgray',
                                          foreground='blue')
                    bookcombo = ttk.Combobox(canvas, font=('arial', 10),
                                             values=['Rates', 'RROD', 'AARP', 'SWR1', 'Go Fast', 'Go Free'],
                                             width=10, foreground='#263238')
                    bookcombo.insert(0, value_book)
                    # book_rooms.append(bookcombo)
                    book_bed_rates.append(bookcombo)
                    universal_bookcombo[0] = bookcombo

                    ratebedscombo = Label(canvas, text="  Bed Class", font=('arial', 10), bg='lightgray',
                                          foreground='blue')
                    bedscombo = ttk.Combobox(canvas, font=('arial', 10),
                                             values=['Bed Class', 'One Bed', 'Two Beds'],
                                             width=10, foreground='#263238')
                    bedscombo.insert(0, value_beds)
                    # book_rooms.append(bedscombo)
                    book_bed_rates.append(bedscombo)
                    universal_bedscombo[0] = bedscombo

                    canvas_nyts = Label(canvas, text="Night", font=('arial', 10), bg='lightgray', foreground='blue')
                    canvas_nyts1 = Label(canvas, text=count_rooms, font=('arial', 9), background='white', padx=12,
                                         pady=1, border=1, relief='sunken')

                    canvas.create_window(10, x + 240, anchor="nw", window=ratebookcombo)
                    canvas.create_window(100, x + 240, anchor="nw", window=bookcombo)
                    canvas.create_window(230, x + 240, anchor="nw", window=ratebedscombo)
                    canvas.create_window(320, x + 240, anchor="nw", window=bedscombo)

                    canvas.create_window(450, x + 240, anchor="nw", window=canvas_nyts)
                    canvas.create_window(500, x + 240, anchor="nw", window=canvas_nyts1)

                    spacing = Label(canvas, text=' ', font=('arial', 9), fg='black',
                                    bg="lightgray")
                    # canvas.create_window(1150, x + 330, anchor="nw", window=spacing)
                    x += 300

            else:
                if temp_siteno == siteno:
                    pass
                else:
                    prop = str(property) + str(y)
                    prop1 = str(property) + str(y)
                    prop2 = str(property) + str(y)

                    prop = Label(canvas, text='  ' + '(' + str(siteno) + ')  ' + hotelname, font=('arial', 12),
                                 fg='black',
                                 bg='lightgray')
                    canvas.create_window(10, x, anchor="nw", window=prop)

                    prop1 = Label(canvas, text='   ' + phoneno + ',  ' + rating, font=('arial', 8), fg='blue',
                                  bg='lightgray')
                    canvas.create_window(10, x + 22, anchor="nw", window=prop1)

                    prop2 = Label(canvas, text='   ' + address, font=('arial', 9), fg='black', bg='lightgray', )
                    canvas.create_window(10, x + 42, anchor="nw", window=prop2)

                    prop3 = Label(canvas, text='  ' + 'Rates', font=('arial', 11), fg='blue', bg='lightgray')
                    canvas.create_window(10, x + 72, anchor="nw", window=prop3)

                    prop4 = Label(canvas, text='One King', font=('arial', 11), fg='blue', bg='lightgray')
                    canvas.create_window(150, x + 72, anchor="nw", window=prop4)

                    prop5 = Label(canvas, text='Taxes', font=('arial', 11), fg='blue', bg="lightgray")
                    canvas.create_window(300, x + 72, anchor="nw", window=prop5)

                    prop6 = Label(canvas, text='Total', font=('arial', 11), fg='blue', bg="lightgray")
                    canvas.create_window(420, x + 72, anchor="nw", window=prop6)

                    prop7 = Label(canvas, text='Two Queens', font=('arial', 11), fg='blue', bg="lightgray")
                    canvas.create_window(650, x + 72, anchor="nw", window=prop7)

                    prop8 = Label(canvas, text='Taxes', font=('arial', 11), fg='blue', bg="lightgray")
                    canvas.create_window(800, x + 72, anchor="nw", window=prop8)

                    prop9 = Label(canvas, text='Total', font=('arial', 11), fg='blue', bg="lightgray")
                    canvas.create_window(920, x + 72, anchor="nw", window=prop9)

                    prop10 = Label(canvas, text='   rrod', font=('arial', 9), fg='blue', bg='lightgray')
                    canvas.create_window(10, x + 100, anchor="nw", window=prop10)

                    prop11 = Label(canvas, text='   swr1', font=('arial', 9), fg='blue', bg='lightgray')
                    canvas.create_window(10, x + 120, anchor="nw", window=prop11)

                    prop12 = Label(canvas, text='   aarp', font=('arial', 9), fg='blue', bg="lightgray")
                    canvas.create_window(10, x + 140, anchor="nw", window=prop12)

                    prop13 = Label(canvas, text='   go fast', font=('arial', 9), fg='blue', bg="lightgray")
                    canvas.create_window(10, x + 160, anchor="nw", window=prop13)

                    prop14 = Label(canvas, text='   go free', font=('arial', 9), fg='blue', bg="lightgray")
                    canvas.create_window(10, x + 180, anchor="nw", window=prop14)

                    # king rates
                    prop15 = Label(canvas, text='$' + str(conversion_numbers1(rrodking)), font=('arial', 9), fg='black',
                                   bg='lightgray')
                    canvas.create_window(150, x + 100, anchor="nw", window=prop15)

                    prop16 = Label(canvas, text='$' + str(conversion_numbers1(swriking)), font=('arial', 9), fg='black',
                                   bg='lightgray')
                    canvas.create_window(150, x + 120, anchor="nw", window=prop16)

                    prop17 = Label(canvas, text='$' + str(conversion_numbers1(arrpking)), font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(150, x + 140, anchor="nw", window=prop17)

                    prop18 = Label(canvas, text='$' + str(conversion_numbers1(fastking)), font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(150, x + 160, anchor="nw", window=prop18)

                    prop19 = Label(canvas, text=str(conversion_numbers(freeking)) + ' pts', font=('arial', 9),
                                   fg='black',
                                   bg="lightgray")
                    canvas.create_window(150, x + 180, anchor="nw", window=prop19)

                    # king taxes rate
                    prop20 = Label(canvas, text='$' + str(conversion_numbers1(rrodking * .05)), font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(300, x + 100, anchor="nw", window=prop20)

                    prop21 = Label(canvas, text='$' + str(conversion_numbers1(swriking * .05)), font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(300, x + 120, anchor="nw", window=prop21)

                    prop22 = Label(canvas, text='$' + str(conversion_numbers1(arrpking * .05)), font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(300, x + 140, anchor="nw", window=prop22)

                    prop23 = Label(canvas, text='$' + str(conversion_numbers1(fastking * .05)), font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(300, x + 160, anchor="nw", window=prop23)

                    prop24 = Label(canvas, text='$' + str(conversion_numbers1(0)), font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(300, x + 180, anchor="nw", window=prop24)

                    # king total rate
                    prop25 = Label(canvas, text='$' + str(
                        conversion_numbers1((rrodking * count_rooms) + ((rrodking * count_rooms) * .05))),
                                   font=('arial', 9), fg='black', bg='lightgray')
                    canvas.create_window(420, x + 100, anchor="nw", window=prop25)

                    prop26 = Label(canvas, text='$' + str(
                        conversion_numbers1((swriking * count_rooms) + ((swriking * count_rooms) * .05))),
                                   font=('arial', 9), fg='black', bg='lightgray')
                    canvas.create_window(420, x + 120, anchor="nw", window=prop26)

                    prop27 = Label(canvas, text='$' + str(
                        conversion_numbers1((arrpking * count_rooms) + ((arrpking * count_rooms) * .05))),
                                   font=('arial', 9), fg='black', bg="lightgray")
                    canvas.create_window(420, x + 140, anchor="nw", window=prop27)

                    prop28 = Label(canvas, text='$' + str(
                        conversion_numbers1((fastking * count_rooms) + ((fastking * count_rooms) * .05))) + ' + ' + str(
                        conversion_numbers((gofastpoints * count_rooms))) + ' pts', font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(420, x + 160, anchor="nw", window=prop28)

                    prop29 = Label(canvas, text=str(conversion_numbers((freeking * count_rooms))) + ' pts',
                                   font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(420, x + 180, anchor="nw", window=prop29)

                    # two queens beds
                    prop29 = Label(canvas, text='$' + str(conversion_numbers1(rrodqueens)),
                                   font=('arial', 9), fg='black', bg='lightgray')
                    canvas.create_window(650, x + 100, anchor="nw", window=prop29)

                    prop30 = Label(canvas, text='$' + str(conversion_numbers1(swriqueens)), font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(650, x + 120, anchor="nw", window=prop30)

                    prop31 = Label(canvas, text='$' + str(conversion_numbers1(arrpqueens)), font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(650, x + 140, anchor="nw", window=prop31)

                    prop32 = Label(canvas, text='$' + str(conversion_numbers1(fastqueens)), font=('arial', 9),
                                   fg='black',
                                   bg="lightgray")
                    canvas.create_window(650, x + 160, anchor="nw", window=prop32)

                    prop33 = Label(canvas, text=str(conversion_numbers(freeking)) + ' pts', font=('arial', 9),
                                   fg='black',
                                   bg="lightgray")
                    canvas.create_window(650, x + 180, anchor="nw", window=prop33)

                    # two queens beds taxes
                    prop34 = Label(canvas, text='$' + str(conversion_numbers1(rrodqueens * .05)), font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(800, x + 100, anchor="nw", window=prop34)

                    prop35 = Label(canvas, text='$' + str(conversion_numbers1(swriqueens * .05)), font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(800, x + 120, anchor="nw", window=prop35)

                    prop36 = Label(canvas, text='$' + str(conversion_numbers1(arrpqueens * .05)), font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(800, x + 140, anchor="nw", window=prop36)

                    prop37 = Label(canvas, text='$' + str(conversion_numbers1(fastqueens * .05)), font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(800, x + 160, anchor="nw", window=prop37)

                    prop38 = Label(canvas, text='$' + str(conversion_numbers1(0)), font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(800, x + 180, anchor="nw", window=prop38)

                    # two queens beds total taxes
                    prop39 = Label(canvas, text='$' + str(
                        conversion_numbers1((rrodqueens * count_rooms) + ((rrodqueens * count_rooms) * .05))),
                                   font=('arial', 9), fg='black', bg='lightgray')
                    canvas.create_window(920, x + 100, anchor="nw", window=prop39)

                    prop40 = Label(canvas, text='$' + str(
                        conversion_numbers1((swriqueens * count_rooms) + ((swriqueens * count_rooms) * .05))),
                                   font=('arial', 9),
                                   fg='black', bg='lightgray')
                    canvas.create_window(920, x + 120, anchor="nw", window=prop40)

                    prop41 = Label(canvas, text='$' + str(
                        conversion_numbers1((arrpqueens * count_rooms) + ((arrpqueens * count_rooms) * .05))),
                                   font=('arial', 9),
                                   fg='black', bg="lightgray")
                    canvas.create_window(920, x + 140, anchor="nw", window=prop41)

                    prop42 = Label(canvas,
                                   text='$' + str(conversion_numbers1(
                                       (fastqueens * count_rooms) + ((fastqueens * count_rooms) * .05))) + ' + ' + str(
                                       conversion_numbers((gofastpoints * count_rooms))) + ' pts', font=('arial', 9),
                                   fg='black',
                                   bg="lightgray")
                    canvas.create_window(920, x + 160, anchor="nw", window=prop42)

                    prop43 = Label(canvas, text=str(conversion_numbers((freeking * count_rooms))) + ' pts',
                                   font=('arial', 9), fg='black',
                                   bg="lightgray")
                    canvas.create_window(920, x + 180, anchor="nw", window=prop43)

                    ratebookcombo = Label(canvas, text="  Rate Code", font=('arial', 10), bg='lightgray',
                                          foreground='blue')
                    bookcombo = ttk.Combobox(canvas, font=('arial', 10),
                                             values=['Rates', 'RROD', 'AARP', 'SWR1', 'Go Fast', 'Go Free'],
                                             width=10, foreground='#263238')
                    bookcombo.insert(0, 'Rates')
                    # book_rooms.append(bookcombo)
                    book_bed_rates.append(bookcombo)

                    ratebedscombo = Label(canvas, text="  Bed Class", font=('arial', 10), bg='lightgray',
                                          foreground='blue')
                    bedscombo = ttk.Combobox(canvas, font=('arial', 10),
                                             values=['Bed Class', 'One Bed', 'Two Beds'],
                                             width=10, foreground='#263238')
                    bedscombo.insert(0, 'Bed Class')
                    # book_rooms.append(bedscombo)
                    book_bed_rates.append(bedscombo)

                    canvas_nyts = Label(canvas, text="Night", font=('arial', 10), bg='lightgray', foreground='blue')

                    canvas_nyts1 = Label(canvas, text=count_rooms, font=('arial', 9),
                                         background='white', padx=12, pady=1, border=1, relief='sunken')

                    canvas.create_window(10, x + 240, anchor="nw", window=ratebookcombo)
                    canvas.create_window(100, x + 240, anchor="nw", window=bookcombo)
                    canvas.create_window(230, x + 240, anchor="nw", window=ratebedscombo)
                    canvas.create_window(320, x + 240, anchor="nw", window=bedscombo)

                    canvas.create_window(450, x + 240, anchor="nw", window=canvas_nyts)
                    canvas.create_window(500, x + 240, anchor="nw", window=canvas_nyts1)

                    spacing = Label(canvas, text=' ', font=('arial', 9), fg='black',
                                    bg="lightgray")
                    # canvas.create_window(1150, x + 330, anchor="nw", window=spacing)
                    x += 300

            data_search += 1

        my_frame1.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
        my_frame1.place(x=200, y=176)
        resv_forget.append(canvas)
        resv_forget.append(my_frame1)
        # site_scrool1_forget.append(canvas)
        site_scrool1_forget.append(my_frame1)

        back_confirm_forget.append(canvas)
        back_confirm_forget.append(my_frame1)
        # print('back confirm forget ',back_confirm_forget)

        '''confirmed_dates = Button(window, text='Next', font=('arial', 10), fg='black', bg='lightgray',
                                 activeforeground="#00695C", activebackground="lightgray", padx=15, pady=2, width=5,
                                 # wraplength= 1,
                                 # state='disable',
                                 command=confirmed_dates2)
        confirmed_dates.place(x=310, y=600)
        resv_forget.append(confirmed_dates)
        site_scrool1_forget.append(confirmed_dates)

        modified_dates = Button(window, text='Back', font=('arial', 10), fg='black', bg='lightgray',
                                activeforeground="#00695C", activebackground="lightgray", padx=15, pady=2, width=5,
                                # wraplength= 1,
                                # state='disable',
                                command=modified_dates1)
        modified_dates.place(x=210, y=600)
        resv_forget.append(modified_dates)
        site_scrool1_forget.append(modified_dates)'''

        change_propety = Button(window, text='Change Property', font=('arial', 9), fg='black', bg='lightgray',
                                activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                                # wraplength= 1,
                                # state='disable',
                                command=modified_dates1)
        change_propety.place(x=210, y=615)
        resv_forget.append(change_propety)
        site_scrool1_forget.append(change_propety)

        next_button = Button(window, text='Next', font=('arial', 9), fg='black', bg='lightgray',
                             activeforeground="#00695C", activebackground="lightgray", padx=80, pady=1, width=1,
                             # wraplength= 1,
                             # state='disable',
                             command=confirmed_dates2)
        next_button.place(x=210, y=585)
        resv_forget.append(next_button)
        site_scrool1_forget.append(next_button)

    else:
        pass


def confirmed_dates1():
    print('Confirmed Dates1')
    site_scrolls1()


def modified_dates1(event=None):
    site_condition[0] = 'False'
    # text = 'Back'
    # bookrates = book_rooms[-2]8.get()
    # bookbeds = book_rooms[-1]9.get()
    print('Modified Dates')
    # print('site scrool loops ', site_scrool1_forget)
    site_scrolls()


def no_of_nyts_post():
    print('no of nights')
    correct_date = 'True'
    try:
        for x in range(0, len(book2_rooms)):
            checkinmonth1 = book2_rooms[0].get()
            checkindays = book2_rooms[1].get()
            checkinyear = book2_rooms[2].get()
            checkoutmonth1 = book2_rooms[3].get()
            checkoutdays = book2_rooms[4].get()
            checkoutyear = book2_rooms[5].get()

        checkinmonth = datetime.datetime.strptime(checkinmonth1, '%B').month
        checkoutmonth = datetime.datetime.strptime(checkoutmonth1, '%B').month
        dateins = date(int(checkinyear), int(checkinmonth), int(checkindays))

        dateouts = date(int(checkoutyear), int(checkoutmonth), int(checkoutdays))
        bookrooms = dateouts - dateins
        bookrooms = bookrooms.days

        if bookrooms == 0:
            bookrooms = 1
        else:
            pass
        # check in - toda date
        today = date.today()
        today_checkin = dateins - today

        if today_checkin.days >= 0:
            # print('correct check in dates ', today_checkin.days)
            pass
        else:
            # print('incorrect check in dates ', today_checkin.days)
            correct_date = 'False'

        today_checkout = dateouts - dateins
        # print('check out dates ', today_checkout.days)
        if today_checkout.days <= 366:
            if today_checkout.days >= 0:
                # print('correct check out dates ', today_checkout.days)
                pass
            else:
                # print('incorrect check out dates ', today_checkout.days)
                correct_date = 'False'
        else:
            # print('incorrect check out dates ', today_checkout.days)
            correct_date = 'False'

        return bookrooms, correct_date

    except ValueError:
        pass
        # print('ValueError: invalid literal for int() with base 10: Year')


def confirmed_dates2(event=None):
    print('Confirmed Two Dates')
    rates_list = ['RROD', 'AARP', 'SWR1', 'Go Fast', 'Go Free']
    bed_list = ['One Bed', 'Two Beds']
    rates_list1 = ''
    bed_list1 = ''
    loop = 0
    for x in book_bed_rates:
        if x.get() in ['RROD', 'AARP', 'SWR1', 'Go Fast', 'Go Free']:
            rates_list1 = x
            loop += 1
        else:
            pass
    for x in book_bed_rates:
        if x.get() in ['One Bed', 'Two Beds']:
            bed_list1 = x
            loop += 1
        else:
            pass

    if loop == 2:
        book_rooms.append('1')
        book_rooms.append('1')
        if len(book_rooms) > 10:
            book_rooms.remove(book_rooms[10])
            book_rooms.remove(book_rooms[10])
        book_rooms[8] = rates_list1
        book_rooms[9] = bed_list1
        booked()
    else:
        pass
        # print('back to site scrolls 1')


def booked():
    print('booked')
    # print('booked ', book_get_site[-1])
    bookrooms = 0
    booked_value1 = 0
    # value_resv_stay = resv_stay(label53.get())
    # print('reservation valueeeeeeeeeeeeeeeeeee ', value_resv_stay)
    # levelcolor = value_resv_stay[1]
    levelcolor = 'lightblue'
    # print('reservation level color ', value_resv_stay[1])

    '''for x in range(0, len(book_rooms)):
        print('reservation booked ', book_rooms[x])'''

    for x in range(0, len(book_rooms)):
        # print('reservation booked ', book_rooms[x].get())
        checkinmonth1 = book_rooms[0].get()
        checkindays = book_rooms[1].get()
        checkinyear = book_rooms[2].get()
        checkoutmonth1 = book_rooms[3].get()
        checkoutdays = book_rooms[4].get()
        checkoutyear = book_rooms[5].get()
        bookadults = book_rooms[6].get()
        bookchild = book_rooms[7].get()
        try:
            bookrates = book_rooms[8].get()
        except:
            bookrates = book_rooms[8]
        try:
            bookbeds = book_rooms[9].get()
        except:
            bookbeds = book_rooms[9]

    value_month = ['Month', 'January', 'February', 'March', 'April', 'May', 'June', 'July',
                   'August', 'September', 'October', 'November', 'December']
    valuedates = 0
    for x in range(2023, 2028):
        if str(checkinyear) == str(x):
            valuedates += 1
            # print('correct number')

    for x in range(2023, 2028):
        if str(checkoutyear) == str(x):
            valuedates += 1
            # print('correct number')

    for x in range(1, len(value_month)):
        if str(checkinmonth1) == value_month[x]:
            valuedates += 1
            # print('correct number')
    for x in range(1, len(value_month)):
        if str(checkoutmonth1) == value_month[x]:
            valuedates += 1
            # print('correct number')
    for x in range(1, 32):
        if str(checkindays) == str(x):
            valuedates += 1
            # print('correct number')
    for x in range(1, 32):
        if str(checkoutdays) == str(x):
            valuedates += 1
            # print('correct number')

    if valuedates == 6:
        # number of nyts
        checkinmonth = datetime.datetime.strptime(checkinmonth1, '%B').month
        checkoutmonth = datetime.datetime.strptime(checkoutmonth1, '%B').month
        booked_checkin_dates1 = wyn_tscroll_dates_GUI.book_dates(checkinyear, checkinmonth, checkindays,
                                                                 checkoutyear, checkoutmonth, checkoutdays)
        if booked_checkin_dates1 == 1:
            valuedates = 1
        else:
            pass

    if valuedates == 6:

        dateins = date(int(checkinyear), int(checkinmonth), int(checkindays))
        dateouts = date(int(checkoutyear), int(checkoutmonth), int(checkoutdays))
        bookrooms = dateouts - dateins
        bookrooms = bookrooms.days
        if bookrooms == 0:
            bookrooms = 1
        else:
            pass

        # check in dates
        check_ins = datetime.datetime(int(checkinyear), int(checkinmonth), int(checkindays))
        check_ins1 = check_ins.strftime("%a, %b %d -")
        # print(check_ins1)

        # check out dates
        check_out = datetime.datetime(int(checkoutyear), int(checkoutmonth), int(checkoutdays))
        check_out1 = check_out.strftime(" %a, %b %d %Y")

        newcheck_ins = datetime.datetime(int(checkinyear), int(checkinmonth), int(checkindays))
        newcheck_ins1 = newcheck_ins.strftime('"%d/%m/%Y"')

        # check out dates
        newcheck_out = datetime.datetime(int(checkoutyear), int(checkoutmonth), int(checkoutdays))
        newcheck_out1 = newcheck_out.strftime('"%d/%m/%Y"')

    else:
        # messagebox.showinfo(title='Check-out ', message='Invalid Dates')
        check_ins1 = 'Wrong Date Input'
        check_out1 = ''

    # rate names
    ratenames = ''
    if bookrates == 'RROD':
        ratenames = 'Wyndham rewards flexible rate'
    elif bookrates == 'SWR1':
        ratenames = 'Wyndham rewards member rate'
    elif bookrates == 'AARP':
        ratenames = 'Wyndham rewards aarp rate'
    elif bookrates == 'Go Fast':
        ratenames = 'Wyndham rewards go fast rate'
    elif bookrates == 'Go Free':
        ratenames = 'Wyndham rewards go free rate'
    else:
        messagebox.showinfo(title='Bookrates ', message='Invalid Book Rates')

    # one or two beds
    no_bed = 0
    bedsize = ''
    if bookbeds == 'One Bed':
        bedsize = 'One King Bed'
        no_bed = 1
    if bookbeds == 'Two Beds':
        bedsize = 'Two Queen Beds'
        no_bed = 1
    if no_bed == 0:
        messagebox.showinfo(title='Bookbeds ', message='Invalid Book Beds')

    bookfirst = ''
    booklast = ''
    bookmiddle = ''
    booksite = ''
    phone_home_list = []
    for x in book_sites1:
        book_site = x

    if valuedates != 6:
        messagebox.showinfo(title='Booked Date ', message='Invalid Dates')
        pass
    elif bookrates == 'Rates':
        pass
    elif bookbeds == 'Bed Class':
        pass
    else:

        for data_search in range(2, data_number3.max_row + 1):
            siteno = data_number3.cell(data_search, 1).value
            hotelname = data_number3.cell(data_search, 2).value
            if str(book_site) == str(siteno):
                site_booksite = hotelname
        if int(bookrooms) == 1:
            night_message = 'night'
        else:
            night_message = 'nights'
        answer = 'yes'
        '''answer = messagebox.askquestion(title='Book a Reservation', message='I have you Confirmed in ' + site_booksite +
                                                                            ' on ' + str(checkinmonth1) + ' ' + str(
            checkindays) + ', ' + str(checkinyear) + ' for ' +
                                                                            str(bookrooms) + ' ' + night_message)'''
        if answer == 'yes':
            for data_search1 in range(2, data_number.max_row + 1):
                acctno = data_number.cell(data_search1, 1).value
                phone_data = data_number.cell(data_search1, 2).value
                firstname = data_number.cell(data_search1, 4).value
                lastname = data_number.cell(data_search1, 3).value
                middlename = data_number.cell(data_search1, 5).value
                phone_home1 = data_number.cell(data_search1, 15).value
                home_data = data_number.cell(data_search1, 17).value

                # print('phone home ', phone_home1)
                if str(26) == str(acctno):
                    resv_mbrno = acctno
                    bookfirst = firstname
                    booklast = lastname
                    bookmiddle = middlename
                    phone_home_list.append(acctno)
                    if str(phone_home1) == 'Phone':
                        savecell = phone_data
                        # print('account no. and phone home ', savecell)
                        # phone_home_list.append(savecell)
                    elif str(phone_home1) == 'Home':
                        savecell = home_data
                        # phone_home_list.append(savecell)
                        # print('account no. and phone home ', savecell)
            count = 1
            total_amount = 0
            total_amount1 = 0
            taxes = 0
            payment = 0
            newgofast = 0
            newgofree = 0
            address = ''

            for data_search in range(2, data_number3.max_row + 1):
                siteno = data_number3.cell(data_search, 1).value
                hotelname = data_number3.cell(data_search, 2).value
                address = data_number3.cell(data_search, 3).value
                phoneno = data_number3.cell(data_search, 4).value
                hotelstar = data_number3.cell(data_search, 5).value
                rrodking = data_number3.cell(data_search, 8).value
                swriking = data_number3.cell(data_search, 10).value
                arrpking = data_number3.cell(data_search, 12).value
                fastking = data_number3.cell(data_search, 14).value
                gofastpoints = data_number3.cell(data_search, 16).value
                freeking = data_number3.cell(data_search, 19).value
                rrodqueens = data_number3.cell(data_search, 9).value
                swriqueens = data_number3.cell(data_search, 11).value
                arrpqueens = data_number3.cell(data_search, 13).value
                fastqueens = data_number3.cell(data_search, 15).value
                # print('data search ', data_search, count)
                # print('book ratesssssssssss ',bookrates)
                if count == 1:
                    if bookrates == 'RROD':

                        if bookbeds == 'One Bed':
                            bookrooms = int(bookrooms)
                            total_amount = '$ ' + str(conversion_numbers1(rrodking))
                            total_amount1 = 'USD ' + str(conversion_numbers1(rrodking * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((rrodking * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((rrodking * bookrooms) * .05 + (rrodking * bookrooms)))

                            newtotal_amount = rrodking
                            newtotal_amount1 = rrodking * bookrooms
                            newtaxes_new = (rrodking * bookrooms) * .05
                            newpayment = ((rrodking * bookrooms) * .05) + (rrodking * bookrooms)
                            newtaxes = rrodking * .05
                            newgofast = 0
                            newgofreetotal = 0

                        else:
                            bookrooms = int(bookrooms)
                            total_amount = '$ ' + str(conversion_numbers1(rrodqueens))
                            total_amount1 = 'USD ' + str(conversion_numbers1(rrodqueens * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((rrodqueens * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((rrodqueens * bookrooms) * .05 + (rrodqueens * bookrooms)))

                            newtotal_amount = rrodqueens
                            newtotal_amount1 = rrodqueens * bookrooms
                            newtaxes_new = (rrodqueens * bookrooms) * .05
                            newpayment = ((rrodqueens * bookrooms) * .05) + (rrodqueens * bookrooms)
                            newtaxes = rrodqueens * .05
                            newgofast = 0
                            newgofreetotal = 0

                    elif bookrates == 'SWR1':
                        bookrooms = int(bookrooms)
                        if bookbeds == 'One Bed':
                            total_amount = '$ ' + str(conversion_numbers1(swriking))
                            total_amount1 = 'USD ' + str(conversion_numbers1(swriking * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((swriking * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((swriking * bookrooms) * .05 + (swriking * bookrooms)))

                            newtotal_amount = swriking
                            newtotal_amount1 = swriking * bookrooms
                            newtaxes_new = (swriking * bookrooms) * .05
                            newpayment = ((swriking * bookrooms) * .05) + (swriking * bookrooms)
                            newtaxes = swriking * .05
                            newgofast = 0
                            newgofreetotal = 0
                        else:
                            bookrooms = int(bookrooms)
                            total_amount = '$ ' + str(conversion_numbers1(swriqueens))
                            total_amount1 = 'USD ' + str(conversion_numbers1(swriqueens * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((swriqueens * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((swriqueens * bookrooms) * .05 + (swriqueens * bookrooms)))

                            newtotal_amount = swriqueens
                            newtotal_amount1 = swriqueens * bookrooms
                            newtaxes_new = (swriqueens * bookrooms) * .05
                            newpayment = ((swriqueens * bookrooms) * .05) + (swriqueens * bookrooms)
                            newtaxes = swriqueens * .05
                            newgofast = 0
                            newgofreetotal = 0

                    elif bookrates == 'AARP':
                        bookrooms = int(bookrooms)
                        if bookbeds == 'One Bed':
                            # print('type of number ', arrpking, )
                            total_amount = '$ ' + str(conversion_numbers1(arrpking))
                            total_amount1 = 'USD ' + str(conversion_numbers1(arrpking * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((arrpking * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((arrpking * bookrooms) * .05 + (arrpking * bookrooms)))

                            newtotal_amount = arrpking
                            newtotal_amount1 = arrpking * bookrooms
                            newtaxes_new = (arrpking * bookrooms) * .05
                            newpayment = ((arrpking * bookrooms) * .05) + (arrpking * bookrooms)
                            newtaxes = arrpking * .05
                            newgofast = 0
                            newgofreetotal = 0
                        else:
                            bookrooms = int(bookrooms)
                            total_amount = '$ ' + str(conversion_numbers1(arrpqueens))
                            total_amount1 = 'USD ' + str(conversion_numbers1(arrpqueens * bookrooms))
                            taxes = 'USD ' + str(conversion_numbers1((arrpqueens * bookrooms) * .05))
                            payment = 'USD ' + str(
                                conversion_numbers1((arrpqueens * bookrooms) * .05 + (arrpqueens * bookrooms)))

                            newtotal_amount = arrpqueens
                            newtotal_amount1 = arrpqueens * bookrooms
                            newtaxes_new = (arrpqueens * bookrooms) * .05
                            newpayment = ((arrpqueens * bookrooms) * .05) + (arrpqueens * bookrooms)
                            newtaxes = arrpqueens * .05
                            newgofast = 0
                            newgofreetotal = 0

                    elif bookrates == 'Go Fast':
                        bookrooms = int(bookrooms)
                        if bookbeds == 'One Bed':
                            total_amount = '$' + str(conversion_numbers1(fastking)) + ' + ' + str(
                                conversion_numbers(gofastpoints)) + ' pts'
                            total_amount1 = 'USD ' + str(conversion_numbers1(fastking * bookrooms)) + ' + ' + str(
                                conversion_numbers(int(gofastpoints) * bookrooms)) + ' pts'
                            taxes = 'USD ' + str(conversion_numbers1((fastking * bookrooms) * .05))
                            payment = 'USD ' + str(conversion_numbers1(
                                (fastking * bookrooms) * .05 + (fastking * bookrooms))) + ' + ' + str(
                                conversion_numbers(gofastpoints * bookrooms)) + ' pts'

                            newtotal_amount = fastking
                            newtotal_amount1 = fastking * bookrooms
                            newtaxes_new = (fastking * bookrooms) * .05
                            newpayment = ((fastking * bookrooms) * .05) + (fastking * bookrooms)
                            newtaxes = fastking * .05
                            newgofast = gofastpoints
                            newgofreetotal = int(gofastpoints) * int(bookrooms)

                        else:
                            bookrooms = int(bookrooms)
                            total_amount = '$' + str(conversion_numbers1(fastqueens)) + ' + ' + str(
                                conversion_numbers(gofastpoints)) + ' pts'
                            total_amount1 = 'USD ' + str(conversion_numbers1(fastqueens * bookrooms)) + ' + ' + str(
                                conversion_numbers(int(gofastpoints) * bookrooms)) + ' pts'
                            taxes = 'USD ' + str(conversion_numbers1((fastqueens * bookrooms) * .05))
                            payment = 'USD ' + str(conversion_numbers1(
                                (fastqueens * bookrooms) * .05 + (fastqueens * bookrooms))) + ' + ' + str(
                                conversion_numbers(int(gofastpoints) * bookrooms)) + ' pts'

                            newtotal_amount = fastqueens
                            newtotal_amount1 = fastqueens * bookrooms
                            newtaxes_new = (fastqueens * bookrooms) * .05
                            newpayment = ((fastqueens * bookrooms) * .05) + (fastqueens * bookrooms)
                            newtaxes = fastqueens * .05
                            newgofast = conversion_numbers(gofastpoints)
                            newgofreetotal = conversion_numbers(int(gofastpoints) * int(bookrooms))

                    elif bookrates == 'Go Free':

                        bookrooms = int(bookrooms)
                        total_amount = str(conversion_numbers(freeking)) + ' pts'
                        total_amount1 = str(conversion_numbers(int(freeking) * int(bookrooms))) + ' pts'
                        taxes = 'USD ' + str(conversion_numbers1(0))
                        payment = str(conversion_numbers(int(freeking) * int(bookrooms))) + ' pts'

                        newtotal_amount = 0
                        newtotal_amount1 = 0
                        newtaxes_new = 0
                        newpayment = 0
                        newtaxes = 0
                        newgofast = freeking
                        newgofreetotal = int(freeking) * int(bookrooms)

                    else:
                        pass

                    if str(book_site) == str(siteno):
                        count = 2
                        bookrooms = int(bookrooms)
                        newaddress = address
                        newhotelname = hotelname
                        newphone = phoneno

                        newtotal_amount_new = newtotal_amount
                        newtotal_amount1_new = newtotal_amount1
                        newtaxes_new1 = newtaxes_new
                        newpayment1 = newpayment

                        newtaxes1 = newtaxes
                        newgofast1 = newgofast
                        newgofreetotal1 = newgofreetotal

                        bookrooms = int(bookrooms)
                        rrod_king1 = Label(window, text='$' + str(
                            conversion_numbers1((rrodking * bookrooms) * .05 + (rrodking * bookrooms))),
                                           font=('arial', 10), bg="lightgray", foreground='#263238',
                                           padx=0,
                                           pady=10)

                        swri_king1 = Label(window, text='$' + str(
                            conversion_numbers1((swriking * bookrooms) * .05 + (swriking * bookrooms))),
                                           font=('arial', 10), bg="lightgray", foreground='#263238',
                                           padx=0,
                                           pady=10)

                        arrp_king1 = Label(window, text='$' + str(
                            conversion_numbers1((arrpking * bookrooms) * .05 + (arrpking * bookrooms))),
                                           font=('arial', 10), bg="lightgray", foreground='#263238',
                                           padx=0,
                                           pady=10)

                        fast_king1 = Label(window, text='$' + str(
                            conversion_numbers1((fastking * bookrooms) * .05 + (fastking * bookrooms))) + ' + ' + str(
                            conversion_numbers(gofastpoints * bookrooms)) + ' pts',
                                           font=('arial', 10), bg="lightgray", foreground='#263238', padx=0, pady=10)

                        free_king1 = Label(window, text=str(conversion_numbers(int(freeking) * bookrooms)) + ' pts',
                                           font=('arial', 10), bg="lightgray", foreground='#263238', padx=0, pady=10)

                        # two queens beds total

                        rrod_queens = Label(window, text='$' + str(
                            conversion_numbers1((rrodqueens * bookrooms) * .05 + (rrodqueens * bookrooms))),
                                            font=('arial', 10), bg="lightgray", foreground='#263238',
                                            padx=0,
                                            pady=10)

                        swri_queens = Label(window, text='$' + str(
                            conversion_numbers1((swriqueens * bookrooms) * .05 + (swriqueens * bookrooms))),
                                            font=('arial', 10), bg="lightgray", foreground='#263238',
                                            padx=0,
                                            pady=10)

                        arrp_queens = Label(window, text='$' + str(
                            conversion_numbers1((arrpqueens * bookrooms) * .05 + (arrpqueens * bookrooms))),
                                            font=('arial', 10), bg="lightgray", foreground='#263238',
                                            padx=0,
                                            pady=10)

                        fast_queens = Label(window,
                                            text='$' + str(conversion_numbers1(
                                                (fastqueens * bookrooms) * .05 + (
                                                        fastqueens * bookrooms))) + ' + ' + str(
                                                conversion_numbers(gofastpoints * bookrooms)) + ' pts',
                                            font=('arial', 10), bg="lightgray", foreground='#263238', padx=0, pady=10)

                        free_queen = Label(window, text=str(conversion_numbers(int(freeking) * bookrooms)) + ' pts',
                                           font=('arial', 10), bg="lightgray", foreground='#263238', padx=0, pady=10)

                        # temp_resv1 = data_number2.max_row + len(temp_resv)
                        new_member = resv_mbrno
                        data_input1 = 0

                        my_frame = Frame(window)
                        canvas = Canvas(my_frame)

                        canvas = Canvas(my_frame, width=373, height=440, background="lightgray")
                        canvas.pack(fill="both", expand=True)

                        vsb = Scrollbar(my_frame, orient="vertical", command=canvas.yview, width=25)
                        hsb = Scrollbar(my_frame, orient="horizontal", command=canvas.xview, width=25)

                        window.grid_rowconfigure(0, weight=1)
                        window.grid_columnconfigure(0, weight=1)
                        canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
                        canvas.grid(row=0, column=0, sticky="nsew")

                        vsb.grid(row=0, column=1, sticky="ns")
                        hsb.grid(row=1, column=0, sticky="ew")

                        a = 0
                        b = 0

                        space_color = 0
                        resv_head = Label(canvas, text="  Reservation Details  ", font=('arial', 15),
                                          bg='lightblue', foreground='blue', padx=0, pady=10)
                        canvas.create_window(10, 50, anchor="nw", window=resv_head)

                        data_value = data_number4.max_row
                        # points conversion
                        points2 = ''
                        remove_coma = [str(x) for x in str(newgofast1)]
                        for x in remove_coma:
                            if x == ',':
                                pass
                            else:
                                points2 += str(x)
                        total_points2 = ''
                        remove_coma1 = [str(x) for x in str(newgofreetotal1)]
                        for x in remove_coma1:
                            if x == ',':
                                pass
                            else:
                                total_points2 += str(x)

                        room_price1 = conversion_numbers1(newtotal_amount_new)
                        total_amount1 = conversion_numbers1(newtotal_amount1_new)
                        taxes1 = conversion_numbers1(newtaxes_new1)
                        payment1 = conversion_numbers1(newpayment1)
                        points1 = conversion_numbers(points2)
                        total_points1 = conversion_numbers(total_points2)

                        def_booked_get_value[0] = siteno
                        def_booked_get_value[1] = newhotelname
                        def_booked_get_value[2] = check_ins1
                        def_booked_get_value[3] = check_out1
                        def_booked_get_value[4] = bookadults
                        def_booked_get_value[5] = bookchild
                        def_booked_get_value[6] = bookfirst
                        def_booked_get_value[7] = booklast
                        def_booked_get_value[8] = bedsize
                        def_booked_get_value[9] = ratenames
                        def_booked_get_value[10] = bookrates
                        def_booked_get_value[11] = points1
                        def_booked_get_value[12] = room_price1
                        def_booked_get_value[13] = total_points1
                        def_booked_get_value[14] = total_amount1
                        def_booked_get_value[15] = taxes1
                        def_booked_get_value[16] = payment1
                        def_booked_get_value[17] = bookrooms

                        confirm = Label(canvas, text=" Pending ", font=('arial', 11), fg='white', bg='#FFB300')

                        hotels = Label(canvas, text='(' + str(siteno) + ')' + '  ' + str(newhotelname),
                                       font=('arial', 11, 'bold'), fg='black', bg='lightgray')
                        dates = Label(canvas, text=str(check_ins1) + str(check_out1), font=('arial', 10, 'bold'),
                                      fg='black', bg='lightgray')
                        adults = Label(canvas, text=str(bookadults) + ' adults ' + str(bookchild) + ' child',
                                       font=('arial', 10), fg='black', bg='lightgray')

                        fullname = Label(canvas, text=str(cap_name(bookfirst)) + ' ' + str(cap_name(booklast)),
                                         font=('arial', 10), fg='blue', bg='lightgray')

                        rooms1 = Label(canvas, text=bedsize, font=('arial', 10), fg='black', bg='lightgray')
                        rate_rooms = Label(canvas, text=str(ratenames) + '  ' + str(bookrates), font=('arial', 10),
                                           fg='blue', bg='lightgray')

                        if bookrates == 'Go Free':
                            room_payment = Label(canvas, text=str(points1) + ' pts ' + 'avg / night ',
                                                 font=('arial', 10),
                                                 fg='black', bg='lightgray')
                        elif bookrates == 'Go Fast':
                            room_payment = Label(canvas,
                                                 text=str(points1) + ' pts + $' + str(room_price1) + ' avg / night ',
                                                 font=('arial', 10),
                                                 fg='black', bg='lightgray')
                        else:
                            room_payment = Label(canvas, text='$' + str(room_price1) + ' avg / night ',
                                                 font=('arial', 10),
                                                 fg='black', bg='lightgray')

                        if bookrates == 'Go Free':
                            nyts_payment = Label(canvas, text='1 room ' + str(bookrooms) + ' night',
                                                 font=('arial', 10), fg='black', bg='lightgray')
                            nyts_payment1 = Label(canvas, text=str(total_points1) + ' pts', font=('arial', 10),
                                                  fg='black',
                                                  bg='lightgray')
                            tax_fees = Label(canvas, text=str('taxes and fees '), font=('arial', 10), fg='black',
                                             bg='lightgray')
                            tax_fees1 = Label(canvas, text='USD ' + str(taxes1), font=('arial', 10), fg='black',
                                              bg='lightgray')
                            tax_total = Label(canvas, text="tax and total payment", font=('arial', 10), fg='black',
                                              bg='lightgray')
                            tax_total1 = Label(canvas, text=str(total_points1) + ' pts', font=('arial', 10), fg='black',
                                               bg='lightgray')


                        elif bookrates == 'Go Fast':
                            nyts_payment = Label(canvas, text='1 room ' + str(bookrooms) + ' night',
                                                 font=('arial', 10), fg='black', bg='lightgray')
                            nyts_payment1 = Label(canvas,
                                                  text='USD ' + str(total_amount1) + ' + ' + total_points1 + ' pts',
                                                  font=('arial', 10),
                                                  fg='black',
                                                  bg='lightgray')
                            tax_fees = Label(canvas, text=str('taxes and fees '), font=('arial', 10), fg='black',
                                             bg='lightgray')
                            tax_fees1 = Label(canvas, text='USD ' + str(taxes1), font=('arial', 10), fg='black',
                                              bg='lightgray')
                            tax_total = Label(canvas, text="tax and total payment", font=('arial', 10), fg='black',
                                              bg='lightgray')
                            tax_total1 = Label(canvas, text='USD ' + str(payment1) + ' + ' + total_points1 + ' pts',
                                               font=('arial', 10),
                                               fg='black',
                                               bg='lightgray')

                        else:
                            nyts_payment = Label(canvas, text='1 room ' + str(bookrooms) + ' night',
                                                 font=('arial', 10), fg='black', bg='lightgray')
                            nyts_payment1 = Label(canvas, text='USD ' + str(total_amount1), font=('arial', 10),
                                                  fg='black',
                                                  bg='lightgray')
                            tax_fees = Label(canvas, text=str('taxes and fees '), font=('arial', 10), fg='black',
                                             bg='lightgray')
                            tax_fees1 = Label(canvas, text='USD ' + str(taxes1), font=('arial', 10), fg='black',
                                              bg='lightgray')
                            tax_total = Label(canvas, text="tax and total payment", font=('arial', 10), fg='black',
                                              bg='lightgray')
                            tax_total1 = Label(canvas, text='USD ' + str(payment1), font=('arial', 10), fg='black',
                                               bg='lightgray')

                        space = Label(canvas, text='', font=('arial', 10), fg='black', bg='lightgray')

                        canvas.create_window(23, 115, anchor="nw", window=confirm)
                        canvas.create_window(23, 143, anchor="nw", window=hotels)
                        canvas.create_window(23, 168, anchor="nw", window=dates)
                        canvas.create_window(23, 193, anchor="nw", window=adults)
                        # canvas.create_window(23, 218 , anchor="nw", window=fullname)
                        canvas.create_window(23, 218, anchor="nw", window=rooms1)
                        canvas.create_window(23, 243, anchor="nw", window=rate_rooms)
                        canvas.create_window(23, 268, anchor="nw", window=room_payment)
                        canvas.create_window(23, 348, anchor="nw", window=nyts_payment)
                        canvas.create_window(190, 348, anchor="nw", window=nyts_payment1)
                        canvas.create_window(23, 378, anchor="nw", window=tax_fees)
                        canvas.create_window(190, 378, anchor="nw", window=tax_fees1)
                        canvas.create_window(23, 408, anchor="nw", window=tax_total)
                        canvas.create_window(190, 408, anchor="nw", window=tax_total1)

                        '''stay_buttons = Button(canvas, text="Modify", font=("arial", 10), fg="black", bg="#F0F0F0",
                                              activeforeground="green", activebackground="#F0F0F0", border=2,
                                              relief=RAISED, padx=6,
                                              pady=2, width=7,
                                              # wraplength= 1,
                                              # state='disable',
                                              command=site_scrools2)
                        canvas.create_window(160, 470 + a, anchor="nw", window=stay_buttons)
                        canvas.create_window(10, 510 + a, anchor="nw", window=space)'''

                        '''stay_labels = Label(canvas, bg="lightgray", padx=8, pady=7,width=8)
                        canvas.create_window(160, 470 + a, anchor="nw", window=stay_labels)'''

                        '''cxl_buttons = Button(canvas, text="Cancel", font=("arial", 10), fg="black", bg="#F0F0F0",
                                             activeforeground="green", activebackground="#F0F0F0", border=2,
                                             relief=RAISED, padx=6,
                                             pady=2, width=7,
                                             # wraplength= 1,
                                             state='disable',
                                             command=site_scrolls)
                        canvas.create_window(200, 470, anchor="nw", window=cxl_buttons)
                        canvas.create_window(10, 510, anchor="nw", window=space)'''

                        my_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
                        my_frame.place(x=870, y=176)
                        resv_forget.append(my_frame)

                        view_post1_payment[0] = points1
                        view_post1_payment[1] = room_price1
                        view_post1_payment[2] = total_points1
                        view_post1_payment[3] = payment1
                        post1()


def def_booked_value():
    change_shift_resv[0] = 1
    change_shift_resv[1] = 0
    change_shift_resv[2] == 0
    print('define booked value')
    resvno = view_resvno[0].get()
    # print('reservation number ', resvno)
    lastname = ''
    firstname = ''
    for data_search in range(2, data_number.max_row + 1):
        accntno = data_number.cell(data_search, 1).value
        last_name = data_number.cell(data_search, 3).value
        first_name = data_number.cell(data_search, 4).value
        if str(accntno) == str(resvno):
            lastname = last_name
            firstname = first_name

    '''for x in def_booked_get_value:
        print('add reservation value ', x)'''
    bookrates = def_booked_get_value[10]
    my_frame = Frame(window)
    canvas = Canvas(my_frame)

    canvas = Canvas(my_frame, width=373, height=440, background="lightgray")
    canvas.pack(fill="both", expand=True)

    vsb = Scrollbar(my_frame, orient="vertical", command=canvas.yview, width=25)
    hsb = Scrollbar(my_frame, orient="horizontal", command=canvas.xview, width=25)

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)
    canvas.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)
    canvas.grid(row=0, column=0, sticky="nsew")

    vsb.grid(row=0, column=1, sticky="ns")
    hsb.grid(row=1, column=0, sticky="ew")

    resv_head = Label(canvas, text="  Reservation Details  ", font=('arial', 15),
                      bg='lightblue', foreground='blue', padx=0, pady=10)
    canvas.create_window(10, 50, anchor="nw", window=resv_head)

    confirm = Label(canvas, text=" Pending ", font=('arial', 11), fg='white', bg='#FFB300')

    hotels = Label(canvas, text='(' + str(def_booked_get_value[0]) + ')' + '  ' + str(def_booked_get_value[1]),
                   font=('arial', 11, 'bold'), fg='black', bg='lightgray')
    dates = Label(canvas, text=str(def_booked_get_value[2]) + str(def_booked_get_value[3]), font=('arial', 10, 'bold'),
                  fg='black', bg='lightgray')
    adults = Label(canvas, text=str(def_booked_get_value[4]) + ' adults ' + str(def_booked_get_value[5]) + ' child',
                   font=('arial', 10), fg='black', bg='lightgray')

    fullname = Label(canvas, text=str(cap_name(firstname)) + ' ' + str(cap_name(lastname)),
                     font=('arial', 10), fg='blue', bg='lightgray')

    rooms1 = Label(canvas, text=def_booked_get_value[8], font=('arial', 10), fg='black', bg='lightgray')
    rate_rooms = Label(canvas, text=str(def_booked_get_value[9]) + '  ' + str(def_booked_get_value[10]),
                       font=('arial', 10),
                       fg='blue', bg='lightgray')

    if bookrates == 'Go Free':
        room_payment = Label(canvas, text=str(def_booked_get_value[11]) + ' pts ' + 'avg / night ', font=('arial', 10),
                             fg='black', bg='lightgray')
    elif bookrates == 'Go Fast':
        room_payment = Label(canvas, text=str(def_booked_get_value[11]) + ' pts + $' + str(
            def_booked_get_value[12]) + ' avg / night ',
                             font=('arial', 10),
                             fg='black', bg='lightgray')
    else:
        room_payment = Label(canvas, text='$' + str(def_booked_get_value[12]) + ' avg / night ', font=('arial', 10),
                             fg='black', bg='lightgray')

    if bookrates == 'Go Free':
        nyts_payment = Label(canvas, text='1 room ' + str(def_booked_get_value[17]) + ' night',
                             font=('arial', 10), fg='black', bg='lightgray')
        nyts_payment1 = Label(canvas, text=str(def_booked_get_value[13]) + ' pts', font=('arial', 10),
                              fg='black',
                              bg='lightgray')
        tax_fees = Label(canvas, text=str('taxes and fees '), font=('arial', 10), fg='black',
                         bg='lightgray')
        tax_fees1 = Label(canvas, text='USD ' + str(def_booked_get_value[15]), font=('arial', 10), fg='black',
                          bg='lightgray')
        tax_total = Label(canvas, text="tax and total payment", font=('arial', 10), fg='black',
                          bg='lightgray')
        tax_total1 = Label(canvas, text=str(def_booked_get_value[13]) + ' pts', font=('arial', 10), fg='black',
                           bg='lightgray')


    elif bookrates == 'Go Fast':
        nyts_payment = Label(canvas, text='1 room ' + str(def_booked_get_value[17]) + ' night',
                             font=('arial', 10), fg='black', bg='lightgray')
        nyts_payment1 = Label(canvas,
                              text='USD ' + str(def_booked_get_value[14]) + ' + ' + def_booked_get_value[13] + ' pts',
                              font=('arial', 10),
                              fg='black',
                              bg='lightgray')
        tax_fees = Label(canvas, text=str('taxes and fees '), font=('arial', 10), fg='black',
                         bg='lightgray')
        tax_fees1 = Label(canvas, text='USD ' + str(def_booked_get_value[15]), font=('arial', 10), fg='black',
                          bg='lightgray')
        tax_total = Label(canvas, text="tax and total payment", font=('arial', 10), fg='black',
                          bg='lightgray')
        tax_total1 = Label(canvas,
                           text='USD ' + str(def_booked_get_value[16]) + ' + ' + str(def_booked_get_value[13]) + ' pts',
                           font=('arial', 10),
                           fg='black', bg='lightgray')

    else:
        nyts_payment = Label(canvas, text='1 room ' + str(def_booked_get_value[17]) + ' night',
                             font=('arial', 10), fg='black', bg='lightgray')
        nyts_payment1 = Label(canvas, text='USD ' + str(def_booked_get_value[14]), font=('arial', 10), fg='black',
                              bg='lightgray')
        tax_fees = Label(canvas, text=str('taxes and fees '), font=('arial', 10), fg='black',
                         bg='lightgray')
        tax_fees1 = Label(canvas, text='USD ' + str(def_booked_get_value[15]), font=('arial', 10), fg='black',
                          bg='lightgray')
        tax_total = Label(canvas, text="tax and total payment", font=('arial', 10), fg='black',
                          bg='lightgray')
        tax_total1 = Label(canvas, text='USD ' + str(def_booked_get_value[16]), font=('arial', 10), fg='black',
                           bg='lightgray')

    space = Label(canvas, text='', font=('arial', 10), fg='black', bg='lightgray')

    canvas.create_window(23, 115, anchor="nw", window=confirm)
    canvas.create_window(23, 143, anchor="nw", window=hotels)
    canvas.create_window(23, 168, anchor="nw", window=dates)
    canvas.create_window(23, 193, anchor="nw", window=adults)
    canvas.create_window(23, 218, anchor="nw", window=fullname)
    canvas.create_window(23, 243, anchor="nw", window=rooms1)
    canvas.create_window(23, 268, anchor="nw", window=rate_rooms)
    canvas.create_window(23, 293, anchor="nw", window=room_payment)
    canvas.create_window(23, 348, anchor="nw", window=nyts_payment)
    canvas.create_window(190, 348, anchor="nw", window=nyts_payment1)
    canvas.create_window(23, 378, anchor="nw", window=tax_fees)
    canvas.create_window(190, 378, anchor="nw", window=tax_fees1)
    canvas.create_window(23, 408, anchor="nw", window=tax_total)
    canvas.create_window(190, 408, anchor="nw", window=tax_total1)

    '''stay_buttons = Button(canvas, text="Modify", font=("arial", 10), fg="black", bg="#F0F0F0",
                          activeforeground="green", activebackground="#F0F0F0", border=2, relief=RAISED, padx=6,
                          pady=2, width=7,
                          # wraplength= 1,
                          # state='disable',
                          command=site_scrools2)
    canvas.create_window(160, 470, anchor="nw", window=stay_buttons)
    canvas.create_window(10, 510, anchor="nw", window=space)'''

    '''cxl_buttons = Button(canvas, text="Cancel", font=("arial", 10), fg="black", bg="#F0F0F0",
                         activeforeground="green", activebackground="#F0F0F0", border=2, relief=RAISED, padx=6,
                         pady=2, width=7,
                         # wraplength= 1,
                         state='disable',
                         command=site_scrolls)
    canvas.create_window(200, 470, anchor="nw", window=cxl_buttons)
    canvas.create_window(10, 510, anchor="nw", window=space)'''

    my_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
    my_frame.place(x=870, y=176)
    resv_forget.append(my_frame)


def temp_booked():
    print('temp booked')
    nonyts = def_booked_get_value[17]
    if int(nonyts) == 1:
        night_message = 'night'
    else:
        night_message = 'nights'
    answer = messagebox.askquestion(title='Book a Reservation',
                                    message='I have you Confirmed in ' + def_booked_get_value[1] +
                                            ' on ' + str(book_rooms[0].get()) + ' ' + str(
                                        book_rooms[1].get()) + ', ' + str(book_rooms[2].get()) +
                                            ' for ' + str(nonyts) + ' ' + night_message)

    if answer == 'yes':
        num, num1 = 0, 0
        for x in def_booked_get_value:
            # pass
            # print('booked get value ',num ,x)
            num += 1
        # print('length of booked get value ', len(def_booked_get_value))
        # print('length of booked rooms ', len(book_rooms))
        for x in range(0, len(book_rooms)):
            # print('reservation temporary booked ',num1 ,book_rooms[x].get())
            checkinmonth1 = book_rooms[0].get()
            checkindays = book_rooms[1].get()
            checkinyear = book_rooms[2].get()
            checkoutmonth1 = book_rooms[3].get()
            checkoutdays = book_rooms[4].get()
            checkoutyear = book_rooms[5].get()
            bookadults = book_rooms[6].get()
            bookchild = book_rooms[7].get()
            bookrates = book_rooms[8].get()
            bookbeds = book_rooms[9].get()
            num1 += 1
        checkinmonth = datetime.datetime.strptime(checkinmonth1, '%B').month
        checkoutmonth = datetime.datetime.strptime(checkoutmonth1, '%B').month
        date1 = date(int(checkinyear), int(checkinmonth), int(checkindays))
        date2 = date(int(checkoutyear), int(checkoutmonth), int(checkoutdays))
        date1_in = str(date1.strftime('"%d/%m/%Y"'))
        date2_out = str(date2.strftime('"%d/%m/%Y"'))

        # print('Reservataion Number ', view_resvno[0].get())
        # print('Number of row data number4 ', data_number4.max_row)
        accountno = view_resvno[0].get()
        for data_search1 in range(2, data_number.max_row + 1):
            acctno = data_number.cell(data_search1, 1).value
            phone_data = data_number.cell(data_search1, 2).value
            lastname = data_number.cell(data_search1, 3).value
            firstname = data_number.cell(data_search1, 4).value
            middlename = data_number.cell(data_search1, 5).value
            home_data = data_number.cell(data_search1, 17).value
            if str(acctno) == str(accountno):
                savecell = phone_data
                booklast = lastname
                bookfirst = firstname
                bookmiddle = middlename
        # print('universal site no ', universal_siteno[0])
        siteno = universal_siteno[0]
        for data_search in range(2, data_number3.max_row + 1):
            no_site = data_number3.cell(data_search, 1).value
            homeadd = data_number3.cell(data_search, 3).value
            phone_hotel = data_number3.cell(data_search, 4).value
            if str(no_site) == str(siteno):
                newaddress = homeadd
                savecell1 = phone_hotel

        if book_rooms[9].get() == 'One Bed':
            bedsize = 'One King Bed'
        else:
            bedsize = 'Two Queen Beds'

        taxes_rate = def_booked_get_value[12]
        taxes_rate2 = ''
        for x in taxes_rate:
            if x == ',':
                pass
            else:
                taxes_rate2 += str(x)
        taxes_rate1 = conversion_numbers1(float(taxes_rate2) * float(.05))
        room_rate = def_booked_get_value[12]
        points = def_booked_get_value[11]
        total_room_rate = def_booked_get_value[14]
        total_taxes = def_booked_get_value[15]
        total_pts = def_booked_get_value[13]
        total_amount = def_booked_get_value[16]
        nonyts = def_booked_get_value[17]
        get_data_search = 0
        for data_search in range(2, data_number4.max_row + 1):
            resvno = data_number4.cell(data_search, 2).value
            if resvno == None:
                get_data_search = data_search
                break
        # print('get data search ', get_data_search)
        if get_data_search == 0:
            data_input = data_number4.max_row + 1
        else:
            data_input = get_data_search
        data_input1 = data_number2.max_row + 1
        # print('reservation number ', data_input1)
        data_number4.cell(data_input, 1).value, data_number2.cell(data_input1, 1).value = view_resvno[0].get(), \
            int(view_resvno[0].get())
        data_number4.cell(data_input, 2).value, data_number2.cell(data_input1,
                                                                  2).value = data_number2.max_row - 1, data_number2.max_row - 1
        data_number4.cell(data_input, 3).value, data_number2.cell(data_input1, 3).value = savecell, savecell
        data_number4.cell(data_input, 4).value, data_number2.cell(data_input1, 4).value = booklast, booklast
        data_number4.cell(data_input, 5).value, data_number2.cell(data_input1, 5).value = bookfirst, bookfirst
        data_number4.cell(data_input, 6).value, data_number2.cell(data_input1, 6).value = bookmiddle, bookmiddle
        data_number4.cell(data_input, 7).value, data_number2.cell(data_input1, 7).value = def_booked_get_value[1], \
            def_booked_get_value[1]
        data_number4.cell(data_input, 8).value, data_number2.cell(data_input1, 8).value = newaddress, newaddress
        data_number4.cell(data_input, 9).value, data_number2.cell(data_input1, 9).value = savecell1, savecell1
        data_number4.cell(data_input, 10).value, data_number2.cell(data_input1, 10).value = book_rooms[8].get(), \
            book_rooms[8].get()
        data_number4.cell(data_input, 11).value, data_number2.cell(data_input1,
                                                                   11).value = room_rate, conversion_numbers2(room_rate)
        data_number4.cell(data_input, 12).value, data_number2.cell(data_input1,
                                                                   12).value = taxes_rate1, conversion_numbers2(
            taxes_rate1)
        data_number4.cell(data_input, 13).value, data_number2.cell(data_input1, 13).value = points, conversion_numbers2(
            points)
        data_number4.cell(data_input, 14).value, data_number2.cell(data_input1,
                                                                   14).value = total_room_rate, conversion_numbers2(
            total_room_rate)
        data_number4.cell(data_input, 15).value, data_number2.cell(data_input1,
                                                                   15).value = total_taxes, conversion_numbers2(
            total_taxes)
        data_number4.cell(data_input, 16).value, data_number2.cell(data_input1,
                                                                   16).value = total_pts, conversion_numbers2(total_pts)
        data_number4.cell(data_input, 17).value, data_number2.cell(data_input1,
                                                                   17).value = total_amount, conversion_numbers2(
            total_amount)
        data_number4.cell(data_input, 18).value, data_number2.cell(data_input1, 18).value = nonyts, nonyts
        data_number4.cell(data_input, 19).value, data_number2.cell(data_input1, 19).value = bookadults, bookadults
        data_number4.cell(data_input, 20).value, data_number2.cell(data_input1, 20).value = bookchild, bookchild
        data_number4.cell(data_input, 21).value, data_number2.cell(data_input1, 21).value = bedsize, bedsize
        data_number4.cell(data_input, 22).value, data_number2.cell(data_input1, 22).value = def_booked_get_value[
            2], date1_in
        data_number4.cell(data_input, 23).value, data_number2.cell(data_input1, 23).value = def_booked_get_value[
            3], date2_out
        data_number4.cell(data_input, 24).value, data_number2.cell(data_input1, 24).value = def_booked_get_value[9], \
            def_booked_get_value[9]
        data_number4.cell(data_input, 25).value, data_number2.cell(data_input1, 25).value = 'no', 'no'
        data_number4.cell(data_input, 26).value, data_number2.cell(data_input1, 26).value = 'no', 'no'
        data_number4.cell(data_input, 27).value, data_number2.cell(data_input1, 27).value = 'not ready', 'not ready'
        data_number4.cell(data_input, 28).value, data_number2.cell(data_input1, 28).value = siteno, siteno
        data_number4.cell(data_input, 29).value, data_number2.cell(data_input1, 29).value = 'no', 'no'
        data_number4.cell(data_input, 30).value = 'confirmed'

        # print('Reservation Confirmed', data_number4.max_row)
        post()
        wb.save('data_wyn_gui.xlsx')
    else:
        pass
#end of reservation

window = Tk(className='vhinz reservation system')
#window.title('vhinz reservation system')
window.geometry("1400x690")
photo = PhotoImage(file='wyndham_logo.png')
photo1 = PhotoImage(file='calendar1.png')
square = PhotoImage(file='square.png')
vhinz = PhotoImage(file='melvin_photo.png')
#window['bg']='green'
#window.configure(bg='blue')

labelb = Label(window, background='black', padx=631, pady=33, bd=0, relief=SOLID)
labelb.place(x=9, y=0)

label = Label(window, background='blue', padx=628, pady=30)
label.place(x=10, y=1)

label1 = Label(window, image=photo, bg='white')
label1.place(x=10, y=1)

label4 = Label(window, background='#FFFFFF', padx=80, pady=30)
label4.place(x=170, y=1)

label2 = Label(window, text='E ', font=('Courier New', 15, 'italic'), fg='white', bg='blue')
label2.place(x=200, y=21)

label3 = Label(window, text='DESK', font=('arial', 15), fg='white', bg='blue')
label3.place(x=220, y=21)

label4 = Label(window, text='Member Information', font=('arial', 18, 'bold'), fg='white', bg='blue')
label4.place(x=440, y=20)

label5 = Label(window, text='Find Call', font=('arial', 11), fg='white', bg='blue')
label5.place(x=1100, y=28)

label7 = Label(window, image=square, fg='blue', bg='blue')
label7.place(x=400, y=25)

label8 = Label(window, image=square, fg='blue', bg='blue')
label8.place(x=390, y=34)

member_search1(0)
# member search define


label52 = Label(window, text='Member Search', font=('arial', 15), fg='darkgray')
label52.place(x=10, y=100)

label53 = Entry(window, font=('arial', 11))

# label53.insert(0,'Spongebob')
label53.place(x=10, y=150)

buttond50g = Button(window,text="Enroll Member",font=("arial", 11),fg="black", bg="lightgray",
                    padx=53, pady=3,width=6, command=add_member)
buttond50g.place(x=10, y=250)

labelc50h = Entry(window, font=('arial', 11), width=20)
labelc50h.place(x=10, y=310)

buttond50i = Button(window,
                    text="Stay Find",
                    # command=click,
                    font=("arial", 11),
                    fg="black",
                    bg="lightgray",
                    activeforeground="black",
                    activebackground="lightgray",
                    padx=53,
                    pady=3,
                    width=6,
                    state='disable',
                    command=stay_history)

buttond50i.place(x=10, y=340)

reservation_main = Button(window,
                    text="Reservation",
                    # command=click,
                    font=("arial", 11),
                    fg="black",
                    bg="lightgray",
                    activeforeground="black",
                    activebackground="lightgray",
                    padx=53,
                    pady=3,
                    width=6,
                    #state='disable',
                    command=reservation)
reservation_main.place(x=10, y=410)

labela50j = Label(window, text='Recent Contacts', font=('arial', 17), fg='darkgray')
labela50j.place(x=10, y=460)

# main search or page
label13 = Label(window, background='lightgray', padx=267, pady=225, border=1, relief=RIDGE)
label13.place(x=200, y=175)
# add_forget.append(label13)

label14 = Label(window, background='lightgray', padx=265, pady=205, border=1, relief=RIDGE)
label14.place(x=735, y=215)
# add_forget.append(label14)

# Personal Information
label15 = Label(window, text="Personal Information",
                font=('arial', 15),
                fg='blue',
                bg='lightgray',
                padx=5, pady=0)
label15.place(x=205, y=195)
# add_forget.append(label15)

label16a = Label(window, text="Title", font=('arial', 11),
                 bg='lightgray',
                 padx=0,
                 pady=10, )
label16a.place(x=210, y=240)

label16 = ttk.Combobox(window, font=('arial', 10), values=["Select", "Mr.", "Ms.", "Mrs."], width=10)
label16.place(x=250, y=250)

label17 = tkinter.Label(window, text="Gender", font=('arial', 11), bg='lightgray')
label17.place(x=500, y=250)

label17a = ttk.Combobox(window, font=('arial', 10), values=["Select", "Male", "Female"], width=10)
label17a.place(x=560, y=250)

label19 = Label(window, text="First Name", font=('arial', 11),
                bg='lightgray',
                padx=0,
                pady=10)
label19.place(x=210, y=280)
label20 = Entry(window, font=('arial', 11))
label20.place(x=210, y=315)

label21 = Label(window, text="Middle Name", font=('arial', 11),
                bg='lightgray',
                padx=0,
                pady=10)
label21.place(x=450, y=280)
label22 = Entry(window, font=('arial', 11))
label22.place(x=450, y=315)

label23 = Label(window, text="Last Name", font=('arial', 11),
                bg='lightgray',
                padx=0,
                pady=10)

label23.place(x=210, y=340)
label24 = Entry(window, font=('arial', 11), width=35)
# label24.insert(END,'last name')
label24.place(x=210, y=375)

label25 = Label(window, text="Version (Language) Preference*", font=('arial', 11),
                bg='lightgray',
                padx=0,
                pady=10)
label25.place(x=210, y=400)
label26 = ttk.Combobox(window, font=('arial', 10), values=["Select", "English", "Filipino"])
label26.place(x=210, y=435)

label27 = Label(window, text="Birthday", font=('arial', 11),
                bg='lightgray',
                padx=0,
                pady=10)
label27.place(x=450, y=400)

label28 = ttk.Combobox(window, font=('arial', 10), values=["                 Month   ", "January", "February",
                                                           "March", "April", "May", "June",
                                                           "July", "August", "September", "October",
                                                           "November", "December"])
label28.place(x=450, y=435)

label29 = ttk.Combobox(window, font=('arial', 10),
                       values=["   day ", 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15,
                               16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31],
                       width=5)
label29.place(x=625, y=435)

label30 = Label(window, text="Phone Number",
                font=('arial', 15),
                fg='blue',
                bg='lightgray',
                padx=5, pady=0)
label30.place(x=205, y=480)

label31 = Label(window, text="Select Primary Phone", font=('arial', 11),
                bg='lightgray',
                padx=0,
                pady=10)
label31.place(x=210, y=515)

label31a = ttk.Combobox(window, font=('arial', 11), values=["Select", "Phone", "Home"], width=10)
label31a.place(x=450, y=523)

label34 = Label(window, text="Phone Number", font=('arial', 11),
                bg='lightgray',
                padx=0,
                pady=10)
label34.place(x=210, y=550)
label35 = Entry(window, font=('arial', 11), width=20)
label35.place(x=210, y=585)

label36 = Label(window, text="Home Number", font=('arial', 11),
                bg='lightgray',
                padx=0,
                pady=10)
label36.place(x=450, y=550)
label37 = Entry(window, font=('arial', 11), width=20)
label37.place(x=450, y=585)

#home address
homeadd = Label(window, text="Home Address",font=('arial', 15),fg='blue',bg='#F0F0F0')
homeadd.place(x=745, y=185)

homenumb = Label(window, text="Number", font=('arial', 11),bg='lightgray')
homenumb.place(x=748, y=225)

entry_homenumb = Entry(window, font=('arial', 11), width=10)
entry_homenumb.place(x=750, y=250)

homest = Label(window, text="Street", font=('arial', 11),bg='lightgray')
homest.place(x=870, y=225)

entry_homest = Entry(window, font=('arial', 11), width=20)
entry_homest.place(x=870, y=250)

homebarangay = Label(window, text="Barangay", font=('arial', 11),bg='lightgray' )
homebarangay.place(x=1065, y=225)

entry_homebarangay = Entry(window, font=('arial', 11), width=20)
entry_homebarangay.place(x=1065, y=250)

homecity = Label(window, text="City", font=('arial', 11),bg='lightgray')
homecity.place(x=748, y=285)

entry_homecity = Entry(window, font=('arial', 11), width=25)
entry_homecity.place(x=750, y=310)

homeprov = Label(window, text="Province", font=('arial', 11),bg='lightgray')
homeprov.place(x=1030, y=285)

entry_homeprov = Entry(window, font=('arial', 11), width=25)
entry_homeprov.place(x=1030, y=310)

homecount = Label(window, text="Country", font=('arial', 11),bg='lightgray')
homecount.place(x=748, y=345)

entry_homecount = Entry(window, font=('arial', 11), width=25)
entry_homecount.place(x=750, y=370)

homezip = Label(window, text="Country Code", font=('arial', 11),bg='lightgray')
homezip.place(x=1030, y=345)

entry_homezip = Entry(window, font=('arial', 11), width=15)
entry_homezip.place(x=1030, y=370)

#membership status
'''label38 = Label(window, text="Membership Status",font=('arial', 15),fg='blue',bg='lightgray')
label38.place(x=745, y=435)'''

label39 = Label(window, text="Status", font=('arial', 15),fg='blue', bg='lightgray')
label39.place(x=747, y=477)

label40 = ttk.Combobox(window, font=('arial', 11), values=["Select ", "Active", "Inactive", "Closed"], width=10)
label40.place(x=825, y=481)

label41 = Label(window, text="Reactivate Date", font=('arial', 11),bg='lightgray')
label41.place(x=995, y=480)

label43 = Label(window, text='Deactivate Date', font=('arial', 11), bg='lightgray')
label43.place(x=750, y=520)

label45 = Label(window, text='Activity Date', font=('arial', 11), bg='lightgray')
label45.place(x=995, y=520)

label47 = Label(window, text='Expiration', font=('arial', 15), fg='blue', bg='lightgray')
label47.place(x=747, y=555)
label48 = Label(window, text='Points Balance', font=('arial', 11), bg='lightgray')
label48.place(x=995, y=560)
label49 = Label(window, text='Points Forfeit', font=('arial', 11), bg='lightgray')
label49.place(x=750, y=600)
label50 = Label(window, text='Points Inactive', font=('arial', 11), bg='lightgray')
label50.place(x=995, y=600)

'''label50a = Label(window, font=('arial', 11), bg='lightgray')
label50a.place(x=900, y=425)
label50b = Label(window, font=('arial', 11), bg='lightgray')
label50b.place(x=900, y=465)
label50c = Label(window, font=('arial', 11), bg='lightgray')
label50c.place(x=900, y=505)'''

email_label = Label(window, text='Email Address', font=('arial', 15), fg='blue', bg='lightgray')
email_label.place(x=745, y=415)
email_entry = Entry(window, font=('arial', 11), width=30)
email_entry.place(x=920, y=419)

main_button = Button(window,
                     text="Update",
                     # command=click,
                     font=("arial", 11),
                     fg="black",
                     bg="lightgray",
                     activeforeground="black",
                     activebackground="lightgray",
                     padx=30,
                     pady=3,
                     width= 10,
                     state='disable',
                     command=update)
main_button.place(x=1110, y=177)

button_search = Button(window,
                       text="Search",
                       # command=click,
                       font=("arial", 11),
                       fg="black",
                       bg="lightgray",
                       activeforeground="black",
                       activebackground="lightgray",
                       padx=53,
                       pady=3,
                       width=6,
                       command=member_search)
button_search.place(x=10, y=190)

valuex = 12
valuey = 495
for data_search in range(2, data_number.max_row + 1):
    member_no = str(data_number.cell(data_search, 1).value)
    first_name = data_number.cell(data_search, 4).value
    recent_contacts = data_number.cell(data_search, 20).value
    if data_search < 7:
        recent_contacts = Label(window, text=str(recent_contacts), font=('arial', 11, 'bold'), fg='darkgray')
        recent_contacts.place(x=valuex, y=valuey)
        valuey += 22

# exit program
exit_prog = Label(window, text='Logout', font=('arial', 11), fg='white', bg='blue')
exit_prog.place(x=1190, y=28)

exit_prog.bind("<Button-1>", exit_program)
exit_prog.bind("<Enter>", exit_program_yellow)
exit_prog.bind("<Leave>", exit_program_white)

#wb.save('data_wyn_gui.xlsx')
window.mainloop()